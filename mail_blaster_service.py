"""mail-blaster：素材提交子功能。

从独立的本地工具（~/mail-blaster，SQLite + 本地文件）移植过来，适配本项目：
  - SQLite  → Postgres（database.py 的连接池）
  - 本地文件 → BYTEA（Render 无持久盘，web/worker 也不共享文件系统）
  - 自建认证 → app.py 现有的 feature_required('mail_blaster')
  - 后台线程 → task_queue + worker.py（web 是 free 套餐，空闲 15 分钟休眠会把线程一起杀掉）

语义：一个收件人 ← N 个发件账号，各带 1 张图，发出 N 封信。
客户的归档流程要求「一素材一邮件」，所以是一条素材一封信，而不是一封信带 N 张图。

表全部用 mb_ 前缀，避免和主库那 30 多张表撞名。
"""

from __future__ import annotations

import base64
import csv
import hashlib
import io
import json
import logging
import mimetypes
import os
import re
import smtplib
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr, formatdate, make_msgid
from html import escape, unescape

from openpyxl import load_workbook
from PIL import Image
import psycopg2
from psycopg2.extras import execute_values

import crypto_util
import database as db
import name_utils

logger = logging.getLogger(__name__)

COOLDOWN_DAYS = 7
DEFAULT_DAILY_LIMIT = 10
SMTP_TIMEOUT = 25
IMAP_TIMEOUT = 30
LOCAL_SMTP_HOST = os.environ.get("MAIL_BLASTER_LOCAL_SMTP_HOST", "127.0.0.1")
LOCAL_SMTP_PORT = int(os.environ.get("MAIL_BLASTER_LOCAL_SMTP_PORT", "1025"))
LOCAL_SMTP_USE_TLS = os.environ.get("MAIL_BLASTER_LOCAL_SMTP_USE_TLS", "false").lower() == "true"
LOCAL_SMTP_USE_SSL = os.environ.get("MAIL_BLASTER_LOCAL_SMTP_USE_SSL", "false").lower() == "true"

# ---- 建联的两道时间护栏 ----
# 冷启动收件人互不相识，密集投递最容易触发频控和垃圾判定，
# 所以节奏比素材提交（3–8 秒）慢一个数量级。
OUTREACH_MIN_GAP_SECONDS = float(os.environ.get("MAIL_BLASTER_OUTREACH_MIN_GAP") or 20)
OUTREACH_MAX_GAP_SECONDS = float(os.environ.get("MAIL_BLASTER_OUTREACH_MAX_GAP") or 45)
SEND_WINDOW = os.environ.get("MAIL_BLASTER_SEND_WINDOW", "08:00-22:00")

# 配额和发送窗口都按这个时区算。
# 全项目其它地方（app.py 的 AT TIME ZONE、worker.py 的 utcnow()+8h、
# APScheduler 的 timezone）都是北京时间，只有 quota_state 原来用的
# date_trunc('day', NOW()) 跟着 Postgres 服务器时区走 —— Render 上是 UTC，
# 于是「今日配额」实际在北京时间早上 8 点重置、08:00-22:00 的窗口
# 实际是北京时间 16:00-06:00。两道护栏看起来在工作，却整整偏 8 小时。
SEND_TZ_NAME = os.environ.get("MAIL_BLASTER_TZ", "Asia/Shanghai")

MAX_IMAGE_BYTES = 1_500_000
MAX_IMAGE_EDGE = 1600
JPEG_QUALITY = 85

# 附件：整批共用一组，每封信原样带上（正文里的图不走这里，那是 mb_images 的内联图）。
# 限额按「编码前」算，但门槛是照编码后定的 —— MIME 的 base64 会把体积撑大 1/3，
# 15MB 的原始附件到收件方那边是 20MB 出头，刚好压在多数邮箱 25MB 的收信上限内。
MAX_ATTACHMENT_BYTES = 8 * 1024 * 1024
MAX_ATTACHMENT_TOTAL_BYTES = 15 * 1024 * 1024
MAX_ATTACHMENT_COUNT = 5
# 这些后缀绝大多数邮件网关直接拒信（整封退回，不是只丢附件），
# 与其让整批信在 SMTP 那步失败，不如在上传时就说清楚。
BLOCKED_ATTACHMENT_EXTS = {
    "exe", "com", "bat", "cmd", "scr", "pif", "msi", "msp", "cpl", "jar",
    "js", "jse", "vbs", "vbe", "wsf", "wsh", "ps1", "reg", "lnk", "hta", "dll",
}


# --------------------------------------------------------------------------- #
# 建表
# --------------------------------------------------------------------------- #

# 建好之后的完整表清单，用来做启动快速路径
_TABLES = {
    "mb_sender_accounts", "mb_images", "mb_jobs", "mb_items", "mb_templates",
    "mb_history", "mb_entity_phones", "mb_phone_pool", "mb_license_ocr",
    "mb_suppression",
    # 收信 / 会话线 / 报价账本
    "mb_imap_cursors", "mb_inbox_messages", "mb_inbox_extractions",
    "mb_threads", "mb_quotes",
    # 附件（建联整批共用 / 素材按行挂）
    "mb_attachments", "mb_job_attachments", "mb_item_attachments",
}
# 最后一批补上的列。表都在但列不全（老版本建的）时仍然要跑一遍 DDL。
# 新增列务必登记到这里，否则线上库走快速路径直接 return，DDL 一条都不跑，
# 故障现象是 worker 任务里的 UndefinedColumn 而不是启动报错，很难联想。
_LATEST_COLUMNS = {
    ("mb_jobs", "ocr_status"), ("mb_jobs", "replace_domain_enabled"),
    ("mb_jobs", "replacement_domain"), ("mb_sender_accounts", "auth_mode"),
    ("mb_jobs", "mode"), ("mb_jobs", "sender_account_id"),
    ("mb_templates", "mode"), ("mb_history", "mode"),
    ("mb_sender_accounts", "purpose"), ("mb_sender_accounts", "imap_host"),
}


def _schema_is_current() -> bool:
    """一次查询判断表和列是否都齐了。

    每条 DDL 都是一次到远程库的往返（Render 上约 1 秒），19 条就是 20 秒。
    free 实例休眠唤醒时要重跑这一整套，会把 5 秒的健康检查直接拖超时——
    web 服务因此挂过一次。所以先用一次查询走快速路径。
    """
    try:
        rows = db.query_all(
            "SELECT table_name, column_name, is_nullable FROM information_schema.columns "
            "WHERE table_schema = 'public' AND table_name = ANY(%s)",
            (sorted(_TABLES),))
    except Exception:
        return False        # 查不了就老老实实跑 DDL
    have_tables = {r["table_name"] for r in rows}
    have_columns = {(r["table_name"], r["column_name"]) for r in rows}
    if not (_TABLES <= have_tables and _LATEST_COLUMNS <= have_columns):
        return False
    # 约束变更 information_schema.columns 的存在性查不出来，得单独看一眼。
    # 漏了这条，线上库会一直走快速路径、永远跑不到 DROP NOT NULL 那句。
    nullable = next((r for r in rows if r["table_name"] == "mb_sender_accounts"
                     and r["column_name"] == "daily_limit"), None)
    return bool(nullable) and nullable.get("is_nullable") == "YES"


def ensure_schema() -> None:
    """幂等建表。app.py 模块级调用一次即可——worker.py 会 `from app import ...`，
    所以这段在 web 和 worker 两个进程里都会执行到。

    已经建好时走快速路径直接返回，只花一次查询。
    """
    if _schema_is_current():
        return

    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_sender_accounts (
            id SERIAL PRIMARY KEY,
            email VARCHAR(255) UNIQUE NOT NULL,
            display_name VARCHAR(255),
            signature_name VARCHAR(255),
            provider VARCHAR(64) NOT NULL DEFAULT 'custom',
            smtp_host VARCHAR(255) NOT NULL,
            smtp_port INTEGER NOT NULL,
            smtp_username VARCHAR(255),
            encrypted_password TEXT,
            use_ssl BOOLEAN NOT NULL DEFAULT FALSE,
            use_tls BOOLEAN NOT NULL DEFAULT TRUE,
            enabled BOOLEAN NOT NULL DEFAULT TRUE,
            sort_order INTEGER NOT NULL DEFAULT 0,
            daily_limit INTEGER,        -- NULL = 不限量，0 = 今天停发
            status VARCHAR(20) NOT NULL DEFAULT 'draft',
            last_test_at TIMESTAMP,
            last_error TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 图片存库不存盘：Render 无持久磁盘，且 web / worker 是两个独立容器，
    # web 写到本地的图 worker 根本读不到。按 sha256 去重，同一张图不重复占空间。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_images (
            id SERIAL PRIMARY KEY,
            sha256 CHAR(64) UNIQUE NOT NULL,
            content BYTEA NOT NULL,
            mime VARCHAR(64) NOT NULL DEFAULT 'image/png',
            byte_size INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 附件同理存库不存盘。文件名不进这张表：同一份 PDF 换个名字上传两次
    # 仍然只该占一份空间，名字是「这一批怎么叫它」，挂在 mb_job_attachments 上。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_attachments (
            id SERIAL PRIMARY KEY,
            sha256 CHAR(64) UNIQUE NOT NULL,
            content BYTEA NOT NULL,
            mime VARCHAR(128) NOT NULL DEFAULT 'application/octet-stream',
            byte_size INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_jobs (
            id SERIAL PRIMARY KEY,
            user_id INTEGER,
            recipient VARCHAR(255) NOT NULL DEFAULT '',
            subject_tpl TEXT NOT NULL DEFAULT '',
            body_tpl TEXT NOT NULL DEFAULT '',
            signature_tpl TEXT NOT NULL DEFAULT '',
            status VARCHAR(20) NOT NULL DEFAULT 'draft',
            paused_reason TEXT,
            task_id VARCHAR(100),
            created_at TIMESTAMP DEFAULT NOW(),
            finished_at TIMESTAMP
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_items (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL REFERENCES mb_jobs(id) ON DELETE CASCADE,
            seq INTEGER NOT NULL DEFAULT 0,
            sender_account_id INTEGER REFERENCES mb_sender_accounts(id) ON DELETE SET NULL,
            recipient VARCHAR(255),
            image_id INTEGER REFERENCES mb_images(id) ON DELETE SET NULL,
            image_name VARCHAR(255),
            vars_json TEXT,
            from_display VARCHAR(255),
            signature_name VARCHAR(255),
            subject TEXT,
            body_html TEXT,
            status VARCHAR(20) NOT NULL DEFAULT 'pending',
            error TEXT,
            smtp_response TEXT,
            message_id TEXT,
            sent_at TIMESTAMP
        )
    """)
    # 建联的附件挂在批次上，整批共用一组：建联是「同一份物料群发给一串达人」，
    # 按行挂会逼着名单 Excel 多一列文件名，而粘贴名单那条入口根本带不了文件。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_job_attachments (
            id SERIAL PRIMARY KEY,
            job_id INTEGER NOT NULL REFERENCES mb_jobs(id) ON DELETE CASCADE,
            attachment_id INTEGER NOT NULL REFERENCES mb_attachments(id) ON DELETE CASCADE,
            filename VARCHAR(255) NOT NULL DEFAULT 'attachment',
            seq INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_mb_job_attachments_job "
               "ON mb_job_attachments (job_id, seq)")
    # 素材提交反过来，附件按行挂：这里附件是「素材的另一种形态」——
    # 有些素材是图（内联进正文），有些本来就是 PDF/压缩包（只能当附件走）。
    # 一行一份素材，所以粒度必须是行，不能像建联那样整批共用。
    # 刻意不加 UNIQUE(item_id, attachment_id)：mb_attachments 按 sha256 去重，
    # 同样字节换个文件名会指向同一个 attachment_id，那是两条合法的附件行。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_item_attachments (
            id SERIAL PRIMARY KEY,
            item_id INTEGER NOT NULL REFERENCES mb_items(id) ON DELETE CASCADE,
            attachment_id INTEGER NOT NULL REFERENCES mb_attachments(id) ON DELETE CASCADE,
            filename VARCHAR(255) NOT NULL DEFAULT 'attachment',
            seq INTEGER NOT NULL DEFAULT 0,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    db.execute("CREATE INDEX IF NOT EXISTS idx_mb_item_attachments_item "
               "ON mb_item_attachments (item_id, seq)")
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_templates (
            id SERIAL PRIMARY KEY,
            name VARCHAR(255) NOT NULL,
            subject TEXT NOT NULL DEFAULT '',
            body_html TEXT NOT NULL DEFAULT '',
            signature_html TEXT NOT NULL DEFAULT '',
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 跨批次留存：既是发信账，也是「这条素材发过没」和「这个账号还在冷却期没」的依据
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_history (
            id SERIAL PRIMARY KEY,
            recipient VARCHAR(255) NOT NULL,
            material_id VARCHAR(255) NOT NULL DEFAULT '',
            material_name VARCHAR(255) NOT NULL DEFAULT '',
            sender_account_id INTEGER,
            sender_email VARCHAR(255) NOT NULL DEFAULT '',
            job_id INTEGER,
            item_id INTEGER,
            subject TEXT NOT NULL DEFAULT '',
            message_id TEXT NOT NULL DEFAULT '',
            sent_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 建联专用：永不投递的地址。退订请求、硬退信、手动拉黑都进这里。
    # reason 允许空串（手动拉黑常常懒得填），所以判定函数返回 None 才表示「不在名单里」。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_suppression (
            id SERIAL PRIMARY KEY,
            email VARCHAR(255) UNIQUE NOT NULL,
            reason TEXT NOT NULL DEFAULT '',
            source VARCHAR(16) NOT NULL DEFAULT 'manual',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # IMAP 增量拉取游标。uidvalidity 变了说明服务端重建了 UID 空间，要从头回填。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_imap_cursors (
            account_id INTEGER NOT NULL REFERENCES mb_sender_accounts(id) ON DELETE CASCADE,
            folder VARCHAR(64) NOT NULL DEFAULT 'INBOX',
            uidvalidity BIGINT,
            last_uid BIGINT NOT NULL DEFAULT 0,
            last_sync_at TIMESTAMP,
            last_error TEXT,
            PRIMARY KEY (account_id, folder)
        )
    """)
    # 收到的信。dedupe_key 唯一：同一封被重复拉取时靠它幂等
    # （有 Message-ID 就用它，没有则退回 账号:folder:uid）。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_inbox_messages (
            id SERIAL PRIMARY KEY,
            account_id INTEGER REFERENCES mb_sender_accounts(id) ON DELETE SET NULL,
            folder VARCHAR(64) NOT NULL DEFAULT 'INBOX',
            uid BIGINT,
            dedupe_key VARCHAR(512) UNIQUE NOT NULL,
            message_id TEXT,
            in_reply_to TEXT,
            refs TEXT,
            from_email VARCHAR(255) NOT NULL DEFAULT '',
            from_name VARCHAR(255) NOT NULL DEFAULT '',
            to_email VARCHAR(255) NOT NULL DEFAULT '',
            subject TEXT NOT NULL DEFAULT '',
            body_text TEXT NOT NULL DEFAULT '',
            body_html TEXT NOT NULL DEFAULT '',
            received_at TIMESTAMP,
            matched_item_id INTEGER REFERENCES mb_items(id) ON DELETE SET NULL,
            thread_id INTEGER,
            match_method VARCHAR(20),
            match_confidence REAL NOT NULL DEFAULT 0,
            kind VARCHAR(20) NOT NULL DEFAULT 'reply',
            handled BOOLEAN NOT NULL DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # AI 抽取结果：既拆列（筛选排序）又留 raw_json（改 schema 时能重放）
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_inbox_extractions (
            id SERIAL PRIMARY KEY,
            inbox_id INTEGER NOT NULL UNIQUE
                REFERENCES mb_inbox_messages(id) ON DELETE CASCADE,
            intent VARCHAR(24),
            quote_amount NUMERIC(14,2),
            quote_currency VARCHAR(8),
            contacts_json TEXT,
            needs_human BOOLEAN NOT NULL DEFAULT FALSE,
            needs_human_reason TEXT,
            summary TEXT,
            reply_draft TEXT,
            raw_json TEXT,
            model VARCHAR(64),
            status VARCHAR(16) NOT NULL DEFAULT 'pending',
            error TEXT,
            tokens INTEGER NOT NULL DEFAULT 0,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 每个 KOL 一条会话线。状态机：pending → replied → negotiating → won / lost
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_threads (
            id SERIAL PRIMARY KEY,
            kol_email VARCHAR(255) UNIQUE NOT NULL,
            kol_name VARCHAR(255) NOT NULL DEFAULT '',
            account_id INTEGER REFERENCES mb_sender_accounts(id) ON DELETE SET NULL,
            job_id INTEGER REFERENCES mb_jobs(id) ON DELETE SET NULL,
            item_id INTEGER REFERENCES mb_items(id) ON DELETE SET NULL,
            status VARCHAR(20) NOT NULL DEFAULT 'pending',
            last_intent VARCHAR(24),
            next_action TEXT,
            vars_json TEXT,
            first_sent_at TIMESTAMP,
            last_reply_at TIMESTAMP,
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 报价账本：append-only，改价是追加新版本而不是改历史。
    # 议价轮次由 status='countered' 的行数推导，不单独存字段——存了就会漂。
    # source_inbox_id 让每个价格都能追溯到具体是哪封回信说的。
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_quotes (
            id SERIAL PRIMARY KEY,
            thread_id INTEGER NOT NULL REFERENCES mb_threads(id) ON DELETE CASCADE,
            version INTEGER NOT NULL,
            status VARCHAR(16) NOT NULL DEFAULT 'proposed',
            amount NUMERIC(14,2),
            currency VARCHAR(8) NOT NULL DEFAULT 'USD',
            source_inbox_id INTEGER REFERENCES mb_inbox_messages(id) ON DELETE SET NULL,
            note TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (thread_id, version)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_entity_phones (
            entity_key VARCHAR(255) PRIMARY KEY,
            company_name VARCHAR(512) NOT NULL DEFAULT '',
            country_code VARCHAR(4) NOT NULL DEFAULT '',
            phone VARCHAR(64) NOT NULL DEFAULT '',
            source VARCHAR(16) NOT NULL DEFAULT 'pool',
            assigned_at TIMESTAMP DEFAULT NOW()
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_phone_pool (
            id SERIAL PRIMARY KEY,
            country_code VARCHAR(4) NOT NULL,
            phone VARCHAR(64) NOT NULL,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (country_code, phone)
        )
    """)
    db.execute("""
        CREATE TABLE IF NOT EXISTS mb_license_ocr (
            id SERIAL PRIMARY KEY,
            image_sha256 CHAR(64) UNIQUE NOT NULL,
            company_name VARCHAR(512) NOT NULL DEFAULT '',
            registration_id VARCHAR(255) NOT NULL DEFAULT '',
            registered_address TEXT NOT NULL DEFAULT '',
            country_code VARCHAR(4) NOT NULL DEFAULT '',
            raw_json TEXT NOT NULL DEFAULT '',
            model VARCHAR(64) NOT NULL DEFAULT '',
            status VARCHAR(16) NOT NULL DEFAULT 'done',
            error TEXT,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)
    # 老库补列：表建好之后新增的字段走这里。Postgres 支持 ADD COLUMN IF NOT EXISTS，
    # 反复执行是幂等的。
    for stmt in (
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS ocr_status VARCHAR(20) NOT NULL DEFAULT 'none'",
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS ocr_report TEXT",
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS ocr_task_id VARCHAR(100)",
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS replace_domain_enabled BOOLEAN NOT NULL DEFAULT FALSE",
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS replacement_domain VARCHAR(255) NOT NULL DEFAULT ''",
        # OAuth2（XOAUTH2）：微软已对部分账号禁用密码直连 SMTP，只能走 OAuth
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS auth_mode VARCHAR(16) NOT NULL DEFAULT 'password'",
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS encrypted_client_id TEXT",
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS encrypted_refresh_token TEXT",
        # KOL 建联：素材提交是「一个收件人 ← N 个发件账号」，建联是「一个发件账号 → N 个收件人」，
        # 轴反过来了。mode 区分两者，sender_account_id 存建联整批共用的那个账号。
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS mode VARCHAR(16) NOT NULL DEFAULT 'material'",
        "ALTER TABLE mb_jobs ADD COLUMN IF NOT EXISTS sender_account_id INTEGER "
        "REFERENCES mb_sender_accounts(id) ON DELETE SET NULL",
        "ALTER TABLE mb_templates ADD COLUMN IF NOT EXISTS mode VARCHAR(16) NOT NULL DEFAULT 'material'",
        "ALTER TABLE mb_history ADD COLUMN IF NOT EXISTS mode VARCHAR(16) NOT NULL DEFAULT 'material'",
        # 模板唯一键 (name) → (mode, name)：两套模板库各自独立命名。
        # 不换的话建联存一个叫「默认」的模板会顶掉素材提交的同名模板。
        "ALTER TABLE mb_templates DROP CONSTRAINT IF EXISTS mb_templates_name_key",
        "CREATE UNIQUE INDEX IF NOT EXISTS uq_mb_templates_mode_name ON mb_templates(mode, name)",
        # 账号用途：素材提交和 KOL 建联用的是两套完全不同的邮箱
        # （素材是 Outlook OAuth2，建联是阿里云企业邮箱），混在一个下拉里很容易选错，
        # 而且选错的后果是建联信从收不了回信的号发出去。默认 both 保持老账号行为不变。
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS "
        "purpose VARCHAR(16) NOT NULL DEFAULT 'both'",
        # 收信（建联要靠它收回信）。空 imap_host 表示这个号只发不收。
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS imap_host VARCHAR(255)",
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS imap_port INTEGER",
        "ALTER TABLE mb_sender_accounts ADD COLUMN IF NOT EXISTS "
        "imap_ssl BOOLEAN NOT NULL DEFAULT TRUE",
        # daily_limit 允许为 NULL = 不限量。阿里云企业邮这类自有企业邮箱
        # 不需要我们再兜一个日发信上限，原来的 NOT NULL DEFAULT 10 会让
        # 每批建联发到第 10 封就暂停。0 仍然表示「今天停发」。
        "ALTER TABLE mb_sender_accounts ALTER COLUMN daily_limit DROP NOT NULL",
        "ALTER TABLE mb_sender_accounts ALTER COLUMN daily_limit DROP DEFAULT",
    ):
        db.execute(stmt)

    for stmt in (
        "CREATE INDEX IF NOT EXISTS idx_mb_items_job ON mb_items(job_id, seq)",
        "CREATE INDEX IF NOT EXISTS idx_mb_history_dedupe ON mb_history(recipient, material_id)",
        "CREATE INDEX IF NOT EXISTS idx_mb_history_sender ON mb_history(recipient, sender_account_id, sent_at)",
        "CREATE INDEX IF NOT EXISTS idx_mb_history_sent_at ON mb_history(sent_at DESC)",
        # quota_state 每封信查一次，条件是 (sender_account_id, sent_at)。
        # idx_mb_history_sender 首列是 recipient，服务不了这个查询。
        "CREATE INDEX IF NOT EXISTS idx_mb_history_quota ON mb_history(sender_account_id, sent_at)",
        # 回信匹配靠 In-Reply-To → mb_items.message_id，这是唯一强信号
        "CREATE INDEX IF NOT EXISTS idx_mb_items_msgid ON mb_items(message_id)",
        "CREATE INDEX IF NOT EXISTS idx_mb_inbox_thread ON mb_inbox_messages(thread_id, received_at)",
        "CREATE INDEX IF NOT EXISTS idx_mb_inbox_unhandled "
        "ON mb_inbox_messages(handled, received_at DESC)",
        "CREATE INDEX IF NOT EXISTS idx_mb_quotes_thread ON mb_quotes(thread_id, version)",
        "CREATE INDEX IF NOT EXISTS idx_mb_threads_status ON mb_threads(status, updated_at DESC)",
    ):
        db.execute(stmt)


# --------------------------------------------------------------------------- #
# 发件账号池
# --------------------------------------------------------------------------- #

PROVIDERS = [
    {"key": "aliyun_qiye", "label": "阿里云企业邮箱", "smtp_host": "smtp.qiye.aliyun.com",
     "smtp_port": 465, "use_ssl": True, "use_tls": False, "domains": [],
     "imap_host": "imap.qiye.aliyun.com", "imap_port": 993, "imap_ssl": True,
     "note": "企业自有域名，无法按域名自动识别，需手动选。密码填邮箱登录密码；"
             "开了安全设置就填客户端专用密码。收信要在管理后台确认 IMAP 服务已开启。"},
    {"key": "tencent_exmail", "label": "腾讯企业邮箱", "smtp_host": "smtp.exmail.qq.com",
     "smtp_port": 465, "use_ssl": True, "use_tls": False, "domains": [],
     "imap_host": "imap.exmail.qq.com", "imap_port": 993, "imap_ssl": True,
     "note": "需要在「安全设置」里开启客户端专用密码，密码填那个而不是登录密码。"},
    {"key": "feishu", "label": "飞书邮箱", "smtp_host": "smtp.feishu.cn",
     "smtp_port": 465, "use_ssl": True, "use_tls": False, "domains": [],
     "imap_host": "imap.feishu.cn", "imap_port": 993, "imap_ssl": True,
     "note": "在飞书邮箱设置里生成客户端专用密码。"},
    {"key": "outlook", "label": "Outlook / Hotmail", "smtp_host": "smtp-mail.outlook.com",
     "smtp_port": 587, "use_ssl": False, "use_tls": True,
     "domains": ["outlook.com", "hotmail.com", "live.com", "msn.com"],
     "imap_host": "outlook.office365.com", "imap_port": 993, "imap_ssl": True,
     "note": "微软正在收紧密码直连。报 535 5.7.139 说明该账号已被禁用密码登录，只能改走 OAuth2。"
             "注意 OAuth2 账号收信需要 IMAP.AccessAsUser.All 权限，"
             "现有授权只申请了 SMTP.Send，收不了信。"},
    {"key": "gmail", "label": "Gmail", "smtp_host": "smtp.gmail.com",
     "smtp_port": 587, "use_ssl": False, "use_tls": True,
     "domains": ["gmail.com", "googlemail.com"],
     "imap_host": "imap.gmail.com", "imap_port": 993, "imap_ssl": True,
     "note": "Google 已永久关闭「登录密码直连 SMTP」，报 535 5.7.8 就是这个原因。"
             "必须先开两步验证，再去 myaccount.google.com/apppasswords 生成 16 位应用专用密码（填的时候去掉空格）。"},
    {"key": "qq", "label": "QQ 邮箱", "smtp_host": "smtp.qq.com",
     "smtp_port": 465, "use_ssl": True, "use_tls": False, "domains": ["qq.com", "foxmail.com"],
     "imap_host": "imap.qq.com", "imap_port": 993, "imap_ssl": True,
     "note": "在设置-账户里开启 SMTP 并生成授权码，密码填授权码而不是 QQ 密码。"},
    {"key": "163", "label": "网易 163", "smtp_host": "smtp.163.com",
     "smtp_port": 465, "use_ssl": True, "use_tls": False, "domains": ["163.com"],
     "imap_host": "imap.163.com", "imap_port": 993, "imap_ssl": True,
     "note": "在网页版设置里开启 SMTP 服务并获取授权码，密码填授权码。"},
    {"key": "custom", "label": "自定义", "smtp_host": "", "smtp_port": 587,
     "use_ssl": False, "use_tls": True, "domains": [],
     "imap_host": "", "imap_port": 993, "imap_ssl": True,
     "note": "自己填 SMTP 服务器和端口。465 一般配 SSL，587 一般配 STARTTLS。"},
]
_BY_KEY = {p["key"]: p for p in PROVIDERS}
_BY_DOMAIN = {d: p for p in PROVIDERS for d in p["domains"]}

# 允许「Header From 和认证账号不一致」的服务商，也就是能用来做域名替换的。
# 微软/Exchange 会直接拒（5.2.252 SendAsDenied，见 friendly_smtp_error），
# Gmail 同理；实测网易 163 放行。勾了替换域名的批次只从这里面挑号——
# 否则是整批发出去、一封封失败之后才发现这个池子里的号根本不能替换。
DOMAIN_REPLACEMENT_PROVIDERS = {"163"}

# 批量粘贴录入时的默认每日上限（单个添加的表单里留空仍然表示不限量）。
# 免费邮箱的日发信量只有几十封，跑满会直接触发风控封号，
# 而批量粘贴那条路一个数字都没填，落到「不限量」上最危险。
BULK_IMPORT_DAILY_LIMIT = {"163": 30, "qq": 30, "gmail": 50, "outlook": 50}


def guess_provider(email: str) -> dict:
    domain = email.rsplit("@", 1)[-1].strip().lower() if "@" in email else ""
    return _BY_DOMAIN.get(domain, _BY_KEY["custom"])


def _titleize(word: str) -> str:
    if word.islower() or word.isupper():
        return word.capitalize()
    return word[:1].upper() + word[1:]


def name_from_email(email: str) -> str:
    """amy.chen01@x.com -> Amy Chen；amychen01@x.com 也认

    数字当分隔符（amy01chen），连写的没有分隔符可依，交给 name_utils 靠词表猜边界；
    猜不动就整段留着（Amychen），不硬拆。
    """
    local = (email or "").split("@", 1)[0].strip()
    if not local:
        return email or ""
    cleaned = re.sub(r"\d+$", "", local) or local
    parts = [p for p in re.split(r"[._\-+\d]+", cleaned) if p] or [cleaned]
    if len(parts) == 1:
        parts = list(name_utils.split_name_token(parts[0]))
    return " ".join(_titleize(p) for p in parts)


def derive_names(account: dict) -> tuple[str, str]:
    display = (account.get("display_name") or "").strip() or name_from_email(account.get("email") or "")
    signature = (account.get("signature_name") or "").strip() or display
    return display, signature


def serialize_account(row: dict) -> dict:
    """给前端：绝不含密码明文或密文。"""
    display, signature = derive_names(row)
    return {
        "id": row["id"], "email": row["email"],
        "display_name": row.get("display_name") or "",
        "signature_name": row.get("signature_name") or "",
        "effective_display_name": display,
        "effective_signature_name": signature,
        "provider": row.get("provider") or "custom",
        "smtp_host": row["smtp_host"], "smtp_port": row["smtp_port"],
        "smtp_username": row.get("smtp_username") or "",
        "use_ssl": bool(row["use_ssl"]), "use_tls": bool(row["use_tls"]),
        "enabled": bool(row["enabled"]), "sort_order": row.get("sort_order") or 0,
        # None = 不限量，0 = 今天停发，别用 or 把两者都吃掉
        "daily_limit": (None if row.get("daily_limit") is None else int(row["daily_limit"])),
        "auth_mode": row.get("auth_mode") or "password",
        "purpose": row.get("purpose") or "both",
        "imap_host": row.get("imap_host") or "",
        "imap_port": row.get("imap_port"),
        "imap_ssl": bool(row.get("imap_ssl", True)),
        # 能不能收信：要有 IMAP 地址，且不是只申请了 SMTP.Send 的 OAuth2 号
        "can_receive": bool(row.get("imap_host"))
                       and (row.get("auth_mode") or "password") == "password",
        "has_client_id": bool(row.get("encrypted_client_id")),
        "has_refresh_token": bool(row.get("encrypted_refresh_token")),
        "status": row.get("status") or "draft",
        "last_test_at": row["last_test_at"].isoformat() if row.get("last_test_at") else None,
        "last_error": row.get("last_error"),
        "has_password": bool(row.get("encrypted_password")),
    }


def usable_account(acc: dict) -> bool:
    """能不能拿来发信。和前端 mail_blaster_common.js 里的 usable() 是同一条规则，
    改一处要同步改另一处——前端只是为了少一次往返，服务端这份才是准的。"""
    if not acc.get("enabled") or acc.get("status") != "ready":
        return False
    if (acc.get("auth_mode") or "password") == "xoauth2":
        return bool(acc.get("has_client_id") and acc.get("has_refresh_token"))
    return bool(acc.get("has_password"))


def list_accounts(only_sendable: bool = False, purpose: str = "") -> list[dict]:
    sql = "SELECT * FROM mb_sender_accounts WHERE 1=1"
    args: list = []
    if only_sendable:
        # 密码模式要有密码，OAuth 模式要有 refresh_token
        sql += (" AND enabled = TRUE AND status = 'ready'"
                " AND (encrypted_password IS NOT NULL OR encrypted_refresh_token IS NOT NULL)")
    if purpose in ("material", "outreach"):
        # both 是通用号，两边都列
        sql += " AND purpose IN (%s, 'both')"
        args.append(purpose)
    sql += " ORDER BY sort_order ASC, id ASC"
    return [serialize_account(dict(r)) for r in db.query_all(sql, tuple(args))]


def get_account(account_id: int) -> dict | None:
    row = db.query_one("SELECT * FROM mb_sender_accounts WHERE id = %s", (account_id,))
    return dict(row) if row else None


def _normalize_account(payload: dict) -> dict:
    email = (payload.get("email") or "").strip().lower()
    if "@" not in email:
        raise ValueError(f"邮箱格式不正确：{email or '(空)'}")
    key = (payload.get("provider") or "").strip() or guess_provider(email)["key"]
    preset = _BY_KEY.get(key, _BY_KEY["custom"])
    host = (payload.get("smtp_host") or "").strip() or preset["smtp_host"]
    if not host:
        raise ValueError(f"{email}：没有 SMTP 服务器地址，请手动填写")
    try:
        port = int(payload.get("smtp_port") or preset["smtp_port"])
    except (TypeError, ValueError):
        raise ValueError(f"{email}：SMTP 端口不是数字") from None
    # SSL / STARTTLS 各自独立判断。以前是「两个 key 都不在才取预设」，
    # 于是表单只传一个的时候，另一个会被静默当成 False。
    use_ssl = bool(payload["use_ssl"]) if "use_ssl" in payload else preset["use_ssl"]
    use_tls = bool(payload["use_tls"]) if "use_tls" in payload else preset["use_tls"]
    # 带了 refresh_token 就走 OAuth2，否则密码直连
    auth_mode = (payload.get("auth_mode") or "").strip()
    if not auth_mode:
        has_token = bool((payload.get("refresh_token") or "").strip()) or bool(
            payload.get("encrypted_refresh_token"))
        auth_mode = "xoauth2" if has_token else "password"

    # 用 is None 判断而不是 or：0 是有意义的值。
    # daily_limit=0 表示「这个账号今天一封都别发」，用 or 会被静默改成默认的 10。
    # 留空（None）表示**不限量**——阿里云企业邮这类自有企业邮箱不需要我们再兜一个
    # 日发信上限，硬套默认的 10 会让每批建联发到第 10 封就暂停。
    limit_raw = payload.get("daily_limit")
    order_raw = payload.get("sort_order")
    try:
        daily_limit = None if limit_raw is None or limit_raw == "" else int(limit_raw)
        sort_order = 0 if order_raw is None or order_raw == "" else int(order_raw)
    except (TypeError, ValueError):
        raise ValueError(f"{email}：每日上限和排序必须是数字") from None
    if daily_limit is not None and daily_limit < 0:
        raise ValueError(f"{email}：每日上限不能是负数（留空=不限量，0=今天停发）")

    purpose = (payload.get("purpose") or "").strip()
    if purpose not in ("material", "outreach", "both"):
        purpose = "both"

    # 收信配置。imap_host 留空表示这个号只发不收。
    imap_host = (payload.get("imap_host") or "").strip()
    if "imap_host" not in payload and preset.get("imap_host"):
        imap_host = preset["imap_host"]
    if imap_host:
        try:
            imap_port = int(payload.get("imap_port") or preset.get("imap_port") or 993)
        except (TypeError, ValueError):
            raise ValueError(f"{email}：IMAP 端口不是数字") from None
        imap_ssl = (bool(payload["imap_ssl"]) if "imap_ssl" in payload
                    else bool(preset.get("imap_ssl", True)))
    else:
        imap_port, imap_ssl = None, True

    return {
        "email": email, "provider": key, "smtp_host": host, "smtp_port": port,
        "auth_mode": auth_mode, "purpose": purpose,
        "display_name": (payload.get("display_name") or "").strip() or None,
        "signature_name": (payload.get("signature_name") or "").strip() or None,
        "smtp_username": (payload.get("smtp_username") or "").strip() or None,
        "use_ssl": use_ssl, "use_tls": use_tls,
        "imap_host": imap_host or None, "imap_port": imap_port, "imap_ssl": imap_ssl,
        "enabled": bool(payload.get("enabled", True)),
        "sort_order": sort_order,
        "daily_limit": daily_limit,
    }


def create_account(payload: dict) -> dict:
    data = _normalize_account(payload)
    if db.query_one("SELECT id FROM mb_sender_accounts WHERE email = %s", (data["email"],)):
        raise ValueError(f"账号已存在：{data['email']}")
    secrets = {
        "pwd": crypto_util.encrypt(payload["app_password"].strip())
               if (payload.get("app_password") or "").strip() else None,
        "cid": crypto_util.encrypt(payload["client_id"].strip())
               if (payload.get("client_id") or "").strip() else None,
        "rtok": crypto_util.encrypt(payload["refresh_token"].strip())
                if (payload.get("refresh_token") or "").strip() else None,
    }
    new_id = db.execute_and_fetch_id("""
        INSERT INTO mb_sender_accounts
            (email, display_name, signature_name, provider, smtp_host, smtp_port,
             smtp_username, use_ssl, use_tls, enabled, sort_order, daily_limit, auth_mode,
             purpose, imap_host, imap_port, imap_ssl,
             encrypted_password, encrypted_client_id, encrypted_refresh_token)
        VALUES (%(email)s, %(display_name)s, %(signature_name)s, %(provider)s, %(smtp_host)s,
                %(smtp_port)s, %(smtp_username)s, %(use_ssl)s, %(use_tls)s, %(enabled)s,
                %(sort_order)s, %(daily_limit)s, %(auth_mode)s,
                %(purpose)s, %(imap_host)s, %(imap_port)s, %(imap_ssl)s,
                %(pwd)s, %(cid)s, %(rtok)s)
        RETURNING id
    """, {**data, **secrets})
    return serialize_account(get_account(new_id))


def update_account(account_id: int, payload: dict) -> dict:
    current = get_account(account_id)
    if current is None:
        raise ValueError(f"账号 {account_id} 不存在")
    merged = {**current, **{k: v for k, v in payload.items() if v is not None}}
    for k in ("display_name", "signature_name", "smtp_username"):
        if k in payload:
            merged[k] = payload[k]
    data = _normalize_account(merged)

    sets = ", ".join(f"{k} = %({k})s" for k in data)
    args = dict(data)
    changed = False
    for field, column, key in (("app_password", "encrypted_password", "pwd"),
                               ("client_id", "encrypted_client_id", "cid"),
                               ("refresh_token", "encrypted_refresh_token", "rtok")):
        value = (payload.get(field) or "").strip()
        if value:
            sets += f", {column} = %({key})s"
            args[key] = crypto_util.encrypt(value)
            changed = True
    if changed:
        # 换了凭据就要重测，不能继续顶着 ready 的状态
        sets += ", status = 'draft', last_error = NULL"
    args["aid"] = account_id
    db.execute(f"UPDATE mb_sender_accounts SET {sets} WHERE id = %(aid)s", args)
    return serialize_account(get_account(account_id))


def delete_account(account_id: int) -> None:
    db.execute("DELETE FROM mb_sender_accounts WHERE id = %s", (account_id,))


_CLIENT_ID_RE = re.compile(
    r"^(?:[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}"
    r"|[\w.\-]+\.apps\.googleusercontent\.com)$")

# 「一串没有空格的字母数字混排」——授权码长这样，人名不长这样。
# 用来认出 邮箱----登录密码----授权码 这种三段格式（网易/QQ 的号常见）。
_AUTH_CODE_RE = re.compile(r"^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9]{12,64}$")


def parse_import_line(line: str) -> dict:
    """把一行拆成 create_account 的 payload。

    四段格式（买来的 Outlook 号常见）：
        邮箱----密码----clientid----refreshtoken
    填了后两段的自动走 OAuth2；只填 邮箱----密码 就是密码直连。
    第五段可选，是发件人显示名。用 ---- 时按位置取值；用逗号等分隔符时，
    只有第三段长得像 client_id 才当 OAuth，否则仍按老规矩当显示名。

    三段格式（网易/QQ 的号常见）：
        邮箱----登录密码----授权码
    这类邮箱的 SMTP 只认授权码，登录密码填进去必然 535。第三段长得像授权码
    （无空格的字母数字混排）时就取它当密码，并在返回里标出来——原来的行为是
    把它当发件人显示名，等于把凭据印在每封信的 From 上，还顺手用了会失败的那个。
    """
    positional = "----" in line
    if positional:
        parts = line.split("----")
    elif re.search(r"[,\t|;]", line):
        parts = re.split(r"[,\t|;]", line)
    else:
        parts = line.split()
    parts = [p.strip() for p in parts if p.strip()]
    if not parts:
        return {}

    email, rest = parts[0], parts[1:]
    password = rest[0] if rest else ""
    tail = rest[1:]

    client_id = refresh_token = ""
    if len(tail) >= 2 and (positional or _CLIENT_ID_RE.match(tail[0])):
        client_id, refresh_token, tail = tail[0], tail[1], tail[2:]

    note = ""
    if positional and not client_id and len(tail) == 1 and _AUTH_CODE_RE.match(tail[0]):
        password, tail = tail[0], []
        note = "第三段按授权码用，登录密码已忽略"

    return {
        "email": email, "app_password": password,
        "client_id": client_id, "refresh_token": refresh_token,
        "display_name": " ".join(tail),
        "provider": guess_provider(email)["key"],
        "_note": note,
    }


def bulk_import(text: str) -> dict:
    created, skipped, errors, notes = [], [], [], []
    for lineno, raw in enumerate((text or "").splitlines(), 1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        payload = parse_import_line(line)
        if not payload:
            continue
        note = payload.pop("_note", "")
        if "daily_limit" not in payload:
            limit = BULK_IMPORT_DAILY_LIMIT.get(payload.get("provider") or "")
            if limit is not None:
                payload["daily_limit"] = limit
        try:
            created.append(create_account(payload))
            if note:
                notes.append(f"{payload['email']}：{note}")
        except ValueError as exc:
            (skipped if "已存在" in str(exc) else errors).append(
                payload["email"] if "已存在" in str(exc) else f"第 {lineno} 行：{exc}")
        except Exception as exc:
            errors.append(f"第 {lineno} 行：{exc}")
    return {"created": created, "skipped": skipped, "errors": errors, "notes": notes}


def open_smtp(account: dict) -> smtplib.SMTP:
    """建立并登录 SMTP。密码模式和 XOAUTH2 模式都走这里，由 _do_login 分发。"""
    host, port = account["smtp_host"], account["smtp_port"]
    client = (smtplib.SMTP_SSL(host, port, timeout=SMTP_TIMEOUT) if account["use_ssl"]
              else smtplib.SMTP(host, port, timeout=SMTP_TIMEOUT))
    try:
        client.ehlo()
        if account["use_tls"] and not account["use_ssl"]:
            client.starttls()
            client.ehlo()
        _do_login(client, account)
    except Exception:
        try:
            client.close()
        except Exception:
            pass
        raise
    return client


def friendly_smtp_error(exc: Exception) -> str:
    """把 SMTP 报错翻成人话，附上原文。错误原文一律保留，不吞。"""
    raw = str(exc)
    low = raw.lower()
    hint = ""
    if "sendasdenied" in low or "not allowed to send as" in low or "5.2.252" in raw:
        hint = ("SMTP 服务商拒绝代发：当前认证账号没有权限以替换后的发件地址发送。"
                "Outlook/Exchange 通常要求管理员授予 Send As 权限，或改用允许该域名发信的账号/企业邮。")
    elif "5.7.139" in raw or "basic authentication is disabled" in low:
        hint = "微软已对该账号禁用密码直连 SMTP，需改用 OAuth2 或换企业邮箱。"
    elif "535" in raw and "5.7.8" in raw:
        hint = ("Gmail 已永久关闭「登录密码直连 SMTP」。需先开两步验证，"
                "再到 myaccount.google.com/apppasswords 生成 16 位应用专用密码（去掉空格）。")
    elif "535" in raw and "5.7.3" in raw:
        hint = ("OAuth2 登录被拒：access_token 换到了但服务端不认。多半是 client_id 对应的应用"
                "没申请 SMTP.Send 权限，或该账号所在租户没开 Authenticated SMTP。")
    elif "535" in raw:
        hint = "认证失败：密码错了，或该邮箱还没开启 SMTP 服务 / 生成授权码。"
    elif isinstance(exc, TimeoutError) or "timed out" in low:
        hint = "连接超时：检查 SMTP 地址端口，或服务器出站是否被挡。"
    elif "getaddrinfo" in low:
        hint = "SMTP 服务器地址解析不了，检查有没有写错。"
    elif "550" in raw and ("spam" in low or "频率" in raw):
        hint = "触发了服务商频控，建议降低每日配额、隔日再发。"
    elif "dt:spm" in low or ("554" in raw and "163" in low):
        hint = ("网易反垃圾把这封判成了垃圾邮件（DT:SPM）。多半是发件账号信誉不够"
                "——新号、境外 IP 发信、正文带链接都会加重。换一个有正常使用历史的号，"
                "或改用企业邮箱。报错里那个 u.163.com 链接是网易的自助申诉页。")
    return f"{raw}\n\n提示：{hint}" if hint else raw


def test_account(account_id: int) -> dict:
    account = get_account(account_id)
    if account is None:
        raise ValueError(f"账号 {account_id} 不存在")
    status, error = "ready", None
    try:
        client = open_smtp(account)
        try:
            client.quit()
        except Exception:
            pass
    except Exception as exc:
        status, error = "failed", friendly_smtp_error(exc)
    db.execute(
        "UPDATE mb_sender_accounts SET status = %s, last_error = %s, last_test_at = NOW() WHERE id = %s",
        (status, error, account_id))
    return serialize_account(get_account(account_id))


# --------------------------------------------------------------------------- #
# 收信（IMAP）
# --------------------------------------------------------------------------- #
# 建联要靠这条链路收回信。素材提交那批 Outlook 号走的是 OAuth2，
# 而当初申请的 scope 只有 SMTP.Send，收不了信——见 MICROSOFT_SCOPE。
# 建联用的阿里云企业邮箱是密码认证，不受这个限制。

def friendly_imap_error(exc: Exception) -> str:
    text = str(exc)
    low = text.lower()
    if "authentication" in low or ("login" in low and "fail" in low) or "auth" in low:
        return (f"IMAP 认证被拒：{text}\n\n"
                "阿里云企业邮箱：确认管理后台已开启 IMAP 服务；"
                "开了安全设置的话密码要填客户端专用密码而不是登录密码。")
    if "timed out" in low or "timeout" in low:
        return f"连接 IMAP 服务器超时：{text}"
    if "name or service not known" in low or "nodename nor servname" in low:
        return f"IMAP 服务器地址解析不了，检查填的是不是对的：{text}"
    if "ssl" in low or "certificate" in low:
        return f"TLS/SSL 握手失败，检查端口和 SSL 开关是否匹配（993 配 SSL）：{text}"
    return text


def open_imap(account: dict):
    """连上收件箱并登录。异常往外抛，由调用方决定怎么记。"""
    import imaplib

    host = (account.get("imap_host") or "").strip()
    if not host:
        raise ValueError(f"{account['email']}：没有配 IMAP 服务器，这个号只能发不能收")
    if (account.get("auth_mode") or "password") != "password":
        raise ValueError(
            f"{account['email']} 是 OAuth2 账号，收信需要 IMAP.AccessAsUser.All 权限，"
            "而现有 refresh_token 只申请了 SMTP.Send（见 MICROSOFT_SCOPE）。"
            "要用这个号收信得重新走一次授权。")
    port = int(account.get("imap_port") or 993)
    password = crypto_util.decrypt(account.get("encrypted_password")) or ""
    if not password:
        raise ValueError(f"{account['email']}：没有密码，没法登录 IMAP")
    username = (account.get("smtp_username") or account["email"]).strip()

    client = (imaplib.IMAP4_SSL(host, port, timeout=IMAP_TIMEOUT)
              if account.get("imap_ssl", True)
              else imaplib.IMAP4(host, port, timeout=IMAP_TIMEOUT))
    try:
        client.login(username, password)
    except Exception:
        try:
            client.logout()
        except Exception:
            pass
        raise
    return client


def test_imap(account_id: int) -> dict:
    """测试收信。和 test_account（测发信）分开：一个号可能能发不能收。"""
    account = get_account(account_id)
    if account is None:
        raise ValueError(f"账号 {account_id} 不存在")
    try:
        client = open_imap(account)
    except Exception as exc:
        return {"ok": False, "error": friendly_imap_error(exc)}
    try:
        typ, data = client.select("INBOX", readonly=True)
        if typ != "OK":
            return {"ok": False, "error": f"打不开收件箱：{data}"}
        count = int(data[0]) if data and data[0] else 0
        return {"ok": True, "info": f"收件箱里有 {count} 封邮件"}
    except Exception as exc:
        return {"ok": False, "error": friendly_imap_error(exc)}
    finally:
        try:
            client.logout()
        except Exception:
            pass


# --------------------------------------------------------------------------- #
# 图片：存 BYTEA，按 sha256 去重
# --------------------------------------------------------------------------- #

_MAGIC = [(b"\x89PNG\r\n\x1a\n", "png", "image/png"), (b"\xff\xd8\xff", "jpg", "image/jpeg"),
          (b"GIF87a", "gif", "image/gif"), (b"GIF89a", "gif", "image/gif"),
          (b"RIFF", "webp", "image/webp"), (b"BM", "bmp", "image/bmp")]


def sniff_image(blob: bytes) -> tuple[str, str]:
    for magic, ext, mime in _MAGIC:
        if blob.startswith(magic):
            return ext, mime
    return "png", "image/png"


def store_image(blob: bytes) -> int:
    """存图并返回 image_id。同样的字节只存一份。"""
    digest = hashlib.sha256(blob).hexdigest()
    row = db.query_one("SELECT id FROM mb_images WHERE sha256 = %s", (digest,))
    if row:
        return row["id"]
    _, mime = sniff_image(blob)
    return db.execute_and_fetch_id(
        "INSERT INTO mb_images (sha256, content, mime, byte_size) VALUES (%s, %s, %s, %s) "
        "ON CONFLICT (sha256) DO UPDATE SET sha256 = EXCLUDED.sha256 RETURNING id",
        (digest, psycopg2.Binary(blob), mime, len(blob)))


def load_image(image_id: int) -> tuple[bytes, str] | None:
    row = db.query_one("SELECT content, mime FROM mb_images WHERE id = %s", (image_id,))
    if not row:
        return None
    return bytes(row["content"]), row["mime"]


# --------------------------------------------------------------------------- #
# 附件（整批共用一组，与正文内联图无关）
# --------------------------------------------------------------------------- #

def _clean_filename(name: str) -> str:
    """只留文件名本身。浏览器一般不会带路径，但 IE 系和某些客户端会传完整路径，
    而这个名字会原样进 Content-Disposition。"""
    name = (name or "").replace("\\", "/").rsplit("/", 1)[-1].strip()
    name = re.sub(r"[\r\n\t]", "", name)
    return name[:200] or "attachment"


def check_attachment(filename: str, blob: bytes) -> None:
    """单个附件的准入检查。不合格抛 ValueError，消息直接给用户看。"""
    name = _clean_filename(filename)
    if not blob:
        raise ValueError(f"{name} 是空文件")
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext in BLOCKED_ATTACHMENT_EXTS:
        raise ValueError(f".{ext} 类型的附件会被邮件网关整封退回，请压成 .zip 再传")
    if len(blob) > MAX_ATTACHMENT_BYTES:
        raise ValueError(f"{name} 有 {len(blob) / 1048576:.1f}MB，"
                         f"单个附件最大 {MAX_ATTACHMENT_BYTES // 1048576}MB")


def store_attachment(blob: bytes, filename: str) -> dict:
    """存附件返回 {id, filename, mime, byte_size}。同样的字节只存一份。"""
    filename = _clean_filename(filename)
    check_attachment(filename, blob)
    mime = mimetypes.guess_type(filename)[0] or "application/octet-stream"
    digest = hashlib.sha256(blob).hexdigest()
    row = db.query_one("SELECT id FROM mb_attachments WHERE sha256 = %s", (digest,))
    if row:
        attachment_id = row["id"]
    else:
        attachment_id = db.execute_and_fetch_id(
            "INSERT INTO mb_attachments (sha256, content, mime, byte_size) "
            "VALUES (%s, %s, %s, %s) "
            "ON CONFLICT (sha256) DO UPDATE SET sha256 = EXCLUDED.sha256 RETURNING id",
            (digest, psycopg2.Binary(blob), mime, len(blob)))
    return {"id": attachment_id, "filename": filename, "mime": mime, "byte_size": len(blob)}


def load_attachment(attachment_id: int) -> tuple[bytes, str] | None:
    row = db.query_one("SELECT content, mime FROM mb_attachments WHERE id = %s",
                       (attachment_id,))
    if not row:
        return None
    return bytes(row["content"]), row["mime"]


def set_job_attachments(job_id: int, specs: list[dict]) -> list[dict]:
    """把一批附件挂到批次上（整组覆盖）。specs 是 [{id, filename?}, ...]。"""
    db.execute("DELETE FROM mb_job_attachments WHERE job_id = %s", (job_id,))
    if not specs:
        return []
    if len(specs) > MAX_ATTACHMENT_COUNT:
        raise ValueError(f"最多带 {MAX_ATTACHMENT_COUNT} 个附件，现在选了 {len(specs)} 个")

    rows, total = [], 0
    for seq, spec in enumerate(specs):
        try:
            attachment_id = int(spec.get("id"))
        except (TypeError, ValueError):
            raise ValueError("附件 id 不对，请重新上传")
        meta = db.query_one(
            "SELECT id, mime, byte_size FROM mb_attachments WHERE id = %s", (attachment_id,))
        if meta is None:
            raise ValueError(f"附件 {attachment_id} 已经不在了，请重新上传")
        total += meta["byte_size"] or 0
        rows.append((job_id, attachment_id,
                     _clean_filename(spec.get("filename") or f"attachment-{seq + 1}"), seq))
    if total > MAX_ATTACHMENT_TOTAL_BYTES:
        raise ValueError(f"附件合计 {total / 1048576:.1f}MB，超过单封上限 "
                         f"{MAX_ATTACHMENT_TOTAL_BYTES // 1048576}MB（编码后还会再大三成）")

    with db.get_db_cursor(commit=True) as cur:
        execute_values(cur, "INSERT INTO mb_job_attachments "
                            "(job_id, attachment_id, filename, seq) VALUES %s", rows)
    return list_job_attachments(job_id)


def list_job_attachments(job_id: int) -> list[dict]:
    """批次的附件清单（不含内容），给页面和预览用。"""
    rows = db.query_all("""
        SELECT ja.attachment_id AS id, ja.filename, a.mime, a.byte_size
        FROM mb_job_attachments ja JOIN mb_attachments a ON a.id = ja.attachment_id
        WHERE ja.job_id = %s ORDER BY ja.seq ASC
    """, (job_id,))
    return [dict(r) for r in rows]


def load_job_attachments(job_id: int) -> list[dict]:
    """带内容的附件清单，发信时用。"""
    rows = db.query_all("""
        SELECT ja.filename, a.mime, a.content
        FROM mb_job_attachments ja JOIN mb_attachments a ON a.id = ja.attachment_id
        WHERE ja.job_id = %s ORDER BY ja.seq ASC
    """, (job_id,))
    return [{"filename": r["filename"], "mime": r["mime"], "content": bytes(r["content"])}
            for r in rows]


# ---- 素材提交：附件按行挂 ----
# 建联那三个是 job 级，下面这组是 item 级。语义不同，刻意分开而不是加个参数：
# 一个批次里两种粒度不会同时生效，混在一个函数里只会让调用点看不出用的是哪种。

def list_item_attachments_map(item_ids: list[int]) -> dict[int, list[dict]]:
    """一次查完一批行的附件清单（不含内容）。

    load_job 会被 /status 每 2.5 秒轮询一次，逐行查的话 500 行的批次
    每轮就是 500 趟往返（Render 上一趟约 1 秒）。这里固定一条 SQL。
    """
    out: dict[int, list[dict]] = {int(i): [] for i in item_ids}
    if not item_ids:
        return out
    rows = db.query_all("""
        SELECT ia.item_id, ia.attachment_id AS id, ia.filename, a.mime, a.byte_size
        FROM mb_item_attachments ia JOIN mb_attachments a ON a.id = ia.attachment_id
        WHERE ia.item_id = ANY(%s) ORDER BY ia.item_id, ia.seq ASC
    """, (list(item_ids),))
    for r in rows:
        out.setdefault(r["item_id"], []).append(
            {"id": r["id"], "filename": r["filename"],
             "mime": r["mime"], "byte_size": r["byte_size"]})
    return out


def count_item_attachments(item_id: int) -> int:
    """只问「这一行有没有附件」，不读字节。

    发送前的可发性判断用这个：load_item_attachments 会把最多 15MB 拽进内存，
    而判断只需要一个数——何况这一行很可能接着就因为没选发件账号而失败。
    """
    row = db.query_one("SELECT COUNT(*) AS c FROM mb_item_attachments WHERE item_id = %s",
                       (item_id,))
    return int((row or {}).get("c") or 0)


def load_item_attachments(item_id: int) -> list[dict]:
    """带内容的附件清单，发这一行时用。"""
    rows = db.query_all("""
        SELECT ia.filename, a.mime, a.content
        FROM mb_item_attachments ia JOIN mb_attachments a ON a.id = ia.attachment_id
        WHERE ia.item_id = %s ORDER BY ia.seq ASC
    """, (item_id,))
    return [{"filename": r["filename"], "mime": r["mime"], "content": bytes(r["content"])}
            for r in rows]


def sync_item_attachments(job_id: int, specs_by_item: dict[int, list[dict]]) -> None:
    """把「哪一行带哪些附件」整体写回，只动真正变了的行。

    刻意是整批签名而不是 set_item_attachments(item_id, specs)：逐行
    delete + insert + 每个附件查一次元信息，500 行就是 1500+ 趟 Render 往返
    （每趟约 1 秒），点一下预览要转二十分钟。整批形态能做到 5 次查询封顶，
    而且绝大多数请求里没有一行变过，第 3 步就直接返回了。

    specs 是 [{id, filename}, ...]。只信这两个字段，byte_size 一律回库里重读 ——
    大小校验采信前端传的值，等于把上限交给调用方自己填。
    """
    if not specs_by_item:
        return

    # 1) 圈定归属。item_id 是前端传来的，必须限定在本批次内，
    #    且已发出的行不能再改附件（和 sync_job 的 status <> 'sent' 一致）。
    owned = {r["id"] for r in db.query_all(
        "SELECT id FROM mb_items WHERE job_id = %s AND status <> 'sent'", (job_id,))}
    item_ids = sorted(i for i in specs_by_item if i in owned)
    if not item_ids:
        return

    # 2) 现状
    current: dict[int, list[tuple[int, str]]] = {i: [] for i in item_ids}
    for r in db.query_all(
            "SELECT item_id, attachment_id, filename FROM mb_item_attachments "
            "WHERE item_id = ANY(%s) ORDER BY item_id, seq ASC", (item_ids,)):
        current[r["item_id"]].append((r["attachment_id"], r["filename"]))

    # 3) 规整，挑出真正变了的行。顺序变化也算变 —— seq 决定附件在信里的排列。
    wanted: dict[int, list[tuple[int, str]]] = {}
    changed: list[int] = []
    for item_id in item_ids:
        specs = specs_by_item.get(item_id) or []
        if len(specs) > MAX_ATTACHMENT_COUNT:
            raise ValueError(f"每行最多带 {MAX_ATTACHMENT_COUNT} 个附件，"
                             f"有一行选了 {len(specs)} 个")
        norm = []
        for seq, spec in enumerate(specs):
            try:
                attachment_id = int(spec.get("id"))
            except (TypeError, ValueError):
                raise ValueError("附件 id 不对，请重新上传")
            norm.append((attachment_id,
                         _clean_filename(spec.get("filename") or f"attachment-{seq + 1}")))
        wanted[item_id] = norm
        if norm != current[item_id]:
            changed.append(item_id)
    if not changed:
        return

    # 4) 只校验变了的行，元信息一次查完
    need = sorted({aid for iid in changed for aid, _ in wanted[iid]})
    sizes: dict[int, int] = {}
    if need:
        for r in db.query_all(
                "SELECT id, byte_size FROM mb_attachments WHERE id = ANY(%s)", (need,)):
            sizes[r["id"]] = r["byte_size"] or 0
        missing = [a for a in need if a not in sizes]
        if missing:
            raise ValueError(f"附件 {missing[0]} 已经不在了，请重新上传")
    for item_id in changed:
        total = sum(sizes[aid] for aid, _ in wanted[item_id])
        if total > MAX_ATTACHMENT_TOTAL_BYTES:
            raise ValueError(f"有一行的附件合计 {total / 1048576:.1f}MB，超过单封上限 "
                             f"{MAX_ATTACHMENT_TOTAL_BYTES // 1048576}MB（编码后还会再大三成）")

    # 5) 一个事务里删了再插。分成两次提交的话，中间失败会把这些行原本存着的
    #    绑定清空 —— 前端 Map 里还留着，用户要等发送时全被跳过才发现丢了。
    rows = [(item_id, attachment_id, name, seq)
            for item_id in changed
            for seq, (attachment_id, name) in enumerate(wanted[item_id])]
    with db.get_db_cursor(commit=True) as cur:
        cur.execute("DELETE FROM mb_item_attachments WHERE item_id = ANY(%s)", (changed,))
        if rows:
            execute_values(cur, "INSERT INTO mb_item_attachments "
                                "(item_id, attachment_id, filename, seq) VALUES %s", rows)


def build_attachment_part(att: dict) -> MIMEBase:
    maintype, _, subtype = (att.get("mime") or "application/octet-stream").partition("/")
    part = MIMEBase(maintype or "application", subtype or "octet-stream")
    part.set_payload(att["content"])
    encoders.encode_base64(part)
    name = att.get("filename") or "attachment"
    try:
        name.encode("ascii")
        part.add_header("Content-Disposition", "attachment", filename=name)
    except UnicodeEncodeError:
        # 中文文件名走 RFC 2231。直接塞进 filename="…" 的话，
        # 收件端看到的是乱码或者被截断成 "attachment"。
        part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", name))
    return part


def prepare_image(blob: bytes) -> tuple[bytes, str]:
    """太大就压缩，避免被服务商拒信。返回 (二进制, MIME 子类型)。"""
    try:
        with Image.open(io.BytesIO(blob)) as im:
            fmt = (im.format or "JPEG").upper()
            needs_resize = max(im.size) > MAX_IMAGE_EDGE
            needs_recompress = len(blob) > MAX_IMAGE_BYTES
            if not needs_resize and not needs_recompress:
                return blob, {"JPEG": "jpeg", "PNG": "png", "GIF": "gif",
                              "WEBP": "webp"}.get(fmt, "jpeg")
            if fmt == "GIF":                       # 动图别动，压了就没了
                return blob, "gif"
            im.load()
            if needs_resize:
                im.thumbnail((MAX_IMAGE_EDGE, MAX_IMAGE_EDGE), Image.LANCZOS)
            buf = io.BytesIO()
            has_alpha = im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info)
            if fmt == "PNG" and has_alpha:
                im.save(buf, format="PNG", optimize=True)
                return buf.getvalue(), "png"
            im.convert("RGB").save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
            return buf.getvalue(), "jpeg"
    except Exception:
        return blob, "jpeg"                        # 认不出的格式原样发，交给收件端


# --------------------------------------------------------------------------- #
# 模板渲染 + 组装邮件
# --------------------------------------------------------------------------- #

DEFAULT_SUBJECT_TPL = "{{sender_name}} - 素材提交"
DEFAULT_BODY_TPL = "Hi,\n\n这是本次的素材，请查收。\n\n{{image}}\n\n{{signature}}"
DEFAULT_SIGNATURE_TPL = "Best regards,\n{{sender_name}}\n{{sender_email}}"

# 配对表里逐行手填的自定义参数，也是 Excel 要识别的列。
# 加字段：这里加一项 + COLUMN_ALIASES 里补别名，前端表格列和占位符会自动跟上。
ITEM_FIELDS = [("name", "name"), ("id", "id"), ("number", "number")]

PLACEHOLDERS = [
    ("{{sender_name}}", "发件账号的显示名（自动推导，可手改）"),
    ("{{sender_email}}", "发件邮箱地址"),
    ("{{recipient}}", "收件邮箱地址"),
    ("{{name}}", "自定义参数，配对表里逐行手填"),
    ("{{id}}", "自定义参数，配对表里逐行手填"),
    ("{{number}}", "自定义参数，配对表里逐行手填"),
    ("{{index}}", "第几封（从 1 开始）"),
    ("{{total}}", "本批共几封"),
    ("{{image_name}}", "图片文件名"),
    ("{{image}}", "图片插入位置（不写则自动放正文末尾）"),
    ("{{signature}}", "落款签名（不写则自动放正文末尾）"),
]

# ---- KOL 建联 ----
# 素材提交是内部往来，建联是冷启动外联，两者的默认文案和护栏都不一样。

DEFAULT_OUTREACH_SUBJECT_TPL = "Collaboration with {{name}}"
DEFAULT_OUTREACH_BODY_TPL = (
    "Hi {{name}},\n\n"
    "I'm {{sender_name}}. I came across your content on {{platform}} and really "
    "enjoyed it — we'd love to explore a collaboration with you.\n\n"
    "Would you be open to sharing your rates and availability?\n\n"
    "{{signature}}"
)
DEFAULT_OUTREACH_SIGNATURE_TPL = "Best regards,\n{{sender_name}}\n{{sender_email}}"

# 退订出口。模板里没写 {{unsubscribe}} 也会自动补到正文末尾——
# 这个出口不能因为谁忘了写占位符就消失。
DEFAULT_UNSUBSCRIBE_TEXT = (
    "If you'd rather not hear from us, just reply with \"unsubscribe\" "
    "and we won't contact you again."
)

# 建联的内置占位符。没有 {{image}}/{{image_name}}（建联的信不带图），
# 也没有 {{unsubscribe}}（自动追加，没有「插入位置」可选）。
# 名单里的每一列会额外生成一个占位符，由页面动态渲染。
OUTREACH_PLACEHOLDERS = [
    ("{{sender_name}}", "发件账号的显示名（自动推导，可手改）"),
    ("{{sender_email}}", "发件邮箱地址"),
    ("{{recipient}}", "收件邮箱地址"),
    ("{{index}}", "第几封（从 1 开始）"),
    ("{{total}}", "本批共几封"),
    ("{{signature}}", "落款签名（不写则自动放正文末尾）"),
]

# 名单里叫这些名字的列会和内置变量撞车，解析时自动改名（recipient → recipient_1）
RESERVED_KEYS = {
    "sender_name", "sender_email", "recipient", "index", "total",
    "image", "image_name", "signature", "unsubscribe",
}


def render(text: str, variables: dict) -> str:
    """简单的 {{key}} 替换。刻意不用 Jinja——正文里出现 {% 就会炸。"""
    out = text or ""
    for key, value in variables.items():
        out = out.replace("{{" + key + "}}", str(value if value is not None else ""))
    return out


def _looks_like_html(text: str) -> bool:
    return bool(re.search(r"<(p|br|div|table|img|a|span|h[1-6]|ul|ol)\b", text or "", re.I))


def to_html(text: str) -> str:
    if _looks_like_html(text):
        return text
    return escape(text or "").replace("\n", "<br>\n")


def html_to_text(html: str) -> str:
    """给 text/plain 降级版用的粗略去标签。"""
    text = re.sub(r"(?is)<(script|style).*?</\1>", "", html or "")
    text = re.sub(r"(?i)<br\s*/?>[ \t]*\n?", "\n", text)
    text = re.sub(r"(?i)</(p|div|tr|h[1-6]|li)>[ \t]*\n?", "\n", text)
    text = re.sub(r"(?i)<img[^>]*alt=\"([^\"]*)\"[^>]*>", r"[图片: \1]", text)
    text = re.sub(r"(?i)<img[^>]*>", "[图片]", text)
    text = re.sub(r"<[^>]+>", "", text)
    # 正文进 to_html() 时被 escape 过，这里必须还原，
    # 否则纯文本版里会出现 &#x27; &quot; 这种给机器看的东西
    text = unescape(text)
    return re.sub(r"\n{3,}", "\n\n", text).strip()


def normalize_replacement_domain(value: str) -> str:
    """用户填的是域名，不是 URL / 邮箱。这里收紧一点，避免误发。"""
    domain = (value or "").strip().lower().rstrip(".")
    if not domain:
        return ""
    if any(x in domain for x in ("://", "/", "?", "#", "@")) or re.search(r"\s", domain):
        raise ValueError("替换域名只填域名本身，例如 tiktok.com，不要带 http、路径或邮箱")
    if re.fullmatch(r"\d+(?:\.\d+){3}", domain):
        raise ValueError("替换域名不能填 IP")
    if len(domain) > 253 or "." not in domain:
        raise ValueError("替换域名格式不对，例如 tiktok.com")
    labels = domain.split(".")
    for label in labels:
        if (not label or len(label) > 63 or label.startswith("-") or label.endswith("-")
                or not re.fullmatch(r"[a-z0-9-]+", label)):
            raise ValueError("替换域名格式不对，例如 tiktok.com")
    return domain


def replacement_domain_for_job(job: dict) -> str:
    if not job.get("replace_domain_enabled"):
        return ""
    domain = normalize_replacement_domain(job.get("replacement_domain") or "")
    if not domain:
        raise ValueError("已勾选替换域名，请填写目标域名")
    return domain


def replace_sender_domain(email: str, domain: str) -> str:
    if not domain:
        return email
    local = (email or "").rsplit("@", 1)[0]
    if not local or local == email:
        raise ValueError("发件账号邮箱格式不对，无法替换域名")
    return f"{local}@{domain}"


def rewrite_url_domain(url: str, domain: str) -> str:
    if not domain:
        return url
    parsed = urllib.parse.urlsplit(url)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return url
    userinfo = ""
    hostport = parsed.netloc
    if "@" in hostport:
        userinfo, hostport = hostport.rsplit("@", 1)
        userinfo += "@"
    port = ""
    if hostport.startswith("["):
        end = hostport.find("]")
        if end != -1:
            port = hostport[end + 1:]
    elif ":" in hostport:
        port = ":" + hostport.rsplit(":", 1)[-1]
    return urllib.parse.urlunsplit((parsed.scheme, userinfo + domain + port,
                                    parsed.path, parsed.query, parsed.fragment))


def rewrite_body_domains(html: str, domain: str) -> str:
    if not domain:
        return html

    def repl_attr(match):
        prefix, quote, url = match.groups()
        return prefix + quote + rewrite_url_domain(url, domain) + quote

    html = re.sub(r"(?i)(\b(?:href|src)\s*=\s*)(['\"])(https?://[^'\"\s<>]+)\2",
                  repl_attr, html or "")

    def repl_bare(match):
        url, tail = match.groups()
        return rewrite_url_domain(url, domain) + tail

    return re.sub(r"(?i)\b(https?://[^\s<'\"]+?)([).,;!?，。；！？、]*)(?=(?:\s|<|$))",
                  repl_bare, html)


def image_html(src: str, alt: str) -> str:
    return (f'<div style="margin:16px 0"><img src="{escape(src, quote=True)}" '
            f'alt="{escape(alt or "", quote=True)}" '
            'style="max-width:600px;width:100%;height:auto;display:block;border:0"></div>')


def _open_local_smtp():
    """打开本地直投 SMTP，不做账号登录。适合 Mailpit / Mailhog / 本机 Postfix 测试。"""
    client = (_SMTP_SSL(LOCAL_SMTP_HOST, LOCAL_SMTP_PORT, timeout=SMTP_TIMEOUT)
              if LOCAL_SMTP_USE_SSL else _SMTP(LOCAL_SMTP_HOST, LOCAL_SMTP_PORT,
                                               timeout=SMTP_TIMEOUT))
    try:
        client.ehlo()
        if LOCAL_SMTP_USE_TLS and not LOCAL_SMTP_USE_SSL:
            client.starttls()
            client.ehlo()
    except Exception:
        try:
            client.close()
        except Exception:
            pass
        raise
    return client


def send_local_test_email(*, sender_name: str, sender_email: str, to_email: str,
                          subject: str, body_html: str,
                          replacement_domain: str = "") -> dict:
    """给本机 SMTP / Mailpit 发一封测试邮件，不走账号池。"""
    sender_email = (sender_email or "").strip().lower()
    if "@" not in sender_email:
        raise ValueError("发件人邮箱格式不对")
    replacement_domain = normalize_replacement_domain(replacement_domain)
    header_sender_email = replace_sender_domain(sender_email, replacement_domain)
    body_html = rewrite_body_domains(body_html or "", replacement_domain)

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject or ""
    msg["From"] = formataddr((sender_name or "", header_sender_email))
    msg["To"] = to_email
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain=header_sender_email.rsplit("@", 1)[-1])
    plain = MIMEText(html_to_text(body_html), "plain", "utf-8")
    html_part = MIMEText(body_html, "html", "utf-8")
    msg.attach(plain)
    msg.attach(html_part)

    client = _open_local_smtp()
    try:
        client.sendmail(header_sender_email, [to_email], msg.as_string())
        answer = getattr(client, "last_data_response", None)
        if answer:
            code, text = answer
            smtp_response = f"{code} {text.decode('utf-8', 'replace').strip()}"
        else:
            smtp_response = ""
    finally:
        try:
            client.quit()
        except Exception:
            pass
    return {
        "subject": msg["Subject"],
        "body_html": body_html,
        "sender_email": header_sender_email,
        "smtp_response": smtp_response,
        "message_id": msg["Message-ID"],
        "smtp_host": LOCAL_SMTP_HOST,
        "smtp_port": LOCAL_SMTP_PORT,
    }


def unsubscribe_html(text: str) -> str:
    """退订出口渲染成正文末尾的灰色小字块。escape 掉，退订文案不允许带 HTML。"""
    return ('<div style="margin:24px 0 0;padding-top:12px;border-top:1px solid #e5e5e7;'
            'font-size:12px;line-height:1.6;color:#86868b">'
            f'{escape(text, quote=False)}</div>')


def compose(*, subject_tpl, body_tpl, signature_tpl, sender_name, sender_email,
            signature_name, recipient, index, total, image_name="", image_src="",
            extra_vars=None, unsubscribe_text="") -> tuple[str, str]:
    """渲染出 (主题, 正文 HTML)。

    image_src 传空表示这封信没有图，此时不自动补 {{image}}，模板里写了也替换成空。
    extra_vars 是该行的自定义参数；与内置变量同名时以内置的为准，
    避免名单里一列叫 recipient 就把收件人顶掉。
    unsubscribe_text 只有建联会传。模板里没写 {{unsubscribe}} 就自动补到末尾——
    这个出口不能因为谁忘了写占位符就消失。素材提交传空，此时残留的占位符替换成空串。
    """
    has_image = bool(image_src)
    base_vars = {**(extra_vars or {}), "sender_name": sender_name, "sender_email": sender_email,
                 "recipient": recipient, "index": index, "total": total, "image_name": image_name}
    subject = render(subject_tpl or "", base_vars).strip()
    signature_html = to_html(render(signature_tpl or "",
                                    {**base_vars, "sender_name": signature_name}))
    raw = body_tpl or ""
    if has_image and "{{image}}" not in raw:
        if "{{signature}}" in raw:
            raw = raw.replace("{{signature}}", "{{image}}\n\n{{signature}}", 1)
        else:
            raw = raw.rstrip() + "\n\n{{image}}\n\n{{signature}}"
    elif "{{signature}}" not in raw:
        raw = raw.rstrip() + "\n\n{{signature}}"
    if unsubscribe_text and "{{unsubscribe}}" not in raw:
        raw = raw.rstrip() + "\n\n{{unsubscribe}}"

    body = to_html(render(raw, base_vars))
    body = body.replace("{{image}}", image_html(image_src, image_name) if has_image else "")
    body = body.replace("{{signature}}", signature_html)
    body = body.replace("{{unsubscribe}}",
                        unsubscribe_html(unsubscribe_text) if unsubscribe_text else "")
    html = ('<div style="font-family:-apple-system,BlinkMacSystemFont,\'Segoe UI\','
            "Helvetica,Arial,sans-serif;font-size:15px;line-height:1.7;color:#1d1d1f\">"
            f"{body}</div>")
    return subject, html


class _RecordsDataResponse:
    """记下服务器对 DATA 的最终应答。sendmail() 会把它丢掉，
    但排查「回了 250 却没人收到」时它是唯一线索——里面一般带 queue id。"""
    last_data_response = None

    def data(self, msg):
        result = super().data(msg)
        self.last_data_response = result
        return result


class _SMTP(_RecordsDataResponse, smtplib.SMTP):
    pass


class _SMTP_SSL(_RecordsDataResponse, smtplib.SMTP_SSL):
    pass


def _open_smtp_recording(account: dict):
    """和 open_smtp 同逻辑，但用能记录 DATA 应答的子类。"""
    host, port = account["smtp_host"], account["smtp_port"]
    client = (_SMTP_SSL(host, port, timeout=SMTP_TIMEOUT) if account["use_ssl"]
              else _SMTP(host, port, timeout=SMTP_TIMEOUT))
    try:
        client.ehlo()
        if account["use_tls"] and not account["use_ssl"]:
            client.starttls()
            client.ehlo()
        _do_login(client, account)
    except Exception:
        try:
            client.close()
        except Exception:
            pass
        raise
    return client


def send_one_email(*, account: dict, to_email: str, subject_tpl: str, body_tpl: str,
                   signature_tpl: str, from_display: str, signature_name: str,
                   image_bytes: bytes | None, image_name: str, index: int, total: int,
                   extra_vars: dict | None = None, replacement_domain: str = "",
                   unsubscribe_text: str = "",
                   attachments: list[dict] | None = None) -> dict:
    """真正发一封。异常往外抛，由调用方决定怎么记。"""
    replacement_domain = normalize_replacement_domain(replacement_domain)
    header_sender_email = replace_sender_domain(account["email"], replacement_domain)
    envelope_sender_email = account["email"]
    message_id_domain = account["email"].rsplit("@", 1)[-1]
    has_image = bool(image_bytes)
    cid = make_msgid(domain="mailblaster.local")[1:-1] if has_image else ""
    subject, html = compose(
        subject_tpl=subject_tpl, body_tpl=body_tpl, signature_tpl=signature_tpl,
        sender_name=from_display, sender_email=header_sender_email, signature_name=signature_name,
        recipient=to_email, index=index, total=total, image_name=image_name,
        image_src=f"cid:{cid}" if has_image else "", extra_vars=extra_vars,
        unsubscribe_text=unsubscribe_text)
    html = rewrite_body_domains(html, replacement_domain)

    # 先把「正文」这一坨拼好（内联图算正文的一部分），
    # 有附件时再拿 mixed 把它整个包起来。
    plain = MIMEText(html_to_text(html), "plain", "utf-8")
    html_part = MIMEText(html, "html", "utf-8")
    if has_image:
        content = MIMEMultipart("related")
        alt = MIMEMultipart("alternative")
        alt.attach(plain)
        alt.attach(html_part)
        content.attach(alt)
        data, subtype = prepare_image(image_bytes)
        img = MIMEImage(data, _subtype=subtype)
        img.add_header("Content-ID", f"<{cid}>")
        img.add_header("Content-Disposition", "inline", filename=image_name or f"image.{subtype}")
        content.attach(img)
    else:
        content = MIMEMultipart("alternative")
        content.attach(plain)
        content.attach(html_part)

    if attachments:
        # 有附件时最外层必须是 mixed。related/alternative 是「同一份正文的多种表示」
        # 的容器，把 PDF 挂进去，部分客户端会把它当成正文的候选版本而不显示成附件。
        root = MIMEMultipart("mixed")
        root.attach(content)
        for att in attachments:
            root.attach(build_attachment_part(att))
    else:
        root = content

    root["Subject"] = subject
    root["From"] = formataddr((from_display, header_sender_email))
    root["To"] = to_email
    root["Date"] = formatdate(localtime=True)
    # 自签 Message-ID：回信/对账时拿它比对。已实测阿里云企业邮会原样保留。
    root["Message-ID"] = make_msgid(domain=message_id_domain)
    if unsubscribe_text:
        # 用 envelope 地址而不是替换后的 header 地址：开了域名替换时，
        # 替换出来的域名不是任何人在读的信箱，退订回信会直接消失。
        #
        # 刻意不加 List-Unsubscribe-Post：RFC 8058 的一键退订要求有一个
        # 无需确认就处理的 URL 端点，我们没有。声明了不兑现比不声明更糟。
        root["List-Unsubscribe"] = f"<mailto:{envelope_sender_email}?subject=unsubscribe>"

    client = _open_smtp_recording(account)
    try:
        client.sendmail(envelope_sender_email, [to_email], root.as_string())
        answer = getattr(client, "last_data_response", None)
        if answer:
            code, text = answer
            smtp_response = f"{code} {text.decode('utf-8', 'replace').strip()}"
        else:
            smtp_response = ""
    finally:
        try:
            client.quit()
        except Exception:
            pass
    return {"subject": subject, "body_html": html, "sender_email": header_sender_email,
            "message_id": root["Message-ID"], "smtp_response": smtp_response}


# --------------------------------------------------------------------------- #
# 模板库
# --------------------------------------------------------------------------- #

def list_templates(mode: str = "material") -> list[dict]:
    return [dict(r) for r in db.query_all(
        "SELECT * FROM mb_templates WHERE mode = %s ORDER BY updated_at DESC", (mode,))]


def save_template(name: str, subject: str, body: str, signature: str,
                  mode: str = "material") -> list[dict]:
    name = (name or "").strip()
    if not name:
        raise ValueError("给模板起个名字")
    db.execute("""
        INSERT INTO mb_templates (mode, name, subject, body_html, signature_html, updated_at)
        VALUES (%s, %s, %s, %s, %s, NOW())
        ON CONFLICT (mode, name) DO UPDATE SET subject = EXCLUDED.subject,
            body_html = EXCLUDED.body_html, signature_html = EXCLUDED.signature_html,
            updated_at = NOW()
    """, (mode, name, subject or "", body or "", signature or ""))
    return list_templates(mode)


def delete_template(template_id: int, mode: str = "material") -> list[dict]:
    db.execute("DELETE FROM mb_templates WHERE id = %s", (template_id,))
    return list_templates(mode)


def defaults_for_page(mode: str = "material") -> dict:
    """页面初始值：有存过的模板就用最近改的那个，没有才退回内置默认。"""
    saved = list_templates(mode)
    if saved:
        t = saved[0]
        return {"subject": t["subject"], "body": t["body_html"],
                "signature": t["signature_html"], "loaded_from": t["name"]}
    if mode == "outreach":
        return {"subject": DEFAULT_OUTREACH_SUBJECT_TPL, "body": DEFAULT_OUTREACH_BODY_TPL,
                "signature": DEFAULT_OUTREACH_SIGNATURE_TPL, "loaded_from": ""}
    return {"subject": DEFAULT_SUBJECT_TPL, "body": DEFAULT_BODY_TPL,
            "signature": DEFAULT_SIGNATURE_TPL, "loaded_from": ""}


# --------------------------------------------------------------------------- #
# 发信历史 / 重传去重 / 账号轮换
# --------------------------------------------------------------------------- #

def record_send(*, recipient, material_id="", material_name="", sender_account_id=None,
                sender_email="", job_id=None, item_id=None, subject="", message_id="",
                mode="material") -> None:
    db.execute("""
        INSERT INTO mb_history (recipient, material_id, material_name, sender_account_id,
                                sender_email, job_id, item_id, subject, message_id, mode)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
    """, ((recipient or "").strip().lower(), material_id or "", material_name or "",
          sender_account_id, sender_email or "", job_id, item_id, subject or "",
          message_id or "", mode or "material"))


def already_sent(pairs: list[tuple[str, str]]) -> set[tuple[str, str]]:
    """哪些 (收件人, 素材id) 已经发过。素材 id 为空的不参与去重
    ——没有 id 就没法判断是不是同一条素材，宁可重发也不误杀。"""
    wanted = {(r.strip().lower(), (m or "").strip()) for r, m in pairs if r and (m or "").strip()}
    if not wanted:
        return set()
    rows = db.query_all(
        "SELECT DISTINCT recipient, material_id FROM mb_history WHERE material_id <> ''")
    have = {(r["recipient"], r["material_id"]) for r in rows}
    return wanted & have


def list_history(limit: int = 500, keyword: str = "", mode: str = "") -> list[dict]:
    sql = "SELECT * FROM mb_history WHERE 1=1"
    args: list = []
    if mode:
        sql += " AND mode = %s"
        args.append(mode)
    if keyword:
        sql += (" AND (material_id ILIKE %s OR material_name ILIKE %s OR subject ILIKE %s"
                " OR sender_email ILIKE %s OR recipient ILIKE %s)")
        args.extend([f"%{keyword}%"] * 5)
    sql += " ORDER BY sent_at DESC, id DESC LIMIT %s"
    args.append(int(limit))
    return [dict(r) for r in db.query_all(sql, tuple(args))]


# --------------------------------------------------------------------------- #
# 抑制名单（只作用于建联）
# --------------------------------------------------------------------------- #
# 冷邮件四道护栏之一。退订请求、硬退信、手动拉黑都进这张表，
# 解析名单时先剔一遍，真正发之前再拦一次——中间隔着人工编辑和 API 调用，
# 只在解析期过滤挡不住直接 POST。

def list_suppressed() -> list[dict]:
    return [dict(r) for r in db.query_all(
        "SELECT * FROM mb_suppression ORDER BY created_at DESC, id DESC")]


def suppress(email: str, reason: str = "", source: str = "manual") -> None:
    email = (email or "").strip().lower()
    if "@" not in email:
        raise ValueError(f"邮箱格式不正确：{email or '(空)'}")
    db.execute("""
        INSERT INTO mb_suppression (email, reason, source) VALUES (%s, %s, %s)
        ON CONFLICT (email) DO UPDATE SET reason = EXCLUDED.reason, source = EXCLUDED.source
    """, (email, reason or "", source or "manual"))


def unsuppress(email: str) -> None:
    db.execute("DELETE FROM mb_suppression WHERE email = %s", ((email or "").strip().lower(),))


def is_suppressed(email: str) -> str | None:
    """在名单里就返回原因，不在返回 None。

    ⚠️ 原因**可能是空字符串**（手动拉黑常常懒得填理由），
    所以调用方必须判 `is not None`，判真值会把「拉黑了但没写理由」当成没拉黑。
    """
    row = db.query_one("SELECT reason FROM mb_suppression WHERE email = %s",
                       ((email or "").strip().lower(),))
    return None if row is None else (row["reason"] or "")


def filter_suppressed(emails: list[str]) -> set[str]:
    """一次查完，返回这批里命中名单的那些。"""
    targets = list({(e or "").strip().lower() for e in emails if e})
    if not targets:
        return set()
    rows = db.query_all(
        "SELECT email FROM mb_suppression WHERE email = ANY(%s)", (targets,))
    return {r["email"] for r in rows}


def cooldown_map(recipients: list[str], days: int = COOLDOWN_DAYS) -> dict[str, dict[int, str]]:
    """{收件人: {账号id: 最近一次发给它的时间}}，只含冷却期内的。一次查完。"""
    targets = tuple({r.strip().lower() for r in recipients if r})
    if not targets:
        return {}
    rows = db.query_all("""
        SELECT recipient, sender_account_id, MAX(sent_at) AS last_at FROM mb_history
        WHERE recipient = ANY(%s) AND sender_account_id IS NOT NULL
          AND sent_at >= NOW() - (%s || ' days')::INTERVAL
        GROUP BY recipient, sender_account_id
    """, (list(targets), str(int(days))))
    out: dict[str, dict[int, str]] = {}
    for r in rows:
        out.setdefault(r["recipient"], {})[r["sender_account_id"]] = (
            r["last_at"].strftime("%Y-%m-%d %H:%M") if r["last_at"] else "")
    return out


def assign_accounts(pool: list[dict], targets: list[str],
                    days: int = COOLDOWN_DAYS) -> list[dict]:
    """给每一行挑发件账号，避开该收件人冷却期内用过的。

    池子不够时**不留空**——留空等于这行发不出去。回退到最久没用过的那个并标出来，
    让人自己决定换不换。
    """
    cooldowns = cooldown_map(targets, days)
    used_this_batch: dict[str, set[int]] = {}
    result = []
    for target in targets:
        key = (target or "").strip().lower()
        blocked = dict(cooldowns.get(key, {}))
        taken = used_this_batch.setdefault(key, set())
        fresh = [a for a in pool if a["id"] not in blocked and a["id"] not in taken]
        if fresh:
            chosen, note = fresh[0], ""
        else:
            candidates = [a for a in pool if a["id"] not in taken] or pool
            if candidates:
                chosen = sorted(candidates, key=lambda a: blocked.get(a["id"], ""))[0]
                last = blocked.get(chosen["id"], "")
                note = (f"可用账号不够，这个 {days} 天内给 {key} 发过"
                        + (f"（{last}）" if last else "") + "，可在下拉里手动换")
            else:
                chosen, note = None, "账号池是空的"
        if chosen:
            taken.add(chosen["id"])
        result.append({"account": chosen, "note": note})
    return result


def quota_state(account_id: int, daily_limit: int | None) -> dict:
    """daily_limit 传 None 表示不限量，此时 remaining 也是 None。
    调用方判耗尽要写 `remaining is not None and remaining <= 0`，
    不能直接判真值——0 和 None 是两件完全不同的事。"""
    row = db.query_one(
        "SELECT COUNT(*) AS c FROM mb_history WHERE sender_account_id = %s "
        "AND sent_at >= %s", (account_id, _today_start_utc()))
    used = row["c"] if row else 0
    if daily_limit is None:
        return {"used": used, "limit": None, "remaining": None}
    return {"used": used, "limit": daily_limit, "remaining": max(0, daily_limit - used)}


# --------------------------------------------------------------------------- #
# 时间护栏：配额的「今日」和发送窗口
# --------------------------------------------------------------------------- #
# 这两样必须钉在同一个时区，否则会各偏各的。见文件头 SEND_TZ_NAME 的注释。

def _send_tz():
    try:
        from zoneinfo import ZoneInfo
        return ZoneInfo(SEND_TZ_NAME)
    except Exception:
        # 认不出时区名就退回东八区，别因为配置写错就整个停发
        return timezone(timedelta(hours=8))


def now_local() -> datetime:
    return datetime.now(_send_tz())


def _today_start_utc() -> datetime:
    """本地时区的今天零点，换算成带时区的时间戳交给 Postgres 比较。

    原来是 date_trunc('day', NOW())，跟着 Postgres 服务器时区走（Render 上是 UTC），
    于是「今日配额」在北京时间早上 8 点才重置。
    """
    local = now_local()
    return local.replace(hour=0, minute=0, second=0, microsecond=0)


def parse_send_window(spec: str = "") -> tuple[int, int] | None:
    """'08:00-22:00' → (480, 1320)，单位是当天第几分钟。

    格式写错按「没有窗口」处理而不是「拒发」——
    不能因为一个配置笔误就让整个功能静默停摆。
    """
    spec = (spec or "").strip()
    if not spec:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})\s*-\s*(\d{1,2}):(\d{2})$", spec)
    if not m:
        logger.warning("mail-blaster: 发送窗口 %r 格式不对，按无窗口处理", spec)
        return None
    h1, m1, h2, m2 = (int(x) for x in m.groups())
    if not (0 <= h1 < 24 and 0 <= h2 <= 24 and 0 <= m1 < 60 and 0 <= m2 < 60):
        logger.warning("mail-blaster: 发送窗口 %r 超出范围，按无窗口处理", spec)
        return None
    return h1 * 60 + m1, h2 * 60 + m2


def outside_send_window(spec: str = None, at: datetime = None) -> str:
    """在窗口内返回空串；不在返回一句中文原因。"""
    window = parse_send_window(SEND_WINDOW if spec is None else spec)
    if window is None:
        return ""
    start, end = window
    at = at or now_local()
    minutes = at.hour * 60 + at.minute
    inside = (start <= minutes < end) if start < end else (minutes >= start or minutes < end)
    if inside:
        return ""
    return (f"现在是 {at.strftime('%H:%M')}（{SEND_TZ_NAME}），"
            f"不在发送窗口 {SEND_WINDOW} 内。冷启动建联刻意避开非工作时间，"
            f"等到窗口内再发。")


# --------------------------------------------------------------------------- #
# Excel 解析：图片嵌在单元格里
# --------------------------------------------------------------------------- #

COLUMN_ALIASES = {
    "image":     ["图片", "图", "素材", "素材图", "图片素材", "image", "img", "picture", "photo"],
    "name":      ["name", "名字", "姓名", "名称", "kol名字", "koc名字", "达人名字", "达人", "昵称"],
    "id":        ["id", "编号", "素材编号", "koc编号", "kol编号", "合约号", "编码"],
    "number":    ["number", "数量", "号码", "序号", "期数", "条数", "num", "no"],
    "recipient": ["收件邮箱", "收件人", "收件邮件", "邮箱", "对方邮箱", "客户邮箱",
                  "recipient", "email", "mail", "to"],
    "status":    ["发送状态", "状态", "发送", "已发送", "是否发送", "status", "sent"],
}
_SPECIAL = {"image", "recipient", "status"}
DATA_FIELDS = [f for f in COLUMN_ALIASES if f not in _SPECIAL]
# 一格里可能塞了多个邮箱，逐个抠出来。排除集里要带上全角标点——
# 中文输入法下分隔符常常是 ，；、 而不是半角，漏了会把 'a@x.com；b@y.com'
# 当成一个地址，发信必然失败。
_EMAIL_RE = re.compile(r"[^@\s,;、，；]+@[^@\s,;、，；]+\.[^@\s,;、，；]+")
MAX_SCAN_ROWS = 10


def _norm(text) -> str:
    return re.sub(r"[\s_\-·:：()（）]+", "", str(text or "")).strip().lower()


def _match_column(cell_value) -> str | None:
    key = _norm(cell_value)
    if not key:
        return None
    for field, aliases in COLUMN_ALIASES.items():
        if key in {_norm(a) for a in aliases}:
            return field
    return None


def _image_bytes_of(image) -> bytes | None:
    """openpyxl 的 Image.ref 可能是 BytesIO / 文件对象 / PIL 图 / 裸 bytes，逐个试。"""
    ref = getattr(image, "ref", None)
    if ref is None:
        return None
    if isinstance(ref, bytes):
        return ref
    if hasattr(ref, "getvalue"):
        return ref.getvalue()
    if hasattr(ref, "read"):
        try:
            ref.seek(0)
        except Exception:
            pass
        return ref.read()
    if hasattr(ref, "save"):
        buf = io.BytesIO()
        ref.save(buf, format=getattr(ref, "format", None) or "PNG")
        return buf.getvalue()
    return None


def _anchor_row(image) -> int | None:
    """图片锚定在第几行（1 基）。OneCellAnchor 和 TwoCellAnchor 都有 _from。"""
    frm = getattr(getattr(image, "anchor", None), "_from", None)
    return int(frm.row) + 1 if frm is not None and hasattr(frm, "row") else None


def _find_header(ws):
    """判定条件是「认出至少两列」而不是写死某两个字段：用户不一定所有列都有，
    但只认出一列就当表头会把普通数据行误判成表头。"""
    for row_idx in range(1, min(MAX_SCAN_ROWS, ws.max_row or 1) + 1):
        mapping = {}
        for col_idx, cell in enumerate(ws[row_idx], start=1):
            field = _match_column(cell.value)
            if field and field not in mapping:
                mapping[field] = col_idx
        if len(mapping) >= 2:
            return row_idx, mapping
    return None, None


def parse_material_xlsx(data: bytes) -> dict:
    """解析素材 Excel。缺图/缺邮箱/已发过的行都会保留并注明原因，**不静默丢行**
    ——否则用户会以为漏发是工具吞了数据。

    缺图不是错：附件是素材的另一种形态，这类行走 notices 而不是 errors，
    照常入队，等第 3 步挂上附件就能发。"""
    try:
        wb = load_workbook(io.BytesIO(data))
    except Exception as exc:
        raise ValueError(f"这个文件打不开，确认是 .xlsx 格式（不是 .xls / .csv）：{exc}") from exc

    ws = wb.active
    header_row, columns = _find_header(ws)
    if header_row is None:
        raise ValueError(
            f"没找到表头。需要有一行至少认出两列（图片 / {' / '.join(DATA_FIELDS)} / 收件邮箱 / 发送状态）。"
            f"已扫描前 {MAX_SCAN_ROWS} 行。")

    by_row: dict[int, list[bytes]] = {}
    orphans = 0
    for image in getattr(ws, "_images", []):
        row = _anchor_row(image)
        blob = _image_bytes_of(image)
        if blob is None:
            continue
        if row is None or row <= header_row:
            orphans += 1
            continue
        by_row.setdefault(row, []).append(blob)

    present = [f for f in DATA_FIELDS if f in columns]
    rows, errors, notices = [], [], []
    if orphans:
        errors.append(f"有 {orphans} 张图没有锚定到数据行（可能浮在表头上方或悬空），已忽略")

    def cell(row_idx, field):
        if field not in columns:
            return ""
        v = ws.cell(row=row_idx, column=columns[field]).value
        return "" if v is None else str(v).strip()

    for row_idx in range(header_row + 1, (ws.max_row or header_row) + 1):
        variables = {f: cell(row_idx, f) for f in present}
        recipient_raw, status_raw = cell(row_idx, "recipient"), cell(row_idx, "status")
        blobs = by_row.get(row_idx, [])
        if not any(variables.values()) and not blobs and not recipient_raw:
            continue

        label = next((v for v in variables.values() if v), "") or str(row_idx)

        # 发送状态列非空 = 发过了，丢弃。空才入队。
        if status_raw:
            rows.append({"vars": variables, "image_bytes": None, "image_name": "",
                         "recipient": recipient_raw.lower(), "excel_row": row_idx,
                         "skip_reason": f"发送状态列已填「{status_raw}」，视为发过了"})
            continue

        emails = _EMAIL_RE.findall(recipient_raw)
        if recipient_raw and not emails:
            errors.append(f"第 {row_idx} 行（{label}）收件邮箱格式不对：{recipient_raw!r}")
        if len(emails) > 1:
            errors.append(f"第 {row_idx} 行有 {len(emails)} 个邮箱，只用第一个（{emails[0]}）。"
                          "一行一个收件人，要发给多个人请拆成多行。")
        if not blobs:
            # 不算错误：附件是素材的另一种形态，这一行照常入队，
            # 等用户在第 3 步给它挂个 PDF 就能发。走 notices 而不是 errors ——
            # errors 在前端是红色 errbox，一半行是附件素材时会糊一屏红。
            # 带上行号：这里还不知道这一行会不会因为缺邮箱被丢掉，
            # 由 create_job_from_excel 按真正入队的行筛一遍再给前端。
            notices.append({"row": row_idx,
                            "text": f"第 {row_idx} 行（{label}）没有图片，"
                                    "在第 3 步点 📎 给它传一个附件就能发。"})
            rows.append({"vars": dict(variables), "recipient": (emails[0].lower() if emails else ""),
                         "image_bytes": None, "image_name": "",
                         "excel_row": row_idx, "skip_reason": ""})
            continue

        # 一行叠了多张图 = 多份素材。全部展开成独立的信，而不是只取第一张。
        # 实测用户的表里第 2 行叠了 10 份不同的执照——只取第一张会静默丢掉 9 份。
        # 每份素材一封信，本来就是这个功能的语义。
        if len(blobs) > 1:
            errors.append(
                f"第 {row_idx} 行叠了 {len(blobs)} 张图，已展开成 {len(blobs)} 封"
                "（每份素材一封信）。如果其中有误放的，在下面的表里删掉对应行即可。")

        recipient = emails[0].lower() if emails else ""
        for k, blob in enumerate(blobs):
            ext = sniff_image(blob)[0]
            suffix = f"-{k + 1}" if len(blobs) > 1 else ""
            rows.append({"vars": dict(variables), "recipient": recipient,
                         "image_bytes": blob, "image_name": f"{label}{suffix}.{ext}",
                         "excel_row": row_idx, "skip_reason": ""})

    return {"rows": rows, "errors": errors, "notices": notices, "header_row": header_row,
            "fields": present, "has_recipient": "recipient" in columns,
            "has_status": "status" in columns, "sheet": ws.title}


# --------------------------------------------------------------------------- #
# KOL 建联名单：粘贴 / Excel 两个入口，一个核心
# --------------------------------------------------------------------------- #
# 和素材提交的区别：素材那边列名是**闭集**（认不出就丢），
# 建联这边认不出的列恰恰是最有用的——每一列都变成一个模板占位符。

# 锚定匹配，判断「整格就是一个邮箱」。
# 刻意不叫 _EMAIL_RE：那个名字上面已经占用了，是非锚定的 findall 版本，
# 供 parse_material_xlsx 从一格里抠出多个邮箱用。同名会静默覆盖并搞坏素材提交。
_EMAIL_EXACT_RE = re.compile(r"^[^@\s,;、，；]+@[^@\s,;、，；]+\.[^@\s,;、，；]+$")

_KOL_DELIMS = ("\t", ",", ";", "|")


def _cell_text(value) -> str:
    """Excel 单元格转字符串。整数别带 .0 尾巴——
    粉丝数 180000 渲染成 180000.0 会出现在每一封信里。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def sniff_delimiter(text: str) -> str:
    """只看第一行。默认 tab——名单绝大多数是从表格里复制出来的。"""
    first = (text or "").strip().splitlines()[0] if (text or "").strip() else ""
    for d in _KOL_DELIMS:
        if first.count(d):
            return d
    return "\t"


def normalize_key(header: str, taken: set[str]) -> str:
    """列名 → 占位符名。中文原样保留（{{key}} 是纯字符串替换，不是 Jinja）。"""
    key = re.sub(r"[{}]", "", str(header or "")).strip()
    key = re.sub(r"\s+", "_", key)
    if not key:
        key = "col"
    if key in RESERVED_KEYS:      # 和内置变量撞名，改名并在页面上显示改后的名字
        key = f"{key}_1"
    base, n = key, 2
    while key in taken:
        key = f"{base}_{n}"
        n += 1
    taken.add(key)
    return key


def _build_kol_rows(headers: list[str], body_rows: list[list[str]], *,
                    first_lineno: int) -> dict:
    """名单解析的唯一核心。粘贴和 Excel 都走这里，
    所以两条路的报错文案、去重行为在构造上就一致，不靠自觉。

    headers 传空列表表示无表头，列名自动生成 col1..colN。
    """
    width = max([len(headers)] + [len(r) for r in body_rows] or [0])
    if headers:
        raw_headers = list(headers) + [""] * (width - len(headers))
    else:
        raw_headers = [f"col{i + 1}" for i in range(width)]

    # 邮箱列：先按别名认（复用素材那套词汇表），认不出就挑 @ 最多的那列
    email_idx = None
    for i, h in enumerate(raw_headers):
        if _match_column(h) == "recipient":
            email_idx = i
            break
    if email_idx is None:
        best, best_hits = None, 0
        for i in range(width):
            hits = sum(1 for r in body_rows
                       if i < len(r) and _EMAIL_EXACT_RE.match((r[i] or "").strip()))
            if hits > best_hits:
                best, best_hits = i, hits
        email_idx = best if best is not None else 0

    taken: set[str] = set()
    columns, key_of = [], {}
    for i, h in enumerate(raw_headers):
        if i == email_idx:
            continue
        key = normalize_key(h, taken)
        key_of[i] = key
        columns.append(key)

    rows, errors, duplicates, seen = [], [], [], set()
    for offset, cells in enumerate(body_rows):
        lineno = first_lineno + offset
        email = (cells[email_idx] if email_idx < len(cells) else "").strip()
        if not email and not any((c or "").strip() for c in cells):
            continue                                   # 整行空，静默跳过
        if not email:
            errors.append(f"第 {lineno} 行：没有邮箱，这一行不会发信")
            continue
        if not _EMAIL_EXACT_RE.match(email):
            errors.append(f"第 {lineno} 行：邮箱格式不对 —— '{email}'")
            continue
        email = email.lower()
        if email in seen:
            duplicates.append(email)
            continue
        seen.add(email)
        rows.append({"email": email,
                     "vars": {key_of[i]: (cells[i].strip() if i < len(cells) else "")
                              for i in key_of},
                     "lineno": lineno})

    return {"rows": rows, "columns": columns,
            "email_column": raw_headers[email_idx] if headers else None,
            "errors": errors, "duplicates": sorted(set(duplicates))}


def parse_kol_list(text: str) -> dict:
    """粘贴的名单。首行不含邮箱就当表头。"""
    lines = [ln for ln in (text or "").splitlines() if ln.strip()]
    if not lines:
        return {"rows": [], "columns": [], "email_column": None,
                "errors": [], "duplicates": []}
    delim = sniff_delimiter(text)

    def split(line):
        return [c.strip().strip('"').strip() for c in line.split(delim)]

    first = split(lines[0])
    has_header = not any(_EMAIL_EXACT_RE.match(c) for c in first)
    headers = first if has_header else []
    body = [split(ln) for ln in (lines[1:] if has_header else lines)]
    return _build_kol_rows(headers, body, first_lineno=2 if has_header else 1)


def _find_kol_header(rows: list[list[str]]) -> int | None:
    """表头行下标；返回 None 表示这张表没有表头。

    刻意不复用 _find_header()——它要求「至少认出 2 个已知列」，
    那是素材提交的承重逻辑；建联名单的列名可以全是自定义的，会被它一票否决。
    """
    for i, cells in enumerate(rows[:MAX_SCAN_ROWS]):
        if not any((c or "").strip() for c in cells):
            continue
        # 第一个非空行里出现了邮箱 → 它是数据行，这张表没表头
        return None if any(_EMAIL_EXACT_RE.match((c or "").strip()) for c in cells) else i
    return None


def parse_kol_xlsx(data: bytes) -> dict:
    """上传的 Excel 名单。和粘贴走同一个核心。"""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:
        raise ValueError(f"这个文件打不开，确认是 .xlsx 吗？（{exc}）") from None
    ws = wb.active
    grid = [[_cell_text(c) for c in row]
            for row in ws.iter_rows(values_only=True)]
    if not any(any(c for c in r) for r in grid):
        raise ValueError("表是空的")
    header_idx = _find_kol_header(grid)
    if header_idx is None:
        first_data = next((i for i, r in enumerate(grid) if any(c for c in r)), 0)
        out = _build_kol_rows([], grid[first_data:], first_lineno=first_data + 1)
    else:
        out = _build_kol_rows(grid[header_idx], grid[header_idx + 1:],
                              first_lineno=header_idx + 2)
    out["sheet"] = ws.title
    return out


KOL_TEMPLATE_HEADERS = ["邮箱", "达人名字", "平台", "粉丝数"]


def build_kol_template_xlsx() -> bytes:
    """名单模板。示例行放说明页，**不能放数据页**——
    留在数据页会被当成真实收件人发出去。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "KOL名单"
    ws.append(KOL_TEMPLATE_HEADERS)
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col, width in zip("ABCD", (34, 22, 16, 14)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    doc = wb.create_sheet("填写说明")
    for line in [
        ["怎么填"],
        [""],
        ["1. 第一行是表头，除「邮箱」外的每一列都会变成一个模板占位符。"],
        ["   比如「达人名字」这列，在邮件模板里写 {{达人名字}} 就会替换成该行的值。"],
        ["2. 列可以自由增删改名，中文列名没问题。想加「国家」「合作过没」就直接加一列。"],
        ["3. 「邮箱」这列必须有，可以叫 邮箱 / email / mail / 收件邮箱，位置不限。"],
        ["4. 邮箱重复的只保留第一条；格式不对的会被指出行号，不会静默丢掉。"],
        ["5. 列名如果和内置变量撞车（recipient、sender_name、index 等），"],
        ["   会自动改名成 recipient_1 这样，页面上会显示改后的名字。"],
        [""],
        ["内置占位符（不用在名单里写，系统自动填）"],
        ["{{sender_name}}", "发件账号的显示名"],
        ["{{sender_email}}", "发件邮箱地址"],
        ["{{recipient}}", "收件邮箱地址"],
        ["{{index}} / {{total}}", "第几封 / 本批共几封"],
        ["{{signature}}", "落款签名，不写则自动放正文末尾"],
        [""],
        ["示例（照这样填在「KOL名单」页）"],
        KOL_TEMPLATE_HEADERS,
        ["amy@example.com", "Amy", "TikTok", "125000"],
        ["ben@example.com", "Ben", "YouTube", "82000"],
    ]:
        doc.append(line)
    doc["A1"].font = Font(bold=True, size=13)
    doc["A11"].font = Font(bold=True)
    doc["A18"].font = Font(bold=True)
    doc.column_dimensions["A"].width = 40
    doc.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------- #
# 号码池 → 主体的固定映射
# --------------------------------------------------------------------------- #

COUNTRY_BY_HEADER = {"thailand": "TH", "philippines": "PH", "singapore": "SG",
                     "malaysia": "MY", "vietnam": "VN", "indonesia": "ID", "china": "CN"}


def seed_phone_pool_from_csv(path: str) -> dict:
    """把 CSV 号码池导进库。重复导入不会产生重复行（UNIQUE 约束 + ON CONFLICT）。"""
    if not os.path.exists(path):
        return {"inserted": 0, "error": f"找不到文件 {path}"}
    with open(path, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh))
    if not rows:
        return {"inserted": 0, "error": "文件是空的"}

    picked: dict[str, str] = {}     # 国家 -> 用哪个表头列（E.164 优先）
    for header in rows[0]:
        if not header:
            continue
        low = header.lower()
        country = next((c for k, c in COUNTRY_BY_HEADER.items() if k in low), None)
        if not country:
            continue
        if country not in picked or "e.164" in low or "e164" in low:
            picked[country] = header

    pairs = []
    for country, header in picked.items():
        for r in rows:
            phone = (r.get(header) or "").strip()
            if phone:
                pairs.append((country, phone))
    if not pairs:
        return {"inserted": 0, "total_in_csv": 0, "countries": sorted(picked)}

    # 一次批量插进去。逐条 INSERT 的话 600 个号码就是 600 次往返远程库，
    # 在 Render 上要跑一两分钟。
    with db.get_db_cursor(commit=True) as cur:
        execute_values(cur,
                       "INSERT INTO mb_phone_pool (country_code, phone) VALUES %s "
                       "ON CONFLICT (country_code, phone) DO NOTHING", pairs)
        inserted = max(0, cur.rowcount)
    return {"inserted": inserted, "total_in_csv": len(pairs),
            "countries": sorted(picked)}


def _entity_key(registration_id: str = "", company_name: str = "") -> str:
    """主体标识：信用代码优先，它才唯一；企业名称可能有全半角/空格差异。"""
    key = re.sub(r"\s+", "", (registration_id or "").strip().upper())
    return key or re.sub(r"\s+", "", (company_name or "").strip()).upper()


def assign_phone(*, registration_id="", company_name="", country_code="") -> dict:
    """给主体分配联系电话。已分配过就原样返回，不重新挑——号码每次变来变去，
    对账时同一家子公司会被当成两家。挑不出来就留空 + 说明，**不编号码**。"""
    key = _entity_key(registration_id, company_name)
    if not key:
        return {"phone": "", "reused": False,
                "reason": "没有企业名称也没有信用代码，无法确定是哪个主体"}

    existing = db.query_one("SELECT * FROM mb_entity_phones WHERE entity_key = %s", (key,))
    if existing:
        return {"phone": existing["phone"], "reused": True, "reason": ""}

    country = (country_code or "").upper()[:2]
    free = db.query_one("""
        SELECT p.phone FROM mb_phone_pool p
        WHERE p.country_code = %s
          AND NOT EXISTS (SELECT 1 FROM mb_entity_phones e
                          WHERE e.country_code = p.country_code AND e.phone = p.phone)
        ORDER BY p.id LIMIT 1
    """, (country,))
    if not free:
        have = db.query_all("SELECT DISTINCT country_code FROM mb_phone_pool ORDER BY 1")
        names = "、".join(r["country_code"] for r in have) or "（号码池是空的）"
        total = db.query_one("SELECT COUNT(*) AS c FROM mb_phone_pool WHERE country_code = %s",
                             (country,))
        reason = (f"{country} 的号码已全部分配完，需要补充号码池" if total and total["c"]
                  else f"号码池里没有 {country or '未知地区'} 的号码，现有：{names}")
        return {"phone": "", "reused": False, "reason": reason}

    db.execute("""
        INSERT INTO mb_entity_phones (entity_key, company_name, country_code, phone, source)
        VALUES (%s, %s, %s, %s, 'pool') ON CONFLICT (entity_key) DO NOTHING
    """, (key, (company_name or "").strip(), country, free["phone"]))
    return {"phone": free["phone"], "reused": False, "reason": ""}


def list_phone_assignments() -> list[dict]:
    return [dict(r) for r in db.query_all(
        "SELECT * FROM mb_entity_phones ORDER BY country_code, company_name, entity_key")]


def set_phone(*, registration_id="", company_name="", phone="", country_code="") -> None:
    key = _entity_key(registration_id, company_name)
    if not key:
        raise ValueError("要指定号码，至少得有企业名称或统一社会信用代码")
    db.execute("""
        INSERT INTO mb_entity_phones (entity_key, company_name, country_code, phone, source)
        VALUES (%s, %s, %s, %s, 'manual')
        ON CONFLICT (entity_key) DO UPDATE SET phone = EXCLUDED.phone,
            country_code = EXCLUDED.country_code, company_name = EXCLUDED.company_name,
            source = 'manual', assigned_at = NOW()
    """, (key, (company_name or "").strip(), (country_code or "").upper()[:2], phone.strip()))


# --------------------------------------------------------------------------- #
# 营业执照 OCR（千问 / DashScope，复用项目已有的 DASHSCOPE_API_KEY）
# --------------------------------------------------------------------------- #

OCR_ENDPOINT = "https://dashscope.aliyuncs.com/compatible-mode/v1/chat/completions"
OCR_MODEL = "qwen-vl-max"
OCR_TIMEOUT = 60

OCR_PROMPT = """# Role
You extract registry data from Southeast Asian company registration documents.

# Documents you will see
These are NOT Chinese 营业执照. Two formats dominate — find the labelled field, never
the surrounding prose:

**Thailand — DBD หนังสือรับรอง (Certificate of Incorporation)**
- company_name: the value of numbered item `1. ชื่อบริษัท` (e.g. "บริษัท จอร์นนี่ บางกอก จำกัด").
  ⚠️ The sentence beginning `ขอรับรองว่าบริษัทนี้ ได้จดทะเบียนเป็นนิติบุคคล...` is boilerplate
  legal text that appears on EVERY such certificate. It is NEVER the company name.
  If you are about to return a sentence containing `ขอรับรอง` or `ประมวลกฎหมาย`, you have
  grabbed the wrong block — go back and read item 1.
- registration_id: the digits after `ทะเบียนนิติบุคคลเลขที่` (13 digits, e.g. 0115566030652).
  Not the `ที่ E...` document number in the top-left corner.
- registered_address: the value of numbered item `5. สำนักงานแห่งใหญ่ ตั้งอยู่เลขที่ ...`
- detected_country_code: TH

**Indonesia — OSS IZIN USAHA / IUMK / NIB**
- company_name: value after `Nama Pemilik Usaha` (or `Nama Perusahaan`), e.g. "PT DAMAR SEJAHTERA"
- registration_id: value after `Nomor Induk Berusaha` (NIB)
- registered_address: value after `Alamat Rumah` (or `Alamat Usaha` / `Alamat Perusahaan`),
  including city and province
- detected_country_code: ID

**Other formats** (Philippines SEC/DTI, Singapore ACRA, Malaysia SSM, Vietnam, Chinese 营业执照):
find the field explicitly labelled as the entity name / registration number / registered address.

# Hard rules
1. Return the VALUE of a labelled field, never the label, never a full sentence, never
   boilerplate that would appear identically on another company's document.
2. NEVER invent, guess, complete, or "correct" a value. If a field is missing, illegible,
   redacted, or you are not certain, return "" for it. An empty string is always better than
   a plausible-looking wrong value — a wrong registration number is worse than no number.
3. Transcribe digits one by one. Do not normalise, reformat, or drop leading zeros.
4. company_name: keep the original script and the legal form
   (บริษัท ... จำกัด / PT ... / ... Pte Ltd / ... Sdn Bhd). Do not translate.
5. If the image is not a company registration document at all, return all fields as "".

# Output
Return ONLY a raw JSON object. No markdown fences, no commentary.

{"company_name": "string", "registration_id": "string",
 "registered_address": "string", "detected_country_code": "string"}"""

# 统一社会信用代码带校验位。OCR 认错一位数字时值看起来完全正常，
# 只有校验位能发现——而错的代码比空值危险得多。
_USCC_CHARS = "0123456789ABCDEFGHJKLMNPQRTUWXY"
_USCC_WEIGHTS = [1, 3, 9, 27, 19, 26, 16, 17, 20, 29, 25, 13, 8, 24, 10, 30, 28]


def check_uscc(code: str):
    """True/False = 校验通过与否；None = 不是 18 位 USCC（境外注册号没有统一校验规则）。"""
    code = (code or "").strip().upper()
    if len(code) != 18 or any(c not in _USCC_CHARS for c in code):
        return None
    total = sum(_USCC_CHARS.index(c) * w for c, w in zip(code[:17], _USCC_WEIGHTS))
    expect = 31 - (total % 31)
    return _USCC_CHARS[0 if expect == 31 else expect] == code[17]


class OCRError(RuntimeError):
    pass


def _ocr_call(image_bytes: bytes) -> dict:
    key = (os.environ.get("DASHSCOPE_API_KEY") or "").strip()
    if not key:
        raise OCRError("未配置 DASHSCOPE_API_KEY，无法做营业执照识别")
    _, mime = sniff_image(image_bytes)
    data_url = f"data:{mime};base64,{base64.b64encode(image_bytes).decode()}"
    payload = {"model": OCR_MODEL, "temperature": 0, "messages": [{"role": "user", "content": [
        {"type": "image_url", "image_url": {"url": data_url}},
        {"type": "text", "text": OCR_PROMPT}]}]}
    req = urllib.request.Request(
        OCR_ENDPOINT, data=json.dumps(payload).encode(),
        headers={"Authorization": f"Bearer {key}", "Content-Type": "application/json"},
        method="POST")
    try:
        with urllib.request.urlopen(req, timeout=OCR_TIMEOUT) as resp:
            body = json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", "replace")
        hint = ""
        if exc.code in (401, 403):
            hint = "　DASHSCOPE_API_KEY 不对或没权限"
        elif exc.code == 429:
            hint = "　被限流，稍后重试"
        raise OCRError(f"千问接口返回 {exc.code}：{raw[:300]}{hint}") from exc
    except urllib.error.URLError as exc:
        raise OCRError(f"连不上千问接口：{exc.reason}") from exc

    try:
        text = body["choices"][0]["message"]["content"]
        if isinstance(text, list):
            text = "".join(p.get("text", "") for p in text if isinstance(p, dict))
    except (KeyError, IndexError, TypeError) as exc:
        raise OCRError(f"接口返回的结构不认识：{json.dumps(body, ensure_ascii=False)[:300]}") from exc

    start, end = text.find("{"), text.rfind("}")
    if start < 0 or end < 0:
        raise OCRError(f"模型没有返回 JSON：{text[:200]}")
    try:
        parsed = json.loads(text[start:end + 1])
    except json.JSONDecodeError as exc:
        raise OCRError(f"模型返回的 JSON 解析失败：{text[:200]}") from exc

    return {"company_name": str(parsed.get("company_name") or "").strip(),
            "registration_id": str(parsed.get("registration_id") or "").strip(),
            "registered_address": str(parsed.get("registered_address") or "").strip(),
            "country_code": str(parsed.get("detected_country_code") or "").strip().upper()[:2],
            "_raw": json.dumps(parsed, ensure_ascii=False)}


def extract_license(image_bytes: bytes) -> dict:
    """识别一张营业执照。同一张图（按 sha256）只调一次接口。"""
    digest = hashlib.sha256(image_bytes).hexdigest()
    cached = db.query_one(
        "SELECT * FROM mb_license_ocr WHERE image_sha256 = %s AND status = 'done'", (digest,))
    if cached:
        return {k: cached[k] for k in
                ("company_name", "registration_id", "registered_address", "country_code")}
    try:
        result = _ocr_call(image_bytes)
    except OCRError as exc:
        db.execute("""
            INSERT INTO mb_license_ocr (image_sha256, model, status, error)
            VALUES (%s, %s, 'failed', %s)
            ON CONFLICT (image_sha256) DO UPDATE SET status = 'failed', error = EXCLUDED.error
        """, (digest, OCR_MODEL, str(exc)))
        raise
    db.execute("""
        INSERT INTO mb_license_ocr (image_sha256, company_name, registration_id,
            registered_address, country_code, raw_json, model, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, 'done')
        ON CONFLICT (image_sha256) DO UPDATE SET company_name = EXCLUDED.company_name,
            registration_id = EXCLUDED.registration_id,
            registered_address = EXCLUDED.registered_address,
            country_code = EXCLUDED.country_code, raw_json = EXCLUDED.raw_json,
            status = 'done', error = NULL
    """, (digest, result["company_name"], result["registration_id"],
          result["registered_address"], result["country_code"], result["_raw"], OCR_MODEL))
    result.pop("_raw", None)
    return result


# --------------------------------------------------------------------------- #
# 批次：建、读、发
# --------------------------------------------------------------------------- #

MIN_GAP_SECONDS = 3.0
MAX_GAP_SECONDS = 8.0


def _parse_ocr_report(raw) -> dict:
    """统一成 {total, notes}。老格式是裸的 notes 列表，兼容一下。"""
    if not raw:
        return {"total": 0, "notes": []}
    try:
        data = json.loads(raw)
    except (TypeError, ValueError):
        return {"total": 0, "notes": []}
    if isinstance(data, list):
        return {"total": len(data), "notes": data}
    return {"total": data.get("total") or 0, "notes": data.get("notes") or []}


def _serialize_item(row: dict, attachments: list[dict] | None = None) -> dict:
    """attachments 由调用方一次性批量查好传进来 —— 这个函数是在列表推导里
    逐行调的，一旦让它自己查库，500 行的批次每次 load_job 就是 500 趟往返。"""
    try:
        variables = json.loads(row["vars_json"]) if row.get("vars_json") else {}
    except (TypeError, ValueError):
        variables = {}
    return {
        "id": row["id"], "seq": row["seq"], "sender_account_id": row["sender_account_id"],
        "recipient": row.get("recipient") or "", "vars": variables,
        "image_id": row.get("image_id"), "image_name": row.get("image_name") or "",
        "image_url": f"/api/mail-blaster/images/{row['image_id']}" if row.get("image_id") else None,
        "from_display": row.get("from_display") or "",
        "signature_name": row.get("signature_name") or "",
        "subject": row.get("subject"), "status": row["status"], "error": row.get("error"),
        "smtp_response": row.get("smtp_response") or "", "message_id": row.get("message_id") or "",
        "sent_at": row["sent_at"].isoformat() if row.get("sent_at") else None,
        # 素材提交里附件是「素材的另一种形态」：有图或有附件才发得出去。
        # 建联批次的附件挂在 job 上，这里恒为空数组。
        "attachments": attachments or [],
    }


def load_job(job_id: int) -> dict:
    job = db.query_one("SELECT * FROM mb_jobs WHERE id = %s", (job_id,))
    if job is None:
        raise ValueError(f"批次 {job_id} 不存在")
    items = db.query_all("SELECT * FROM mb_items WHERE job_id = %s ORDER BY seq ASC", (job_id,))
    # 建联的附件挂在 job 上，mb_item_attachments 里恒无它的行 —— 别为它跑这条查询。
    # /status 是 2 秒一轮，500 行的 ANY(%s) 白跑一趟在 Render 上就是一秒。
    att_map = ({} if (job.get("mode") or "material") == "outreach"
               else list_item_attachments_map([r["id"] for r in items]))
    counts = {"total": len(items), "sent": 0, "failed": 0, "pending": 0, "sending": 0, "skipped": 0}
    for it in items:
        counts[it["status"]] = counts.get(it["status"], 0) + 1
    return {
        "job": {"id": job["id"], "recipient": job["recipient"], "status": job["status"],
                "mode": job.get("mode") or "material",
                "sender_account_id": job.get("sender_account_id"),
                "paused_reason": job["paused_reason"], "task_id": job["task_id"],
                "ocr_status": job.get("ocr_status") or "none",
                "ocr_report": _parse_ocr_report(job.get("ocr_report")),
                "replace_domain_enabled": bool(job.get("replace_domain_enabled")),
                "replacement_domain": job.get("replacement_domain") or "",
                "subject_tpl": job["subject_tpl"], "body_tpl": job["body_tpl"],
                "signature_tpl": job["signature_tpl"]},
        "items": [_serialize_item(dict(r), att_map.get(r["id"])) for r in items],
        "counts": counts,
        # 这个是 job 级的（建联用）。素材批次恒为空，它的附件在每个 item 里。
        "attachments": list_job_attachments(job_id),
    }


def sync_job(job_id: int, data: dict) -> None:
    """把前端改过的收件人 / 模板 / 配对写回。

    只写 payload 里**出现过**的键：无条件覆盖的话，一个空 body 的请求
    就能把整批的模板和收件人清成空串。
    """
    cleaners = {"recipient": lambda v: (v or "").strip(),
                "subject_tpl": lambda v: v or "", "body_tpl": lambda v: v or "",
                "signature_tpl": lambda v: v or "",
                "replace_domain_enabled": lambda v: bool(v),
                "replacement_domain": lambda v: (v or "").strip().lower().rstrip(".")}
    sets, args = [], {}
    for key, clean in cleaners.items():
        if key in data:
            sets.append(f"{key} = %({key})s")
            args[key] = clean(data[key])
    if sets:
        if args.get("replace_domain_enabled"):
            args["replacement_domain"] = normalize_replacement_domain(args.get("replacement_domain") or "")
            if not args["replacement_domain"]:
                raise ValueError("已勾选替换域名，请填写目标域名")
        elif "replacement_domain" in args:
            args["replacement_domain"] = normalize_replacement_domain(args["replacement_domain"])
        args["jid"] = job_id
        db.execute(f"UPDATE mb_jobs SET {', '.join(sets)} WHERE id = %(jid)s", args)

    item_attachments: dict[int, list[dict]] = {}
    for item in data.get("items") or []:
        fields = ["sender_account_id = %(acc)s", "from_display = %(disp)s",
                  "signature_name = %(sig)s"]
        params = {"acc": item.get("sender_account_id") or None,
                  "disp": (item.get("from_display") or "").strip(),
                  "sig": (item.get("signature_name") or "").strip(),
                  "iid": item.get("id"), "jid": job_id}
        if "vars" in item:
            fields.append("vars_json = %(vars)s")
            params["vars"] = json.dumps(item.get("vars") or {}, ensure_ascii=False)
        if "recipient" in item:
            fields.append("recipient = %(rcpt)s")
            params["rcpt"] = (item.get("recipient") or "").strip().lower()
        # 附件跟 vars / recipient 一样「出现才写」：不带这个键的请求
        # （建联页的 templatePayload、以及任何只改模板的调用）不能把附件清空。
        # 这里只收集，循环外一次写完 —— 逐行写会把 sync_job 的往返次数翻三倍。
        if "attachments" in item and item.get("id") is not None:
            item_attachments[int(item["id"])] = item.get("attachments") or []
        db.execute(
            f"UPDATE mb_items SET {', '.join(fields)} "
            "WHERE id = %(iid)s AND job_id = %(jid)s AND status <> 'sent'", params)

    sync_item_attachments(job_id, item_attachments)


def _mark_failed(item_id: int, message: str) -> None:
    db.execute("UPDATE mb_items SET status = 'failed', error = %s WHERE id = %s",
               (message, item_id))


def _mark_skipped(item_id: int, reason: str) -> None:
    """跳过 ≠ 失败：不是发送出了问题，是我们主动决定不发。别混进「失败」计数，
    否则用户会去点重发。"""
    db.execute("UPDATE mb_items SET status = 'skipped', error = %s WHERE id = %s",
               (reason, item_id))


def _mark_sent(item_id: int, result: dict) -> None:
    db.execute("""
        UPDATE mb_items SET status = 'sent', error = NULL, sent_at = NOW(), subject = %s,
            body_html = %s, smtp_response = %s, message_id = %s WHERE id = %s
    """, (result["subject"], result["body_html"], result.get("smtp_response") or "",
          result.get("message_id") or "", item_id))


def send_item(job_id: int, item_id: int, attachments: list[dict] | None = None) -> str:
    """发一封。返回结果状态字符串，异常都记进该行不往外抛。

    attachments 传 None 就自己去库里取，按 mode 分叉：建联取批次级的一组，
    素材取这一行自己的。建联整批发送时由 run_job 取一次传进来 ——
    15MB 的附件乘 500 个收件人，逐封重取就是 7GB 的库往返；素材按行挂，
    各是各的，只能逐行取（每行 ≤15MB，取完就丢，内存有界）。

    投递与记账刻意分成两段：SMTP 一旦收下这封信，它就已经在对方队列里、不可撤回，
    此后任何写库失败都不能把这行标成 failed —— 用户看到红色会去点重发，
    收件人就收到两封。宁可让状态停在 sending 由人来判断。
    """
    row = db.query_one("SELECT * FROM mb_items WHERE id = %s AND job_id = %s", (item_id, job_id))
    job = db.query_one("SELECT * FROM mb_jobs WHERE id = %s", (job_id,))
    if row is None or job is None:
        return "missing"
    total = (db.query_one("SELECT COUNT(*) AS c FROM mb_items WHERE job_id = %s",
                          (job_id,)) or {}).get("c", 0)
    item = dict(row)
    recipient = (item["recipient"] or job["recipient"] or "").strip()
    is_outreach = (job.get("mode") or "material") == "outreach"

    # 抑制名单闸门。放在标 sending 之前：命中是「主动决定不发」，
    # 不是发送失败，所以走 _mark_skipped 而不是 _mark_failed。
    # ⚠️ 必须判 is not None —— 原因可能是空字符串（手动拉黑常常不写理由）。
    if is_outreach:
        reason = is_suppressed(recipient)
        if reason is not None:
            _mark_skipped(item_id, f"在抑制名单里，未发送。原因：{reason or '未填'}")
            return "skipped"

    # 素材提交的可发性闸门：附件是「素材的另一种形态」，图和附件至少要有一个。
    # 两个都没有 = 没有可提交的内容，这是主动决定不发而不是发送失败，
    # 所以同样走 _mark_skipped（不进失败计数、不画「重发」按钮）。
    # 只数不读：判断用不着把最多 15MB 的附件拽进内存。
    if not is_outreach and not item.get("image_id"):
        has_attachment = (len(attachments) > 0 if attachments is not None
                          else count_item_attachments(item_id) > 0)
        if not has_attachment:
            _mark_skipped(item_id, f"{NO_CONTENT_SKIP_MARK} 这一行既没有图片也没有附件，未发送。"
                                   "在「逐封确认」表里点 📎 传一个文件，再点一次发送即可。")
            return "skipped"

    db.execute("UPDATE mb_items SET status = 'sending', error = NULL WHERE id = %s", (item_id,))

    # 第一段：投递。这里失败才算这封信没发出去。
    try:
        if "@" not in recipient:
            raise ValueError("收件邮箱没填或格式不对")
        account_id = item["sender_account_id"] or job.get("sender_account_id")
        account = get_account(account_id) if account_id else None
        if account is None:
            raise ValueError("还没选发件账号")

        blob = None
        if item.get("image_id"):
            loaded = load_image(item["image_id"])
            blob = loaded[0] if loaded else None

        try:
            extra_vars = json.loads(item["vars_json"]) if item["vars_json"] else {}
        except (TypeError, ValueError):
            extra_vars = {}

        auto_display, auto_signature = derive_names(account)
        replacement_domain = replacement_domain_for_job(job)
        result = send_one_email(
            account=account, to_email=recipient,
            subject_tpl=job["subject_tpl"], body_tpl=job["body_tpl"],
            signature_tpl=job["signature_tpl"],
            from_display=item["from_display"] or auto_display,
            signature_name=item["signature_name"] or auto_signature,
            image_bytes=blob, image_name=item["image_name"] or "",
            index=item["seq"] + 1, total=total, extra_vars=extra_vars,
            replacement_domain=replacement_domain,
            unsubscribe_text=DEFAULT_UNSUBSCRIBE_TEXT if is_outreach else "",
            # 建联整批共用一组，素材按行挂，各是各的
            attachments=(attachments if attachments is not None
                         else (load_job_attachments(job_id) if is_outreach
                               else load_item_attachments(item_id))))
    except Exception as exc:
        _mark_failed(item_id, friendly_smtp_error(exc))
        return "failed"

    # 第二段：记账。信已经发出去了，这里再失败也绝不能标 failed。
    try:
        _mark_sent(item_id, result)
        record_send(recipient=recipient, material_id=extra_vars.get("id") or "",
                    material_name=extra_vars.get("name") or "",
                    sender_account_id=account["id"],
                    sender_email=result.get("sender_email") or account["email"],
                    job_id=job_id, item_id=item_id, subject=result.get("subject") or "",
                    message_id=result.get("message_id") or "",
                    mode="outreach" if is_outreach else "material")
    except Exception:
        logger.exception("mail-blaster: 第 %s 封已投递但记账失败", item_id)
    return "sent"


# 因配额/窗口而暂停的行，error 里带这个前缀。
# 它们和「抑制名单命中」不同：那是永久决定，这只是「今天先不发」，
# 所以重跑时要能捡回来 —— 否则暂停提示里那句「明天再点一次发送即可继续」是骗人的。
PAUSE_SKIP_MARK = "⏸"

# 素材行「既没图也没附件」时 error 里带这个前缀。同样是可恢复的跳过：
# 用户在第 3 步补个附件再点发送就该捡回来，所以 run_job 要认得它。
NO_CONTENT_SKIP_MARK = "📎"


def _pause_job(job_id: int, reason: str, remaining_ids: list[int], skip_reason: str = "") -> None:
    """整批暂停：剩余的标 skipped 并把原因写在批次上。

    状态仍然置 'done' 而不是新增一个 'paused' —— 前端的轮询终止条件和
    暂停提示框都是按 done + paused_reason 判断的，改状态会让页面一直转圈。
    """
    if remaining_ids:
        db.execute("UPDATE mb_items SET status = 'skipped', error = %s WHERE id = ANY(%s)",
                   (skip_reason or reason, list(remaining_ids)))
    db.execute("UPDATE mb_jobs SET status = 'done', finished_at = NOW(), paused_reason = %s "
               "WHERE id = %s", (reason, job_id))


def run_job(job_id: int, progress=None) -> dict:
    """整批发送。由 worker 调用——web 是 free 套餐会休眠，后台线程会被连同进程杀掉。"""
    import random
    import time

    job = db.query_one("SELECT mode, sender_account_id FROM mb_jobs WHERE id = %s", (job_id,))
    is_outreach = ((job or {}).get("mode") or "material") == "outreach"
    db.execute("UPDATE mb_jobs SET status = 'sending', paused_reason = NULL WHERE id = %s",
               (job_id,))
    pending = [r["id"] for r in db.query_all(
        "SELECT id FROM mb_items WHERE job_id = %s AND status NOT IN ('sent','skipped') "
        "ORDER BY seq ASC", (job_id,))]
    # 上一轮「暂时性」跳过的行捡回来重跑。
    # 建联是因配额/窗口暂停的（⏸），素材是缺图缺附件的（📎，用户补完附件再点发送就该发）。
    # 抑制名单命中的两边都捡不回来 —— 那是永久决定，没有前缀。
    resume_mark = PAUSE_SKIP_MARK if is_outreach else NO_CONTENT_SKIP_MARK
    resume = [r["id"] for r in db.query_all(
        "SELECT id FROM mb_items WHERE job_id = %s AND status = 'skipped' "
        "AND error LIKE %s ORDER BY seq ASC", (job_id, f"{resume_mark}%"))]
    if resume:
        db.execute("UPDATE mb_items SET status = 'pending', error = NULL "
                   "WHERE id = ANY(%s)", (resume,))
        pending = [r["id"] for r in db.query_all(
            "SELECT id FROM mb_items WHERE job_id = %s AND status NOT IN ('sent','skipped') "
            "ORDER BY seq ASC", (job_id,))]

    # 建联的节奏刻意慢一个数量级：冷启动收件人互不相识，
    # 密集投递最容易触发频控和垃圾判定。
    lo, hi = ((OUTREACH_MIN_GAP_SECONDS, OUTREACH_MAX_GAP_SECONDS) if is_outreach
              else (MIN_GAP_SECONDS, MAX_GAP_SECONDS))
    account = (get_account((job or {}).get("sender_account_id"))
               if is_outreach and (job or {}).get("sender_account_id") else None)
    # None 表示这个账号不限量，quota_state 会据此跳过配额检查
    daily_limit = serialize_account(account)["daily_limit"] if account else None

    sent = failed = skipped = 0
    paused = ""
    # 建联整批共用同一组附件，取一次带着走：15MB × 500 个收件人逐封重取就是 7GB 库往返。
    # 素材是按行挂的，各是各的，只能让 send_item 逐行取（取完就丢，内存有界）。
    attachments = load_job_attachments(job_id) if is_outreach else None
    for i, item_id in enumerate(pending):
        # 窗口和配额每轮都查。20–45 秒间隔下 200 封要跑两小时，
        # 只在开头查一次会直接冲过 22:00；配额也要重查，
        # 这样同一个账号上的并发批次才会互相尊重。
        if is_outreach:
            paused = outside_send_window()
            if paused:
                _pause_job(job_id, paused, pending[i:],
                           f"{PAUSE_SKIP_MARK} 不在发送窗口内，未发送")
                break
            if account:
                q = quota_state(account["id"], daily_limit)
                if q["remaining"] is not None and q["remaining"] <= 0:
                    paused = (f"{account['email']} 今日配额已用满"
                              f"（{q['used']}/{q['limit']}），剩余 {len(pending) - i} 封未发。"
                              f"明天再点一次发送即可继续。")
                    _pause_job(job_id, paused, pending[i:],
                               f"{PAUSE_SKIP_MARK} 今日配额已用满"
                               f"（{q['used']}/{q['limit']}），未发送")
                    break

        outcome = send_item(job_id, item_id, attachments=attachments)
        if outcome == "sent":
            sent += 1
        elif outcome == "failed":
            failed += 1
        elif outcome == "skipped":
            skipped += 1
        if progress:
            progress(f"已处理 {i + 1}/{len(pending)}　成功 {sent}　失败 {failed}")
        if i < len(pending) - 1:
            time.sleep(random.uniform(lo, hi))

    if not paused:
        db.execute("UPDATE mb_jobs SET status = 'done', finished_at = NOW() WHERE id = %s",
                   (job_id,))
    return {"total": len(pending), "sent": sent, "failed": failed,
            "skipped": skipped + (len(pending) - sent - failed - skipped if paused else 0),
            "paused_reason": paused}


def reset_stuck_items() -> int:
    """复位上次进程被杀时卡在 sending 的行。

    刻意标成 failed 而不是 pending：那封信可能**已经送达**，自动重发会让收件人收到两封。
    """
    n = db.execute("""
        UPDATE mb_items SET status = 'failed',
            error = '进程在发送途中退出，这封信是否已投递未知。请先到发件箱确认再决定是否重发——直接重发可能让对方收到两封。'
        WHERE status = 'sending'
    """)
    db.execute("UPDATE mb_jobs SET status = 'done', finished_at = NOW() WHERE status = 'sending'")
    return n or 0


def build_previews(job_id: int) -> list[dict]:
    """逐封渲染预览。图片走 /api/mail-blaster/images/<id>，不用把 BYTEA 塞进 JSON。"""
    state = load_job(job_id)
    job, items = state["job"], state["items"]
    previews = []
    replacement_domain = replacement_domain_for_job(job)
    # 建联的退订块必须在预览里就看得见，否则这道护栏要等信发出去才现身
    unsubscribe_text = (DEFAULT_UNSUBSCRIBE_TEXT
                        if (job.get("mode") or "material") == "outreach" else "")
    is_material = (job.get("mode") or "material") != "outreach"
    for idx, item in enumerate(items, 1):
        account_id = item["sender_account_id"] or job.get("sender_account_id")
        account = get_account(account_id) if account_id else None
        if account is None:
            previews.append({**item, "error": "还没选发件账号"})
            continue
        recipient = item["recipient"] or job["recipient"]
        display = item["from_display"] or derive_names(account)[0]
        signature = item["signature_name"] or derive_names(account)[1]
        sender_email = replace_sender_domain(account["email"], replacement_domain)
        subject, html = compose(
            subject_tpl=job["subject_tpl"], body_tpl=job["body_tpl"],
            signature_tpl=job["signature_tpl"], sender_name=display,
            sender_email=sender_email, signature_name=signature,
            recipient=recipient, index=idx, total=len(items),
            image_name=item["image_name"], image_src=item["image_url"] or "",
            extra_vars=item["vars"] or None, unsubscribe_text=unsubscribe_text)
        html = rewrite_body_domains(html, replacement_domain)
        previews.append({**item, "from_line": f"{display} <{sender_email}>",
                         "to_line": recipient, "subject": subject, "html": html,
                         # 每封预览都带上附件清单：预览是「这封信长什么样」，
                         # 附件漏在预览外面，人就是要等发出去才发现忘了传。
                         # 建联整批共用一组，素材按行挂各是各的。
                         "attachments": (item["attachments"] if is_material
                                         else state["attachments"])})
    return previews


def _fill_from_license(row: dict) -> dict | None:
    """name / id / number 有空缺时，拿这一行的图当营业执照识别补上。

    识别不出就留空，绝不猜。电话来自公司自有号码池，按主体做固定映射
    ——同一家子公司永远是同一个号，不是每次随机取。
    """
    variables = row["vars"]
    missing = [f for f in ("name", "id", "number") if not (variables.get(f) or "").strip()]
    if not missing or not row.get("image_bytes"):
        return None

    note = {"row": row["excel_row"], "filled": {}, "warn": "", "error": ""}
    try:
        lic = extract_license(row["image_bytes"])
    except OCRError as exc:
        note["error"] = str(exc).split("\n")[0][:160]
        return note

    if "name" in missing and lic["company_name"]:
        variables["name"] = lic["company_name"]
        note["filled"]["name"] = lic["company_name"]
    if "id" in missing and lic["registration_id"]:
        variables["id"] = lic["registration_id"]
        note["filled"]["id"] = lic["registration_id"]
        # OCR 认错一位数字时值看起来完全正常，校验位是唯一能自动发现的手段
        valid = check_uscc(lic["registration_id"])
        if valid is False:
            note["warn"] = (f"信用代码 {lic['registration_id']} 校验位不对，"
                            "大概率识别错了一位，请对着执照核一遍")
        elif valid is None:
            note["warn"] = "境外注册号没有校验规则，识别结果请人工核对"

    if "number" in missing:
        picked = assign_phone(registration_id=variables.get("id") or lic["registration_id"],
                              company_name=variables.get("name") or lic["company_name"],
                              country_code=lic["country_code"])
        if picked["phone"]:
            variables["number"] = picked["phone"]
            note["filled"]["number"] = picked["phone"] + (
                "（沿用已分配）" if picked["reused"] else "（新分配）")
        elif picked["reason"]:
            note["error"] = (note["error"] + " " + picked["reason"]).strip()

    if not lic["company_name"] and not lic["registration_id"] and not note["error"]:
        note["error"] = "这张图没识别出营业执照信息，请手填"
    return note


def create_job_from_excel(*, file_bytes: bytes, fallback_recipient: str = "",
                          subject_tpl: str = "", body_tpl: str = "", signature_tpl: str = "",
                          user_id=None, run_ocr: bool = True,
                          replace_domain_enabled: bool = False,
                          replacement_domain: str = "") -> dict:
    """从 Excel 建批次。已发过的行和缺邮箱的行不入队，都逐条点名。

    缺图的行**照常入队**：附件是素材的另一种形态，用户会在第 3 步给它挂文件。
    图和附件都没有的判断留到 send_item —— 建批次这一刻还没到挂附件那一步。"""
    # 发件账号这一关先过。解析 Excel 要抽图、要走 OCR 判断，几十行就是几秒钟，
    # 跑完才告诉人「没有能用的号」是白等——何况这两个错都跟 Excel 内容无关。
    replacement_domain = normalize_replacement_domain(replacement_domain)
    if replace_domain_enabled and not replacement_domain:
        raise ValueError("已勾选替换域名，请填写目标域名")
    pool = list_accounts(only_sendable=True, purpose="material")
    if replace_domain_enabled:
        # 收窄到能替换发件域名的服务商。不收窄的话，池子里的 Outlook 号会被
        # 正常配上，然后一封封 SendAsDenied——错误要等整批跑完才看得见。
        pool = [a for a in pool if a["provider"] in DOMAIN_REPLACEMENT_PROVIDERS]
        if not pool:
            raise ValueError(
                "勾了替换域名，但账号池里没有能替换发件域名的账号。"
                "目前只有网易 163 的号支持（微软/Gmail 会拒绝代发），"
                "请先在发件账号池里加一个 163 账号并测试通过。")

    parsed = parse_material_xlsx(file_bytes)

    history_hits = already_sent(
        [(r.get("recipient") or fallback_recipient, (r["vars"].get("id") or ""))
         for r in parsed["rows"] if not r.get("skip_reason")])

    usable, skipped = [], []
    for row in parsed["rows"]:
        target = (row.get("recipient") or fallback_recipient or "").strip().lower()
        material_id = (row["vars"].get("id") or "").strip()
        if row.get("skip_reason"):
            skipped.append({"row": row["excel_row"], "reason": row["skip_reason"]})
            continue
        if (target, material_id) in history_hits:
            skipped.append({"row": row["excel_row"],
                            "reason": f"库里已有发送记录：{material_id} → {target}"})
            continue
        # 没图的行照样入队 —— 附件是素材的另一种形态，用户会在第 3 步给它挂个 PDF。
        # 真正发不出去的判断留给 send_item，那时才知道附件到底挂上没有。
        if not target or "@" not in target:
            skipped.append({"row": row["excel_row"], "reason": "没有收件邮箱"})
            continue
        row["_target"] = target
        usable.append(row)

    if not usable:
        detail = "\n".join(f"第 {s['row']} 行：{s['reason']}" for s in skipped[:6])
        raise ValueError("这个 Excel 里没有需要发送的行。\n"
                         + (detail or "每行需要一个收件邮箱。内容可以是嵌在单元格里的图片，"
                                      "也可以在第 3 步给这一行挂附件。")
                         + ("\n" + "\n".join(parsed["errors"][:4]) if parsed["errors"] else ""))

    # OCR 不在这里跑。web 是 free 套餐且 gunicorn workers=1/threads=1，
    # 一张图识别 3–8 秒，20 行就能把整个工作台堵两分钟。
    # 这里只判断需不需要识别，真正的识别交给 worker。
    needs_ocr = bool(run_ocr) and any(
        r.get("image_bytes") and any(
            not (r["vars"].get(f) or "").strip() for f in ("name", "id", "number"))
        for r in usable)
    job_id = db.execute_and_fetch_id("""
        INSERT INTO mb_jobs (user_id, recipient, subject_tpl, body_tpl, signature_tpl,
            ocr_status, replace_domain_enabled, replacement_domain)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s) RETURNING id
    """, (user_id, fallback_recipient, subject_tpl or DEFAULT_SUBJECT_TPL,
          body_tpl or DEFAULT_BODY_TPL, signature_tpl or DEFAULT_SIGNATURE_TPL,
          'pending' if needs_ocr else 'none', bool(replace_domain_enabled), replacement_domain))

    assignments = assign_accounts(pool, [r["_target"] for r in usable])
    cooldown_notes = []
    for seq, (row, pick) in enumerate(zip(usable, assignments)):
        # 无图行的 image_bytes 是 None，不能进 store_image（它要算 sha256）。
        # image_id 留空，这一行等着用户在第 3 步挂附件。
        image_id = store_image(row["image_bytes"]) if row["image_bytes"] else None
        account = pick["account"]
        if pick["note"]:
            cooldown_notes.append(f"第 {row['excel_row']} 行：{pick['note']}")
        db.execute("""
            INSERT INTO mb_items (job_id, seq, sender_account_id, recipient, image_id,
                image_name, vars_json, from_display, signature_name)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (job_id, seq, account["id"] if account else None, row["_target"], image_id,
              row["image_name"], json.dumps(row["vars"], ensure_ascii=False),
              account["effective_display_name"] if account else "",
              account["effective_signature_name"] if account else ""))

    payload = load_job(job_id)
    payload["pool_size"] = len(pool)
    imported_rows = {r["excel_row"] for r in usable}
    payload["excel"] = {
        "sheet": parsed["sheet"], "header_row": parsed["header_row"],
        "fields": parsed["fields"], "has_recipient": parsed["has_recipient"],
        "has_status": parsed["has_status"], "total_rows": len(parsed["rows"]),
        "imported": len(usable), "skipped": skipped, "cooldown": cooldown_notes,
        "cooldown_days": COOLDOWN_DAYS, "needs_ocr": needs_ocr, "errors": parsed["errors"],
        # 缺图的行也入队了，得让页面把它们点出来 —— 不提示的话用户不会知道
        # 这些行还等着挂附件，会直接点发送然后收获一批「已跳过」。
        # 只留真正入队的：缺图**又**缺邮箱的行已经进 skipped 了，
        # 再在这里说一遍「照常导入」，数字对不上，人会去表里找不存在的行。
        "notices": [n["text"] for n in parsed["notices"] if n["row"] in imported_rows],
        "no_image": sum(1 for r in usable if not r["image_bytes"]),
    }
    return payload


def create_outreach_job(*, sender_account_id: int, rows: list[dict],
                        subject_tpl: str = "", body_tpl: str = "", signature_tpl: str = "",
                        attachments: list[dict] | None = None, user_id=None) -> dict:
    """从解析好的名单建一个建联批次。

    和素材提交的轴反过来：整批共用一个发件账号，每一行是一个收件人 + 一组变量。
    """
    account = get_account(int(sender_account_id)) if sender_account_id else None
    if account is None:
        raise ValueError("请先选一个发件账号")
    acc = serialize_account(account)
    if not usable_account(acc):
        raise ValueError(f"{acc['email']} 还不可用：需要「已启用 + 测试通过 + 有凭据」")

    if not rows:
        raise ValueError("名单是空的")

    # 服务端再过一遍抑制名单。解析期那次挡不住直接 POST，
    # 而且从解析到建批次之间可能有人刚把某个地址拉黑。
    dropped = filter_suppressed([r.get("email") for r in rows])
    usable = [r for r in rows if (r.get("email") or "").strip().lower() not in dropped]
    if not usable:
        raise ValueError("名单里的地址全都在抑制名单里，没有可发送的行")

    job_id = db.execute_and_fetch_id("""
        INSERT INTO mb_jobs (user_id, mode, sender_account_id, recipient,
            subject_tpl, body_tpl, signature_tpl)
        VALUES (%s, 'outreach', %s, '', %s, %s, %s) RETURNING id
    """, (user_id, account["id"],
          subject_tpl or DEFAULT_OUTREACH_SUBJECT_TPL,
          body_tpl or DEFAULT_OUTREACH_BODY_TPL,
          signature_tpl or DEFAULT_OUTREACH_SIGNATURE_TPL))

    # 用 execute_values 批量插。素材那边逐行 execute 是因为每行都要先 store_image，
    # 这里没有这个约束，500 行名单逐行插就是 500 次 Render 往返。
    with db.get_db_cursor(commit=True) as cur:
        execute_values(cur, """
            INSERT INTO mb_items (job_id, seq, sender_account_id, recipient,
                vars_json, from_display, signature_name)
            VALUES %s
        """, [(job_id, seq, account["id"], (r["email"] or "").strip().lower(),
               json.dumps(r.get("vars") or {}, ensure_ascii=False),
               acc["effective_display_name"], acc["effective_signature_name"])
              for seq, r in enumerate(usable)])

    # 附件要有 job_id 才挂得上，所以只能放在建批次之后。这里抛错会留下一个
    # 没挂上附件的 draft 批次 —— 它不会被发送（发送要另外点一次），
    # 而前端每次预览/发送都重建批次，所以下一次点会得到一个干净的新批次。
    set_job_attachments(job_id, attachments or [])

    payload = load_job(job_id)
    payload["quota"] = quota_state(account["id"], acc["daily_limit"])
    payload["account"] = acc
    payload["list"] = {
        "total_rows": len(rows), "imported": len(usable),
        "suppressed": sorted(dropped),
    }
    return payload


def run_ocr_for_job(job_id: int, progress=None) -> dict:
    """在 worker 里给整批补 name / id / number。

    和 _fill_from_license 的区别是这里从库里读 item（图片在 BYTEA 里），
    识别完写回 vars_json，报告落在 mb_jobs.ocr_report 供页面轮询。
    单行失败不影响其它行。
    """
    items = db.query_all(
        "SELECT id, seq, image_id, vars_json FROM mb_items WHERE job_id = %s ORDER BY seq",
        (job_id,))
    # 只有真正需要识别的行才计入总数，否则进度条会卡在中间不动
    todo_total = sum(
        1 for r in items
        if r["image_id"] and any(
            not ((json.loads(r["vars_json"]) if r["vars_json"] else {}).get(f) or "").strip()
            for f in ("name", "id", "number")))
    db.execute("UPDATE mb_jobs SET ocr_status = 'running', ocr_report = %s WHERE id = %s",
               (json.dumps({"total": todo_total, "notes": []}, ensure_ascii=False), job_id))

    notes = []
    for n, row in enumerate(items, 1):
        try:
            variables = json.loads(row["vars_json"]) if row["vars_json"] else {}
        except (TypeError, ValueError):
            variables = {}
        missing = [f for f in ("name", "id", "number") if not (variables.get(f) or "").strip()]
        if not missing or not row["image_id"]:
            continue

        note = {"row": row["seq"] + 1, "item_id": row["id"], "filled": {}, "warn": "", "error": ""}
        try:
            loaded = load_image(row["image_id"])
            if loaded is None:
                raise ValueError("图片读不出来")
            lic = extract_license(loaded[0])

            if "name" in missing and lic["company_name"]:
                variables["name"] = lic["company_name"]
                note["filled"]["name"] = lic["company_name"]
            if "id" in missing and lic["registration_id"]:
                variables["id"] = lic["registration_id"]
                note["filled"]["id"] = lic["registration_id"]
                # OCR 认错一位数字时值看起来完全正常，校验位是唯一能自动发现的手段
                valid = check_uscc(lic["registration_id"])
                if valid is False:
                    note["warn"] = (f"信用代码 {lic['registration_id']} 校验位不对，"
                                    "大概率识别错了一位，请对着执照核一遍")
                elif valid is None:
                    note["warn"] = "境外注册号没有校验规则，识别结果请人工核对"

            if "number" in missing:
                picked = assign_phone(
                    registration_id=variables.get("id") or lic["registration_id"],
                    company_name=variables.get("name") or lic["company_name"],
                    country_code=lic["country_code"])
                if picked["phone"]:
                    variables["number"] = picked["phone"]
                    note["filled"]["number"] = picked["phone"] + (
                        "（沿用已分配）" if picked["reused"] else "（新分配）")
                elif picked["reason"]:
                    note["error"] = picked["reason"]

            if not lic["company_name"] and not lic["registration_id"] and not note["error"]:
                note["error"] = "这张图没识别出营业执照信息，请手填"

            if note["filled"]:
                db.execute("UPDATE mb_items SET vars_json = %s WHERE id = %s",
                           (json.dumps(variables, ensure_ascii=False), row["id"]))
        except Exception as exc:                       # noqa: BLE001 单行失败不毁整批
            note["error"] = f"{type(exc).__name__}: {exc}"[:160]

        notes.append(note)
        # 每行立刻落盘，前端轮询才能一行一行地看到结果填进去
        db.execute("UPDATE mb_jobs SET ocr_report = %s WHERE id = %s",
                   (json.dumps({"total": todo_total, "notes": notes}, ensure_ascii=False), job_id))
        if progress:
            progress(f"识别中 {len(notes)}/{todo_total}")

    db.execute("UPDATE mb_jobs SET ocr_status = 'done', ocr_report = %s WHERE id = %s",
               (json.dumps({"total": todo_total, "notes": notes}, ensure_ascii=False), job_id))
    return {"processed": len(notes),
            "filled": sum(1 for x in notes if x["filled"]),
            "failed": sum(1 for x in notes if x["error"])}


# --------------------------------------------------------------------------- #
# OAuth2 / XOAUTH2
#
# 微软已对很多 outlook.com 账号关闭「密码直连 SMTP」（报 535 5.7.139），
# 这类账号只能拿 refresh_token 现换 access_token 再用 SASL XOAUTH2 登录。
# 只用标准库，不引第三方依赖。
# --------------------------------------------------------------------------- #

MICROSOFT_SCOPE = "https://outlook.office.com/SMTP.Send offline_access"
GOOGLE_SCOPE = "https://mail.google.com/"
GOOGLE_TOKEN_URL = "https://oauth2.googleapis.com/token"
OAUTH_TIMEOUT = 20
TOKEN_EXPIRY_MARGIN = 300      # 提前 5 分钟当过期，避免拿到手就正好失效

_token_cache: dict[tuple[str, str], tuple[str, float]] = {}
_token_lock = __import__("threading").Lock()


class OAuthError(RuntimeError):
    """换令牌失败。消息里带上端点返回的原文，不吞错。"""


def _oauth_flavor(provider: str | None) -> str:
    """按服务商决定用哪家的令牌端点。默认微软——这个格式就是 Outlook 号在用。"""
    return "google" if (provider or "") == "gmail" else "microsoft"


def _token_url(flavor: str) -> str:
    if flavor == "google":
        return GOOGLE_TOKEN_URL
    tenant = (os.environ.get("MAIL_BLASTER_MS_TENANT") or "common").strip() or "common"
    return f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"


def _explain_oauth(body: str) -> str:
    low = body.lower()
    if "invalid_grant" in low:
        return "refresh_token 已失效或被撤销，需要重新授权拿一个新的"
    if "invalid_client" in low:
        return "client_id 不对，或该应用要求 client_secret"
    if "unauthorized_client" in low:
        return "该 client_id 没有被授权使用这个 scope（SMTP.Send）"
    if "invalid_scope" in low:
        return "scope 不被接受，检查应用的 API 权限配置"
    return ""


def _request_access_token(flavor: str, client_id: str, refresh_token: str) -> tuple[str, float]:
    import time
    form = {"grant_type": "refresh_token", "client_id": client_id,
            "refresh_token": refresh_token,
            "scope": GOOGLE_SCOPE if flavor == "google" else MICROSOFT_SCOPE}
    secret = (os.environ.get("MAIL_BLASTER_GOOGLE_CLIENT_SECRET") or "").strip()
    if flavor == "google" and secret:
        form["client_secret"] = secret

    req = urllib.request.Request(
        _token_url(flavor), data=urllib.parse.urlencode(form).encode(),
        headers={"Content-Type": "application/x-www-form-urlencoded"}, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=OAUTH_TIMEOUT) as resp:
            payload = json.loads(resp.read())
    except urllib.error.HTTPError as exc:
        raw = exc.read().decode("utf-8", "replace")
        hint = _explain_oauth(raw)
        raise OAuthError(f"换 access_token 失败（{exc.code}）：{raw[:300]}"
                         + (f"\n\n提示：{hint}" if hint else "")) from exc
    except urllib.error.URLError as exc:
        raise OAuthError(f"连不上令牌端点：{exc.reason}") from exc

    token = payload.get("access_token")
    if not token:
        raise OAuthError(f"令牌端点没有返回 access_token：{json.dumps(payload)[:300]}")
    # 微软 v2 端点在 refresh 成功时可能返回一个**新的** refresh_token 并让旧的失效。
    # 不写回库的话，下一次换令牌就会拿一个已作废的去换，账号直接掉线。
    # 之前一直没暴露只是因为 SMTP.Send 这个 scope 的令牌通常不轮换。
    return token, time.time() + float(payload.get("expires_in") or 3600), \
        (payload.get("refresh_token") or "").strip()


def _rotate_refresh_token(old_token: str, new_token: str) -> None:
    """令牌轮换后写回库。找不到对应行就算了——缓存里那份这一轮仍然可用，
    但下一轮会拿作废的令牌去换，所以这里失败要留日志。"""
    if not new_token or new_token == old_token:
        return
    try:
        rows = db.query_all(
            "SELECT id, encrypted_refresh_token FROM mb_sender_accounts "
            "WHERE encrypted_refresh_token IS NOT NULL AND auth_mode = 'xoauth2'")
        hit = [r["id"] for r in rows
               if crypto_util.decrypt(r["encrypted_refresh_token"]) == old_token]
        if not hit:
            logger.warning("mail-blaster: 令牌轮换了但库里找不到对应账号，下次换令牌可能失败")
            return
        db.execute("UPDATE mb_sender_accounts SET encrypted_refresh_token = %s "
                   "WHERE id = ANY(%s)", (crypto_util.encrypt(new_token), hit))
        logger.info("mail-blaster: refresh_token 已轮换并写回账号 %s", hit)
    except Exception:
        logger.exception("mail-blaster: refresh_token 轮换写回失败")


def get_access_token(*, provider: str | None, client_id: str, refresh_token: str) -> str:
    """带缓存地换 access_token。群发时不会每封都换。

    缓存键用 (client_id, refresh_token) 而不是邮箱——同一个 client_id 配不同
    refresh_token 也能正确区分。网络请求刻意放在锁外，不阻塞其它账号。
    """
    import time
    key = (client_id, refresh_token)
    now = time.time()
    with _token_lock:
        cached = _token_cache.get(key)
        if cached and cached[1] - TOKEN_EXPIRY_MARGIN > now:
            return cached[0]
    token, expires_at, rotated = _request_access_token(
        _oauth_flavor(provider), client_id, refresh_token)
    with _token_lock:
        _token_cache[key] = (token, expires_at)
        if rotated and rotated != refresh_token:
            # 新令牌也进缓存，这样下游拿新令牌来问时不用再跑一趟网络
            _token_cache[(client_id, rotated)] = (token, expires_at)
    _rotate_refresh_token(refresh_token, rotated)
    return token


def invalidate_token(client_id: str, refresh_token: str) -> None:
    with _token_lock:
        _token_cache.pop((client_id, refresh_token), None)


def build_xoauth2(username: str, access_token: str) -> str:
    """SASL XOAUTH2 明文，smtplib 会代做 base64。"""
    return f"user={username}\1auth=Bearer {access_token}\1\1"


def _login_xoauth2(client, username: str, account: dict) -> None:
    client_id = crypto_util.decrypt(account.get("encrypted_client_id"))
    refresh_token = crypto_util.decrypt(account.get("encrypted_refresh_token"))
    if not client_id or not refresh_token:
        raise ValueError("该账号是 OAuth2 模式，但缺 client_id 或 refresh_token")
    token = get_access_token(provider=account.get("provider"),
                             client_id=client_id, refresh_token=refresh_token)
    auth_string = build_xoauth2(username, token)
    try:
        # 服务端回 334 挑战时按 XOAUTH2 约定回一个空行，让它把失败原因落到 535
        client.auth("XOAUTH2", lambda challenge=None: "" if challenge else auth_string)
    except smtplib.SMTPAuthenticationError:
        # 被拒的 token 不留在缓存里；这条连接多半已被服务端关掉，不在上面重试
        invalidate_token(client_id, refresh_token)
        raise


def _login_password(client, username: str, account: dict) -> None:
    password = crypto_util.decrypt(account.get("encrypted_password"))
    if not password:
        raise ValueError("该账号还没有填密码")
    client.login(username, password)


def _do_login(client, account: dict) -> None:
    username = (account.get("smtp_username") or "").strip() or account["email"]
    if (account.get("auth_mode") or "password") == "xoauth2":
        _login_xoauth2(client, username, account)
    else:
        _login_password(client, username, account)
