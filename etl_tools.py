"""
Excel ETL helpers: Thai row filter, engagement aggregation, DataFrame export.
"""
import io
import logging
import re
from dataclasses import dataclass
from typing import List, Optional, Tuple

import pandas as pd

from thai_utils import is_thai_content

logger = logging.getLogger(__name__)

# 功能4：兼容列名
DATE_ALIASES = ("date", "post_date", "日期", "day")
ENGAGEMENT_ALIASES = ("engagement", "互动量", "互动", "total_engagement")

MAX_SYNC_EXCEL_ROWS = 50000

URL_COLUMN_ALIASES = (
    "post_url", "url", "link", "video_url", "permalink", "URL", "Link",
    "帖子链接", "视频链接", "链接", "作品链接", "主页链接", "短链",
    "Video URL", "video link", "Video Link", "tt_link", "ins_link", "yt_link",
)


def _looks_like_url(value) -> bool:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return False
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return False
    lower = s.lower()
    if lower.startswith("http://") or lower.startswith("https://"):
        return True
    if lower.startswith("www."):
        return True
    domains = (
        "tiktok.com", "instagram.com", "youtube.com", "youtu.be",
        "facebook.com", "fb.watch", "fb.com", "twitter.com", "x.com",
    )
    return any(d in lower for d in domains)


def resolve_url_column(df: pd.DataFrame, url_column: Optional[str] = None) -> str:
    """解析 Excel 中的链接列：显式列名 > 常见别名 > 按内容自动推断。"""
    if url_column:
        key = url_column.strip()
        if key in df.columns:
            return key
        for col in df.columns:
            if str(col).strip() == key or str(col).strip().lower() == key.lower():
                return col

    for alias in URL_COLUMN_ALIASES:
        for col in df.columns:
            name = str(col).strip()
            if name == alias or name.lower() == alias.lower():
                return col

    best_col = None
    best_score = -1
    for col in df.columns:
        score = sum(1 for v in df[col] if _looks_like_url(v))
        if score > best_score:
            best_score = score
            best_col = col
    if best_col is not None and best_score > 0:
        return best_col

    headers = ", ".join(str(c) for c in df.columns[:25])
    raise ValueError(f"未找到链接列，请在「链接列名」填写正确列名。当前表头: {headers}")


def _column_url_score(df: pd.DataFrame, col) -> int:
    score = sum(1 for v in df[col] if _looks_like_url(v))
    name = str(col).strip().lower()
    if any(k in name for k in ("链接", "link", "url", "video", "视频", "permalink", "post")):
        score += 2
    return score


def _collect_hyperlink_urls(file_bytes: bytes) -> List[str]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []
    urls: List[str] = []
    seen = set()
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    target = None
                    if getattr(cell, "hyperlink", None) is not None:
                        target = getattr(cell.hyperlink, "target", None)
                    if target and _looks_like_url(target):
                        norm = _normalize_url_cell(target)
                        if norm and norm not in seen:
                            seen.add(norm)
                            urls.append(norm)
                        continue
                    val = cell.value
                    if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
                        m = re.search(r"https?://[^\"')\\s]+", val, re.I)
                        if m:
                            norm = _normalize_url_cell(m.group(0))
                            if norm and norm not in seen:
                                seen.add(norm)
                                urls.append(norm)
        wb.close()
    except Exception as e:
        logger.warning("读取 Excel 超链接失败: %s", e)
    return urls


@dataclass
class ExcelUrlParseResult:
    df: pd.DataFrame
    url_column: str
    sheet_name: str
    header_row: int
    urls: List[str]


def _hyperlink_rows_by_excel_row(file_bytes: bytes, sheet_name: str) -> dict:
    """返回 Excel 行号(1-based) -> 规范化 URL。"""
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}
    out: dict = {}
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False, read_only=False)
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
        for row in ws.iter_rows():
            for cell in row:
                target = None
                if getattr(cell, "hyperlink", None) is not None:
                    target = getattr(cell.hyperlink, "target", None)
                if target and _looks_like_url(target):
                    out[cell.row] = _normalize_url_cell(target)
                    break
                val = cell.value
                if isinstance(val, str) and val.strip().upper().startswith("=HYPERLINK("):
                    m = re.search(r"https?://[^\"')\\s]+", val, re.I)
                    if m:
                        out[cell.row] = _normalize_url_cell(m.group(0))
                        break
        wb.close()
    except Exception as e:
        logger.warning("读取 Sheet 超链接失败: %s", e)
    return out


def _inject_row_hyperlinks(df: pd.DataFrame, col: str, header_row: int, row_links: dict) -> pd.DataFrame:
    if not row_links:
        return df
    out = df.copy()
    for idx in out.index:
        try:
            pandas_pos = int(idx) if isinstance(idx, (int, float)) else list(out.index).index(idx)
        except Exception:
            pandas_pos = 0
        excel_row = header_row + 2 + pandas_pos
        link = row_links.get(excel_row)
        if link and not _looks_like_url(out.at[idx, col]):
            out.at[idx, col] = link
    return out

def load_best_excel_table(
    file_bytes: bytes,
    url_column: Optional[str] = None,
) -> Tuple[pd.DataFrame, str, str, int]:
    """扫描多 Sheet / 多表头行，返回最可能含链接的数据表。"""
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    best_df = None
    best_col = None
    best_sheet = xl.sheet_names[0] if xl.sheet_names else "Sheet1"
    best_header = 0
    best_score = -1

    for sheet in xl.sheet_names:
        for header_row in range(0, 8):
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=header_row)
            except Exception:
                continue
            if df is None or df.empty:
                continue
            df = df.dropna(axis=1, how="all")
            if df.empty:
                continue
            try:
                col = resolve_url_column(df, url_column)
                score = _column_url_score(df, col)
            except ValueError:
                if url_column:
                    raise
                score = max((_column_url_score(df, c) for c in df.columns), default=0)
                if score <= 0:
                    continue
                col = max(df.columns, key=lambda c: _column_url_score(df, c))
            if score > best_score:
                best_score = score
                best_df = df
                best_col = col
                best_sheet = sheet
                best_header = header_row

    if best_df is not None and best_col is not None and best_score > 0:
        return best_df, best_col, best_sheet, best_header

    df = pd.read_excel(io.BytesIO(file_bytes))
    col = resolve_url_column(df, url_column)
    return df, col, xl.sheet_names[0] if xl.sheet_names else "Sheet1", 0


def parse_excel_urls(file_bytes: bytes, url_column: Optional[str] = None) -> ExcelUrlParseResult:
    df, col, sheet_name, header_row = load_best_excel_table(file_bytes, url_column)
    row_links = _hyperlink_rows_by_excel_row(file_bytes, sheet_name)
    df = _inject_row_hyperlinks(df, col, header_row, row_links)
    urls: List[str] = []
    seen = set()
    for v in df[col]:
        s = _normalize_url_cell(v)
        if s and _looks_like_url(s) and s not in seen:
            seen.add(s)
            urls.append(s)

    if not urls:
        for link in _collect_hyperlink_urls(file_bytes):
            if link not in seen:
                seen.add(link)
                urls.append(link)

    if not urls:
        headers = ", ".join(str(c) for c in df.columns[:25])
        raise ValueError(f"未找到有效链接，请确认表格中有 URL 或超链接。当前表头: {headers}")

    return ExcelUrlParseResult(
        df=df,
        url_column=str(col),
        sheet_name=sheet_name,
        header_row=header_row,
        urls=urls,
    )


def _normalize_url_cell(value) -> str:
    s = "" if value is None or (isinstance(value, float) and pd.isna(value)) else str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    if not s.lower().startswith("http"):
        if s.lower().startswith("www.") or any(
            d in s.lower() for d in ("tiktok.com", "instagram.com", "youtube.com", "youtu.be", "facebook.com", "fb.watch")
        ):
            s = "https://" + s.lstrip("/")
    return s


def _normalize_date_col(df: pd.DataFrame) -> pd.DataFrame:
    for a in DATE_ALIASES:
        if a in df.columns:
            out = df.rename(columns={a: "_date_norm"})
            return out
    raise ValueError(f"未找到日期列，需要其一: {DATE_ALIASES}")


def _normalize_engagement_col(df: pd.DataFrame) -> pd.DataFrame:
    for a in ENGAGEMENT_ALIASES:
        if a in df.columns:
            out = df.rename(columns={a: "_eng_norm"})
            return out
    raise ValueError(f"未找到互动量列，需要其一: {ENGAGEMENT_ALIASES}")


def sum_daily_engagement_from_excels(file_bytes_list: List[bytes]) -> bytes:
    """Concatenate multiple workbooks and sum engagement by date."""
    frames = []
    for i, raw in enumerate(file_bytes_list):
        df = pd.read_excel(io.BytesIO(raw))
        df = _normalize_date_col(df)
        df = _normalize_engagement_col(df)
        df["_date_norm"] = pd.to_datetime(df["_date_norm"], errors="coerce").dt.strftime("%Y-%m-%d")
        df["_eng_norm"] = pd.to_numeric(df["_eng_norm"], errors="coerce").fillna(0)
        frames.append(df[["_date_norm", "_eng_norm"]])
    if not frames:
        raise ValueError("没有有效文件")
    all_df = pd.concat(frames, ignore_index=True)
    summed = all_df.groupby("_date_norm", as_index=False)["_eng_norm"].sum()
    summed = summed.rename(columns={"_date_norm": "date", "_eng_norm": "engagement"})
    summed = summed.sort_values("date")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        summed.to_excel(writer, sheet_name="daily_engagement_sum", index=False)
    buf.seek(0)
    return buf.read()


def filter_thai_rows_excel(file_bytes: bytes, text_column: Optional[str]) -> Tuple[bytes, int, int]:
    """
    Keep rows where text_column passes is_thai_content.
    If text_column is None, auto-pick first object column with most non-null strings.
    Returns (xlsx_bytes, kept_count, dropped_count).
    """
    df = pd.read_excel(io.BytesIO(file_bytes))
    if len(df) > MAX_SYNC_EXCEL_ROWS:
        raise ValueError(f"行数超过上限 {MAX_SYNC_EXCEL_ROWS}")
    col = text_column
    if not col or col not in df.columns:
        col = _guess_text_column(df)
    kept_mask = []
    for val in df[col]:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            s = ""
        else:
            s = str(val).strip()
        if not s or s.lower() == "nan":
            kept_mask.append(False)
        else:
            kept_mask.append(bool(is_thai_content(s)))
    out_df = df.loc[kept_mask].reset_index(drop=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="thai_only", index=False)
    buf.seek(0)
    return buf.read(), len(out_df), len(df) - len(out_df)


def _guess_text_column(df: pd.DataFrame) -> str:
    best_col = None
    best_score = -1
    for c in df.columns:
        ser = df[c]
        if ser.dtype == object or str(ser.dtype).startswith("string"):
            non_null = ser.notna() & (ser.astype(str).str.strip() != "")
            score = int(non_null.sum())
            if score > best_score:
                best_score = score
                best_col = c
    if not best_col:
        raise ValueError("无法自动推断文本列，请指定列名")
    return best_col


def posts_metrics_to_excel_bytes(rows: List[dict], start_date: str, end_date: str) -> bytes:
    """Filter posts by post_date and write core columns."""
    if not rows:
        empty = pd.DataFrame(
            columns=[
                "post_url", "platform", "author", "post_date", "post_content",
                "likes", "comments_count", "shares", "views", "engagement", "thumbnail_url",
            ]
        )
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            empty.to_excel(writer, sheet_name="posts", index=False)
        buf.seek(0)
        return buf.read()
    df = pd.DataFrame(rows)
    if "post_date" in df.columns:
        df["post_date"] = df["post_date"].astype(str).str[:10]
        df = df[(df["post_date"] >= start_date) & (df["post_date"] <= end_date)]
    cols = [
        "post_url", "platform", "author", "post_date", "post_content",
        "likes", "comments_count", "shares", "views", "engagement", "thumbnail_url",
    ]
    for c in cols:
        if c not in df.columns:
            df[c] = ""
    df = df[cols]
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="posts", index=False)
    buf.seek(0)
    return buf.read()


def read_urls_from_excel(file_bytes: bytes, url_column: Optional[str]) -> List[str]:
    return parse_excel_urls(file_bytes, url_column).urls


def comments_to_excel_bytes(
    comment_rows: List[dict],
    beijing_tz_name: str = "Asia/Shanghai",
) -> bytes:
    """
    comment_rows: dicts with keys post_url, comment_id, author, content, created_at (datetime)
    Sheet1: detail. Sheet2: daily new comment count per post_url.
    """
    if not comment_rows:
        detail = pd.DataFrame(columns=["post_url", "comment_id", "author", "content", "created_at", "date"])
        pivot = pd.DataFrame(columns=["post_url", "date", "comment_count"])
    else:
        detail = pd.DataFrame(comment_rows)
        if "created_at" in detail.columns:
            ts = pd.to_datetime(detail["created_at"], errors="coerce", utc=True)
            ts = ts.dt.tz_convert(beijing_tz_name)
            detail["date"] = ts.dt.strftime("%Y-%m-%d")
            detail["created_at"] = ts.dt.strftime("%Y-%m-%d %H:%M:%S")
        else:
            detail["date"] = ""
        pivot = (
            detail.groupby(["post_url", "date"], as_index=False)
            .size()
            .rename(columns={"size": "comment_count"})
        )
        sort_cols = [c for c in ("post_url", "created_at") if c in detail.columns]
        if sort_cols:
            detail = detail.sort_values(sort_cols, na_position="last")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="comments_detail", index=False)
        pivot.to_excel(writer, sheet_name="daily_new_comments", index=False)
    buf.seek(0)
    return buf.read()


def list_text_columns_preview(file_bytes: bytes) -> List[str]:
    try:
        _, _, sheet_name, header_row = load_best_excel_table(file_bytes)
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=header_row, nrows=0)
        return [str(c) for c in df.columns]
    except Exception:
        df = pd.read_excel(io.BytesIO(file_bytes), nrows=0)
        return [str(c) for c in df.columns]
