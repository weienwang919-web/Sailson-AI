from __future__ import annotations

import json
import logging
import os
import re
import secrets as pysecrets
import time
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from typing import Any
from urllib.parse import unquote

import requests
from itsdangerous import BadSignature, URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill

import crypto_util
import database as db
import usage_service

logger = logging.getLogger(__name__)

API_BASE = "https://business-api.tiktok.com/open_api/v1.3"
INVITE_SALT = "tiktok-official-invite"

# 业务口径时区：快照日期、发布日、报表区间这类「自然日」一律按北京时间归属。
# 不能用 DB 的 CURRENT_DATE 或服务器本地时间——线上跑在 UTC，北京 03:30 的每日同步
# 落在 UTC 的前一天，直接用 CURRENT_DATE 会让快照日期整体错位一天。
# 注意：库里的时间戳列存的都是 naive UTC（见 _epoch_to_dt），转换前先当 UTC 处理。
BUSINESS_TZ = timezone(timedelta(hours=8))


def _business_date(value: datetime | None = None) -> date:
    """把 naive-UTC 时间戳归到北京自然日；不传参数则取「今天」。"""
    if value is None:
        return datetime.now(tz=BUSINESS_TZ).date()
    return value.replace(tzinfo=timezone.utc).astimezone(BUSINESS_TZ).date()

VIDEO_FIELDS = [
    "item_id",
    "media_type",
    "is_ad",
    "thumbnail_url",
    "share_url",
    "embed_url",
    "caption",
    "video_duration",
    "likes",
    "comments",
    "shares",
    "favorites",
    "create_time",
    "reach",
    "video_views",
    "total_time_watched",
    "average_time_watched",
    "full_video_watched_rate",
    "new_followers",
    "profile_views",
    "website_clicks",
    "phone_number_clicks",
    "lead_submissions",
    "app_download_clicks",
    "email_clicks",
    "address_clicks",
    "video_view_retention",
    "impression_sources",
    "audience_genders",
    "audience_countries",
    "audience_cities",
    "audience_types",
    "engagement_likes",
]

PROFILE_FIELDS = [
    "is_business_account",
    "profile_image",
    "username",
    "profile_deep_link",
    "display_name",
    "bio_description",
    "is_verified",
    "following_count",
    "followers_count",
    "total_likes",
    "videos_count",
    "video_views",
    "unique_video_views",
    "profile_views",
    "likes",
    "comments",
    "shares",
    "phone_number_clicks",
    "lead_submissions",
    "app_download_clicks",
    "bio_link_clicks",
    "email_clicks",
    "address_clicks",
    "daily_total_followers",
    "daily_new_followers",
    "daily_lost_followers",
    "audience_activity",
    "engaged_audience",
    "audience_ages",
    "audience_genders",
    "audience_countries",
    "audience_cities",
]


_TABLES = {
    "tiktok_official_accounts", "tiktok_official_video_snapshots",
    "tiktok_official_profile_daily_metrics", "tiktok_official_tokens",
    "tiktok_official_invites", "tiktok_spark_invites", "tiktok_spark_tokens",
    "tiktok_advertiser_tokens", "tiktok_ad_spend_daily", "tiktok_ad_video_map",
    "tiktok_official_video_daily_snapshots", "tiktok_matrix_video_exports",
    "tiktok_official_video_publish_window_snapshots",
}
_LATEST_COLUMNS = {
    ("tiktok_official_accounts", "account_alias"),
    ("tiktok_official_accounts", "authorized_by"),
    ("tiktok_official_accounts", "status"),
    ("tiktok_official_accounts", "region"),
    ("tiktok_official_accounts", "account_type"),
    ("tiktok_official_accounts", "needs_follower_boost"),
    ("tiktok_official_tokens", "account_alias"),
    ("tiktok_official_tokens", "authorized_by"),
    ("tiktok_official_tokens", "status"),
    ("tiktok_ad_spend_daily", "ad_id"),
    ("tiktok_ad_spend_daily", "tiktok_item_id"),
    ("tiktok_ad_spend_daily", "video_play_actions"),
    ("tiktok_official_video_snapshots", "task_no"),
    ("tiktok_official_video_snapshots", "kol_campaign"),
    ("tiktok_official_video_snapshots", "spark_code"),
    ("tiktok_official_video_snapshots", "spark_code_start_time"),
    ("tiktok_official_video_snapshots", "spark_code_end_time"),
    ("tiktok_official_video_snapshots", "is_boosted"),
    ("tiktok_official_video_snapshots", "boosted_at"),
    ("tiktok_official_video_snapshots", "boost_status"),
    ("tiktok_spark_tokens", "expires_at"),
    ("tiktok_spark_tokens", "refresh_expires_at"),
    ("tiktok_spark_tokens", "business_id"),
    ("tiktok_spark_invites", "business_id"),
}


def _schema_is_current() -> bool:
    """一次查询判断表和列是否都齐了，省得每次启动都跑 40+ 条 DDL（Render 上每条约 1 秒）。"""
    try:
        rows = db.query_all(
            "SELECT table_name, column_name FROM information_schema.columns "
            "WHERE table_schema = 'public' AND table_name = ANY(%s)",
            (sorted(_TABLES),))
    except Exception:
        return False
    have_tables = {r["table_name"] for r in rows}
    have_columns = {(r["table_name"], r["column_name"]) for r in rows}
    return _TABLES <= have_tables and _LATEST_COLUMNS <= have_columns


def ensure_schema() -> None:
    if _schema_is_current():
        return
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_accounts (
            id SERIAL PRIMARY KEY,
            business_id VARCHAR(128) UNIQUE NOT NULL,
            account_name VARCHAR(255),
            display_name VARCHAR(255),
            username VARCHAR(255),
            profile_deep_link TEXT,
            profile_image TEXT,
            bio_description TEXT,
            is_business_account BOOLEAN,
            is_verified BOOLEAN,
            following_count BIGINT,
            followers_count BIGINT,
            total_likes BIGINT,
            videos_count BIGINT,
            enabled BOOLEAN DEFAULT TRUE,
            notes TEXT,
            raw_json TEXT,
            last_refreshed_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_video_snapshots (
            id SERIAL PRIMARY KEY,
            business_id VARCHAR(128) NOT NULL,
            item_id VARCHAR(128) NOT NULL,
            media_type VARCHAR(32),
            is_ad BOOLEAN,
            thumbnail_url TEXT,
            share_url TEXT,
            embed_url TEXT,
            caption TEXT,
            video_duration DOUBLE PRECISION,
            create_time TIMESTAMP,
            likes BIGINT,
            comments BIGINT,
            shares BIGINT,
            favorites BIGINT,
            reach BIGINT,
            video_views BIGINT,
            total_time_watched DOUBLE PRECISION,
            average_time_watched DOUBLE PRECISION,
            full_video_watched_rate DOUBLE PRECISION,
            new_followers BIGINT,
            profile_views BIGINT,
            website_clicks BIGINT,
            phone_number_clicks BIGINT,
            lead_submissions BIGINT,
            app_download_clicks BIGINT,
            email_clicks BIGINT,
            address_clicks BIGINT,
            video_view_retention TEXT,
            engagement_likes TEXT,
            impression_sources TEXT,
            audience_genders TEXT,
            audience_countries TEXT,
            audience_cities TEXT,
            audience_types TEXT,
            request_id VARCHAR(255),
            log_id VARCHAR(255),
            raw_json TEXT,
            fetched_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (business_id, item_id)
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_profile_daily_metrics (
            id SERIAL PRIMARY KEY,
            business_id VARCHAR(128) NOT NULL,
            metric_date DATE NOT NULL,
            followers_count BIGINT,
            video_views BIGINT,
            unique_video_views BIGINT,
            profile_views BIGINT,
            likes BIGINT,
            comments BIGINT,
            shares BIGINT,
            phone_number_clicks BIGINT,
            lead_submissions BIGINT,
            app_download_clicks BIGINT,
            bio_link_clicks BIGINT,
            email_clicks BIGINT,
            address_clicks BIGINT,
            daily_total_followers BIGINT,
            daily_new_followers BIGINT,
            daily_lost_followers BIGINT,
            engaged_audience BIGINT,
            raw_json TEXT,
            fetched_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (business_id, metric_date)
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_tokens (
            id SERIAL PRIMARY KEY,
            token_type VARCHAR(32) DEFAULT 'account',
            open_id VARCHAR(128) UNIQUE NOT NULL,
            access_token TEXT NOT NULL,
            refresh_token TEXT,
            scope TEXT,
            expires_at TIMESTAMP,
            refresh_expires_at TIMESTAMP,
            raw_json TEXT,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_tt_official_videos_business ON tiktok_official_video_snapshots (business_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_tt_official_videos_create_time ON tiktok_official_video_snapshots (create_time DESC)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_tt_official_profile_business_date ON tiktok_official_profile_daily_metrics (business_id, metric_date)")

    # 矩阵号自助授权：账号别名/授权人/启停状态
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS account_alias VARCHAR(255)")
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS authorized_by VARCHAR(255)")
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS status VARCHAR(16) DEFAULT 'active'")
    db.execute("ALTER TABLE tiktok_official_tokens ADD COLUMN IF NOT EXISTS account_alias VARCHAR(255)")
    db.execute("ALTER TABLE tiktok_official_tokens ADD COLUMN IF NOT EXISTS authorized_by VARCHAR(255)")
    db.execute("ALTER TABLE tiktok_official_tokens ADD COLUMN IF NOT EXISTS status VARCHAR(16) DEFAULT 'active'")

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_invites (
            nonce VARCHAR(64) PRIMARY KEY,
            account_alias VARCHAR(255) NOT NULL,
            authorized_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            expires_at TIMESTAMP,
            used_at TIMESTAMP
        )
        """
    )
    db.execute("ALTER TABLE tiktok_official_invites ALTER COLUMN expires_at DROP NOT NULL")

    # Spark Ads 独立授权小流程（新 App，仅补齐 biz.spark.auth 权限，不进主账号表）
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_spark_invites (
            nonce VARCHAR(64) PRIMARY KEY,
            account_alias VARCHAR(255) NOT NULL,
            business_id VARCHAR(64),
            authorized_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            expires_at TIMESTAMP,
            used_at TIMESTAMP
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_spark_tokens (
            account_alias VARCHAR(255),
            business_id VARCHAR(64),
            open_id VARCHAR(128),
            access_token TEXT NOT NULL,
            refresh_token TEXT,
            scope TEXT,
            authorized_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    # Spark token 原先没存过期时间，导致 access_token 到期后只能等 TikTok 接口报错才发现。
    # 补上跟主账号 token 一样的过期时间跟踪，配合 get_spark_token_info() 里的自动刷新。
    db.execute("ALTER TABLE tiktok_spark_tokens ADD COLUMN IF NOT EXISTS expires_at TIMESTAMP")
    db.execute("ALTER TABLE tiktok_spark_tokens ADD COLUMN IF NOT EXISTS refresh_expires_at TIMESTAMP")
    # account_alias 在 tiktok_official_accounts 里并不保证唯一（不同账号的运营人员可能填了同一个别名），
    # 但 Spark token 原来按 account_alias 做主键去重，导致撞名的两个账号会互相顶掉对方的 token，
    # 生成推广码时用错身份，报 "item_id does not belong to the user"。改成按真正唯一的 business_id 做键。
    db.execute("ALTER TABLE tiktok_spark_tokens ADD COLUMN IF NOT EXISTS business_id VARCHAR(64)")
    db.execute("ALTER TABLE tiktok_spark_invites ADD COLUMN IF NOT EXISTS business_id VARCHAR(64)")
    try:
        db.execute("ALTER TABLE tiktok_spark_tokens DROP CONSTRAINT IF EXISTS tiktok_spark_tokens_pkey")
    except Exception as e:
        logger.warning(f"drop tiktok_spark_tokens 旧主键失败（可能已经不存在）: {e}")
    try:
        db.execute("ALTER TABLE tiktok_spark_tokens ADD CONSTRAINT tiktok_spark_tokens_business_id_key UNIQUE (business_id)")
    except Exception as e:
        logger.warning(f"添加 tiktok_spark_tokens.business_id 唯一约束失败（可能已存在）: {e}")
    # 只回填 account_alias 在主表里唯一对应一个 business_id 的安全情况；
    # 撞名的账号回填不出真实归属，保持 business_id 为空，视为需要重新走一遍 Spark 邀请。
    db.execute(
        """
        UPDATE tiktok_spark_tokens t
        SET business_id = a.business_id
        FROM (
            SELECT account_alias, MIN(business_id) AS business_id
            FROM tiktok_official_accounts
            WHERE account_alias IS NOT NULL AND account_alias <> ''
            GROUP BY account_alias
            HAVING COUNT(*) = 1
        ) a
        WHERE t.account_alias = a.account_alias AND t.business_id IS NULL
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_advertiser_tokens (
            advertiser_id VARCHAR(64) PRIMARY KEY,
            access_token TEXT NOT NULL,
            scope TEXT,
            authorized_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_ad_spend_daily (
            id SERIAL PRIMARY KEY,
            advertiser_id VARCHAR(64) NOT NULL,
            stat_date DATE NOT NULL,
            country_code VARCHAR(8) NOT NULL DEFAULT '',
            spend NUMERIC(14,2) DEFAULT 0,
            impressions BIGINT DEFAULT 0,
            clicks BIGINT DEFAULT 0,
            conversions BIGINT DEFAULT 0,
            cost_per_conversion NUMERIC(14,4),
            ctr NUMERIC(8,4),
            cpc NUMERIC(14,4),
            cpm NUMERIC(14,4),
            raw_json JSONB,
            fetched_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (advertiser_id, stat_date, country_code)
        )
        """
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_tiktok_ad_spend_daily_stat_date ON tiktok_ad_spend_daily (stat_date)"
    )

    # 投流数据改为按 ad_id/tiktok_item_id 精确归因到矩阵视频，而不是整个广告账户的消耗
    # （账户级汇总会把同一个 Business Center 下其他客户业务的投流也算进来，参见 plan 里的验证数据）
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_ad_video_map (
            ad_id VARCHAR(64) PRIMARY KEY,
            advertiser_id VARCHAR(64) NOT NULL,
            campaign_id VARCHAR(64),
            adgroup_id VARCHAR(64),
            tiktok_item_id VARCHAR(64) NOT NULL,
            objective_type VARCHAR(32),
            first_seen_at TIMESTAMP DEFAULT NOW(),
            last_seen_at TIMESTAMP DEFAULT NOW()
        )
        """
    )
    db.execute("CREATE INDEX IF NOT EXISTS idx_ad_video_map_item ON tiktok_ad_video_map (tiktok_item_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_ad_video_map_advertiser ON tiktok_ad_video_map (advertiser_id)")
    db.execute("ALTER TABLE tiktok_ad_spend_daily ADD COLUMN IF NOT EXISTS ad_id VARCHAR(64)")
    db.execute("ALTER TABLE tiktok_ad_spend_daily ADD COLUMN IF NOT EXISTS tiktok_item_id VARCHAR(64)")
    db.execute("ALTER TABLE tiktok_ad_spend_daily ADD COLUMN IF NOT EXISTS video_play_actions BIGINT DEFAULT 0")
    db.execute(
        "ALTER TABLE tiktok_ad_spend_daily DROP CONSTRAINT IF EXISTS tiktok_ad_spend_daily_advertiser_id_stat_date_country_code_key"
    )
    db.execute(
        "CREATE UNIQUE INDEX IF NOT EXISTS uniq_ad_spend_ad_date_country ON tiktok_ad_spend_daily (ad_id, stat_date, country_code)"
    )

    # 矩阵号视频监控看板：账号级国家/账号类型 + 视频级人工标签/Spark授权码
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS region VARCHAR(16)")
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS account_type VARCHAR(16)")

    # 需补粉标记：人工标记该账号需要补充粉丝，永久保留，供官号总览筛选
    db.execute("ALTER TABLE tiktok_official_accounts ADD COLUMN IF NOT EXISTS needs_follower_boost BOOLEAN DEFAULT FALSE")

    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS task_no VARCHAR(128)")
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS kol_campaign VARCHAR(255)")
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS spark_code TEXT")
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS spark_code_start_time TIMESTAMP")
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS spark_code_end_time TIMESTAMP")

    # 投流标记：人工标记该视频已投流，永久保留，供看板筛选
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS is_boosted BOOLEAN DEFAULT FALSE")
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS boosted_at TIMESTAMP")
    # 投流三态（已标记/已投放/已关闭），取代原先的布尔开关，仍然是人工维护、不由投放数据推导。
    # is_boosted 保留不动：老的筛选/导出仍在用，且作为三态的兜底回退，写入时同步维护。
    db.execute("ALTER TABLE tiktok_official_video_snapshots ADD COLUMN IF NOT EXISTS boost_status VARCHAR(16)")
    # 一次性回填：老数据只有布尔值，字面含义是"已标记"，统一落到最保守的 marked 态。
    # 只填 NULL，重复执行安全。
    db.execute(
        "UPDATE tiktok_official_video_snapshots SET boost_status = 'marked' "
        "WHERE is_boosted = TRUE AND boost_status IS NULL"
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_video_daily_snapshots (
            id SERIAL PRIMARY KEY,
            business_id VARCHAR(128) NOT NULL,
            item_id VARCHAR(128) NOT NULL,
            snapshot_date DATE NOT NULL,
            video_views BIGINT,
            likes BIGINT,
            comments BIGINT,
            shares BIGINT,
            favorites BIGINT,
            fetched_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (business_id, item_id, snapshot_date)
        )
        """
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_tt_official_video_daily_snapshots_lookup "
        "ON tiktok_official_video_daily_snapshots (business_id, item_id, snapshot_date)"
    )

    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_matrix_video_exports (
            export_date DATE PRIMARY KEY,
            file_bytes BYTEA NOT NULL,
            generated_at TIMESTAMP DEFAULT NOW()
        )
        """
    )

    # 发布后3/24/48/72小时时间点快照：每条新视频入库时生成4行占位，worker 到期后回填
    db.execute(
        """
        CREATE TABLE IF NOT EXISTS tiktok_official_video_publish_window_snapshots (
            id SERIAL PRIMARY KEY,
            business_id VARCHAR(128) NOT NULL,
            item_id VARCHAR(128) NOT NULL,
            window_hours SMALLINT NOT NULL,
            due_at TIMESTAMP NOT NULL,
            captured_at TIMESTAMP,
            video_views BIGINT,
            likes BIGINT,
            comments BIGINT,
            shares BIGINT,
            favorites BIGINT,
            reach BIGINT,
            total_time_watched DOUBLE PRECISION,
            average_time_watched DOUBLE PRECISION,
            full_video_watched_rate DOUBLE PRECISION,
            impression_sources TEXT,
            engagement_rate DOUBLE PRECISION,
            followers_count_snapshot BIGINT,
            distribution_rate DOUBLE PRECISION,
            created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE (business_id, item_id, window_hours)
        )
        """
    )
    db.execute(
        "CREATE INDEX IF NOT EXISTS idx_tt_official_publish_window_due "
        "ON tiktok_official_video_publish_window_snapshots (due_at) WHERE captured_at IS NULL"
    )


def configured_accounts() -> list[dict[str, Any]]:
    raw = os.environ.get("TIKTOK_OFFICIAL_ACCOUNTS_JSON", "").strip()
    if raw:
        data = json.loads(raw)
        if not isinstance(data, list):
            raise ValueError("TIKTOK_OFFICIAL_ACCOUNTS_JSON 必须是 JSON 数组")
        return [_normalize_account(x) for x in data if x]
    business_id = os.environ.get("TIKTOK_OFFICIAL_BUSINESS_ID", "").strip()
    if business_id:
        return [
            {
                "business_id": business_id,
                "account_name": os.environ.get("TIKTOK_OFFICIAL_ACCOUNT_NAME", "TikTok Official"),
                "enabled": True,
                "notes": "",
            }
        ]
    return []


def sync_configured_accounts() -> list[dict[str, Any]]:
    accounts = configured_accounts()
    for account in accounts:
        db.execute(
            """
            INSERT INTO tiktok_official_accounts (business_id, account_name, enabled, notes, updated_at)
            VALUES (%s, %s, %s, %s, NOW())
            ON CONFLICT (business_id) DO UPDATE SET
                account_name = COALESCE(NULLIF(EXCLUDED.account_name, ''), tiktok_official_accounts.account_name),
                enabled = EXCLUDED.enabled,
                notes = COALESCE(EXCLUDED.notes, tiktok_official_accounts.notes),
                updated_at = NOW()
            """,
            (
                account["business_id"],
                account.get("account_name") or account["business_id"],
                bool(account.get("enabled", True)),
                account.get("notes") or "",
            ),
        )
    return list_accounts()


def list_accounts() -> list[dict[str, Any]]:
    sync_missing = db.query_one("SELECT COUNT(*) AS count FROM tiktok_official_accounts")
    if not sync_missing or int(sync_missing.get("count") or 0) == 0:
        for account in configured_accounts():
            db.execute(
                """
                INSERT INTO tiktok_official_accounts (business_id, account_name, enabled, notes)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (business_id) DO NOTHING
                """,
                (
                    account["business_id"],
                    account.get("account_name") or account["business_id"],
                    bool(account.get("enabled", True)),
                    account.get("notes") or "",
                ),
            )
    return db.query_all(
        """
        SELECT a.*, t.expires_at AS token_expires_at, t.status AS token_status,
               t.refresh_expires_at AS token_refresh_expires_at
        FROM tiktok_official_accounts a
        LEFT JOIN tiktok_official_tokens t ON t.open_id = a.business_id
        ORDER BY a.enabled DESC, a.account_name, a.id
        """
    ) or []


def set_account_meta(business_id: str, region: str | None = None, account_type: str | None = None) -> None:
    db.execute(
        """
        UPDATE tiktok_official_accounts SET
            region = %s,
            account_type = %s,
            updated_at = NOW()
        WHERE business_id = %s
        """,
        (region or None, account_type or None, business_id),
    )


def set_account_needs_boost(business_id: str, needs_boost: bool) -> None:
    db.execute(
        """
        UPDATE tiktok_official_accounts SET
            needs_follower_boost = %s,
            updated_at = NOW()
        WHERE business_id = %s
        """,
        (bool(needs_boost), business_id),
    )


def set_account_enabled(business_id: str, enabled: bool) -> None:
    db.execute(
        """
        UPDATE tiktok_official_accounts SET
            enabled = %s,
            updated_at = NOW()
        WHERE business_id = %s
        """,
        (bool(enabled), business_id),
    )


def refresh_official_accounts(
    business_ids: list[str] | None = None,
    profile_days: int = 30,
    max_pages: int = 5,
    progress_hook=None,
) -> dict[str, Any]:
    sync_configured_accounts()
    accounts = list_accounts()
    if business_ids:
        wanted = set(business_ids)
        accounts = [a for a in accounts if a.get("business_id") in wanted]
    accounts = [a for a in accounts if a.get("enabled", True)]
    if not accounts:
        raise RuntimeError("未配置可用官号，请设置 TIKTOK_OFFICIAL_ACCOUNTS_JSON")

    total_videos = 0
    ok_count = 0
    failed: list[dict[str, Any]] = []
    for idx, account in enumerate(accounts, start=1):
        bid = account["business_id"]
        name = account.get("account_name") or bid
        try:
            token = get_access_token(bid)
            if not token:
                raise RuntimeError(f"{name} 缺少 access_token，请先点击 TikTok 账号授权")
            if progress_hook:
                progress_hook(f"正在刷新主页数据：{name} ({idx}/{len(accounts)})")
            profile_data, profile_meta = fetch_profile(token, bid, profile_days)
            upsert_profile(bid, profile_data, profile_meta)

            if progress_hook:
                progress_hook(f"正在刷新视频列表：{name}")
            videos, video_meta = fetch_videos(token, bid, max_pages=max_pages)
            for video in videos:
                upsert_video(bid, video, video_meta)
            total_videos += len(videos)
            db.execute(
                "UPDATE tiktok_official_accounts SET last_refreshed_at = NOW(), updated_at = NOW() WHERE business_id = %s",
                (bid,),
            )
            ok_count += 1
        except Exception as exc:
            # 单个账号失败（token 失效/接口报错等）不应连累同批次其它账号，记录下来继续处理下一个
            logger.warning("⚠️ TikTok 官号刷新失败，跳过该账号继续：%s (%s) - %s", name, bid, exc)
            failed.append({"business_id": bid, "account_name": name, "error": str(exc)[:300]})
            if progress_hook:
                progress_hook(f"跳过失败账号：{name}（{str(exc)[:60]}）")

    if ok_count == 0 and failed:
        raise RuntimeError(failed[0]["error"] if len(failed) == 1 else f"全部 {len(failed)} 个账号刷新失败，例如：{failed[0]['error']}")

    return {"accounts": ok_count, "videos": total_videos, "failed": failed, "failed_count": len(failed)}


def _chunks(items: list, size: int):
    for i in range(0, len(items), size):
        yield items[i:i + size]


def enqueue_daily_sync(
    create_task_fn,
    *,
    update_task_params_fn,
    after_enqueue_fn=None,
    batch_size: int = 5,
    profile_days: int = 30,
    max_pages: int = 5,
) -> list[str]:
    """每日矩阵账号批量拉取：把 status='active' 的账号按 batch_size 分批，创建 task_queue 记录复用现有刷新链路。

    不在调度线程里同步跑完，交给 worker 的线程池/DB worker 消费，账号一多也不会拖垮主调度线程。
    """
    accounts = [
        a for a in list_accounts()
        if a.get("enabled", True) and (a.get("status") or "active") == "active"
    ]
    if not accounts:
        logger.info("TikTok 官号矩阵每日同步：无启用中的账号，跳过")
        return []

    session_id = f"tiktok_official_daily_sync_{_business_date().isoformat()}"
    task_ids = []
    for batch in _chunks(accounts, max(1, batch_size)):
        task_id = pysecrets.token_hex(16)
        business_ids = [a["business_id"] for a in batch]
        create_task_fn(task_id, None, session_id, function_type="tiktok_official_refresh")
        params = {
            "source": "tiktok_official_daily_sync",
            "trigger_type": "scheduled",
            "session_id": session_id,
            "business_ids": business_ids,
            "profile_days": profile_days,
            "max_pages": max_pages,
        }
        update_task_params_fn(task_id, params)
        if after_enqueue_fn:
            after_enqueue_fn(task_id, params)
        task_ids.append(task_id)
    return task_ids


def enqueue_publish_window_capture(
    create_task_fn,
    *,
    update_task_params_fn,
    after_enqueue_fn=None,
    max_items: int = 200,
) -> list[str]:
    """扫描到期未采集的时间点占位行（发布后3/24/48/72小时 + 发布日定格），按账号分组建 task_queue 任务交给 worker 消费。"""
    due_rows = db.query_all(
        """
        SELECT business_id, item_id, window_hours
        FROM tiktok_official_video_publish_window_snapshots
        WHERE due_at <= NOW() AND captured_at IS NULL
        ORDER BY due_at
        LIMIT %s
        """,
        (max_items,),
    ) or []
    if not due_rows:
        return []

    by_account: dict[str, list[dict[str, Any]]] = {}
    for row in due_rows:
        by_account.setdefault(row["business_id"], []).append(
            {"item_id": row["item_id"], "window_hours": row["window_hours"]}
        )

    task_ids = []
    for business_id, targets in by_account.items():
        task_id = pysecrets.token_hex(16)
        session_id = f"tiktok_official_publish_window_capture_{business_id}_{task_id[:8]}"
        create_task_fn(task_id, None, session_id, function_type="tiktok_official_publish_window_capture")
        params = {
            "source": "tiktok_official_publish_window_capture",
            "trigger_type": "scheduled",
            "session_id": session_id,
            "business_id": business_id,
            "targets": targets,
        }
        update_task_params_fn(task_id, params)
        if after_enqueue_fn:
            after_enqueue_fn(task_id, params)
        task_ids.append(task_id)
    return task_ids


def enqueue_video_discovery(
    create_task_fn,
    *,
    update_task_params_fn,
    after_enqueue_fn=None,
) -> list[str]:
    """每30分钟轻量轮询：只拉每个启用账号视频列表第1页，尽早发现新发布视频，
    让 upsert_video() 尽快建好 3/24/48/72h 占位行，避免因发现延迟导致"3h"数据实际是发布后6-9小时的状态。"""
    accounts = [
        a for a in list_accounts()
        if a.get("enabled", True) and (a.get("status") or "active") == "active"
    ]
    if not accounts:
        return []

    session_id = f"tiktok_official_video_discovery_{pysecrets.token_hex(8)}"
    task_ids = []
    for account in accounts:
        task_id = pysecrets.token_hex(16)
        business_id = account["business_id"]
        create_task_fn(task_id, None, session_id, function_type="tiktok_official_video_discovery")
        params = {
            "source": "tiktok_official_video_discovery",
            "trigger_type": "scheduled",
            "session_id": session_id,
            "business_id": business_id,
        }
        update_task_params_fn(task_id, params)
        if after_enqueue_fn:
            after_enqueue_fn(task_id, params)
        task_ids.append(task_id)
    return task_ids


def fetch_videos(token: str, business_id: str, max_pages: int = 5) -> tuple[list[dict[str, Any]], dict[str, str]]:
    videos: list[dict[str, Any]] = []
    cursor = None
    meta: dict[str, str] = {}
    for _ in range(max(1, max_pages)):
        params: dict[str, Any] = {
            "business_id": business_id,
            "fields": json.dumps(VIDEO_FIELDS, separators=(",", ":")),
            "max_count": 20,
        }
        if cursor is not None:
            params["cursor"] = cursor
        payload, headers = _request(token, "/business/video/list/", params)
        meta = {"request_id": payload.get("request_id") or "", "log_id": headers.get("X-Tt-Logid") or ""}
        data = payload.get("data") or {}
        videos.extend(data.get("videos") or [])
        if not data.get("has_more"):
            break
        cursor = data.get("cursor")
        if cursor is None:
            break
    return videos, meta


def fetch_profile(token: str, business_id: str, days: int = 30) -> tuple[dict[str, Any], dict[str, str]]:
    days = max(1, min(int(days or 30), 60))
    end = _business_date() - timedelta(days=1)
    start = end - timedelta(days=days - 1)
    params = {
        "business_id": business_id,
        "fields": json.dumps(PROFILE_FIELDS, separators=(",", ":")),
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
    }
    payload, headers = _request(token, "/business/get/", params)
    return payload.get("data") or {}, {
        "request_id": payload.get("request_id") or "",
        "log_id": headers.get("X-Tt-Logid") or "",
    }


def build_account_auth_url(public_base: str, state: str | None = None) -> str:
    from urllib.parse import urlencode

    app_id = (os.environ.get("TIKTOK_APP_ID") or os.environ.get("TIKTOK_CLIENT_KEY") or "").strip()
    if not app_id:
        raise RuntimeError("TIKTOK_APP_ID 未配置")
    scopes = [
        "user.info.basic",
        "user.info.username",
        "user.info.stats",
        "user.info.profile",
        "user.account.type",
        "user.insights",
        "video.list",
        "video.insights",
        "biz.ads.recommend",
    ]
    params = {
        "client_key": app_id,
        "scope": ",".join(scopes),
        "response_type": "code",
        "redirect_uri": f"{public_base.rstrip('/')}/tiktok/account/callback",
        "state": state or "tiktok_account",
    }
    return "https://www.tiktok.com/v2/auth/authorize?" + urlencode(params)


_dev_secret_key_cache: str | None = None


def _invite_signing_key() -> str:
    """签名密钥：优先复用 Flask 的 SECRET_KEY；未配置时退化为进程内随机密钥（仅供本地开发，重启后旧邀请链接失效）。"""
    global _dev_secret_key_cache
    key = os.environ.get("SECRET_KEY", "").strip()
    if key:
        return key
    if _dev_secret_key_cache is None:
        logger.warning("⚠️ SECRET_KEY 未配置，邀请链接签名将使用临时开发密钥，进程重启后已生成的邀请链接会失效")
        _dev_secret_key_cache = "dev-" + pysecrets.token_hex(32)
    return _dev_secret_key_cache


def _invite_serializer() -> URLSafeTimedSerializer:
    return URLSafeTimedSerializer(_invite_signing_key(), salt=INVITE_SALT)


def create_invite(account_alias: str, authorized_by: str | None = None, ttl_seconds: int | None = 24 * 3600) -> dict[str, Any]:
    """生成一条一次性授权邀请：写入 tiktok_official_invites，返回签名后的 state token。

    ttl_seconds 为 None 表示永不过期（仍然是一次性的，被使用后立即核销失效）。
    """
    account_alias = (account_alias or "").strip()
    if not account_alias:
        raise ValueError("account_alias 不能为空")
    expires_at = None
    if ttl_seconds is not None:
        ttl_seconds = max(300, min(int(ttl_seconds), 365 * 24 * 3600))
        expires_at = datetime.utcnow() + timedelta(seconds=ttl_seconds)
    nonce = pysecrets.token_hex(16)
    db.execute(
        """
        INSERT INTO tiktok_official_invites (nonce, account_alias, authorized_by, created_at, expires_at)
        VALUES (%s, %s, %s, NOW(), %s)
        """,
        (nonce, account_alias, authorized_by, expires_at),
    )
    state = _invite_serializer().dumps({"nonce": nonce, "account_alias": account_alias, "authorized_by": authorized_by})
    return {"state": state, "nonce": nonce, "account_alias": account_alias, "expires_at": expires_at}


def verify_and_consume_invite(state: str) -> dict[str, Any]:
    """校验授权回调带回的 state：签名有效、（若设置了有效期则）未过期、对应 invite 存在且未被使用过，成功后立即标记为已使用。

    过期时间完全由 tiktok_official_invites.expires_at 决定（NULL 代表永不过期），签名本身不做时间校验。
    失败时抛出 ValueError，message 可直接展示给用户。
    """
    try:
        payload = _invite_serializer().loads(state)
    except BadSignature as exc:
        raise ValueError("授权链接无效，请联系管理员重新生成 / Invalid link, please contact admin for a new one") from exc

    nonce = payload.get("nonce")
    row = db.query_one("SELECT * FROM tiktok_official_invites WHERE nonce = %s", (nonce,))
    if not row:
        raise ValueError("授权链接无效，请联系管理员重新生成 / Invalid link, please contact admin for a new one")
    if row.get("used_at"):
        raise ValueError("该授权链接已被使用过，请联系管理员重新生成一条新链接 / This link has already been used. Please contact admin for a new link.")
    if row.get("expires_at") and row["expires_at"] < datetime.utcnow():
        raise ValueError("授权链接已过期，请联系管理员重新生成 / This link has expired. Please contact admin for a new link.")

    updated = db.execute(
        "UPDATE tiktok_official_invites SET used_at = NOW() WHERE nonce = %s AND used_at IS NULL",
        (nonce,),
    )
    if not updated:
        # 并发场景下两次回调几乎同时到达，只有第一次能抢到这行 UPDATE
        raise ValueError("该授权链接已被使用过，请联系管理员重新生成一条新链接 / This link has already been used. Please contact admin for a new link.")

    return {
        "nonce": nonce,
        "account_alias": row.get("account_alias") or payload.get("account_alias"),
        "authorized_by": row.get("authorized_by") or payload.get("authorized_by"),
    }


def release_invite(nonce: str) -> None:
    """把已核销的邀请重新置为未使用，供 code 换 token 失败后重试同一条链接（避免因换 token 的瞬时失败而永久烧掉这条邀请）。"""
    db.execute("UPDATE tiktok_official_invites SET used_at = NULL WHERE nonce = %s", (nonce,))


def build_invite_link(account_alias: str, public_base: str, authorized_by: str | None = None, ttl_seconds: int | None = 24 * 3600) -> dict[str, Any]:
    invite = create_invite(account_alias, authorized_by=authorized_by, ttl_seconds=ttl_seconds)
    invite["url"] = build_account_auth_url(public_base, state=invite["state"])
    return invite


def list_invites(limit: int = 20, public_base: str | None = None) -> list[dict[str, Any]]:
    """返回最近的授权邀请记录。若传入 public_base，会为尚未使用且未过期的邀请重新签名出可用链接
    （签名只依赖 nonce，不校验签发时间，因此重新生成的 state 和当初发出的链接效果一致）。
    """
    rows = db.query_all(
        """
        SELECT nonce, account_alias, authorized_by, created_at, expires_at, used_at
        FROM tiktok_official_invites
        ORDER BY created_at DESC
        LIMIT %s
        """,
        (limit,),
    ) or []
    if public_base:
        now = datetime.utcnow()
        for row in rows:
            expired = bool(row.get("expires_at") and row["expires_at"] < now)
            if row.get("used_at") or expired:
                continue
            state = _invite_serializer().dumps({
                "nonce": row["nonce"],
                "account_alias": row.get("account_alias"),
                "authorized_by": row.get("authorized_by"),
            })
            row["url"] = build_account_auth_url(public_base, state=state)
    return rows


# ==================== Spark Ads 独立授权小流程 ====================
# 用另一个申请到 biz.spark.auth 权限的 TikTok App，单独走一遍账号持有人授权，
# 只用来补齐"生成 Spark 授权码"所需的权限，不写入 tiktok_official_accounts/tiktok_official_tokens 主表。
# 按 business_id（跟主表的真正唯一键一致）建表，与主表完全隔离，靠 business_id 对应回同一个真实账号。
# account_alias 只是展示用的别名，不保证唯一，不能用来做身份键（同名撞车会导致互相顶掉对方的 token）。

def create_spark_invite(business_id: str, account_alias: str | None = None, authorized_by: str | None = None, ttl_seconds: int | None = 3 * 24 * 3600) -> dict[str, Any]:
    business_id = (business_id or "").strip()
    if not business_id:
        raise ValueError("business_id 不能为空")
    account_alias = (account_alias or "").strip()
    expires_at = None
    if ttl_seconds is not None:
        ttl_seconds = max(300, min(int(ttl_seconds), 365 * 24 * 3600))
        expires_at = datetime.utcnow() + timedelta(seconds=ttl_seconds)
    nonce = pysecrets.token_hex(16)
    db.execute(
        """
        INSERT INTO tiktok_spark_invites (nonce, account_alias, business_id, authorized_by, created_at, expires_at)
        VALUES (%s, %s, %s, %s, NOW(), %s)
        """,
        (nonce, account_alias, business_id, authorized_by, expires_at),
    )
    state = _invite_serializer().dumps({"nonce": nonce, "account_alias": account_alias, "business_id": business_id, "authorized_by": authorized_by, "flow": "spark"})
    return {"state": state, "nonce": nonce, "account_alias": account_alias, "business_id": business_id, "expires_at": expires_at}


def verify_and_consume_spark_invite(state: str) -> dict[str, Any]:
    try:
        payload = _invite_serializer().loads(state)
    except BadSignature as exc:
        raise ValueError("授权链接无效，请联系管理员重新生成 / Invalid link, please contact admin for a new one") from exc

    nonce = payload.get("nonce")
    row = db.query_one("SELECT * FROM tiktok_spark_invites WHERE nonce = %s", (nonce,))
    if not row:
        raise ValueError("授权链接无效，请联系管理员重新生成 / Invalid link, please contact admin for a new one")
    if row.get("used_at"):
        raise ValueError("该授权链接已被使用过，请联系管理员重新生成一条新链接 / This link has already been used. Please contact admin for a new link.")
    if row.get("expires_at") and row["expires_at"] < datetime.utcnow():
        raise ValueError("授权链接已过期，请联系管理员重新生成 / This link has expired. Please contact admin for a new link.")

    updated = db.execute(
        "UPDATE tiktok_spark_invites SET used_at = NOW() WHERE nonce = %s AND used_at IS NULL",
        (nonce,),
    )
    if not updated:
        raise ValueError("该授权链接已被使用过，请联系管理员重新生成一条新链接 / This link has already been used. Please contact admin for a new link.")

    return {
        "nonce": nonce,
        "account_alias": row.get("account_alias") or payload.get("account_alias"),
        "business_id": row.get("business_id") or payload.get("business_id"),
        "authorized_by": row.get("authorized_by") or payload.get("authorized_by"),
    }


def release_spark_invite(nonce: str) -> None:
    db.execute("UPDATE tiktok_spark_invites SET used_at = NULL WHERE nonce = %s", (nonce,))


def build_spark_auth_url(public_base: str, state: str | None = None) -> str:
    from urllib.parse import urlencode

    app_id = (os.environ.get("TIKTOK_SPARK_APP_ID") or "").strip()
    if not app_id:
        raise RuntimeError("TIKTOK_SPARK_APP_ID 未配置")
    scopes_raw = (os.environ.get("TIKTOK_SPARK_SCOPES") or "user.info.basic,biz.spark.auth,video.list").strip()
    scopes = [s.strip() for s in scopes_raw.split(",") if s.strip()]
    params = {
        "client_key": app_id,
        "scope": ",".join(scopes),
        "response_type": "code",
        "redirect_uri": f"{public_base.rstrip('/')}/tiktok/spark/callback",
        "state": state or "tiktok_spark",
    }
    return "https://www.tiktok.com/v2/auth/authorize?" + urlencode(params)


def build_spark_invite_link(business_id: str, account_alias: str | None, public_base: str, authorized_by: str | None = None, ttl_seconds: int | None = 3 * 24 * 3600) -> dict[str, Any]:
    invite = create_spark_invite(business_id, account_alias=account_alias, authorized_by=authorized_by, ttl_seconds=ttl_seconds)
    invite["url"] = build_spark_auth_url(public_base, state=invite["state"])
    return invite


def list_spark_invites(limit: int = 20, public_base: str | None = None) -> list[dict[str, Any]]:
    rows = db.query_all(
        """
        SELECT nonce, account_alias, business_id, authorized_by, created_at, expires_at, used_at
        FROM tiktok_spark_invites
        ORDER BY created_at DESC
        LIMIT %s
        """,
        (limit,),
    ) or []
    if public_base:
        now = datetime.utcnow()
        for row in rows:
            expired = bool(row.get("expires_at") and row["expires_at"] < now)
            if row.get("used_at") or expired:
                continue
            state = _invite_serializer().dumps({
                "nonce": row["nonce"],
                "account_alias": row.get("account_alias"),
                "business_id": row.get("business_id"),
                "authorized_by": row.get("authorized_by"),
                "flow": "spark",
            })
            row["url"] = build_spark_auth_url(public_base, state=state)
    return rows


def build_spark_invite_batch(public_base: str, authorized_by: str | None = None) -> list[dict[str, Any]]:
    """给所有账号批量生成 Spark 授权邀请链接（已授权过的账号不重复生成，只标记状态）。

    按 business_id 判断"是否已授权"——account_alias 在主表里不保证唯一，两个不同账号
    如果撞了同一个别名，按别名判断会把没授权过的那个也误标成"已授权"。

    批量分发出去的链接不知道对方什么时候会点，统一用永不过期，避免还没发到人手上就失效。
    """
    accounts = db.query_all(
        "SELECT business_id, account_alias, account_name, region FROM tiktok_official_accounts "
        "WHERE account_alias IS NOT NULL AND account_alias <> '' ORDER BY account_name"
    ) or []
    authorized = {
        row["business_id"]
        for row in db.query_all("SELECT business_id FROM tiktok_spark_tokens WHERE business_id IS NOT NULL") or []
    }
    result = []
    for account in accounts:
        business_id = account["business_id"]
        alias = account["account_alias"]
        if business_id in authorized:
            result.append({
                "business_id": business_id,
                "account_alias": alias,
                "account_name": account.get("account_name"),
                "region": account.get("region"),
                "status": "已授权",
                "url": "",
            })
            continue
        invite = build_spark_invite_link(business_id, alias, public_base, authorized_by=authorized_by, ttl_seconds=None)
        result.append({
            "business_id": business_id,
            "account_alias": alias,
            "account_name": account.get("account_name"),
            "region": account.get("region"),
            "status": "待授权",
            "url": invite["url"],
        })
    return result


def build_spark_invite_batch_xlsx(public_base: str, authorized_by: str | None = None) -> bytes:
    rows = build_spark_invite_batch(public_base, authorized_by=authorized_by)
    wb = Workbook()
    ws = wb.active
    ws.title = "Spark授权链接"
    ws.append(["账号别名", "账号名称", "地区", "business_id", "状态", "Spark 授权链接"])
    for row in rows:
        ws.append([
            row["account_alias"],
            row.get("account_name"),
            row.get("region"),
            row.get("business_id"),
            row["status"],
            row["url"],
        ])
    ws.freeze_panes = "A2"
    _autosize_columns(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exchange_spark_code(code: str, redirect_uri: str, business_id: str, account_alias: str | None = None, authorized_by: str | None = None) -> dict[str, Any]:
    """用 Spark App 的授权 code 换 access_token，按 business_id 存入独立的 tiktok_spark_tokens。"""
    app_id = (os.environ.get("TIKTOK_SPARK_APP_ID") or "").strip()
    app_secret = (os.environ.get("TIKTOK_SPARK_APP_SECRET") or "").strip()
    if not app_id:
        raise RuntimeError("TIKTOK_SPARK_APP_ID 未配置")
    if not app_secret:
        raise RuntimeError("TIKTOK_SPARK_APP_SECRET 未配置")

    payload = {
        "client_key": app_id,
        "client_secret": app_secret,
        "code": unquote(code),
        "grant_type": "authorization_code",
        "redirect_uri": redirect_uri,
    }
    token_url = "https://open.tiktokapis.com/v2/oauth/token/"
    resp = requests.post(
        token_url,
        data=payload,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cache-Control": "no-cache",
        },
        timeout=60,
    )
    data = _parse_token_response(resp)
    if not data.get("access_token") or not data.get("open_id"):
        wrapped = data.get("data") if isinstance(data.get("data"), dict) else {}
        data = {**data, **wrapped}
    if not data.get("access_token") or not data.get("open_id"):
        raise RuntimeError(f"TikTok token 返回缺少 access_token/open_id: {json.dumps(data, ensure_ascii=False)[:500]}")

    save_spark_token(data, business_id=business_id, account_alias=account_alias, authorized_by=authorized_by)
    return data


def save_spark_token(token_data: dict[str, Any], business_id: str, account_alias: str | None = None, authorized_by: str | None = None) -> dict[str, Any]:
    business_id = (business_id or "").strip()
    if not business_id:
        raise ValueError("business_id 不能为空")
    account_alias = (account_alias or "").strip() or None
    open_id = str(token_data.get("open_id") or "").strip() or None
    access_token = str(token_data.get("access_token") or "").strip()
    if not access_token:
        raise ValueError("token_data 缺少 access_token")
    scope = token_data.get("scope")
    if isinstance(scope, list):
        scope = ",".join(scope)
    now = datetime.utcnow()
    expires_at = _seconds_from_now(now, token_data.get("expires_in"))
    refresh_expires_at = _seconds_from_now(now, token_data.get("refresh_expires_in"))

    db.execute(
        """
        INSERT INTO tiktok_spark_tokens (business_id, account_alias, open_id, access_token, refresh_token, scope, authorized_by, expires_at, refresh_expires_at, created_at, updated_at)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
        ON CONFLICT (business_id) DO UPDATE SET
            account_alias = COALESCE(EXCLUDED.account_alias, tiktok_spark_tokens.account_alias),
            open_id = COALESCE(EXCLUDED.open_id, tiktok_spark_tokens.open_id),
            access_token = EXCLUDED.access_token,
            refresh_token = COALESCE(EXCLUDED.refresh_token, tiktok_spark_tokens.refresh_token),
            scope = EXCLUDED.scope,
            authorized_by = COALESCE(EXCLUDED.authorized_by, tiktok_spark_tokens.authorized_by),
            expires_at = EXCLUDED.expires_at,
            refresh_expires_at = COALESCE(EXCLUDED.refresh_expires_at, tiktok_spark_tokens.refresh_expires_at),
            updated_at = NOW()
        """,
        (
            business_id,
            account_alias,
            open_id,
            crypto_util.encrypt(access_token),
            crypto_util.encrypt(token_data.get("refresh_token")),
            scope,
            authorized_by,
            expires_at,
            refresh_expires_at,
        ),
    )
    return {"business_id": business_id, "account_alias": account_alias, "open_id": open_id, "scope": scope}


def refresh_spark_token(business_id: str) -> str:
    """用存量 refresh_token 换新的 Spark access_token，成功后重新加密落库，返回新的明文 access_token。"""
    row = db.query_one(
        "SELECT refresh_token, account_alias, authorized_by FROM tiktok_spark_tokens WHERE business_id = %s",
        (business_id,),
    )
    refresh_token = crypto_util.decrypt((row or {}).get("refresh_token"))
    if not refresh_token:
        raise RuntimeError(f"business_id={business_id} 没有可用的 refresh_token，需要重新走一遍 Spark 授权")

    app_id = (os.environ.get("TIKTOK_SPARK_APP_ID") or "").strip()
    app_secret = (os.environ.get("TIKTOK_SPARK_APP_SECRET") or "").strip()
    if not app_id or not app_secret:
        raise RuntimeError("TIKTOK_SPARK_APP_ID / TIKTOK_SPARK_APP_SECRET 未配置")

    resp = requests.post(
        "https://open.tiktokapis.com/v2/oauth/token/",
        data={
            "client_key": app_id,
            "client_secret": app_secret,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
        },
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cache-Control": "no-cache",
        },
        timeout=60,
    )
    data = _parse_token_response(resp)
    if not data.get("access_token"):
        wrapped = data.get("data") if isinstance(data.get("data"), dict) else {}
        data = {**data, **wrapped}
    if not data.get("access_token"):
        raise RuntimeError(f"TikTok Spark refresh_token 返回缺少 access_token: {json.dumps(data, ensure_ascii=False)[:500]}")

    save_spark_token(data, business_id=business_id, account_alias=(row or {}).get("account_alias"), authorized_by=(row or {}).get("authorized_by"))
    logger.info(f"TikTok Spark token 已刷新: business_id={business_id}")
    return data["access_token"]


def get_spark_token_info(business_id: str | None, auto_refresh: bool = True) -> dict[str, Any] | None:
    """Spark token 是在另一个 App 下授权的，open_id 跟主流程的 business_id 不是同一个值。
    调用 TikTok 接口时 body 里的 business_id 必须跟 access_token 自己的身份一致，
    不能沿用主流程那个 business_id，否则 TikTok 会报 access token 不合法。

    按 business_id（而不是 account_alias）做键：account_alias 在主表里不保证唯一，
    按别名查会把撞名账号的 token 张冠李戴，报 "item_id does not belong to the user"。
    """
    if not business_id:
        return None
    row = db.query_one(
        "SELECT access_token, open_id, refresh_token, expires_at FROM tiktok_spark_tokens WHERE business_id = %s",
        (business_id,),
    )
    if not row or not row.get("access_token"):
        return None

    expires_at = row.get("expires_at")
    now = datetime.utcnow()
    # 历史遗留行没存过 expires_at（迁移前授权的），一律当作需要刷新处理，而不是当成永不过期。
    needs_refresh = expires_at is None or expires_at < now + timedelta(hours=1)
    if auto_refresh and needs_refresh and row.get("refresh_token"):
        try:
            return {"access_token": refresh_spark_token(business_id), "open_id": row.get("open_id")}
        except Exception as exc:
            if expires_at and expires_at < now:
                raise RuntimeError(f"Spark token 刷新失败且旧 token 已过期，需重新授权或稍后重试: {exc}") from exc
            logger.warning(f"TikTok Spark token 刷新失败，旧 token 尚未过期，暂用旧 token 兜底: business_id={business_id} err={exc}")

    return {"access_token": crypto_util.decrypt(row["access_token"]), "open_id": row.get("open_id")}


def list_spark_tokens(limit: int = 100) -> list[dict[str, Any]]:
    return db.query_all(
        """
        SELECT business_id, account_alias, open_id, scope, authorized_by, created_at, updated_at
        FROM tiktok_spark_tokens
        ORDER BY updated_at DESC
        LIMIT %s
        """,
        (limit,),
    ) or []


def exchange_business_code(code: str, authorized_by: str | None = None) -> dict[str, Any]:
    """用广告主授权（Business Portal）的 code 换 access_token，按 advertiser_id 存入 tiktok_advertiser_tokens。

    跟 exchange_spark_code/exchange_account_code 是完全不同的接口体系：这是 TikTok
    Marketing/Business API 自己的 OAuth（POST .../oauth2/access_token/，JSON body），
    不是 Login Kit 的 v2/oauth/token/（form body）。同样用 Spark App 的 app_id/secret，
    因为投流相关权限（Ads Management/Reporting等）是挂在 Spark App 下的。
    """
    app_id = (os.environ.get("TIKTOK_SPARK_APP_ID") or "").strip()
    app_secret = (os.environ.get("TIKTOK_SPARK_APP_SECRET") or "").strip()
    if not app_id:
        raise RuntimeError("TIKTOK_SPARK_APP_ID 未配置")
    if not app_secret:
        raise RuntimeError("TIKTOK_SPARK_APP_SECRET 未配置")

    resp = requests.post(
        "https://business-api.tiktok.com/open_api/v1.3/oauth2/access_token/",
        json={
            "app_id": app_id,
            "secret": app_secret,
            "auth_code": unquote(code),
        },
        headers={"Content-Type": "application/json"},
        timeout=60,
    )
    try:
        payload = resp.json()
    except Exception:
        raise RuntimeError(f"广告主授权换 token 失败，响应非 JSON（HTTP {resp.status_code}）: {resp.text[:500]}")
    if payload.get("code") != 0:
        raise RuntimeError(f"广告主授权换 token 失败: {json.dumps(payload, ensure_ascii=False)[:500]}")

    data = payload.get("data") or {}
    access_token = str(data.get("access_token") or "").strip()
    advertiser_ids = data.get("advertiser_ids") or []
    if not access_token or not advertiser_ids:
        raise RuntimeError(f"广告主授权返回缺少 access_token/advertiser_ids: {json.dumps(data, ensure_ascii=False)[:500]}")

    scope = data.get("scope")
    if isinstance(scope, list):
        scope = ",".join(str(s) for s in scope)

    save_advertiser_tokens(advertiser_ids, access_token, scope=scope, authorized_by=authorized_by)
    return {"advertiser_ids": advertiser_ids, "scope": scope}


def save_advertiser_tokens(advertiser_ids: list[str], access_token: str, scope: Any = None, authorized_by: str | None = None) -> None:
    encrypted = crypto_util.encrypt(access_token)
    for advertiser_id in advertiser_ids:
        db.execute(
            """
            INSERT INTO tiktok_advertiser_tokens (advertiser_id, access_token, scope, authorized_by, created_at, updated_at)
            VALUES (%s, %s, %s, %s, NOW(), NOW())
            ON CONFLICT (advertiser_id) DO UPDATE SET
                access_token = EXCLUDED.access_token,
                scope = EXCLUDED.scope,
                authorized_by = COALESCE(EXCLUDED.authorized_by, tiktok_advertiser_tokens.authorized_by),
                updated_at = NOW()
            """,
            (str(advertiser_id), encrypted, scope, authorized_by),
        )


def list_advertiser_tokens(limit: int = 100) -> list[dict[str, Any]]:
    return db.query_all(
        """
        SELECT advertiser_id, scope, authorized_by, created_at, updated_at
        FROM tiktok_advertiser_tokens
        ORDER BY updated_at DESC
        LIMIT %s
        """,
        (limit,),
    ) or []


def get_advertiser_token_info(advertiser_id: str | None) -> dict[str, Any] | None:
    if not advertiser_id:
        return None
    row = db.query_one(
        "SELECT access_token FROM tiktok_advertiser_tokens WHERE advertiser_id = %s",
        (advertiser_id,),
    )
    if not row or not row.get("access_token"):
        return None
    return {"access_token": crypto_util.decrypt(row["access_token"])}


AD_SPEND_METRICS = [
    "spend",
    "impressions",
    "clicks",
    "conversion",
    "video_play_actions",
]

# 只有这三种 objective 的 campaign 有可能是给矩阵已发布视频加热的 Spark Ads；
# APP_PROMOTION/WEB_CONVERSIONS/LEAD_GENERATION 用的是 App 安装/落地页素材，
# 不会带我们视频的 tiktok_item_id（已用生产数据验证过，91 个有消耗的账户里
# 1509 个 campaign 中 77% 是 APP_PROMOTION，抽查确认跟矩阵无关）。
AD_VIDEO_OBJECTIVE_TYPES = {"VIDEO_VIEWS", "TRAFFIC", "ENGAGEMENT"}

AD_SPEND_MAX_SPAN_DAYS = 30  # TikTok 报表接口限制：dimensions 带 stat_time_day 时单次请求时间跨度不能超过30天


def _chunk_date_range(start_date: date, end_date: date, max_span_days: int) -> list[tuple[date, date]]:
    chunks = []
    cur = start_date
    while cur <= end_date:
        chunk_end = min(cur + timedelta(days=max_span_days - 1), end_date)
        chunks.append((cur, chunk_end))
        cur = chunk_end + timedelta(days=1)
    return chunks


def _chunk_list(items: list[Any], size: int) -> list[list[Any]]:
    return [items[i : i + size] for i in range(0, len(items), size)]


def sync_ad_video_map_for_advertiser(advertiser_id: str) -> list[str]:
    """从 campaign/ad 明细里找出哪些 ad 是在给我们自己矩阵的视频投流，upsert 进 tiktok_ad_video_map。

    只扫 VIDEO_VIEWS/TRAFFIC/ENGAGEMENT 三种 objective 的 campaign（见 AD_VIDEO_OBJECTIVE_TYPES
    注释），逐条 ad 取 tiktok_item_id，跟 tiktok_official_video_snapshots.item_id 做匹配，
    只有真正命中我们自己视频的 ad 才会被记录。返回本次命中的 ad_id 列表。
    """
    info = get_advertiser_token_info(advertiser_id)
    if not info:
        raise RuntimeError(f"advertiser_id={advertiser_id} 没有已保存的 access_token")
    token = info["access_token"]

    campaign_ids: list[str] = []
    objective_by_campaign: dict[str, str] = {}
    page = 1
    while True:
        payload, _headers = _request(
            token, "/campaign/get/", {"advertiser_id": advertiser_id, "page_size": 100, "page": page}
        )
        data = payload.get("data") or {}
        for camp in data.get("list") or []:
            if camp.get("objective_type") in AD_VIDEO_OBJECTIVE_TYPES:
                campaign_ids.append(camp["campaign_id"])
                objective_by_campaign[camp["campaign_id"]] = camp.get("objective_type")
        page_info = data.get("page_info") or {}
        if page >= int(page_info.get("total_page") or 1):
            break
        page += 1

    if not campaign_ids:
        return []

    candidates: list[dict[str, Any]] = []
    for batch in _chunk_list(campaign_ids, 100):
        page = 1
        while True:
            payload, _headers = _request(
                token,
                "/ad/get/",
                {
                    "advertiser_id": advertiser_id,
                    "filtering": json.dumps({"campaign_ids": batch}, separators=(",", ":")),
                    "page_size": 100,
                    "page": page,
                },
            )
            data = payload.get("data") or {}
            for ad in data.get("list") or []:
                tiktok_item_id = ad.get("tiktok_item_id")
                if tiktok_item_id:
                    candidates.append(
                        {
                            "ad_id": ad["ad_id"],
                            "adgroup_id": ad.get("adgroup_id"),
                            "campaign_id": ad.get("campaign_id"),
                            "tiktok_item_id": tiktok_item_id,
                        }
                    )
            page_info = data.get("page_info") or {}
            if page >= int(page_info.get("total_page") or 1):
                break
            page += 1

    if not candidates:
        return []

    candidate_items = list({c["tiktok_item_id"] for c in candidates})
    our_items = {
        r["item_id"]
        for r in db.query_all(
            "SELECT item_id FROM tiktok_official_video_snapshots WHERE item_id = ANY(%s)",
            (candidate_items,),
        )
        or []
    }

    matched_ad_ids: list[str] = []
    for c in candidates:
        if c["tiktok_item_id"] not in our_items:
            continue
        db.execute(
            """
            INSERT INTO tiktok_ad_video_map
                (ad_id, advertiser_id, campaign_id, adgroup_id, tiktok_item_id, objective_type, first_seen_at, last_seen_at)
            VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())
            ON CONFLICT (ad_id) DO UPDATE SET
                advertiser_id = EXCLUDED.advertiser_id,
                campaign_id = EXCLUDED.campaign_id,
                adgroup_id = EXCLUDED.adgroup_id,
                tiktok_item_id = EXCLUDED.tiktok_item_id,
                objective_type = EXCLUDED.objective_type,
                last_seen_at = NOW()
            """,
            (
                c["ad_id"],
                advertiser_id,
                c["campaign_id"],
                c["adgroup_id"],
                c["tiktok_item_id"],
                objective_by_campaign.get(c["campaign_id"]),
            ),
        )
        matched_ad_ids.append(c["ad_id"])
    return matched_ad_ids


def sync_ad_spend_for_advertiser(advertiser_id: str, start_date: date, end_date: date) -> int:
    """拉取单个 advertiser_id 名下已归因到矩阵视频的 ad 在 [start_date, end_date] 的消耗数据。

    ad_id 集合来自 tiktok_ad_video_map 的累计记录（不只是本次新发现的），这样即使某条 ad
    所在的 campaign 后来被暂停/归档、不再出现在 campaign/get 的当前列表里，历史消耗也能续拉。
    没有任何命中的 ad_id 时直接返回 0，不发请求。返回 upsert 的行数。
    """
    mapped = db.query_all(
        "SELECT ad_id, tiktok_item_id FROM tiktok_ad_video_map WHERE advertiser_id = %s",
        (advertiser_id,),
    ) or []
    if not mapped:
        return 0
    item_id_by_ad = {m["ad_id"]: m["tiktok_item_id"] for m in mapped}
    ad_ids = list(item_id_by_ad.keys())

    info = get_advertiser_token_info(advertiser_id)
    if not info:
        raise RuntimeError(f"advertiser_id={advertiser_id} 没有已保存的 access_token")
    token = info["access_token"]

    rows: list[dict[str, Any]] = []
    for chunk_start, chunk_end in _chunk_date_range(start_date, end_date, AD_SPEND_MAX_SPAN_DAYS):
        for batch in _chunk_list(ad_ids, 100):
            page = 1
            while True:
                params = {
                    "advertiser_id": advertiser_id,
                    "report_type": "BASIC",
                    "data_level": "AUCTION_AD",
                    "dimensions": json.dumps(["ad_id", "stat_time_day", "country_code"], separators=(",", ":")),
                    "metrics": json.dumps(AD_SPEND_METRICS, separators=(",", ":")),
                    "filtering": json.dumps(
                        [{"field_name": "ad_ids", "filter_type": "IN", "filter_value": json.dumps(batch)}],
                        separators=(",", ":"),
                    ),
                    "start_date": chunk_start.isoformat(),
                    "end_date": chunk_end.isoformat(),
                    "page_size": 1000,
                    "page": page,
                }
                payload, _headers = _request(token, "/report/integrated/get/", params)
                data = payload.get("data") or {}
                rows.extend(data.get("list") or [])
                page_info = data.get("page_info") or {}
                if page >= int(page_info.get("total_page") or 1):
                    break
                page += 1

    upserted = 0
    for row in rows:
        dims = row.get("dimensions") or {}
        metrics = row.get("metrics") or {}
        stat_date = dims.get("stat_time_day")
        ad_id = dims.get("ad_id")
        if not stat_date or not ad_id:
            continue
        stat_date = str(stat_date)[:10]
        country_code = dims.get("country_code") or ""
        db.execute(
            """
            INSERT INTO tiktok_ad_spend_daily
                (ad_id, advertiser_id, tiktok_item_id, stat_date, country_code, spend, impressions, clicks,
                 conversions, video_play_actions, raw_json, fetched_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (ad_id, stat_date, country_code) DO UPDATE SET
                advertiser_id = EXCLUDED.advertiser_id,
                tiktok_item_id = EXCLUDED.tiktok_item_id,
                spend = EXCLUDED.spend,
                impressions = EXCLUDED.impressions,
                clicks = EXCLUDED.clicks,
                conversions = EXCLUDED.conversions,
                video_play_actions = EXCLUDED.video_play_actions,
                raw_json = EXCLUDED.raw_json,
                fetched_at = NOW()
            """,
            (
                ad_id,
                advertiser_id,
                item_id_by_ad.get(ad_id),
                stat_date,
                country_code,
                metrics.get("spend") or 0,
                int(float(metrics.get("impressions") or 0)),
                int(float(metrics.get("clicks") or 0)),
                int(float(metrics.get("conversion") or 0)),
                int(float(metrics.get("video_play_actions") or 0)),
                json.dumps(row, ensure_ascii=False),
            ),
        )
        upserted += 1
    return upserted


def sync_all_ad_spend(start_date: date, end_date: date) -> dict[str, Any]:
    """遍历全部已授权 advertiser_id：先刷新 ad→视频归因映射，再只拉命中的 ad 的消耗。

    单个账户失败不影响其他账户（沿用旧版账户级同步已验证过的失败隔离模式）。
    """
    tokens = list_advertiser_tokens(limit=1000)
    ok = 0
    failed: list[dict[str, str]] = []
    total_upserted = 0
    for row in tokens:
        advertiser_id = row["advertiser_id"]
        try:
            sync_ad_video_map_for_advertiser(advertiser_id)
            total_upserted += sync_ad_spend_for_advertiser(advertiser_id, start_date, end_date)
            ok += 1
        except Exception as exc:
            logger.error(f"投流消耗同步失败 advertiser_id={advertiser_id}: {exc}")
            failed.append({"advertiser_id": advertiser_id, "error": str(exc)})
    return {"ok": ok, "failed": failed, "total_advertisers": len(tokens), "upserted_rows": total_upserted}


def run_ad_spend_sync_task(task_id: str, params: dict[str, Any], update_task_fn) -> None:
    """worker 每日自触发入口：默认只拉「昨天」一天，账户内部失败已由 sync_all_ad_spend 隔离。"""
    try:
        update_task_fn(task_id, status="processing", progress="正在同步投流消耗数据...")
        target = _business_date() - timedelta(days=1)
        start_str = params.get("start_date") or target.isoformat()
        end_str = params.get("end_date") or target.isoformat()
        start_date = date.fromisoformat(start_str)
        end_date = date.fromisoformat(end_str)
        result = sync_all_ad_spend(start_date, end_date)
        progress = (
            f"完成（{result['ok']}/{result['total_advertisers']} 个账户成功，"
            f"upsert {result['upserted_rows']} 行，失败 {len(result['failed'])} 个）"
        )
        update_task_fn(task_id, status="completed", progress=progress)
    except Exception as exc:
        logger.error(f"❌ TikTok 投流消耗每日同步失败: {exc}")
        update_task_fn(task_id, status="failed", error=str(exc)[:500], progress="失败")


def enqueue_ad_spend_sync(
    create_task_fn,
    *,
    update_task_params_fn,
    after_enqueue_fn=None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[str]:
    """投流消耗每日同步：单个任务遍历全部已授权 advertiser_id（账户级失败隔离已在 sync_all_advertiser_spend 里做了，不需要按账户分批）。"""
    session_id = f"tiktok_ad_spend_sync_{_business_date().isoformat()}"
    task_id = pysecrets.token_hex(16)
    create_task_fn(task_id, None, session_id, function_type="tiktok_official_ad_spend_sync")
    params = {
        "source": "tiktok_ad_spend_sync",
        "trigger_type": "scheduled",
        "session_id": session_id,
        "start_date": start_date,
        "end_date": end_date,
    }
    update_task_params_fn(task_id, params)
    if after_enqueue_fn:
        after_enqueue_fn(task_id, params)
    return [task_id]


def get_ad_spend_summary(date_from: str | None = None, date_to: str | None = None) -> dict[str, Any]:
    """按日汇总全部账户的消耗/曝光/点击/转化，供趋势图使用。"""
    start_date, end_date = _matrix_date_range(date_from, date_to, days=7)
    rows = db.query_all(
        """
        SELECT stat_date,
               SUM(spend)::float AS spend,
               SUM(impressions)::bigint AS impressions,
               SUM(clicks)::bigint AS clicks,
               SUM(conversions)::bigint AS conversions
        FROM tiktok_ad_spend_daily
        WHERE stat_date BETWEEN %s AND %s
        GROUP BY stat_date
        ORDER BY stat_date
        """,
        (start_date, end_date),
    ) or []
    return {"date_from": start_date.isoformat(), "date_to": end_date.isoformat(), "rows": rows}


def get_ad_spend_by_country(date_from: str | None = None, date_to: str | None = None) -> dict[str, Any]:
    """按国家汇总消耗/曝光/转化，country_code 为空字符串表示 TikTok 未归类到具体国家的流量。"""
    start_date, end_date = _matrix_date_range(date_from, date_to, days=7)
    rows = db.query_all(
        """
        SELECT country_code,
               SUM(spend)::float AS spend,
               SUM(impressions)::bigint AS impressions,
               SUM(clicks)::bigint AS clicks,
               SUM(conversions)::bigint AS conversions
        FROM tiktok_ad_spend_daily
        WHERE stat_date BETWEEN %s AND %s
        GROUP BY country_code
        ORDER BY spend DESC
        """,
        (start_date, end_date),
    ) or []
    return {"date_from": start_date.isoformat(), "date_to": end_date.isoformat(), "rows": rows}


def get_ad_spend_by_video(date_from: str | None = None, date_to: str | None = None) -> dict[str, Any]:
    """按视频归因的投流消耗明细，直接回答"哪个账号/哪条视频花了多少钱"。"""
    start_date, end_date = _matrix_date_range(date_from, date_to, days=7)
    rows = db.query_all(
        """
        SELECT s.tiktok_item_id, v.business_id, v.caption, v.thumbnail_url,
               SUM(s.spend)::float AS spend,
               SUM(s.impressions)::bigint AS impressions,
               SUM(s.video_play_actions)::bigint AS video_play_actions
        FROM tiktok_ad_spend_daily s
        JOIN tiktok_official_video_snapshots v ON v.item_id = s.tiktok_item_id
        WHERE s.stat_date BETWEEN %s AND %s
        GROUP BY s.tiktok_item_id, v.business_id, v.caption, v.thumbnail_url
        ORDER BY spend DESC
        """,
        (start_date, end_date),
    ) or []
    return {"date_from": start_date.isoformat(), "date_to": end_date.isoformat(), "rows": rows}


def get_paid_traffic_ratio(date_from: str | None = None, date_to: str | None = None) -> dict[str, Any]:
    """投流流量占比：仅覆盖被投流命中的视频，不是全矩阵口径。

    分子是这些视频在投放报表里的 video_play_actions（TikTok 报表口径的每日增量，直接 SUM 即可）。
    分母是同一批视频在区间内的**总播放增量**——注意 TikTok 的 video_views 本身已经包含了
    广告带来的播放，所以分母不能再写成「付费播放 + video_views 增量」，那样等于把付费播放
    算了两遍，占比会被系统性压低（实测 24.4% vs 修正后 32.3%）。
    真实自然播放 = 总播放增量 - 付费播放。

    快照表存的是「截至当天」的累计播放量，不能直接 SUM 多天，要用区间首尾快照做差。
    起点取 snapshot_date < date_from 的最近一次快照（不是 <=），这样增量覆盖的是
    [date_from, date_to] 闭区间，跟分子的 stat_date BETWEEN 对齐；用 <= 会让分母少算
    date_from 当天的自然增长，而分子却算了当天的投放。
    """
    start_date, end_date = _matrix_date_range(date_from, date_to, days=7)
    paid_row = db.query_one(
        """
        SELECT COALESCE(SUM(spend), 0)::float AS spend,
               COALESCE(SUM(video_play_actions), 0) AS paid_views
        FROM tiktok_ad_spend_daily WHERE stat_date BETWEEN %s AND %s AND tiktok_item_id IS NOT NULL
        """,
        (start_date, end_date),
    ) or {}
    delta_row = db.query_one(
        """
        WITH items AS (
            SELECT DISTINCT tiktok_item_id FROM tiktok_ad_spend_daily
            WHERE stat_date BETWEEN %(from)s AND %(to)s AND tiktok_item_id IS NOT NULL
        ),
        start_snap AS (
            SELECT DISTINCT ON (d.item_id) d.item_id, d.video_views
            FROM tiktok_official_video_daily_snapshots d
            JOIN items i ON i.tiktok_item_id = d.item_id
            WHERE d.snapshot_date < %(from)s
            ORDER BY d.item_id, d.snapshot_date DESC
        ),
        end_snap AS (
            SELECT DISTINCT ON (d.item_id) d.item_id, d.video_views
            FROM tiktok_official_video_daily_snapshots d
            JOIN items i ON i.tiktok_item_id = d.item_id
            WHERE d.snapshot_date <= %(to)s
            ORDER BY d.item_id, d.snapshot_date DESC
        )
        SELECT COALESCE(SUM(GREATEST(e.video_views - COALESCE(s.video_views, 0), 0)), 0) AS total_views_delta
        FROM end_snap e
        LEFT JOIN start_snap s ON s.item_id = e.item_id
        """,
        {"from": start_date, "to": end_date},
    ) or {}
    paid_views = int(paid_row.get("paid_views") or 0)
    total_views_delta = int(delta_row.get("total_views_delta") or 0)
    # 自然播放是推导量：总增量里扣掉付费部分。数据不一致时（付费 > 总增量）兜底为 0。
    organic_views = max(total_views_delta - paid_views, 0)
    ratio = (paid_views / total_views_delta) if total_views_delta else 0.0
    return {
        "spend": float(paid_row.get("spend") or 0),
        "paid_views": paid_views,
        "organic_views": organic_views,
        "total_views_delta": total_views_delta,
        "paid_ratio": ratio,
        "date_from": start_date.isoformat(),
        "date_to": end_date.isoformat(),
    }


def exchange_account_code(code: str, redirect_uri: str, account_alias: str | None = None, authorized_by: str | None = None) -> dict[str, Any]:
    """用 TikTok 账号授权 code 换 access_token，并保存 open_id 作为官号 business_id。"""
    app_id = (os.environ.get("TIKTOK_APP_ID") or os.environ.get("TIKTOK_CLIENT_KEY") or "").strip()
    app_secret = (os.environ.get("TIKTOK_APP_SECRET") or os.environ.get("TIKTOK_CLIENT_SECRET") or "").strip()
    if not app_id:
        raise RuntimeError("TIKTOK_APP_ID 未配置")
    if not app_secret:
        raise RuntimeError("TIKTOK_APP_SECRET 未配置")

    payload = {
        "client_key": app_id,
        "client_secret": app_secret,
        "code": unquote(code),
        "grant_type": "authorization_code",
        "redirect_uri": redirect_uri,
    }
    token_url = "https://open.tiktokapis.com/v2/oauth/token/"
    resp = requests.post(
        token_url,
        data=payload,
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cache-Control": "no-cache",
        },
        timeout=60,
    )
    data = _parse_token_response(resp)
    if not data.get("access_token") or not data.get("open_id"):
        # Some TikTok endpoints wrap token payload in data.
        wrapped = data.get("data") if isinstance(data.get("data"), dict) else {}
        data = {**data, **wrapped}
    if not data.get("access_token") or not data.get("open_id"):
        raise RuntimeError(f"TikTok token 返回缺少 access_token/open_id: {json.dumps(data, ensure_ascii=False)[:500]}")

    account_status = save_account_token(data, account_alias=account_alias, authorized_by=authorized_by)
    data["_account_status"] = account_status
    return data


def save_account_token(token_data: dict[str, Any], account_alias: str | None = None, authorized_by: str | None = None) -> dict[str, Any]:
    open_id = str(token_data.get("open_id") or "").strip()
    access_token = str(token_data.get("access_token") or "").strip()
    if not open_id or not access_token:
        raise ValueError("token_data 缺少 open_id/access_token")
    previous_row = db.query_one(
        "SELECT account_alias FROM tiktok_official_accounts WHERE business_id = %s", (open_id,)
    )
    now = datetime.utcnow()
    expires_at = _seconds_from_now(now, token_data.get("expires_in"))
    refresh_expires_at = _seconds_from_now(now, token_data.get("refresh_expires_in"))
    scope = token_data.get("scope")
    if isinstance(scope, list):
        scope = ",".join(scope)
    db.execute(
        """
        INSERT INTO tiktok_official_tokens (
            token_type, open_id, access_token, refresh_token, scope,
            expires_at, refresh_expires_at, raw_json, account_alias, authorized_by, status, updated_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'active', NOW())
        ON CONFLICT (open_id) DO UPDATE SET
            access_token = EXCLUDED.access_token,
            refresh_token = EXCLUDED.refresh_token,
            scope = EXCLUDED.scope,
            expires_at = EXCLUDED.expires_at,
            refresh_expires_at = EXCLUDED.refresh_expires_at,
            raw_json = EXCLUDED.raw_json,
            account_alias = COALESCE(EXCLUDED.account_alias, tiktok_official_tokens.account_alias),
            authorized_by = COALESCE(EXCLUDED.authorized_by, tiktok_official_tokens.authorized_by),
            status = 'active',
            updated_at = NOW()
        """,
        (
            "account",
            open_id,
            crypto_util.encrypt(access_token),
            crypto_util.encrypt(token_data.get("refresh_token")),
            scope,
            expires_at,
            refresh_expires_at,
            json.dumps(token_data, ensure_ascii=False),
            account_alias,
            authorized_by,
        ),
    )
    db.execute(
        """
        INSERT INTO tiktok_official_accounts (business_id, account_name, enabled, notes, account_alias, authorized_by, status, updated_at)
        VALUES (%s, %s, TRUE, %s, %s, %s, 'active', NOW())
        ON CONFLICT (business_id) DO UPDATE SET
            enabled = TRUE,
            account_alias = COALESCE(EXCLUDED.account_alias, tiktok_official_accounts.account_alias),
            authorized_by = COALESCE(EXCLUDED.authorized_by, tiktok_official_accounts.authorized_by),
            status = 'active',
            updated_at = NOW()
        """,
        (
            open_id,
            account_alias or token_data.get("display_name") or f"TikTok {open_id[-6:]}",
            "OAuth authorized account",
            account_alias,
            authorized_by,
        ),
    )
    return {
        "is_new": previous_row is None,
        "previous_alias": previous_row["account_alias"] if previous_row else None,
    }


def refresh_account_token(open_id: str) -> str:
    """用存量 refresh_token 换新的 access_token，成功后重新加密落库，返回新的明文 access_token。"""
    row = db.query_one(
        "SELECT refresh_token, account_alias, authorized_by FROM tiktok_official_tokens WHERE open_id = %s",
        (open_id,),
    )
    refresh_token = crypto_util.decrypt((row or {}).get("refresh_token"))
    if not refresh_token:
        raise RuntimeError(f"open_id={open_id} 没有可用的 refresh_token，需要重新走一遍 OAuth 授权")

    app_id = (os.environ.get("TIKTOK_APP_ID") or os.environ.get("TIKTOK_CLIENT_KEY") or "").strip()
    app_secret = (os.environ.get("TIKTOK_APP_SECRET") or os.environ.get("TIKTOK_CLIENT_SECRET") or "").strip()
    if not app_id or not app_secret:
        raise RuntimeError("TIKTOK_APP_ID / TIKTOK_APP_SECRET 未配置")

    resp = requests.post(
        "https://open.tiktokapis.com/v2/oauth/token/",
        data={
            "client_key": app_id,
            "client_secret": app_secret,
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
        },
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Cache-Control": "no-cache",
        },
        timeout=60,
    )
    data = _parse_token_response(resp)
    if not data.get("access_token"):
        wrapped = data.get("data") if isinstance(data.get("data"), dict) else {}
        data = {**data, **wrapped}
    if not data.get("access_token"):
        raise RuntimeError(f"TikTok refresh_token 返回缺少 access_token: {json.dumps(data, ensure_ascii=False)[:500]}")

    data.setdefault("open_id", open_id)
    save_account_token(data, account_alias=(row or {}).get("account_alias"), authorized_by=(row or {}).get("authorized_by"))
    logger.info(f"TikTok token 已刷新: open_id={open_id}")
    return data["access_token"]


def get_access_token(business_id: str | None = None, auto_refresh: bool = True) -> str:
    env_token = os.environ.get("TIKTOK_BUSINESS_ACCESS_TOKEN", "").strip()
    if env_token:
        return env_token
    if business_id:
        row = db.query_one(
            """
            SELECT open_id, access_token, refresh_token, expires_at FROM tiktok_official_tokens
            WHERE open_id = %s
            ORDER BY updated_at DESC
            LIMIT 1
            """,
            (business_id,),
        )
    else:
        row = db.query_one(
            """
            SELECT open_id, access_token, refresh_token, expires_at FROM tiktok_official_tokens
            ORDER BY updated_at DESC
            LIMIT 1
            """
        )
    if not row:
        return ""

    expires_at = row.get("expires_at")
    now = datetime.utcnow()
    if auto_refresh and expires_at and expires_at < now + timedelta(hours=1) and row.get("refresh_token"):
        last_exc = None
        for attempt in range(2):  # 刷新接口偶发抖动，重试一次再判定
            try:
                return refresh_account_token(row["open_id"])
            except Exception as exc:
                last_exc = exc
                logger.warning(f"TikTok token 刷新失败(第{attempt + 1}次): open_id={row.get('open_id')} err={exc}")
        if expires_at < now:
            # 旧 token 已经确定过期，回退等于必然在后续接口调用报「access token 已吊销/不正确」，
            # 不如直接抛出，让调用方（如批量刷新）感知到并跳过，而不是掩盖成一个更难排查的下游错误
            raise RuntimeError(f"token 刷新失败且旧 token 已过期，需重新授权或稍后重试: {last_exc}")
        logger.warning(f"TikTok token 刷新失败，旧 token 尚未过期，暂用旧 token 兜底: open_id={row.get('open_id')}")
    return crypto_util.decrypt(row.get("access_token")) or ""


def _parse_token_response(resp: requests.Response) -> dict[str, Any]:
    try:
        data = resp.json()
    except Exception as exc:
        raise RuntimeError(f"TikTok token 返回非 JSON: HTTP {resp.status_code} {resp.text[:300]}") from exc
    if resp.status_code >= 400 or data.get("error") or data.get("code") not in (0, "0", None):
        message = data.get("error_description") or data.get("message") or data.get("error") or resp.text[:300]
        raise RuntimeError(f"TikTok token 错误: {message}")
    return data


def _seconds_from_now(now: datetime, seconds) -> datetime | None:
    try:
        if seconds in (None, ""):
            return None
        return now + timedelta(seconds=int(seconds))
    except Exception:
        return None


def _request(token: str, path: str, params: dict[str, Any], max_retries: int = 3) -> tuple[dict[str, Any], requests.structures.CaseInsensitiveDict]:
    last_exc: Exception | None = None
    for attempt in range(max_retries):
        resp = requests.get(
            f"{API_BASE}{path}",
            params=params,
            headers={"Access-Token": token},
            timeout=60,
        )
        if resp.status_code == 429 or resp.status_code >= 500:
            last_exc = RuntimeError(f"TikTok API 限流/服务端错误: HTTP {resp.status_code} {resp.text[:200]}")
            if attempt < max_retries - 1:
                wait = 2 ** attempt
                logger.warning(f"TikTok API {path} 返回 {resp.status_code}，{wait}s 后重试（第 {attempt + 1}/{max_retries} 次）")
                time.sleep(wait)
                continue
            raise last_exc
        try:
            payload = resp.json()
        except Exception as exc:
            raise RuntimeError(f"TikTok API 返回非 JSON: HTTP {resp.status_code} {resp.text[:300]}") from exc
        if resp.status_code >= 400 or payload.get("code") not in (0, "0", None):
            request_id = payload.get("request_id") or resp.headers.get("X-Tt-Logid") or ""
            raise RuntimeError(f"TikTok API 错误: {payload.get('message') or resp.text[:200]} request_id={request_id}")
        return payload, resp.headers
    raise last_exc or RuntimeError("TikTok API 请求失败")


def upsert_profile(business_id: str, data: dict[str, Any], meta: dict[str, str]) -> None:
    db.execute(
        """
        UPDATE tiktok_official_accounts SET
            display_name = %s,
            username = %s,
            profile_deep_link = %s,
            profile_image = %s,
            bio_description = %s,
            is_business_account = %s,
            is_verified = %s,
            following_count = %s,
            followers_count = %s,
            total_likes = %s,
            videos_count = %s,
            raw_json = %s,
            updated_at = NOW()
        WHERE business_id = %s
        """,
        (
            data.get("display_name"),
            data.get("username"),
            data.get("profile_deep_link"),
            data.get("profile_image"),
            data.get("bio_description"),
            data.get("is_business_account"),
            data.get("is_verified"),
            _to_int(data.get("following_count")),
            _to_int(data.get("followers_count")),
            _to_int(data.get("total_likes")),
            _to_int(data.get("videos_count")),
            json.dumps({"data": data, "meta": meta}, ensure_ascii=False),
            business_id,
        ),
    )
    for metric in data.get("metrics") or []:
        metric_date = metric.get("date")
        if not metric_date:
            continue
        db.execute(
            """
            INSERT INTO tiktok_official_profile_daily_metrics (
                business_id, metric_date, followers_count, video_views, unique_video_views,
                profile_views, likes, comments, shares, phone_number_clicks, lead_submissions,
                app_download_clicks, bio_link_clicks, email_clicks, address_clicks,
                daily_total_followers, daily_new_followers, daily_lost_followers,
                engaged_audience, raw_json, fetched_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
            ON CONFLICT (business_id, metric_date) DO UPDATE SET
                followers_count = EXCLUDED.followers_count,
                video_views = EXCLUDED.video_views,
                unique_video_views = EXCLUDED.unique_video_views,
                profile_views = EXCLUDED.profile_views,
                likes = EXCLUDED.likes,
                comments = EXCLUDED.comments,
                shares = EXCLUDED.shares,
                daily_total_followers = EXCLUDED.daily_total_followers,
                daily_new_followers = EXCLUDED.daily_new_followers,
                daily_lost_followers = EXCLUDED.daily_lost_followers,
                engaged_audience = EXCLUDED.engaged_audience,
                raw_json = EXCLUDED.raw_json,
                fetched_at = NOW()
            """,
            (
                business_id,
                metric_date,
                _to_int(metric.get("followers_count")),
                _to_int(metric.get("video_views")),
                _to_int(metric.get("unique_video_views")),
                _to_int(metric.get("profile_views")),
                _to_int(metric.get("likes")),
                _to_int(metric.get("comments")),
                _to_int(metric.get("shares")),
                _to_int(metric.get("phone_number_clicks")),
                _to_int(metric.get("lead_submissions")),
                _to_int(metric.get("app_download_clicks")),
                _to_int(metric.get("bio_link_clicks")),
                _to_int(metric.get("email_clicks")),
                _to_int(metric.get("address_clicks")),
                _to_int(metric.get("daily_total_followers")),
                _to_int(metric.get("daily_new_followers")),
                _to_int(metric.get("daily_lost_followers")),
                _to_int(metric.get("engaged_audience")),
                json.dumps(metric, ensure_ascii=False),
            ),
        )


# 发布后时间点窗口的唯一真相源：前端不再自己写死这组数字，
# 由 /matrix-video-dashboard 路由注入模板（见 app.py 的 matrix_video_dashboard_page）。
PUBLISH_WINDOW_HOURS = (3, 24, 48, 72)

# 「发布日定格」哨兵：跟 3/24/48/72h 共用同一张表和同一套到期回填机制
# （enqueue_publish_window_capture / run_publish_window_capture 都不关心具体 window 值）。
# 口径 = 发布日次日 03:30 北京时间那次同步的值，也就是发布日完整结束后的定格数据，
# 之后不再变化。总览看板用它，避免每日快照表被当天多次同步反复覆盖的问题。
_PUBLISH_DAY_WINDOW = 0
# 定格采集时刻：发布日次日的北京 03:30，跟每日全量同步对齐。
_PUBLISH_DAY_CAPTURE_HOUR = 3
_PUBLISH_DAY_CAPTURE_MINUTE = 30


def _publish_day_due_at(create_time: datetime) -> datetime:
    """发布日定格行的到期时刻，返回 naive UTC（库里时间戳统一是 naive UTC）。"""
    publish_day = _business_date(create_time)
    # 先按北京墙钟拼出「次日 03:30」，再减 8 小时转回 UTC 存库
    due_bj = datetime(publish_day.year, publish_day.month, publish_day.day) + timedelta(
        days=1, hours=_PUBLISH_DAY_CAPTURE_HOUR, minutes=_PUBLISH_DAY_CAPTURE_MINUTE
    )
    return due_bj - timedelta(hours=8)


def upsert_video(business_id: str, video: dict[str, Any], meta: dict[str, str]) -> None:
    item_id = str(video.get("item_id") or "").strip()
    if not item_id:
        return
    is_new_video = not db.query_one(
        "SELECT 1 FROM tiktok_official_video_snapshots WHERE business_id = %s AND item_id = %s",
        (business_id, item_id),
    )
    db.execute(
        """
        INSERT INTO tiktok_official_video_snapshots (
            business_id, item_id, media_type, is_ad, thumbnail_url, share_url, embed_url,
            caption, video_duration, create_time, likes, comments, shares, favorites,
            reach, video_views, total_time_watched, average_time_watched,
            full_video_watched_rate, new_followers, profile_views, website_clicks,
            phone_number_clicks, lead_submissions, app_download_clicks, email_clicks,
            address_clicks, video_view_retention, engagement_likes, impression_sources,
            audience_genders, audience_countries, audience_cities, audience_types,
            request_id, log_id, raw_json, fetched_at, updated_at
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
            NOW(), NOW()
        )
        ON CONFLICT (business_id, item_id) DO UPDATE SET
            media_type = EXCLUDED.media_type,
            is_ad = EXCLUDED.is_ad,
            thumbnail_url = EXCLUDED.thumbnail_url,
            share_url = EXCLUDED.share_url,
            embed_url = EXCLUDED.embed_url,
            caption = EXCLUDED.caption,
            video_duration = EXCLUDED.video_duration,
            create_time = EXCLUDED.create_time,
            likes = COALESCE(EXCLUDED.likes, tiktok_official_video_snapshots.likes),
            comments = COALESCE(EXCLUDED.comments, tiktok_official_video_snapshots.comments),
            shares = COALESCE(EXCLUDED.shares, tiktok_official_video_snapshots.shares),
            favorites = COALESCE(EXCLUDED.favorites, tiktok_official_video_snapshots.favorites),
            reach = EXCLUDED.reach,
            video_views = EXCLUDED.video_views,
            total_time_watched = EXCLUDED.total_time_watched,
            average_time_watched = EXCLUDED.average_time_watched,
            full_video_watched_rate = EXCLUDED.full_video_watched_rate,
            new_followers = EXCLUDED.new_followers,
            profile_views = EXCLUDED.profile_views,
            website_clicks = EXCLUDED.website_clicks,
            phone_number_clicks = EXCLUDED.phone_number_clicks,
            lead_submissions = EXCLUDED.lead_submissions,
            app_download_clicks = EXCLUDED.app_download_clicks,
            email_clicks = EXCLUDED.email_clicks,
            address_clicks = EXCLUDED.address_clicks,
            video_view_retention = EXCLUDED.video_view_retention,
            engagement_likes = EXCLUDED.engagement_likes,
            impression_sources = EXCLUDED.impression_sources,
            audience_genders = EXCLUDED.audience_genders,
            audience_countries = EXCLUDED.audience_countries,
            audience_cities = EXCLUDED.audience_cities,
            audience_types = EXCLUDED.audience_types,
            request_id = EXCLUDED.request_id,
            log_id = EXCLUDED.log_id,
            raw_json = EXCLUDED.raw_json,
            fetched_at = NOW(),
            updated_at = NOW()
        """,
        (
            business_id,
            item_id,
            video.get("media_type"),
            video.get("is_ad"),
            video.get("thumbnail_url"),
            video.get("share_url"),
            video.get("embed_url"),
            video.get("caption"),
            _to_float(video.get("video_duration")),
            _epoch_to_dt(video.get("create_time")),
            _to_int(video.get("likes")),
            _to_int(video.get("comments")),
            _to_int(video.get("shares")),
            _to_int(video.get("favorites")),
            _to_int(video.get("reach")),
            _to_int(video.get("video_views")),
            _to_float(video.get("total_time_watched")),
            _to_float(video.get("average_time_watched")),
            _to_float(video.get("full_video_watched_rate")),
            _to_int(video.get("new_followers")),
            _to_int(video.get("profile_views")),
            _to_int(video.get("website_clicks")),
            _to_int(video.get("phone_number_clicks")),
            _to_int(video.get("lead_submissions")),
            _to_int(video.get("app_download_clicks")),
            _to_int(video.get("email_clicks")),
            _to_int(video.get("address_clicks")),
            _json(video.get("video_view_retention")),
            _json(video.get("engagement_likes")),
            _json(video.get("impression_sources")),
            _json(video.get("audience_genders")),
            _json(video.get("audience_countries")),
            _json(video.get("audience_cities")),
            _json(video.get("audience_types")),
            meta.get("request_id"),
            meta.get("log_id"),
            json.dumps(video, ensure_ascii=False),
        ),
    )
    db.execute(
        """
        INSERT INTO tiktok_official_video_daily_snapshots (
            business_id, item_id, snapshot_date, video_views, likes, comments, shares, favorites, fetched_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ON CONFLICT (business_id, item_id, snapshot_date) DO UPDATE SET
            video_views = EXCLUDED.video_views,
            likes = COALESCE(EXCLUDED.likes, tiktok_official_video_daily_snapshots.likes),
            comments = COALESCE(EXCLUDED.comments, tiktok_official_video_daily_snapshots.comments),
            shares = COALESCE(EXCLUDED.shares, tiktok_official_video_daily_snapshots.shares),
            favorites = COALESCE(EXCLUDED.favorites, tiktok_official_video_daily_snapshots.favorites),
            fetched_at = NOW()
        """,
        (
            business_id,
            item_id,
            _business_date(),
            _to_int(video.get("video_views")),
            _to_int(video.get("likes")),
            _to_int(video.get("comments")),
            _to_int(video.get("shares")),
            _to_int(video.get("favorites")),
        ),
    )

    create_time = _epoch_to_dt(video.get("create_time"))
    # 只对"真的是最近发布"的视频建时间点占位行——避免新授权账号首次同步时，
    # 把整个历史视频库都当成"刚发布"，生成一堆 due_at 早已过期、一入库就被立刻
    # 抓取的假时间点数据（实际抓到的是老视频的当前状态，却被贴上"3h"之类的标签）。
    if is_new_video and create_time and create_time >= datetime.utcnow() - timedelta(hours=24):
        for window_hours in PUBLISH_WINDOW_HOURS:
            db.execute(
                """
                INSERT INTO tiktok_official_video_publish_window_snapshots (
                    business_id, item_id, window_hours, due_at
                ) VALUES (%s, %s, %s, %s + (%s || ' hours')::interval)
                ON CONFLICT (business_id, item_id, window_hours) DO NOTHING
                """,
                (business_id, item_id, window_hours, create_time, window_hours),
            )
        # 发布日定格行：due_at 不是「发布时间 + N 小时」，而是发布日次日的北京 03:30
        db.execute(
            """
            INSERT INTO tiktok_official_video_publish_window_snapshots (
                business_id, item_id, window_hours, due_at
            ) VALUES (%s, %s, %s, %s)
            ON CONFLICT (business_id, item_id, window_hours) DO NOTHING
            """,
            (business_id, item_id, _PUBLISH_DAY_WINDOW, _publish_day_due_at(create_time)),
        )


def update_video_tags(
    business_id: str,
    item_id: str,
    task_no: str | None = None,
    kol_campaign: str | None = None,
) -> None:
    db.execute(
        """
        UPDATE tiktok_official_video_snapshots SET
            task_no = %s,
            kol_campaign = %s,
            updated_at = NOW()
        WHERE business_id = %s AND item_id = %s
        """,
        (task_no or None, kol_campaign or None, business_id, item_id),
    )


BOOST_STATUS_LABELS = {"marked": "已标记", "delivered": "已投放", "closed": "已关闭"}
BOOST_STATUSES = tuple(BOOST_STATUS_LABELS)


def set_video_boost_status(business_id: str, item_id: str, status: str | None) -> None:
    """设置投流三态（marked/delivered/closed），传空则清除标记。

    人工维护，不由投放消耗数据推导——消耗表只覆盖 VIDEO_VIEWS/TRAFFIC/ENGAGEMENT
    三种 objective 且依赖广告主已授权，拿它当"是否投过"并不可靠。
    is_boosted 同步维护：有任何状态即为 TRUE，供老的筛选和导出继续使用。
    """
    status = (status or "").strip().lower() or None
    if status is not None and status not in BOOST_STATUSES:
        raise ValueError(f"未知的投流状态：{status}（可选 {'/'.join(BOOST_STATUSES)}）")
    is_boosted = status is not None
    db.execute(
        """
        UPDATE tiktok_official_video_snapshots SET
            boost_status = %s,
            is_boosted = %s,
            boosted_at = CASE WHEN %s THEN COALESCE(boosted_at, NOW()) ELSE boosted_at END,
            updated_at = NOW()
        WHERE business_id = %s AND item_id = %s
        """,
        (status, is_boosted, is_boosted, business_id, item_id),
    )


def set_video_boosted(business_id: str, item_id: str, is_boosted: bool) -> None:
    """旧的布尔接口，保留兼容：勾上落到 marked 态，取消则清空。"""
    set_video_boost_status(business_id, item_id, "marked" if is_boosted else None)


_DEFAULT_FLAT_DAYS = 5


def _truncate_flat_series(rows: list[dict[str, Any]], create_date, flat_days: int = _DEFAULT_FLAT_DAYS) -> dict[str, Any]:
    """rows 需按 snapshot_date 升序。逐日算增量，连续 flat_days 天播放量不再增长（delta<=0）后截断序列。"""
    series: list[dict[str, Any]] = []
    prev_views = None
    flat_streak = 0
    stopped_date = None
    for r in rows:
        snapshot_date = r["snapshot_date"]
        views = int(r.get("video_views") or 0)
        delta = (views - prev_views) if prev_views is not None else None
        day_index = (snapshot_date - create_date).days if create_date else None
        series.append({
            "date": snapshot_date,
            "day_index": day_index,
            "video_views": views,
            "delta": delta,
        })
        if delta is not None:
            flat_streak = flat_streak + 1 if delta <= 0 else 0
        if flat_streak >= flat_days and stopped_date is None:
            stopped_date = snapshot_date
        prev_views = views

    if stopped_date is not None:
        series = [s for s in series if s["date"] <= stopped_date]

    return {"series": series, "stopped": stopped_date is not None, "stopped_date": stopped_date}


def get_video_daily_view_series(business_id: str, item_id: str, flat_days: int = _DEFAULT_FLAT_DAYS) -> dict[str, Any]:
    """单个视频：发布日起的每日播放量序列，连续 flat_days 天不增长后截断。"""
    video = db.query_one(
        "SELECT create_time FROM tiktok_official_video_snapshots WHERE business_id = %s AND item_id = %s",
        (business_id, item_id),
    )
    create_date = video["create_time"].date() if (video and video.get("create_time")) else None

    rows = db.query_all(
        """
        SELECT snapshot_date, video_views
        FROM tiktok_official_video_daily_snapshots
        WHERE business_id = %s AND item_id = %s
        ORDER BY snapshot_date
        """,
        (business_id, item_id),
    ) or []

    return _truncate_flat_series(rows, create_date, flat_days)


def get_daily_view_series_bulk(
    pairs: list[tuple[str, str]],
    create_dates: dict[tuple[str, str], Any],
    flat_days: int = _DEFAULT_FLAT_DAYS,
) -> dict[tuple[str, str], dict[str, Any]]:
    """导出用：一次查询多个视频的 daily_snapshots，Python 按 (business_id, item_id) 分组后复用截断逻辑。"""
    if not pairs:
        return {}
    rows = db.query_all(
        """
        SELECT business_id, item_id, snapshot_date, video_views
        FROM tiktok_official_video_daily_snapshots
        WHERE (business_id, item_id) IN %s
        ORDER BY business_id, item_id, snapshot_date
        """,
        (tuple(pairs),),
    ) or []

    grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for r in rows:
        key = (r["business_id"], r["item_id"])
        grouped.setdefault(key, []).append(r)

    result: dict[tuple[str, str], dict[str, Any]] = {}
    for key, group_rows in grouped.items():
        create_time = create_dates.get(key)
        create_date = create_time.date() if create_time else None
        result[key] = _truncate_flat_series(group_rows, create_date, flat_days)
    return result


def authorize_video_for_ads(business_id: str, item_id: str, authorization_days: int = 30) -> dict[str, Any]:
    # 生成 Spark 授权码要求 access_token 带 biz.spark.auth 权限；主账号 token 大多没有这个权限，
    # 优先用该账号在独立 Spark 小流程里授权过的 token（按 business_id 直接查，account_alias 在
    # 主表里不唯一，按别名查会把撞名账号的 token 张冠李戴），没有的话再退回主账号 token（兼容老账号）。
    spark_info = get_spark_token_info(business_id)
    if spark_info and spark_info.get("access_token"):
        token = spark_info["access_token"]
        # 用 Spark App 的 open_id 作为 business_id，跟这个 token 自己的身份对上，
        # 否则 TikTok 会认为 token 跟 business_id 不匹配，报 access token 不合法。
        call_business_id = spark_info.get("open_id") or business_id
    else:
        token = get_access_token(business_id)
        call_business_id = business_id
    if not token:
        raise RuntimeError(f"business_id={business_id} 缺少 access_token，请先完成 TikTok 账号授权")

    resp = requests.post(
        f"{API_BASE}/business/post/authorize/setting/",
        headers={"Access-Token": token, "Content-Type": "application/json"},
        json={
            "business_id": call_business_id,
            "item_id": item_id,
            "is_ad_promotable": True,
            "authorization_days": authorization_days,
        },
        timeout=60,
    )
    try:
        payload = resp.json()
    except Exception as exc:
        raise RuntimeError(f"TikTok Spark 授权接口返回非 JSON: HTTP {resp.status_code} {resp.text[:300]}") from exc
    if resp.status_code >= 400 or payload.get("code") not in (0, "0", None):
        raise RuntimeError(f"TikTok Spark 授权失败: {payload.get('message') or resp.text[:200]}")

    data = payload.get("data") or {}
    auth_code = data.get("auth_code")
    start_time = _parse_spark_time(data.get("auth_code_start_time"))
    end_time = _parse_spark_time(data.get("auth_code_end_time"))
    if not auth_code:
        raise RuntimeError(f"TikTok Spark 授权接口未返回 auth_code: {json.dumps(payload, ensure_ascii=False)[:300]}")

    db.execute(
        """
        UPDATE tiktok_official_video_snapshots SET
            spark_code = %s,
            spark_code_start_time = %s,
            spark_code_end_time = %s,
            updated_at = NOW()
        WHERE business_id = %s AND item_id = %s
        """,
        (auth_code, start_time, end_time, business_id, item_id),
    )
    return {
        "spark_code": auth_code,
        "spark_code_start_time": start_time,
        "spark_code_end_time": end_time,
    }


def _parse_spark_time(value) -> datetime | None:
    if not value:
        return None
    try:
        return datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
    except Exception:
        return None


_HASHTAG_RE = re.compile(r"#([^\s#]+)")


def _extract_hashtags(caption: str | None) -> list[str]:
    if not caption:
        return []
    return _HASHTAG_RE.findall(caption)


def _engagement_sum_sql(alias: str = "v") -> str:
    """互动量口径的唯一定义：赞 + 评 + 转 + 藏。各聚合查询一律引用这里，不要再手抄一遍。"""
    return (
        f"COALESCE({alias}.likes,0) + COALESCE({alias}.comments,0) "
        f"+ COALESCE({alias}.shares,0) + COALESCE({alias}.favorites,0)"
    )


_ENGAGEMENT_RATE_SQL = (
    f"(CASE WHEN v.video_views > 0 "
    f"THEN ({_engagement_sum_sql('v')})::numeric / v.video_views ELSE NULL END)"
)

_MATRIX_VIDEO_SORT_FIELDS = {
    "create_time": "v.create_time",
    "video_views": "v.video_views",
    "likes": "v.likes",
    "comments": "v.comments",
    "shares": "v.shares",
    "favorites": "v.favorites",
    "engagement_rate": _ENGAGEMENT_RATE_SQL,
    "full_video_watched_rate": "lw.full_video_watched_rate",
    "average_time_watched": "lw.average_time_watched",
    "pw_engagement_rate": "lw.pw_engagement_rate",
    "distribution_rate": "lw.distribution_rate",
}


def _matrix_engagement_filter_sql(filters: dict[str, Any]) -> str | None:
    engagement_filter = (filters.get("engagement_filter") or "").strip()
    if engagement_filter == "ge1":
        return f"{_ENGAGEMENT_RATE_SQL} >= 0.01"
    if engagement_filter == "lt1":
        return f"(v.video_views IS NULL OR v.video_views = 0 OR {_ENGAGEMENT_RATE_SQL} < 0.01)"
    return None


def _matrix_views_filter_sql(filters: dict[str, Any]) -> str | None:
    views_filter = (filters.get("views_filter") or "").strip()
    if views_filter == "ge1000":
        return "v.video_views >= 1000"
    if views_filter == "lt1000":
        return "(v.video_views IS NULL OR v.video_views < 1000)"
    if views_filter == "ge500":
        return "v.video_views >= 500"
    if views_filter == "lt500":
        return "(v.video_views IS NULL OR v.video_views < 500)"
    if views_filter == "500to1000":
        return "(v.video_views >= 500 AND v.video_views < 1000)"
    return None


def list_matrix_videos(
    filters: dict[str, Any] | None = None,
    limit: int = 50,
    offset: int = 0,
    sort: str = "create_time",
    order: str = "desc",
) -> dict[str, Any]:
    filters = filters or {}
    where, params = _matrix_account_filters_sql(filters)

    date_from = (filters.get("date_from") or "").strip()
    if date_from:
        where.append("v.create_time >= %s")
        params.append(date_from)

    date_to = (filters.get("date_to") or "").strip()
    if date_to:
        where.append("v.create_time < (%s::date + INTERVAL '1 day')")
        params.append(date_to)

    creator = (filters.get("creator") or "").strip()
    if creator:
        where.append("(a.account_alias ILIKE %s OR a.account_name ILIKE %s OR v.business_id ILIKE %s)")
        like = f"%{creator}%"
        params.extend([like, like, like])

    keyword = (filters.get("keyword") or "").strip()
    if keyword:
        where.append("v.caption ILIKE %s")
        params.append(f"%{keyword}%")

    if filters.get("only_boosted"):
        where.append("v.is_boosted = TRUE")
    elif filters.get("only_unboosted"):
        where.append("(v.is_boosted = FALSE OR v.is_boosted IS NULL)")

    engagement_where = _matrix_engagement_filter_sql(filters)
    if engagement_where:
        where.append(engagement_where)

    views_where = _matrix_views_filter_sql(filters)
    if views_where:
        where.append(views_where)

    where_sql = " AND ".join(where)

    joins_sql = """
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
    """
    base_sql = f"{joins_sql} WHERE {where_sql}"

    # 汇总里的 COUNT(*) 就是分页用的 total（同一个 FROM/WHERE），不再单独跑一条 COUNT 查询。
    summary_row = db.query_one(
        f"""
        SELECT
            COALESCE(SUM(v.video_views), 0) AS total_views,
            COALESCE(SUM({_engagement_sum_sql('v')}), 0) AS total_engagement,
            COUNT(DISTINCT v.business_id) AS account_count,
            COUNT(*) AS video_count
        {base_sql}
        """,
        tuple(params),
    ) or {}
    total = int(summary_row.get("video_count") or 0)
    total_views = int(summary_row.get("total_views") or 0)
    total_engagement = int(summary_row.get("total_engagement") or 0)
    overall_engagement_rate = (total_engagement / total_views) if total_views else 0.0

    sort_field = sort if sort in _MATRIX_VIDEO_SORT_FIELDS else "create_time"
    sort_dir = "ASC" if str(order).strip().lower() == "asc" else "DESC"
    sort_expr = _MATRIX_VIDEO_SORT_FIELDS[sort_field]
    sort_order_sql = f"{sort_expr} {sort_dir} NULLS LAST, v.create_time DESC NULLS LAST"

    rows = db.query_all(
        f"""
        SELECT
            v.*,
            a.account_alias, a.account_name, a.display_name, a.region, a.account_type
        {joins_sql}
        LEFT JOIN LATERAL (
            SELECT full_video_watched_rate, average_time_watched,
                   engagement_rate AS pw_engagement_rate, distribution_rate
            FROM tiktok_official_video_publish_window_snapshots w
            WHERE w.business_id = v.business_id AND w.item_id = v.item_id AND w.captured_at IS NOT NULL
              AND w.window_hours > 0
            ORDER BY w.window_hours DESC
            LIMIT 1
        ) lw ON true
        WHERE {where_sql}
        ORDER BY {sort_order_sql}
        LIMIT %s OFFSET %s
        """,
        tuple(params + [limit, offset]),
    ) or []

    videos = []
    for row in rows:
        views = int(row.get("video_views") or 0)
        engagement = int((row.get("likes") or 0) + (row.get("comments") or 0) + (row.get("shares") or 0) + (row.get("favorites") or 0))
        row["engagement_rate"] = (engagement / views) if views else 0.0
        row["hashtags"] = _extract_hashtags(row.get("caption"))
        videos.append(row)

    if videos:
        pairs = tuple((v["business_id"], v["item_id"]) for v in videos)
        window_rows = db.query_all(
            """
            SELECT business_id, item_id, window_hours, captured_at, video_views,
                   average_time_watched, full_video_watched_rate, engagement_rate, distribution_rate
            FROM tiktok_official_video_publish_window_snapshots
            WHERE (business_id, item_id) IN %s AND window_hours > 0
            ORDER BY business_id, item_id, window_hours
            """,
            (pairs,),
        ) or []
        windows_map: dict[tuple, list] = {}
        for w in window_rows:
            key = (w["business_id"], w["item_id"])
            windows_map.setdefault(key, []).append(w)
        for v in videos:
            v["publish_windows"] = windows_map.get((v["business_id"], v["item_id"]), [])

    return {
        "videos": videos,
        "total": total,
        "summary": {
            "total_views": total_views,
            "total_engagement": total_engagement,
            "overall_engagement_rate": overall_engagement_rate,
            "account_count": int(summary_row.get("account_count") or 0),
            "video_count": int(summary_row.get("video_count") or 0),
        },
    }


# PUBG MOBILE 官方品牌主账号，误加进了官号授权体系，不是矩阵切片账号，矩阵看板全系列查询需排除
_MATRIX_EXCLUDED_BUSINESS_IDS = ["-000UA9EpttJ5qzoWWRbZ7pT7qyGfmhW62C2"]


def _matrix_account_filters_sql(filters: dict[str, Any]) -> tuple[list[str], list[Any]]:
    where = ["1=1"]
    params: list[Any] = []

    where.append("a.enabled = TRUE")
    where.append("a.business_id != ALL(%s)")
    params.append(_MATRIX_EXCLUDED_BUSINESS_IDS)

    regions = [str(v).strip() for v in (filters.get("regions") or []) if str(v).strip()]
    if not regions:
        region = (filters.get("region") or "").strip()
        if region:
            regions = [region]
    if regions:
        where.append("a.region = ANY(%s)")
        params.append(regions)

    account_type = (filters.get("account_type") or "").strip()
    if account_type:
        where.append("a.account_type = %s")
        params.append(account_type)

    account_ids = [str(v).strip() for v in (filters.get("account_ids") or []) if str(v).strip()]
    if account_ids:
        where.append("a.business_id = ANY(%s)")
        params.append(account_ids)

    return where, params


def _matrix_date_range(date_from: str | None, date_to: str | None, days: int = 7) -> tuple[date, date]:
    """解析看板二/三的日期范围输入，缺省时回退到最近 N 天（含首尾两天）。"""
    today = _business_date()

    def _parse(value):
        if not value:
            return None
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None

    parsed_from = _parse(date_from)
    parsed_to = _parse(date_to)
    if parsed_from and parsed_to:
        return parsed_from, parsed_to
    if parsed_from:
        return parsed_from, today
    if parsed_to:
        return parsed_to - timedelta(days=days - 1), parsed_to
    return today - timedelta(days=days - 1), today


def matrix_overview_summary(filters: dict[str, Any] | None = None) -> dict[str, Any]:
    """总览：按「发布日定格值」统计——数字一旦定格就不再随时间变化。

    取值优先用 window_hours=0 的发布日定格快照（发布日次日北京 03:30 采集）。
    该行只在 upsert_video() 里为新发布的视频生成，历史视频没有，回退到日快照表里
    snapshot_date = 北京发布日 的那一行。两者时点是对齐的：北京 D 发布的视频，其
    发布日结束后的首次全量同步跑在北京 D+1 03:30 = UTC D 19:30，而历史 snapshot_date
    用的是 UTC 日期，正好等于 D——所以那一行就是历史视频的发布日定格值。

    video_count 只数「拿得到定格值」的视频，video_total 是筛选命中的全部视频，
    两者的差就是暂时没有定格数据的部分（比如今天刚发、还没到次日凌晨）。
    """
    filters = filters or {}
    where, params = _matrix_account_filters_sql(filters)
    where_sql = " AND ".join(where)
    frozen_views = "COALESCE(w.video_views, d.video_views)"

    row = db.query_one(
        f"""
        SELECT
            COUNT(DISTINCT a.business_id) AS account_count,
            COUNT(v.item_id) AS video_total,
            COUNT({frozen_views}) AS video_count,
            COALESCE(SUM({frozen_views}), 0) AS total_views,
            COALESCE(SUM(CASE
                WHEN w.video_views IS NOT NULL THEN {_engagement_sum_sql('w')}
                WHEN d.video_views IS NOT NULL THEN {_engagement_sum_sql('d')}
                ELSE NULL
            END), 0) AS total_engagement
        FROM tiktok_official_accounts a
        LEFT JOIN tiktok_official_video_snapshots v
               ON v.business_id = a.business_id AND v.create_time IS NOT NULL
        LEFT JOIN tiktok_official_video_publish_window_snapshots w
               ON w.business_id = v.business_id AND w.item_id = v.item_id
              AND w.window_hours = {_PUBLISH_DAY_WINDOW} AND w.captured_at IS NOT NULL
        LEFT JOIN tiktok_official_video_daily_snapshots d
               ON d.business_id = v.business_id AND d.item_id = v.item_id
              AND d.snapshot_date = (v.create_time AT TIME ZONE 'UTC' AT TIME ZONE 'Asia/Shanghai')::date
        WHERE {where_sql}
        """,
        tuple(params),
    ) or {}

    total_views = int(row.get("total_views") or 0)
    total_engagement = int(row.get("total_engagement") or 0)
    return {
        "account_count": int(row.get("account_count") or 0),
        "video_count": int(row.get("video_count") or 0),
        "video_total": int(row.get("video_total") or 0),
        "total_views": total_views,
        "total_engagement": total_engagement,
        "engagement_rate": (total_engagement / total_views) if total_views else 0.0,
    }


def matrix_publish_range_summary(
    filters: dict[str, Any] | None = None, date_from: str | None = None, date_to: str | None = None
) -> dict[str, Any]:
    filters = filters or {}
    range_from, range_to = _matrix_date_range(date_from, date_to, days=7)
    where, params = _matrix_account_filters_sql(filters)
    where.append("v.create_time >= %s")
    where.append("v.create_time < (%s::date + INTERVAL '1 day')")
    params.extend([range_from.isoformat(), range_to.isoformat()])
    where_sql = " AND ".join(where)

    summary_row = db.query_one(
        f"""
        SELECT
            COUNT(*) AS video_count,
            COALESCE(SUM(v.video_views), 0) AS total_views,
            COALESCE(SUM({_engagement_sum_sql('v')}), 0) AS total_engagement
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        WHERE {where_sql}
        """,
        tuple(params),
    ) or {}
    total_views = int(summary_row.get("total_views") or 0)
    total_engagement = int(summary_row.get("total_engagement") or 0)

    daily_rows = db.query_all(
        f"""
        SELECT
            v.create_time::date AS date,
            COUNT(*) AS video_count,
            COALESCE(SUM(v.video_views), 0) AS views,
            COALESCE(SUM({_engagement_sum_sql('v')}), 0) AS engagement
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        WHERE {where_sql}
        GROUP BY v.create_time::date
        ORDER BY v.create_time::date
        """,
        tuple(params),
    ) or []

    daily = []
    for drow in daily_rows:
        d_views = int(drow.get("views") or 0)
        d_engagement = int(drow.get("engagement") or 0)
        d_date = drow.get("date")
        daily.append({
            "date": d_date.isoformat() if hasattr(d_date, "isoformat") else d_date,
            "video_count": int(drow.get("video_count") or 0),
            "views": d_views,
            "engagement": d_engagement,
            "engagement_rate": (d_engagement / d_views) if d_views else 0.0,
        })

    return {
        "date_from": range_from.isoformat(),
        "date_to": range_to.isoformat(),
        "summary": {
            "video_count": int(summary_row.get("video_count") or 0),
            "total_views": total_views,
            "total_engagement": total_engagement,
            "engagement_rate": (total_engagement / total_views) if total_views else 0.0,
        },
        "daily": daily,
    }


def matrix_snapshot_delta_range(
    filters: dict[str, Any] | None = None, date_from: str | None = None, date_to: str | None = None
) -> dict[str, Any]:
    filters = filters or {}
    range_from, range_to = _matrix_date_range(date_from, date_to, days=7)
    where, params = _matrix_account_filters_sql(filters)
    where_sql = " AND ".join(where)

    available_rows = db.query_all(
        f"""
        SELECT DISTINCT d.snapshot_date
        FROM tiktok_official_video_daily_snapshots d
        LEFT JOIN tiktok_official_accounts a ON a.business_id = d.business_id
        WHERE {where_sql} AND d.snapshot_date <= %s
        ORDER BY d.snapshot_date
        """,
        tuple(params + [range_to.isoformat()]),
    ) or []
    available_dates = [r["snapshot_date"] for r in available_rows]

    empty = {
        "date_from": range_from.isoformat(), "date_to": range_to.isoformat(),
        "date_from_actual": None, "date_to_actual": None,
        "new_video_count": 0, "views_delta": 0, "engagement_delta": 0,
        "new_video_views": 0, "existing_video_views_delta": 0,
        "new_video_engagement": 0, "existing_video_engagement_delta": 0,
        "engagement_rate_start": 0.0, "engagement_rate_end": 0.0, "engagement_rate_change_pct": None,
        "daily": [],
    }
    if not available_dates:
        return empty

    date_to_actual = max(available_dates)
    candidates_before_from = [d for d in available_dates if d <= range_from]
    date_from_actual = max(candidates_before_from) if candidates_before_from else min(available_dates)

    # 区间两端（date_from_actual / date_to_actual）的总量与增量拆分，一条查询搞定。
    # 按 (business_id, item_id) 透视成每条视频一行，start 侧为 NULL 即区间内新发布的视频。
    engagement_expr = _engagement_sum_sql("d")
    bounds = db.query_one(
        f"""
        WITH pv AS (
            SELECT d.business_id, d.item_id,
                   MAX(d.video_views) FILTER (WHERE d.snapshot_date = %s) AS v_start,
                   MAX({engagement_expr}) FILTER (WHERE d.snapshot_date = %s) AS e_start,
                   MAX(d.video_views) FILTER (WHERE d.snapshot_date = %s) AS v_end,
                   MAX({engagement_expr}) FILTER (WHERE d.snapshot_date = %s) AS e_end
            FROM tiktok_official_video_daily_snapshots d
            LEFT JOIN tiktok_official_accounts a ON a.business_id = d.business_id
            WHERE {where_sql} AND d.snapshot_date IN (%s, %s)
            GROUP BY d.business_id, d.item_id
        )
        SELECT
            COALESCE(SUM(v_start), 0) AS views_start,
            COALESCE(SUM(e_start), 0) AS engagement_start,
            COALESCE(SUM(v_end), 0) AS views_end,
            COALESCE(SUM(e_end), 0) AS engagement_end,
            COUNT(*) FILTER (WHERE v_start IS NULL AND v_end IS NOT NULL) AS new_video_count,
            COALESCE(SUM(v_end) FILTER (WHERE v_start IS NULL), 0) AS new_video_views,
            COALESCE(SUM(e_end) FILTER (WHERE v_start IS NULL), 0) AS new_video_engagement,
            COALESCE(SUM(v_end - v_start) FILTER (WHERE v_start IS NOT NULL AND v_end IS NOT NULL), 0)
                AS existing_video_views_delta,
            COALESCE(SUM(e_end - e_start) FILTER (WHERE v_start IS NOT NULL AND v_end IS NOT NULL), 0)
                AS existing_video_engagement_delta
        FROM pv
        """,
        tuple(
            [date_from_actual, date_from_actual, date_to_actual, date_to_actual]
            + params
            + [date_from_actual, date_to_actual]
        ),
    ) or {}

    views_start = int(bounds.get("views_start") or 0)
    engagement_start = int(bounds.get("engagement_start") or 0)
    views_end = int(bounds.get("views_end") or 0)
    engagement_end = int(bounds.get("engagement_end") or 0)

    if date_from_actual != date_to_actual:
        new_video_count = int(bounds.get("new_video_count") or 0)
        new_video_views = int(bounds.get("new_video_views") or 0)
        new_video_engagement = int(bounds.get("new_video_engagement") or 0)
        existing_views_delta = int(bounds.get("existing_video_views_delta") or 0)
        existing_engagement_delta = int(bounds.get("existing_video_engagement_delta") or 0)
    else:
        new_video_count = new_video_views = new_video_engagement = 0
        existing_views_delta = existing_engagement_delta = 0

    rate_start = (engagement_start / views_start) if views_start else 0.0
    rate_end = (engagement_end / views_end) if views_end else 0.0
    rate_change_pct = ((rate_end - rate_start) / rate_start * 100) if rate_start else None

    # 逐日增量：一条 LAG() 窗口查询取代原先「按相邻日期对逐次全表扫描」的 2N 次查询。
    # 语义差异（有意为之）：某视频在 D 日缺失但 D-1/D+1 都有时，LAG 会把 D+1 连回 D-1
    # 算成存量增量；旧实现按相邻两日 key 交集比较，会把它误判成 D+1 新发布的视频。
    daily_rows = db.query_all(
        f"""
        WITH filtered AS (
            SELECT d.business_id, d.item_id, d.snapshot_date,
                   COALESCE(d.video_views, 0) AS views,
                   ({engagement_expr}) AS engagement
            FROM tiktok_official_video_daily_snapshots d
            LEFT JOIN tiktok_official_accounts a ON a.business_id = d.business_id
            WHERE {where_sql} AND d.snapshot_date BETWEEN %s AND %s
        ), lagged AS (
            SELECT f.*,
                   LAG(views) OVER w AS prev_views,
                   LAG(engagement) OVER w AS prev_engagement
            FROM filtered f
            WINDOW w AS (PARTITION BY business_id, item_id ORDER BY snapshot_date)
        )
        SELECT snapshot_date,
               COUNT(*) FILTER (WHERE prev_views IS NULL) AS new_video_count,
               COALESCE(SUM(views) FILTER (WHERE prev_views IS NULL), 0) AS new_video_views,
               COALESCE(SUM(engagement) FILTER (WHERE prev_views IS NULL), 0) AS new_video_engagement,
               COALESCE(SUM(views - prev_views) FILTER (WHERE prev_views IS NOT NULL), 0) AS existing_views_delta,
               COALESCE(SUM(engagement - prev_engagement) FILTER (WHERE prev_views IS NOT NULL), 0)
                   AS existing_engagement_delta,
               COALESCE(SUM(views), 0) AS cur_views_total,
               COALESCE(SUM(engagement), 0) AS cur_engagement_total
        FROM lagged
        WHERE snapshot_date > %s
        GROUP BY snapshot_date
        ORDER BY snapshot_date
        """,
        tuple(params + [date_from_actual, date_to_actual, date_from_actual]),
    ) or []

    daily = []
    for row in daily_rows:
        cur_views_total = int(row.get("cur_views_total") or 0)
        cur_engagement_total = int(row.get("cur_engagement_total") or 0)
        d_date = row.get("snapshot_date")
        daily.append({
            "date": d_date.isoformat() if hasattr(d_date, "isoformat") else d_date,
            "new_video_count": int(row.get("new_video_count") or 0),
            "views_delta": int(row.get("new_video_views") or 0) + int(row.get("existing_views_delta") or 0),
            "engagement_delta": int(row.get("new_video_engagement") or 0) + int(row.get("existing_engagement_delta") or 0),
            "engagement_rate": (cur_engagement_total / cur_views_total) if cur_views_total else 0.0,
        })

    return {
        "date_from": range_from.isoformat(),
        "date_to": range_to.isoformat(),
        "date_from_actual": date_from_actual.isoformat() if hasattr(date_from_actual, "isoformat") else date_from_actual,
        "date_to_actual": date_to_actual.isoformat() if hasattr(date_to_actual, "isoformat") else date_to_actual,
        "new_video_count": new_video_count,
        "views_delta": new_video_views + existing_views_delta,
        "engagement_delta": new_video_engagement + existing_engagement_delta,
        "new_video_views": new_video_views,
        "existing_video_views_delta": existing_views_delta,
        "new_video_engagement": new_video_engagement,
        "existing_video_engagement_delta": existing_engagement_delta,
        "engagement_rate_start": rate_start,
        "engagement_rate_end": rate_end,
        "engagement_rate_change_pct": rate_change_pct,
        "daily": daily,
    }


def matrix_cumulative_views_range(
    filters: dict[str, Any] | None = None, date_from: str | None = None, date_to: str | None = None
) -> dict[str, Any]:
    """按视频发布日期截止到某日的累计总播放量（只增不减的曲线，非当日新增）。"""
    filters = filters or {}
    range_from, range_to = _matrix_date_range(date_from, date_to, days=30)
    where, params = _matrix_account_filters_sql(filters)
    where.append("v.create_time IS NOT NULL")
    where.append("v.create_time::date <= %s")
    params.append(range_to.isoformat())
    where_sql = " AND ".join(where)

    rows = db.query_all(
        f"""
        WITH daily AS (
            SELECT v.create_time::date AS date, COALESCE(SUM(v.video_views), 0) AS views
            FROM tiktok_official_video_snapshots v
            LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
            WHERE {where_sql}
            GROUP BY v.create_time::date
        )
        SELECT date, SUM(views) OVER (ORDER BY date) AS cumulative_views
        FROM daily
        ORDER BY date
        """,
        tuple(params),
    ) or []

    daily = [
        {
            "date": r["date"].isoformat() if hasattr(r["date"], "isoformat") else r["date"],
            "cumulative_views": int(r["cumulative_views"] or 0),
        }
        for r in rows
        if r["date"] >= range_from
    ]

    return {
        "date_from": range_from.isoformat(),
        "date_to": range_to.isoformat(),
        "daily": daily,
    }


def build_matrix_export(export_date) -> bytes:
    rows = db.query_all(
        """
        SELECT
            v.*,
            a.account_alias, a.account_name, a.display_name, a.region, a.account_type, a.profile_deep_link
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        WHERE v.create_time::date = %s
          AND v.business_id != ALL(%s)
        ORDER BY COALESCE(a.account_alias, a.account_name, a.display_name, v.business_id), v.create_time DESC NULLS LAST
        """,
        (export_date, _MATRIX_EXCLUDED_BUSINESS_IDS),
    ) or []

    next_day_map: dict[tuple, int | None] = {}
    if rows:
        pairs = [(r["business_id"], r["item_id"]) for r in rows]
        next_day_rows = db.query_all(
            """
            SELECT v.business_id, v.item_id, d.video_views AS next_day_views
            FROM tiktok_official_video_snapshots v
            LEFT JOIN tiktok_official_video_daily_snapshots d
                ON d.business_id = v.business_id
                AND d.item_id = v.item_id
                AND d.snapshot_date = (v.create_time::date + INTERVAL '1 day')::date
            WHERE (v.business_id, v.item_id) IN %s
            """,
            (tuple(pairs),),
        ) or []
        for r in next_day_rows:
            next_day_map[(r["business_id"], r["item_id"])] = r.get("next_day_views")

    headers = [
        "序号", "关联任务编号", "所属矩阵账号", "国家", "账号类型",
        "对应KOL/Campaign", "作品发布时间", "作品标题/文案", "话题标签", "Spark code",
        "主页链接", "视频链接", "播放量（最新）", "次日播放量", "互动率（最新）",
    ]
    ACCOUNT_COL = 3  # "所属矩阵账号"

    wb = Workbook()
    ws = wb.active
    ws.title = "矩阵号视频"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    merge_start_row = None
    prev_account = object()
    for idx, row in enumerate(rows, start=1):
        account_name = row.get("display_name") or row.get("account_alias") or row.get("account_name") or row.get("business_id")
        views = int(row.get("video_views") or 0)
        engagement = int((row.get("likes") or 0) + (row.get("comments") or 0) + (row.get("shares") or 0) + (row.get("favorites") or 0))
        engagement_rate = (engagement / views) if views else 0.0
        ws.append(
            [
                idx,
                row.get("task_no") or "",
                account_name,
                row.get("region") or "",
                row.get("account_type") or "",
                row.get("kol_campaign") or "",
                row.get("create_time"),
                row.get("caption") or "",
                " ".join(_extract_hashtags(row.get("caption"))),
                row.get("spark_code") or "",
                row.get("profile_deep_link") or "",
                row.get("share_url") or "",
                views,
                next_day_map.get((row.get("business_id"), row.get("item_id"))),
                round(engagement_rate, 4),
            ]
        )
        excel_row = idx + 1  # 第1行是表头

        if account_name != prev_account:
            if merge_start_row is not None and excel_row - 1 > merge_start_row:
                ws.merge_cells(start_row=merge_start_row, end_row=excel_row - 1, start_column=ACCOUNT_COL, end_column=ACCOUNT_COL)
                ws.cell(row=merge_start_row, column=ACCOUNT_COL).alignment = Alignment(vertical="center")
            merge_start_row = excel_row
            prev_account = account_name

    last_row = len(rows) + 1
    if merge_start_row is not None and last_row > merge_start_row:
        ws.merge_cells(start_row=merge_start_row, end_row=last_row, start_column=ACCOUNT_COL, end_column=ACCOUNT_COL)
        ws.cell(row=merge_start_row, column=ACCOUNT_COL).alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


_MATRIX_QUERY_EXPORT_MAX_ROWS = 5000


def build_matrix_query_export(filters: dict[str, Any] | None = None) -> bytes:
    """按当前视频明细表筛选条件导出查询结果（不缓存，现算现出），跟固定的"下载报表（前一日）"是两个平行入口。"""
    filters = filters or {}
    where, params = _matrix_account_filters_sql(filters)

    date_from = (filters.get("date_from") or "").strip()
    if date_from:
        where.append("v.create_time >= %s")
        params.append(date_from)

    date_to = (filters.get("date_to") or "").strip()
    if date_to:
        where.append("v.create_time < (%s::date + INTERVAL '1 day')")
        params.append(date_to)

    creator = (filters.get("creator") or "").strip()
    if creator:
        where.append("(a.account_alias ILIKE %s OR a.account_name ILIKE %s OR v.business_id ILIKE %s)")
        like = f"%{creator}%"
        params.extend([like, like, like])

    keyword = (filters.get("keyword") or "").strip()
    if keyword:
        where.append("v.caption ILIKE %s")
        params.append(f"%{keyword}%")

    if filters.get("only_boosted"):
        where.append("v.is_boosted = TRUE")
    elif filters.get("only_unboosted"):
        where.append("(v.is_boosted = FALSE OR v.is_boosted IS NULL)")

    engagement_where = _matrix_engagement_filter_sql(filters)
    if engagement_where:
        where.append(engagement_where)

    views_where = _matrix_views_filter_sql(filters)
    if views_where:
        where.append(views_where)

    where_sql = " AND ".join(where)

    rows = db.query_all(
        f"""
        SELECT
            v.*,
            a.account_alias, a.account_name, a.display_name, a.region, a.account_type, a.profile_deep_link
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        WHERE {where_sql}
        ORDER BY v.create_time DESC NULLS LAST, v.updated_at DESC
        LIMIT %s
        """,
        tuple(params + [_MATRIX_QUERY_EXPORT_MAX_ROWS]),
    ) or []

    window_rows = []
    windows_by_key: dict[tuple[str, str], list[dict[str, Any]]] = {}
    daily_series_map: dict[tuple[str, str], dict[str, Any]] = {}
    if rows:
        pairs = [(r["business_id"], r["item_id"]) for r in rows]
        create_dates = {(r["business_id"], r["item_id"]): r.get("create_time") for r in rows}
        daily_series_map = get_daily_view_series_bulk(pairs, create_dates)
        window_rows = db.query_all(
            """
            SELECT business_id, item_id, window_hours, captured_at, video_views, likes, comments,
                   shares, favorites, reach, total_time_watched, average_time_watched,
                   full_video_watched_rate, engagement_rate, followers_count_snapshot, distribution_rate
            FROM tiktok_official_video_publish_window_snapshots
            WHERE (business_id, item_id) IN %s AND window_hours > 0
            ORDER BY business_id, item_id, window_hours
            """,
            (tuple(pairs),),
        ) or []
        for w in window_rows:
            windows_by_key.setdefault((w["business_id"], w["item_id"]), []).append(w)

    def _latest_captured_window(key):
        captured = [w for w in windows_by_key.get(key, []) if w.get("captured_at") is not None]
        if not captured:
            return None
        return max(captured, key=lambda w: w["window_hours"])

    def _daily_views_at_day_index(key, day_index):
        series = daily_series_map.get(key, {}).get("series") or []
        point = next((p for p in series if p["day_index"] == day_index), None)
        return point["video_views"] if point else None

    headers = [
        "关联任务编号", "所属矩阵账号", "国家", "账号类型",
        "对应KOL/Campaign", "作品发布时间", "作品标题/文案", "话题标签", "Spark code",
        "主页链接", "视频链接", "播放量（最新）", "点赞", "评论", "转发", "收藏", "互动率（最新）",
        "发布后播放量-3h", "发布后播放量-24h", "发布后播放量-48h", "发布后播放量-72h",
        "发布后播放量-96h（每日快照，约第4天）", "发布后播放量-120h（每日快照，约第5天）",
        "完播率（最新窗口）", "平均观看时长（最新窗口）", "发布后互动率（最新窗口）", "Distribution Rate（最新窗口）",
        "是否已投流",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "查询结果"
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        account_name = row.get("display_name") or row.get("account_alias") or row.get("account_name") or row.get("business_id")
        views = int(row.get("video_views") or 0)
        engagement = int((row.get("likes") or 0) + (row.get("comments") or 0) + (row.get("shares") or 0) + (row.get("favorites") or 0))
        engagement_rate = (engagement / views) if views else 0.0
        key = (row["business_id"], row["item_id"])
        windows_for_video = {w["window_hours"]: w for w in windows_by_key.get(key, [])}
        pw_views = []
        for h in (3, 24, 48, 72):
            w = windows_for_video.get(h)
            pw_views.append(w.get("video_views") if (w and w.get("captured_at") is not None) else None)
        daily_pw_views = [_daily_views_at_day_index(key, 4), _daily_views_at_day_index(key, 5)]
        latest_window = _latest_captured_window(key)
        ws.append(
            [
                row.get("task_no") or "",
                account_name,
                row.get("region") or "",
                row.get("account_type") or "",
                row.get("kol_campaign") or "",
                row.get("create_time"),
                row.get("caption") or "",
                " ".join(_extract_hashtags(row.get("caption"))),
                row.get("spark_code") or "",
                row.get("profile_deep_link") or "",
                row.get("share_url") or "",
                views,
                int(row.get("likes") or 0),
                int(row.get("comments") or 0),
                int(row.get("shares") or 0),
                int(row.get("favorites") or 0),
                round(engagement_rate, 4),
                *pw_views,
                *daily_pw_views,
                latest_window.get("full_video_watched_rate") if latest_window else None,
                latest_window.get("average_time_watched") if latest_window else None,
                latest_window.get("engagement_rate") if latest_window else None,
                latest_window.get("distribution_rate") if latest_window else None,
                BOOST_STATUS_LABELS.get(row.get("boost_status") or "", "是" if row.get("is_boosted") else "否"),
            ]
        )
    ws.freeze_panes = "A2"

    window_headers = [
        "所属矩阵账号", "主页链接", "视频链接", "发布后小时数", "采集时间",
        "播放量", "点赞", "评论", "分享", "收藏", "reach",
        "总观看时长", "平均观看时长", "完播率", "互动率",
        "粉丝数（采集时）", "Distribution Rate (views/followers)",
    ]
    ws2 = wb.create_sheet(title="发布后时间点数据")
    ws2.append(window_headers)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font

    if rows:
        account_label_map = {
            (r["business_id"], r["item_id"]): (r.get("display_name") or r.get("account_alias") or r.get("account_name") or r.get("business_id"))
            for r in rows
        }
        link_map = {
            (r["business_id"], r["item_id"]): (r.get("profile_deep_link") or "", r.get("share_url") or "")
            for r in rows
        }
        for w in window_rows:
            key = (w["business_id"], w["item_id"])
            profile_link, video_link = link_map.get(key, ("", ""))
            captured = w.get("captured_at") is not None
            ws2.append(
                [
                    account_label_map.get(key, w["business_id"]),
                    profile_link,
                    video_link,
                    w["window_hours"],
                    w.get("captured_at") if captured else "未到期/未采集",
                    w.get("video_views") if captured else None,
                    w.get("likes") if captured else None,
                    w.get("comments") if captured else None,
                    w.get("shares") if captured else None,
                    w.get("favorites") if captured else None,
                    w.get("reach") if captured else None,
                    w.get("total_time_watched") if captured else None,
                    w.get("average_time_watched") if captured else None,
                    w.get("full_video_watched_rate") if captured else None,
                    w.get("engagement_rate") if captured else None,
                    w.get("followers_count_snapshot") if captured else None,
                    w.get("distribution_rate") if captured else None,
                ]
            )
    ws2.freeze_panes = "A2"

    daily_headers = [
        "所属矩阵账号", "视频ID", "视频链接", "发布日期", "第几天", "日期", "播放量", "较前日新增",
    ]
    ws3 = wb.create_sheet(title="每日播放量")
    ws3.append(daily_headers)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font

    if rows:
        create_dates = {(r["business_id"], r["item_id"]): r.get("create_time") for r in rows}
        account_label_map = {
            (r["business_id"], r["item_id"]): (r.get("display_name") or r.get("account_alias") or r.get("account_name") or r.get("business_id"))
            for r in rows
        }
        link_map = {(r["business_id"], r["item_id"]): (r.get("share_url") or "") for r in rows}
        for r in rows:
            key = (r["business_id"], r["item_id"])
            series = daily_series_map.get(key, {}).get("series") or []
            create_time = create_dates.get(key)
            for point in series:
                ws3.append(
                    [
                        account_label_map.get(key, r["business_id"]),
                        r["item_id"],
                        link_map.get(key, ""),
                        create_time.date() if create_time else None,
                        point["day_index"],
                        point["date"],
                        point["video_views"],
                        point["delta"],
                    ]
                )
    ws3.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def list_videos(business_id: str | None = None, page: int = 1, page_size: int = 50) -> dict[str, Any]:
    where = ""
    params: list[Any] = []
    if business_id:
        where = "WHERE v.business_id = %s"
        params.append(business_id)
    total_row = db.query_one(f"SELECT COUNT(*) AS count FROM tiktok_official_video_snapshots v {where}", tuple(params))
    rows = db.query_all(
        f"""
        SELECT v.*, a.account_name, a.display_name, a.username
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        {where}
        ORDER BY v.create_time DESC NULLS LAST, v.updated_at DESC
        LIMIT %s OFFSET %s
        """,
        tuple(params + [page_size, (page - 1) * page_size]),
    )
    return {"total": int(total_row.get("count") or 0) if total_row else 0, "items": rows or []}


def get_video(item_id: str, business_id: str | None = None) -> dict[str, Any] | None:
    if business_id:
        row = db.query_one(
            "SELECT * FROM tiktok_official_video_snapshots WHERE item_id = %s AND business_id = %s",
            (item_id, business_id),
        )
    else:
        row = db.query_one("SELECT * FROM tiktok_official_video_snapshots WHERE item_id = %s", (item_id,))
    if row:
        for key in (
            "video_view_retention",
            "engagement_likes",
            "impression_sources",
            "audience_genders",
            "audience_countries",
            "audience_cities",
            "audience_types",
            "raw_json",
        ):
            row[key] = _loads(row.get(key))
    return row


def list_profile_metrics(business_id: str | None = None, days: int = 30) -> list[dict[str, Any]]:
    days = max(1, min(int(days or 30), 60))
    params: list[Any] = [_business_date() - timedelta(days=days)]
    where = "WHERE metric_date >= %s"
    if business_id:
        where += " AND business_id = %s"
        params.append(business_id)
    return db.query_all(
        f"""
        SELECT *
        FROM tiktok_official_profile_daily_metrics
        {where}
        ORDER BY metric_date ASC, business_id ASC
        """,
        tuple(params),
    ) or []


def build_export(videos: list[dict[str, Any]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "视频明细"
    video_headers = [
        "发布时间",
        "账号",
        "视频",
        "播放",
        "覆盖",
        "点赞",
        "评论",
        "分享",
        "收藏",
        "平均观看",
        "视频时长",
        "完播率",
        "新增粉",
        "详情",
    ]
    ws.append(video_headers)
    rows = _export_video_rows(videos)
    for row in rows:
        ws.append(
            [
                row.get("create_time"),
                row.get("display_name") or row.get("account_name") or row.get("username") or row.get("business_id"),
                row.get("caption"),
                row.get("video_views"),
                row.get("reach"),
                row.get("likes"),
                row.get("comments"),
                row.get("shares"),
                row.get("favorites"),
                row.get("average_time_watched"),
                row.get("video_duration"),
                row.get("full_video_watched_rate"),
                row.get("new_followers"),
                row.get("share_url") or row.get("item_id"),
            ]
        )
    ws.freeze_panes = "A2"
    _autosize_columns(ws)
    for idx, row in enumerate(rows, start=1):
        _append_video_detail_sheet(wb, row, idx)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _export_video_rows(videos: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    selected: list[tuple[str, str | None]] = []
    for item in videos or []:
        if not isinstance(item, dict):
            continue
        item_id = str(item.get("item_id") or "").strip()
        business_id = str(item.get("business_id") or "").strip() or None
        if item_id:
            selected.append((item_id, business_id))
    if not selected:
        return list_videos(page=1, page_size=10000)["items"]

    clauses = []
    params: list[Any] = []
    for item_id, business_id in selected:
        if business_id:
            clauses.append("(v.item_id = %s AND v.business_id = %s)")
            params.extend([item_id, business_id])
        else:
            clauses.append("v.item_id = %s")
            params.append(item_id)
    rows = db.query_all(
        f"""
        SELECT v.*, a.account_name, a.display_name, a.username
        FROM tiktok_official_video_snapshots v
        LEFT JOIN tiktok_official_accounts a ON a.business_id = v.business_id
        WHERE {" OR ".join(clauses)}
        ORDER BY v.create_time DESC NULLS LAST, v.updated_at DESC
        """,
        tuple(params),
    )
    return rows or []


def _append_video_detail_sheet(wb: Workbook, row: dict[str, Any], index: int) -> None:
    title = _unique_sheet_title(wb, f"{index}-{_chart_label(row)}")
    ws = wb.create_sheet(title)
    account = row.get("display_name") or row.get("account_name") or row.get("username") or row.get("business_id")
    ws["A1"] = "视频详情"
    ws["A2"] = "账号"
    ws["B2"] = account
    ws["A3"] = "视频"
    ws["B3"] = row.get("caption")
    ws["A4"] = "链接"
    ws["B4"] = row.get("share_url") or row.get("item_id")

    engagement_likes = _like_points(_loads(row.get("engagement_likes")) or [], row.get("likes"))
    retention = _series_points(_loads(row.get("video_view_retention")) or [])
    likes_count = _write_like_table(ws, 30, 1, engagement_likes)
    retention_count = _write_series_table(ws, 30, 6, "每秒留存明细", retention, "留存率")
    _add_series_chart(ws, 30, 1, likes_count, "估算每秒点赞量", "估算点赞数", "A6", value_offset=2)
    _add_series_chart(ws, 30, 6, retention_count, "留存折线图", "留存率", "I6")
    _autosize_columns(ws)


def _write_like_table(ws, start_row: int, start_col: int, values) -> int:
    ws.cell(row=start_row, column=start_col, value="点赞时间分布")
    ws.cell(row=start_row + 1, column=start_col, value="second")
    ws.cell(row=start_row + 1, column=start_col + 1, value="点赞分布比例")
    ws.cell(row=start_row + 1, column=start_col + 2, value="估算点赞数")
    if not isinstance(values, list) or not values:
        ws.cell(row=start_row + 2, column=start_col, value="暂无数据")
        return 0
    out_row = start_row + 2
    for second, ratio, estimated_likes in values:
        ws.cell(row=out_row, column=start_col, value=second)
        ws.cell(row=out_row, column=start_col + 1, value=ratio)
        ws.cell(row=out_row, column=start_col + 2, value=estimated_likes)
        out_row += 1
    return len(values)


def _write_series_table(ws, start_row: int, start_col: int, title: str, values, value_header: str) -> int:
    ws.cell(row=start_row, column=start_col, value=title)
    ws.cell(row=start_row + 1, column=start_col, value="second")
    ws.cell(row=start_row + 1, column=start_col + 1, value=value_header)
    if not isinstance(values, list) or not values:
        ws.cell(row=start_row + 2, column=start_col, value="暂无数据")
        return 0
    out_row = start_row + 2
    count = 0
    for item in values:
        second, value = item
        ws.cell(row=out_row, column=start_col, value=second)
        ws.cell(row=out_row, column=start_col + 1, value=value)
        out_row += 1
        count += 1
    if count == 0:
        ws.cell(row=start_row + 2, column=start_col, value="暂无数据")
    return count


def _series_points(values) -> list[tuple[int, float]]:
    if not isinstance(values, list):
        return []
    points: dict[int, float] = {}
    for item in values:
        if not isinstance(item, dict):
            continue
        second = _series_second(item)
        value = _series_value(item)
        if second is None or value is None:
            continue
        points[int(second)] = float(value)
    return sorted(points.items())


def _like_points(values, total_likes) -> list[tuple[int, float, float]]:
    likes = _to_float(total_likes) or 0.0
    return [(second, ratio, ratio * likes) for second, ratio in _series_points(values)]


def _add_series_chart(
    ws,
    start_row: int,
    start_col: int,
    source_len: int,
    title: str,
    value_title: str,
    anchor: str,
    value_offset: int = 1,
) -> None:
    if source_len <= 0:
        return
    data_start = start_row + 1
    data_end = start_row + 1 + source_len
    if data_end <= data_start:
        return
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = value_title
    chart.x_axis.title = "second"
    data_col = start_col + value_offset
    data = Reference(ws, min_col=data_col, max_col=data_col, min_row=data_start, max_row=data_end)
    cats = Reference(ws, min_col=start_col, min_row=data_start + 1, max_row=data_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 12
    chart.width = 24
    ws.add_chart(chart, anchor)


def _unique_sheet_title(wb: Workbook, raw: str) -> str:
    safe = "".join("_" if ch in '[]:*?/\\' else ch for ch in str(raw or "video")).strip() or "video"
    base = safe[:31]
    title = base
    n = 2
    while title in wb.sheetnames:
        suffix = f"_{n}"
        title = base[: 31 - len(suffix)] + suffix
        n += 1
    return title


def run_refresh_task(task_id: str, params: dict[str, Any], update_task_fn) -> None:
    try:
        update_task_fn(task_id, status="processing", progress="TikTok 官号刷新中...")

        def hook(msg):
            update_task_fn(task_id, progress=msg)

        result = refresh_official_accounts(
            business_ids=params.get("business_ids") or None,
            profile_days=int(params.get("profile_days") or 30),
            max_pages=int(params.get("max_pages") or 5),
            progress_hook=hook,
        )
        usage_service.record_usage_event(
            module="tiktok_official_refresh",
            user_id=params.get("user_id"),
            task_id=task_id,
            item_count=int(result.get("videos") or 0),
            crawler_items=0,
            api_calls=int(result.get("accounts") or 0) * 2,
            source="actual",
            detail={"business_ids": params.get("business_ids") or [], "profile_days": params.get("profile_days"), "result": result},
        )
        failed_count = int(result.get("failed_count") or 0)
        progress = f"完成（{failed_count} 个账号刷新失败，已跳过）" if failed_count else "完成"
        update_task_fn(
            task_id,
            status="completed",
            progress=progress,
            result=json.dumps(result, ensure_ascii=False),
        )
    except Exception as exc:
        update_task_fn(task_id, status="failed", error=str(exc)[:500], progress="失败")


def run_publish_window_capture(task_id: str, params: dict[str, Any], update_task_fn) -> None:
    """采集某账号下到期的时间点数据（发布后3/24/48/72小时，以及 window_hours=0 的发布日定格）。

    TikTok `/business/video/list/` 不支持按 item_id 过滤，只能拉该账号视频列表全量再本地匹配。
    """
    business_id = params.get("business_id")
    targets = params.get("targets") or []
    try:
        update_task_fn(task_id, status="processing", progress="正在采集发布后时间点数据...")
        if not business_id or not targets:
            update_task_fn(task_id, status="completed", progress="无待采集目标")
            return

        token = get_access_token(business_id, auto_refresh=True)
        if not token:
            raise RuntimeError(f"{business_id} 缺少 access_token，请先完成 TikTok 账号授权")

        videos, _meta = fetch_videos(token, business_id, max_pages=5)
        video_by_id = {str(v.get("item_id") or ""): v for v in videos}

        account = db.query_one(
            "SELECT followers_count FROM tiktok_official_accounts WHERE business_id = %s",
            (business_id,),
        ) or {}
        followers_count = _to_int(account.get("followers_count"))

        captured = 0
        missing = 0
        for target in targets:
            item_id = str(target.get("item_id") or "")
            window_hours = target.get("window_hours")
            video = video_by_id.get(item_id)
            if not video:
                missing += 1
                continue

            views = _to_int(video.get("video_views"))
            likes = _to_int(video.get("likes"))
            comments = _to_int(video.get("comments"))
            shares = _to_int(video.get("shares"))
            favorites = _to_int(video.get("favorites"))
            engagement = sum(v or 0 for v in (likes, comments, shares, favorites))
            engagement_rate = (engagement / views) if views else None
            distribution_rate = (views / followers_count) if (views is not None and followers_count) else None

            db.execute(
                """
                UPDATE tiktok_official_video_publish_window_snapshots
                SET captured_at = NOW(), video_views = %s, likes = %s, comments = %s, shares = %s,
                    favorites = %s, reach = %s, total_time_watched = %s, average_time_watched = %s,
                    full_video_watched_rate = %s, impression_sources = %s, engagement_rate = %s,
                    followers_count_snapshot = %s, distribution_rate = %s
                WHERE business_id = %s AND item_id = %s AND window_hours = %s AND captured_at IS NULL
                """,
                (
                    views, likes, comments, shares, favorites,
                    _to_int(video.get("reach")),
                    _to_float(video.get("total_time_watched")),
                    _to_float(video.get("average_time_watched")),
                    _to_float(video.get("full_video_watched_rate")),
                    _json(video.get("impression_sources")),
                    engagement_rate,
                    followers_count,
                    distribution_rate,
                    business_id, item_id, window_hours,
                ),
            )
            captured += 1

        progress = f"完成（采集 {captured} 条，{missing} 条视频未匹配到已跳过）" if missing else f"完成（采集 {captured} 条）"
        update_task_fn(task_id, status="completed", progress=progress)
    except Exception as exc:
        logger.warning("⚠️ TikTok 发布后时间点数据采集失败：%s - %s", business_id, exc)
        update_task_fn(task_id, status="failed", error=str(exc)[:500], progress="失败")


def run_video_discovery_task(task_id: str, params: dict[str, Any], update_task_fn) -> None:
    """轻量新视频发现：只拉该账号视频列表第1页，只为尽早 upsert 新视频、建好时间点占位行，
    不做主页数据刷新（那部分仍由每日全量同步负责）。"""
    business_id = params.get("business_id")
    try:
        update_task_fn(task_id, status="processing", progress="正在发现新视频...")
        if not business_id:
            update_task_fn(task_id, status="completed", progress="无目标账号")
            return

        token = get_access_token(business_id, auto_refresh=True)
        if not token:
            raise RuntimeError(f"{business_id} 缺少 access_token，请先完成 TikTok 账号授权")

        videos, video_meta = fetch_videos(token, business_id, max_pages=1)
        for video in videos:
            upsert_video(business_id, video, video_meta)

        update_task_fn(task_id, status="completed", progress=f"完成（扫描 {len(videos)} 条）")
    except Exception as exc:
        logger.warning("⚠️ TikTok 新视频发现失败：%s - %s", business_id, exc)
        update_task_fn(task_id, status="failed", error=str(exc)[:500], progress="失败")


def _append_array_sheet(wb: Workbook, title: str, field: str, headers: list[str]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    rows = db.query_all(f"SELECT item_id, {field} FROM tiktok_official_video_snapshots WHERE {field} IS NOT NULL")
    for row in rows or []:
        values = _loads(row.get(field)) or []
        if not isinstance(values, list):
            continue
        for item in values:
            if isinstance(item, dict):
                ws.append([row.get("item_id"), _series_second(item), _series_value(item)])
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _append_series_chart_sheet(wb: Workbook, title: str, field: str, value_title: str) -> None:
    ws = wb.create_sheet(title)
    rows = db.query_all(
        f"""
        SELECT item_id, caption, {field}
        FROM tiktok_official_video_snapshots
        WHERE {field} IS NOT NULL
        ORDER BY create_time DESC NULLS LAST, updated_at DESC
        LIMIT 8
        """
    )
    series: list[tuple[str, dict[int, float]]] = []
    seconds: set[int] = set()
    for row in rows or []:
        values = _loads(row.get(field)) or []
        if not isinstance(values, list):
            continue
        points: dict[int, float] = {}
        for item in values:
            if not isinstance(item, dict):
                continue
            second = _series_second(item)
            value = _series_value(item)
            if second is None or value is None:
                continue
            points[int(second)] = float(value)
            seconds.add(int(second))
        if points:
            series.append((_chart_label(row), points))

    ws.append(["second"] + [label for label, _points in series])
    for second in sorted(seconds):
        ws.append([second] + [points.get(second) for _label, points in series])
    if not series or not seconds:
        ws["A2"] = f"暂无 {value_title} 分布数据"
        return

    chart = LineChart()
    chart.title = title
    chart.y_axis.title = value_title
    chart.x_axis.title = "second"
    data = Reference(ws, min_col=2, max_col=1 + len(series), min_row=1, max_row=1 + len(seconds))
    cats = Reference(ws, min_col=1, min_row=2, max_row=1 + len(seconds))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 14
    chart.width = 28
    ws.add_chart(chart, "H2")
    ws.freeze_panes = "A2"
    _autosize_columns(ws)


def _series_second(item: dict[str, Any]) -> int | None:
    for key in ("second", "seconds", "time", "time_second"):
        value = item.get(key)
        if value in (None, ""):
            continue
        try:
            return int(float(value))
        except Exception:
            continue
    return None


def _series_value(item: dict[str, Any]) -> float | None:
    for key in ("percentage", "percent", "value", "likes", "like_count", "count", "rate"):
        value = item.get(key)
        if value in (None, ""):
            continue
        try:
            return float(value)
        except Exception:
            continue
    return None


def _series_value_at_second(values, second: int) -> float | None:
    if not isinstance(values, list):
        return None
    for item in values:
        if isinstance(item, dict) and _series_second(item) == second:
            return _series_value(item)
    return None


def _format_series_summary(values) -> str | None:
    if not isinstance(values, list) or not values:
        return None
    points = []
    for item in values[:20]:
        if not isinstance(item, dict):
            continue
        second = _series_second(item)
        value = _series_value(item)
        if second is not None and value is not None:
            points.append(f"{second}s:{value:g}")
    return "; ".join(points) if points else None


def _chart_label(row: dict[str, Any]) -> str:
    caption = str(row.get("caption") or "").strip()
    item_id = str(row.get("item_id") or "").strip()
    label = caption[:18] if caption else item_id
    return label or "video"


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        letter = column_cells[0].column_letter
        max_len = 0
        for cell in column_cells[:80]:
            value = cell.value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 36)


def _normalize_account(raw: dict[str, Any]) -> dict[str, Any]:
    bid = (raw.get("business_id") or raw.get("open_id") or "").strip()
    if not bid:
        raise ValueError("官号配置缺少 business_id/open_id")
    return {
        "business_id": bid,
        "account_name": raw.get("account_name") or raw.get("name") or bid,
        "enabled": raw.get("enabled", True),
        "notes": raw.get("notes") or "",
    }


def _to_int(value) -> int | None:
    try:
        if value in (None, ""):
            return None
        return int(float(value))
    except Exception:
        return None


def _to_float(value) -> float | None:
    try:
        if value in (None, ""):
            return None
        return float(value)
    except Exception:
        return None


def _epoch_to_dt(value) -> datetime | None:
    try:
        if value in (None, ""):
            return None
        ts = float(value)
        if ts > 1e12:
            ts = ts / 1000.0
        return datetime.fromtimestamp(ts, tz=timezone.utc).replace(tzinfo=None)
    except Exception:
        return None


def _json(value) -> str | None:
    if value in (None, ""):
        return None
    return json.dumps(value, ensure_ascii=False)


def _loads(value):
    if not value:
        return None
    try:
        return json.loads(value)
    except Exception:
        return value
