"""竞品雷达 v2：TikTok + Instagram 视频数据采集 + AI 视频内容分析。

输出字段：达人名称 / 视频链接 / 视频时长 / 发布时间 / 播放量 / 点赞数 /
转发量 / 收藏量 / 评论量 / 视频文案 / 营销类型 / 内容概述

对外暴露：
  - detect_platform(url) -> 'TT' | 'IG' | 'UNKNOWN'
  - parse_urls_text(text) -> list[str]
  - run_radar_pipeline(urls, start_date, end_date, apify_token, ...)
  - build_html_table(structured) -> str
  - build_excel(structured) -> openpyxl Workbook
  - RADAR_HEADERS / SCHEMA_VERSION
"""

from __future__ import annotations

import datetime
import logging
import os
import re
import time
from html import escape
from typing import Callable, Iterable
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

SCHEMA_VERSION = "radar_v1"

RADAR_HEADERS = [
    "达人名称",
    "视频链接",
    "视频时长",
    "发布时间",
    "播放量",
    "点赞数",
    "转发量",
    "收藏量",
    "评论量",
    "视频文案",
    "营销类型",
    "内容概述",
]

# 12 列对应的 dict key（与 build_html_table / build_excel / 内部数据结构保持一致）
RADAR_KEYS = [
    "author",
    "video_url",
    "duration",
    "post_time",
    "views",
    "likes",
    "shares",
    "collects",
    "comments",
    "caption",
    "marketing_type",
    "vision_summary",
]

# Apify Actor 配置（可通过环境变量覆盖）
DEFAULT_ACTORS = {
    "TT": os.environ.get("APIFY_TIKTOK_PROFILE_ACTOR_ID", "clockworks/tiktok-scraper"),
    "IG": os.environ.get("APIFY_INSTAGRAM_PROFILE_ACTOR_ID", "apify/instagram-scraper"),
}

ACTOR_TIMEOUT_SECS = int(os.environ.get("RADAR_ACTOR_TIMEOUT_SECS", "480"))
ACTOR_POLL_INTERVAL = int(os.environ.get("RADAR_ACTOR_POLL_INTERVAL", "5"))
POSTS_PER_PROFILE_LIMIT = int(os.environ.get("RADAR_POSTS_PER_PROFILE", "50"))
MAX_VIDEOS_PER_PROFILE = int(os.environ.get("RADAR_MAX_VIDEOS_PER_PROFILE", "120"))

BEIJING_TZ = datetime.timezone(datetime.timedelta(hours=8))

MARKETING_TYPES = [
    "游戏实录",
    "赛事预热",
    "版本更新预告",
    "KOL合作/开箱",
    "福利/抽奖活动",
    "品牌形象宣传",
    "UGC互动征集",
    "其他",
]


# ============================================
# 平台识别 & URL 解析
# ============================================

_TT_HOSTS = ("tiktok.com", "vm.tiktok.com", "m.tiktok.com")
_IG_HOSTS = ("instagram.com",)


def detect_platform(url: str) -> str:
    if not url:
        return "UNKNOWN"
    u = url.lower()
    if any(h in u for h in _TT_HOSTS):
        return "TT"
    if any(h in u for h in _IG_HOSTS):
        return "IG"
    return "UNKNOWN"


def parse_urls_text(text: str) -> list[str]:
    """从用户输入（多行/逗号分隔）解析出 URL 列表，去重去空。"""
    if not text:
        return []
    seen: set[str] = set()
    result: list[str] = []
    for token in re.split(r"[\s,;\n]+", text.strip()):
        token = token.strip()
        if not token:
            continue
        if not (token.startswith("http://") or token.startswith("https://")):
            continue
        if token in seen:
            continue
        seen.add(token)
        result.append(token)
    return result


# ============================================
# Apify 调用（与 sentiment_insight 同套写法）
# ============================================


def _start_actor(actor_id: str, run_input: dict, apify_token: str) -> dict:
    actor_path = actor_id.replace("/", "~")
    api_url = f"https://api.apify.com/v2/acts/{actor_path}/runs"
    headers = {"Authorization": f"Bearer {apify_token}", "Content-Type": "application/json"}
    resp = requests.post(api_url, json=run_input, headers=headers, timeout=30)
    if resp.status_code != 201:
        raise RuntimeError(f"启动 {actor_id} 失败 status={resp.status_code} body={resp.text[:200]}")
    return resp.json().get("data", {})


def _abort_actor_run(run_id: str, apify_token: str) -> None:
    headers = {"Authorization": f"Bearer {apify_token}"}
    try:
        resp = requests.post(
            f"https://api.apify.com/v2/actor-runs/{run_id}/abort",
            headers=headers,
            timeout=15,
        )
        if resp.status_code >= 400:
            logger.warning("⚠️ abort actor run %s 失败 status=%s body=%s", run_id, resp.status_code, resp.text[:200])
    except Exception as exc:
        logger.warning("⚠️ abort actor run %s 异常: %s", run_id, exc)


def _wait_actor(run_id: str, apify_token: str, timeout: int = ACTOR_TIMEOUT_SECS, should_abort: Callable[[], bool] | None = None) -> dict:
    headers = {"Authorization": f"Bearer {apify_token}"}
    api_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    start = time.time()
    while True:
        if should_abort and should_abort():
            _abort_actor_run(run_id, apify_token)
            raise RuntimeError(f"actor run {run_id} 已因任务停止而中断")
        if time.time() - start > timeout:
            _abort_actor_run(run_id, apify_token)
            raise TimeoutError(f"actor run {run_id} 等待超时（{timeout}s）")
        resp = requests.get(api_url, headers=headers, timeout=15)
        if resp.status_code != 200:
            raise RuntimeError(f"查询 run {run_id} 失败 status={resp.status_code}")
        data = resp.json().get("data", {})
        status = data.get("status")
        if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
            return data
        for _ in range(max(1, int(ACTOR_POLL_INTERVAL))):
            if should_abort and should_abort():
                _abort_actor_run(run_id, apify_token)
                raise RuntimeError(f"actor run {run_id} 已因任务停止而中断")
            time.sleep(1)


def _fetch_dataset(dataset_id: str, apify_token: str) -> list[dict]:
    headers = {"Authorization": f"Bearer {apify_token}"}
    api_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    resp = requests.get(api_url, headers=headers, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"获取 dataset {dataset_id} 失败 status={resp.status_code}")
    items = resp.json()
    return items if isinstance(items, list) else []


def _call_actor(
    actor_id: str,
    candidate_inputs: list[dict],
    apify_token: str,
    *,
    should_abort: Callable[[], bool] | None = None,
    allow_input_fallback: bool = True,
) -> list[dict]:
    last_err = None
    inputs = candidate_inputs if allow_input_fallback else candidate_inputs[:1]
    for run_input in inputs:
        if should_abort and should_abort():
            raise RuntimeError("任务已停止，未继续启动 Apify actor")
        try:
            run = _start_actor(actor_id, run_input, apify_token)
            run_id = run.get("id")
            if not run_id:
                raise RuntimeError("actor 返回缺少 run id")
            final = _wait_actor(run_id, apify_token, should_abort=should_abort)
            if final.get("status") != "SUCCEEDED":
                raise RuntimeError(f"run 结束状态={final.get('status')}")
            dataset_id = final.get("defaultDatasetId")
            if not dataset_id:
                raise RuntimeError("run 缺少 defaultDatasetId")
            items = _fetch_dataset(dataset_id, apify_token)
            return items
        except Exception as e:
            last_err = e
            logger.warning(f"⚠️ actor {actor_id} 调用失败（input keys={list(run_input.keys())}）: {e}")
            if should_abort and should_abort():
                raise RuntimeError(f"actor {actor_id} 已停止: {e}") from e
            continue
    raise RuntimeError(f"actor {actor_id} 所有候选 input 均失败: {last_err}")


# ============================================
# 时间/格式化工具
# ============================================


def _to_beijing_dt(value) -> datetime.datetime | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, (int, float)):
            ts = float(value)
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).astimezone(BEIJING_TZ)
        s = str(value).strip()
        if not s:
            return None
        if s.isdigit():
            ts = float(s)
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).astimezone(BEIJING_TZ)
        dt = datetime.datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        return dt.astimezone(BEIJING_TZ)
    except Exception:
        return None


def _fmt_dt(dt: datetime.datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _fmt_duration(seconds) -> str:
    """秒数 -> mm:ss / hh:mm:ss"""
    try:
        if seconds is None or seconds == "":
            return ""
        s = int(float(seconds))
    except Exception:
        return ""
    if s <= 0:
        return ""
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    if h > 0:
        return f"{h:02d}:{m:02d}:{sec:02d}"
    return f"{m:02d}:{sec:02d}"


def _safe_int(value) -> int:
    try:
        if value is None or value == "":
            return 0
        return int(float(value))
    except Exception:
        return 0


def _first_str(item: dict, keys: Iterable[str]) -> str:
    for k in keys:
        v = item.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, dict):
            for sub in ("name", "username", "userName", "fullName", "displayName"):
                vv = v.get(sub)
                if isinstance(vv, str) and vv.strip():
                    return vv.strip()
    return ""


# ============================================
# 平台抓取：TikTok 主页
# ============================================


def _scrape_tiktok_profile(url: str, start_dt_str: str | None, apify_token: str) -> list[dict]:
    actor = DEFAULT_ACTORS["TT"]
    base_input = {
        "profiles": [url],
        "resultsPerPage": POSTS_PER_PROFILE_LIMIT,
        "shouldDownloadVideos": False,
    }
    if start_dt_str:
        base_input["oldestPostDate"] = start_dt_str
    candidates = [base_input, {"profileUrls": [url], "resultsPerPage": POSTS_PER_PROFILE_LIMIT, "shouldDownloadVideos": False}]
    return _call_actor(actor, candidates, apify_token)


def _extract_tiktok_video(item: dict) -> dict:
    """从 TikTok actor 输出抽取一个视频的标准字段。"""
    author_meta = item.get("authorMeta") or {}
    video_meta = item.get("videoMeta") or {}
    author = (
        _first_str(item, ["authorName", "uniqueId"])
        or _first_str(author_meta, ["nickName", "nickname", "name", "uniqueId"])
    )
    if not author and isinstance(author_meta, dict):
        author = author_meta.get("nickname") or author_meta.get("name") or author_meta.get("uniqueId") or ""

    web_url = item.get("webVideoUrl") or item.get("url") or ""
    duration_sec = video_meta.get("duration") or item.get("duration") or item.get("videoDuration")
    create_dt = _to_beijing_dt(item.get("createTimeISO") or item.get("createTime") or item.get("createTimestamp"))

    caption = item.get("text") or item.get("desc") or item.get("title") or ""

    # 直链视频（供视觉分析下载，可能为空）
    direct_video_url = ""
    if isinstance(video_meta, dict):
        direct_video_url = (
            video_meta.get("downloadAddr")
            or video_meta.get("playAddr")
            or video_meta.get("originalDownloadAddr")
            or ""
        )

    return {
        "platform": "TT",
        "author": author or "未知",
        "video_url": web_url,
        "duration": _fmt_duration(duration_sec),
        "duration_sec": _safe_int(duration_sec),
        "post_dt": create_dt,
        "post_time": _fmt_dt(create_dt),
        "post_date": create_dt.date().isoformat() if create_dt else "",
        "views": _safe_int(item.get("playCount")),
        "likes": _safe_int(item.get("diggCount")),
        "shares": _safe_int(item.get("shareCount")),
        "collects": _safe_int(item.get("collectCount")),
        "comments": _safe_int(item.get("commentCount")),
        "caption": (caption or "").strip(),
        "direct_video_url": direct_video_url,
        "cover_url": (video_meta.get("coverUrl") if isinstance(video_meta, dict) else "") or item.get("coverUrl") or "",
    }


# ============================================
# 平台抓取：Instagram 主页
# ============================================


def _scrape_instagram_profile(url: str, apify_token: str) -> list[dict]:
    actor = DEFAULT_ACTORS["IG"]
    candidates = [
        # apify/instagram-scraper（profile -> posts）
        {
            "directUrls": [url],
            "resultsType": "posts",
            "resultsLimit": POSTS_PER_PROFILE_LIMIT,
            "searchType": "user",
            "addParentData": False,
        },
        {
            "username": [_extract_ig_username(url)] if _extract_ig_username(url) else [],
            "resultsType": "posts",
            "resultsLimit": POSTS_PER_PROFILE_LIMIT,
        },
        {"directUrls": [url], "resultsLimit": POSTS_PER_PROFILE_LIMIT},
    ]
    candidates = [c for c in candidates if c]
    return _call_actor(actor, candidates, apify_token)


def _extract_ig_username(url: str) -> str:
    try:
        p = urlparse(url)
        path = (p.path or "").strip("/")
        if not path:
            return ""
        first = path.split("/")[0]
        if first in {"p", "reel", "tv", "explore"}:
            return ""
        return first.lstrip("@")
    except Exception:
        return ""


def _extract_instagram_video(item: dict) -> dict | None:
    """从 IG actor 输出抽取一个 *视频/Reel* 的标准字段。只返回视频；图集/单图返回 None。"""
    media_type = (item.get("type") or item.get("productType") or item.get("mediaType") or "").lower()
    has_video = bool(
        item.get("videoUrl")
        or item.get("video_url")
        or item.get("isVideo")
        or media_type in {"video", "clips", "reel", "reels", "igtv"}
    )
    if not has_video:
        return None

    author = (
        item.get("ownerFullName")
        or item.get("ownerUsername")
        or _first_str(item.get("owner") or {}, ["full_name", "username"])
        or item.get("username")
        or ""
    )

    post_url = item.get("url") or item.get("postUrl") or ""
    if not post_url and item.get("shortCode"):
        post_url = f"https://www.instagram.com/p/{item.get('shortCode')}/"

    duration_sec = item.get("videoDuration") or item.get("video_duration")
    create_dt = _to_beijing_dt(item.get("timestamp") or item.get("takenAtTimestamp") or item.get("taken_at_timestamp"))

    caption = item.get("caption") or item.get("text") or ""

    direct_video_url = item.get("videoUrl") or item.get("video_url") or ""

    return {
        "platform": "IG",
        "author": (author or "未知").strip(),
        "video_url": post_url,
        "duration": _fmt_duration(duration_sec),
        "duration_sec": _safe_int(duration_sec),
        "post_dt": create_dt,
        "post_time": _fmt_dt(create_dt),
        "post_date": create_dt.date().isoformat() if create_dt else "",
        "views": _safe_int(item.get("videoPlayCount") or item.get("videoViewCount") or item.get("playCount")),
        "likes": _safe_int(item.get("likesCount") or item.get("likes_count")),
        # IG 公开数据通常没有转发量
        "shares": 0,
        # IG 公开数据通常没有收藏量
        "collects": 0,
        "comments": _safe_int(item.get("commentsCount") or item.get("comments_count")),
        "caption": (caption or "").strip(),
        "direct_video_url": direct_video_url,
        "cover_url": item.get("displayUrl") or item.get("thumbnailUrl") or "",
    }


# ============================================
# 主流程
# ============================================


def _filter_in_range(video: dict, start_dt: datetime.date | None, end_dt: datetime.date | None) -> bool:
    if not start_dt and not end_dt:
        return True
    if not video.get("post_dt"):
        return False
    d = video["post_dt"].date()
    if start_dt and d < start_dt:
        return False
    if end_dt and d > end_dt:
        return False
    return True


def _scrape_single_profile(
    profile_url: str,
    platform: str,
    apify_token: str,
    start_dt: datetime.date | None,
    end_dt: datetime.date | None,
    start_dt_str: str | None,
) -> dict:
    """抓取并标准化单个主页的视频列表。返回 {profile_url, platform, author, videos:[...] }。"""
    if platform == "TT":
        items = _scrape_tiktok_profile(profile_url, start_dt_str, apify_token)
        videos: list[dict] = []
        for it in items or []:
            if not isinstance(it, dict):
                continue
            row = _extract_tiktok_video(it)
            if not row.get("video_url"):
                continue
            videos.append(row)
    elif platform == "IG":
        items = _scrape_instagram_profile(profile_url, apify_token)
        videos = []
        for it in items or []:
            if not isinstance(it, dict):
                continue
            row = _extract_instagram_video(it)
            if not row or not row.get("video_url"):
                continue
            videos.append(row)
    else:
        return {"profile_url": profile_url, "platform": platform, "author": "", "videos": []}

    # 时间过滤
    videos = [v for v in videos if _filter_in_range(v, start_dt, end_dt)]
    # 按发布时间倒序
    videos.sort(key=lambda v: v.get("post_dt") or datetime.datetime.min.replace(tzinfo=BEIJING_TZ), reverse=True)
    # 顶部限制
    if len(videos) > MAX_VIDEOS_PER_PROFILE:
        videos = videos[:MAX_VIDEOS_PER_PROFILE]

    # 主页代表 author（取出现最多的作者名）
    author_counts: dict[str, int] = {}
    for v in videos:
        a = (v.get("author") or "").strip()
        if a:
            author_counts[a] = author_counts.get(a, 0) + 1
    profile_author = max(author_counts.items(), key=lambda x: x[1])[0] if author_counts else ""

    return {
        "profile_url": profile_url,
        "platform": platform,
        "author": profile_author or _extract_ig_username(profile_url) or "未知",
        "videos": videos,
    }


def run_radar_pipeline(
    urls: list[str],
    apify_token: str,
    *,
    start_date: datetime.date | None = None,
    end_date: datetime.date | None = None,
    enable_vision: bool = False,
    vision_call: Callable[[dict], dict] | None = None,
    progress: Callable[[str], None] | None = None,
) -> dict:
    """主流程：URL 列表 -> 每个主页的视频列表 -> 视觉分析（可选） -> 结构化数据。

    返回：
      {
        "version": SCHEMA_VERSION,
        "profiles": [{
            "profile_url": str,
            "platform": "TT"/"IG",
            "author": str,
            "videos": [
              { ...12 列字段..., "platform":..., "post_date":... }
            ],
        }, ...],
        "total_videos": int,
        "total_vision_done": int,
      }

    vision_call: 接收一个 video dict（含 direct_video_url/caption/platform 等），
                 返回 {"marketing_type": str, "vision_summary": str}。
    """
    def _p(msg: str) -> None:
        if progress:
            try:
                progress(msg)
            except Exception:
                pass
        logger.info(msg)

    profiles: list[dict] = []
    total_videos = 0

    for idx, url in enumerate(urls, 1):
        platform = detect_platform(url)
        if platform not in {"TT", "IG"}:
            _p(f"⚠️ 跳过不支持的平台链接：{url}")
            profiles.append({"profile_url": url, "platform": "UNKNOWN", "author": "", "videos": [], "error": "不支持的平台"})
            continue
        _p(f"🌐 [{idx}/{len(urls)}] 抓取 {platform} 主页：{url}")
        try:
            start_dt_str = start_date.isoformat() if start_date else None
            data = _scrape_single_profile(url, platform, apify_token, start_date, end_date, start_dt_str)
        except Exception as e:
            logger.error(f"❌ 抓取主页失败 {url}: {e}")
            profiles.append({"profile_url": url, "platform": platform, "author": "", "videos": [], "error": str(e)})
            continue
        total_videos += len(data["videos"])
        profiles.append(data)

    _p(f"📦 已采集视频数：{total_videos}")

    total_vision_done = 0
    if enable_vision and vision_call:
        for prof in profiles:
            for v in prof.get("videos", []):
                try:
                    out = vision_call(v) or {}
                    mtype = (out.get("marketing_type") or "").strip()
                    summary = (out.get("vision_summary") or "").strip()
                    if mtype:
                        v["marketing_type"] = mtype
                    if summary:
                        v["vision_summary"] = summary
                    if mtype or summary:
                        total_vision_done += 1
                except Exception as e:
                    logger.warning(f"⚠️ 单条视频视觉分析失败: {e}")
                    continue
        _p(f"🎬 视觉分析完成视频数：{total_vision_done}")

    # 兜底字段
    for prof in profiles:
        for v in prof.get("videos", []):
            v.setdefault("marketing_type", "")
            v.setdefault("vision_summary", "")

    return {
        "version": SCHEMA_VERSION,
        "profiles": profiles,
        "total_videos": total_videos,
        "total_vision_done": total_vision_done,
    }


# ============================================
# HTML 输出
# ============================================


def _short_cell(value: str, limit: int) -> str:
    if not value:
        return ""
    v = re.sub(r"\s+", " ", str(value)).strip()
    if len(v) <= limit:
        return v
    return v[:limit] + "…"


def build_html_table(structured: dict) -> str:
    profiles = (structured or {}).get("profiles") or []
    if not profiles:
        return "<div class='alert alert-warning'>未抓取到任何视频数据。</div>"

    parts: list[str] = []
    # 总览
    total_videos = sum(len(p.get("videos") or []) for p in profiles)
    parts.append(
        f"<div style='margin:8px 0 18px;color:#555;'>共 {len(profiles)} 个主页 · {total_videos} 条视频</div>"
    )

    for prof in profiles:
        platform = prof.get("platform", "")
        author = prof.get("author", "") or "未知"
        videos = prof.get("videos") or []
        prof_url = prof.get("profile_url", "")
        err = prof.get("error")
        head_line = (
            f"<h5 style='margin:18px 0 8px;color:#333;'>"
            f"<span style='display:inline-block;padding:2px 8px;border-radius:6px;background:#FFEBEE;color:#D32F2F;font-size:0.8rem;margin-right:8px;'>{escape(platform)}</span>"
            f"{escape(author)} · {len(videos)} 条视频"
            f"<a href='{escape(prof_url)}' target='_blank' style='margin-left:10px;font-size:0.85rem;color:#888;'>主页 ↗</a>"
            f"</h5>"
        )
        parts.append(head_line)
        if err:
            parts.append(f"<div class='alert alert-warning' style='font-size:0.85rem;'>{escape(err)}</div>")
        if not videos:
            parts.append("<div style='color:#999;font-size:0.85rem;margin-bottom:10px;'>该时间范围内未发现视频。</div>")
            continue

        parts.append("<table class='radar-table'><thead><tr>")
        for h in RADAR_HEADERS:
            parts.append(f"<th>{escape(h)}</th>")
        parts.append("</tr></thead><tbody>")
        for v in videos:
            video_url = v.get("video_url") or ""
            url_cell = (
                f"<a href='{escape(video_url)}' target='_blank' rel='noopener'>{escape(_short_cell(video_url, 40))}</a>"
                if video_url else ""
            )
            parts.append("<tr>")
            parts.append(f"<td>{escape(v.get('author', '') or '')}</td>")
            parts.append(f"<td class='cell-link'>{url_cell}</td>")
            parts.append(f"<td>{escape(v.get('duration', '') or '')}</td>")
            parts.append(f"<td>{escape(v.get('post_time', '') or '')}</td>")
            parts.append(f"<td>{_safe_int(v.get('views', 0))}</td>")
            parts.append(f"<td>{_safe_int(v.get('likes', 0))}</td>")
            parts.append(f"<td>{_safe_int(v.get('shares', 0))}</td>")
            parts.append(f"<td>{_safe_int(v.get('collects', 0))}</td>")
            parts.append(f"<td>{_safe_int(v.get('comments', 0))}</td>")
            parts.append(f"<td class='cell-caption'>{escape(_short_cell(v.get('caption', ''), 140))}</td>")
            parts.append(f"<td>{escape(v.get('marketing_type', '') or '')}</td>")
            parts.append(f"<td class='cell-vision'>{escape(v.get('vision_summary', '') or '')}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table>")

    # 样式由前端模板提供（templates/competitor.html 中 .radar-table 系列）
    return "".join(parts)


# ============================================
# Excel 输出
# ============================================


_INVALID_SHEET_CHARS = re.compile(r"[\\/?*\[\]:]")


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub(" ", name or "").strip() or "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 2
    while cleaned in used:
        tail = f" ({suffix})"
        cleaned = (base[: 31 - len(tail)] + tail).strip()
        suffix += 1
    used.add(cleaned)
    return cleaned


def _write_radar_sheet(ws, profile: dict) -> None:
    header_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    # 顶部主页信息一行
    platform = profile.get("platform", "")
    author = profile.get("author", "")
    profile_url = profile.get("profile_url", "")
    ws.append([f"平台：{platform}", f"达人：{author}", f"主页：{profile_url}"])
    info_font = Font(bold=True, color="555555")
    for cell in ws[1]:
        cell.font = info_font
        cell.alignment = Alignment(vertical="center")

    ws.append([])  # 空行
    ws.append(RADAR_HEADERS)
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for v in profile.get("videos") or []:
        ws.append([
            v.get("author", ""),
            v.get("video_url", ""),
            v.get("duration", ""),
            v.get("post_time", ""),
            _safe_int(v.get("views", 0)),
            _safe_int(v.get("likes", 0)),
            _safe_int(v.get("shares", 0)),
            _safe_int(v.get("collects", 0)),
            _safe_int(v.get("comments", 0)),
            v.get("caption", ""),
            v.get("marketing_type", ""),
            v.get("vision_summary", ""),
        ])

    widths = [16, 50, 10, 18, 12, 12, 12, 12, 12, 50, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数据行 wrap，链接列加超链接
    link_font = Font(color="0563C1", underline="single")
    for row_idx, row in enumerate(ws.iter_rows(min_row=4), start=4):
        for cell in row:
            cell.alignment = wrap
        link_cell = row[1] if len(row) > 1 else None
        url = link_cell.value if link_cell else None
        if url and isinstance(url, str) and url.startswith(("http://", "https://")):
            link_cell.hyperlink = url
            link_cell.font = link_font

    ws.freeze_panes = "A4"


def build_excel(structured: dict) -> Workbook:
    """生成 Excel：每个主页一个 sheet。"""
    wb = Workbook()
    wb.remove(wb.active)

    profiles = (structured or {}).get("profiles") or []
    if not profiles:
        ws = wb.create_sheet("竞品雷达")
        ws.append(RADAR_HEADERS)
        return wb

    used: set[str] = set()
    seq_counter: dict[str, int] = {}
    for prof in profiles:
        platform = (prof.get("platform") or "X").strip() or "X"
        seq_counter[platform] = seq_counter.get(platform, 0) + 1
        seq = seq_counter[platform]
        author = re.sub(r"\s+", " ", (prof.get("author") or "")).strip()
        base = f"{platform}-{seq}"
        raw_name = f"{base} {author[:18]}".strip() if author else base
        sheet_name = _sanitize_sheet_name(raw_name, used)
        ws = wb.create_sheet(sheet_name)
        _write_radar_sheet(ws, prof)

    return wb
