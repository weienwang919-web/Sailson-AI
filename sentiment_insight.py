"""舆情洞察 v2：多平台 Apify 抓取 + AI 翻译/情感/分类。

支持平台：
  - FB  : apify/facebook-comments-scraper（评论时间通常缺失，留空）
  - IG  : apify/instagram-comment-scraper
  - TT  : clockworks/tiktok-comments-scraper
  - YTB : streamers/youtube-comments-scraper（可由环境变量覆盖）
  - X   : apidojo/tweet-scraper（可由环境变量覆盖）

对外暴露：
  - detect_platform(url) -> 'FB' | 'IG' | 'TT' | 'YTB' | 'X' | 'UNKNOWN'
  - run_insight_pipeline(urls, apify_token, ai_call, progress=None)
      -> dict(structured=[...], html=str, total_comments=int, total_tokens=int)
  - build_excel(structured) -> openpyxl Workbook
  - INSIGHT_HEADERS / SCHEMA_VERSION
"""

from __future__ import annotations

import datetime
import hashlib
import inspect
import json
import logging
import os
import re
import time
from typing import Callable, Iterable
from urllib.parse import parse_qsl, urlencode, urljoin, urlparse, urlunparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

SCHEMA_VERSION = "insight_v2"

# Excel/HTML 表头（图中样式）
INSIGHT_HEADERS = [
    "来源序号",
    "平台",
    "发布时间",
    "主贴链接",
    "主贴标题",
    "评论时间",
    "归属时间段",
    "评论点赞数",
    "评论人",
    "内容原文",
    "内容翻译",
    "内容判断（正向/中立/负面）",
    "人工复核（正向/中立/负面）",
    "内容分类",
    "评论ID",
    "评论链接",
    "抓取状态",
]

# Apify Actor 配置（可通过环境变量覆盖）
DEFAULT_ACTORS = {
    "FB": os.environ.get("APIFY_FB_COMMENTS_ACTOR_ID", "apify/facebook-comments-scraper"),
    "IG": os.environ.get("APIFY_IG_COMMENTS_ACTOR_ID", "apify/instagram-comment-scraper"),
    "TT": os.environ.get("APIFY_TT_COMMENTS_ACTOR_ID", "clockworks/tiktok-comments-scraper"),
    "YTB": os.environ.get("APIFY_YT_COMMENTS_ACTOR_ID", "streamers/youtube-comments-scraper"),
    "X": os.environ.get("APIFY_X_TWEET_ACTOR_ID", "apidojo/tweet-scraper"),
}
FB_COOKIE_FALLBACK_ACTOR_ID = os.environ.get(
    "APIFY_FB_COOKIE_COMMENTS_ACTOR_ID",
    "dz_omar/facebook-comment-scraper",
)
FB_COOKIE_FALLBACK_SORT = os.environ.get(
    "FB_COOKIE_FALLBACK_SORT",
    "RANKED_UNFILTERED_CHRONOLOGICAL_REPLIES_INTENT_V1",
)

# 单个 actor 最长等待时间（秒）
ACTOR_TIMEOUT_SECS = int(os.environ.get("INSIGHT_ACTOR_TIMEOUT_SECS", "420"))
ACTOR_POLL_INTERVAL = int(os.environ.get("INSIGHT_ACTOR_POLL_INTERVAL", "5"))
DEFAULT_COMMENTS_PER_POST_LIMIT = int(os.environ.get("INSIGHT_COMMENTS_PER_POST", "500"))
UNLIMITED_COMMENTS_PER_POST_LIMIT = int(os.environ.get("INSIGHT_COMMENTS_UNLIMITED_PER_POST", "50000"))
COMMENTS_PER_POST_LIMIT = DEFAULT_COMMENTS_PER_POST_LIMIT
AI_BATCH_SIZE = int(os.environ.get("INSIGHT_AI_BATCH_SIZE", "30"))
MAX_AI_COMMENTS = int(os.environ.get("INSIGHT_MAX_AI_COMMENTS", "1500"))

BEIJING_TZ = datetime.timezone(datetime.timedelta(hours=8))

# 允许的标签（白名单，用于校验 AI 输出）
SENTIMENT_LABELS = {"正向", "中立", "负面"}
CATEGORY_LABELS = {
    "产品体验",
    "功能建议",
    "账号&充值",
    "外挂作弊",
    "活动运营",
    "客服投诉",
    "其他",
}

# 用于预处理：剥离 [Sticker] / [贴图] 等元标记
_STICKER_PREFIX_RE = re.compile(r"^\s*\[(?:sticker|贴图|表情|emoji)\]\s*", re.IGNORECASE)
# 用于识别"无实质内容"评论（纯 emoji / 重复字符 / 单符号）
_EMOJI_OR_PUNCT_RE = re.compile(
    r"[\u2600-\u27BF\U0001F300-\U0001FAFF\U0001F000-\U0001F2FF\u2300-\u23FF\u2B00-\u2BFF\s\W_]+",
    re.UNICODE,
)


def _truthy(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "unlimited", "none", "no_limit", "不设上限", "不限"}


def normalize_comments_per_post_limit(value=None, unlimited=False) -> int:
    """Normalize a per-link comment scraping cap.

    "No limit" is represented as 50,000 rows because Apify actors still need a
    finite upper bound.
    """
    upper_bound = max(1, UNLIMITED_COMMENTS_PER_POST_LIMIT)
    if _truthy(unlimited):
        return upper_bound
    if value is None or str(value).strip() == "":
        raw_limit = DEFAULT_COMMENTS_PER_POST_LIMIT
    else:
        try:
            raw_limit = int(float(str(value).strip()))
        except (TypeError, ValueError):
            raw_limit = DEFAULT_COMMENTS_PER_POST_LIMIT
    return max(1, min(raw_limit, upper_bound))


def _preprocess_comment(text: str) -> tuple[str, str]:
    """评论预处理。返回 (送给 AI 的净化文本, 调用方应直接使用的翻译占位)。

    - 占位为非空时，跳过 AI 翻译，直接落库。
    - 净化文本去除了 [Sticker] 这类元标记。
    """
    if not text:
        return "", ""
    cleaned = text.strip()
    # 剥离前缀的元标记，比如 "[Sticker] 真好玩" -> "真好玩"
    cleaned = _STICKER_PREFIX_RE.sub("", cleaned).strip()

    # 如果剥离后是空的，说明原文本身只是个表情贴图
    if not cleaned:
        return "", "（贴图/无文字内容）"

    # 纯 emoji / 标点 / 单字母重复（长度短 + 没有有意义字母字数）
    letters = re.sub(r"[\W_0-9]+", "", cleaned, flags=re.UNICODE)
    if not letters:
        return cleaned, "（仅表情/符号）"
    # 灌水识别：同一个字母连续重复 ≥5 次（aaaaa / pubgggggg / hahahaha）
    if re.search(r"(.)\1{4,}", cleaned, flags=re.IGNORECASE | re.UNICODE):
        return cleaned, cleaned
    # 极端单字母重复：去重后只有 1-2 个字符
    if len(set(letters.lower())) <= 2 and len(cleaned) <= 60:
        return cleaned, cleaned
    return cleaned, ""


# ============================================
# 平台识别
# ============================================

_FB_HOSTS = ("facebook.com", "fb.com", "fb.watch", "m.facebook.com", "web.facebook.com")
_IG_HOSTS = ("instagram.com",)
_TT_HOSTS = ("tiktok.com", "vm.tiktok.com", "m.tiktok.com")
_YT_HOSTS = ("youtube.com", "youtu.be", "m.youtube.com")
_X_HOSTS = ("twitter.com", "x.com", "mobile.twitter.com", "mobile.x.com")


def detect_platform(url: str) -> str:
    if not url:
        return "UNKNOWN"
    try:
        host = (urlparse(url.strip()).hostname or "").lower()
    except Exception:
        host = url.strip().lower()
    if any(host == h or host.endswith("." + h) or host == h.split(".", 1)[-1] for h in _FB_HOSTS):
        # 兼容 m./web. 子域
        pass
    # 简单匹配
    u = url.lower()
    if any(h in u for h in _FB_HOSTS):
        return "FB"
    if any(h in u for h in _IG_HOSTS):
        return "IG"
    if any(h in u for h in _TT_HOSTS):
        return "TT"
    if any(h in u for h in _YT_HOSTS):
        return "YTB"
    if any(h in u for h in _X_HOSTS):
        return "X"
    return "UNKNOWN"


# ============================================
# Apify 调用
# ============================================


def _start_actor(actor_id: str, run_input: dict, apify_token: str) -> dict:
    """启动 Apify actor，返回 run 数据（含 id）。"""
    actor_path = actor_id.replace("/", "~")
    api_url = f"https://api.apify.com/v2/acts/{actor_path}/runs"
    headers = {"Authorization": f"Bearer {apify_token}", "Content-Type": "application/json"}
    resp = requests.post(api_url, json=run_input, headers=headers, timeout=30)
    if resp.status_code != 201:
        raise RuntimeError(f"启动 {actor_id} 失败 status={resp.status_code} body={resp.text[:200]}")
    return resp.json().get("data", {})


def _wait_actor(run_id: str, apify_token: str, timeout: int = ACTOR_TIMEOUT_SECS) -> dict:
    headers = {"Authorization": f"Bearer {apify_token}"}
    api_url = f"https://api.apify.com/v2/actor-runs/{run_id}"
    start = time.time()
    while True:
        if time.time() - start > timeout:
            raise TimeoutError(f"actor run {run_id} 等待超时（{timeout}s）")
        resp = requests.get(api_url, headers=headers, timeout=15)
        if resp.status_code != 200:
            raise RuntimeError(f"查询 run {run_id} 失败 status={resp.status_code}")
        data = resp.json().get("data", {})
        status = data.get("status")
        if status in {"SUCCEEDED", "FAILED", "ABORTED", "TIMED-OUT"}:
            return data
        time.sleep(ACTOR_POLL_INTERVAL)


def _fetch_dataset(dataset_id: str, apify_token: str) -> list[dict]:
    headers = {"Authorization": f"Bearer {apify_token}"}
    api_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items"
    resp = requests.get(api_url, headers=headers, timeout=60)
    if resp.status_code != 200:
        raise RuntimeError(f"获取 dataset {dataset_id} 失败 status={resp.status_code}")
    items = resp.json()
    return items if isinstance(items, list) else []


def _cookie_values_from_actor_input(value) -> list[str]:
    values: list[str] = []
    if isinstance(value, list):
        for entry in value:
            values.extend(_cookie_values_from_actor_input(entry))
        return values
    if not isinstance(value, dict):
        return values
    for key, nested in value.items():
        key_lower = str(key).lower()
        if key_lower in {"customcookies", "cookies"} and isinstance(nested, list):
            for cookie in nested:
                if not isinstance(cookie, dict):
                    continue
                cookie_value = str(cookie.get("value") or "")
                if len(cookie_value) >= 4:
                    values.append(cookie_value)
        else:
            values.extend(_cookie_values_from_actor_input(nested))
    return values


def _redact_sensitive_text(text: str, actor_input: dict | None = None) -> str:
    redacted = str(text or "")
    for value in _cookie_values_from_actor_input(actor_input or {}):
        redacted = redacted.replace(value, "[REDACTED]")
    return redacted


def _redact_actor_input(value):
    if isinstance(value, list):
        return [_redact_actor_input(entry) for entry in value]
    if not isinstance(value, dict):
        return value

    redacted = {}
    for key, nested in value.items():
        key_lower = str(key).lower()
        if key_lower in {"customcookies", "cookies"} and isinstance(nested, list):
            redacted[key] = [
                {**cookie, "value": "[REDACTED]"} if isinstance(cookie, dict) else "[REDACTED]"
                for cookie in nested
            ]
        elif key_lower in {"cookie", "authorization", "access_token", "token", "api_key", "apikey", "password", "secret"}:
            redacted[key] = "[REDACTED]" if nested else nested
        else:
            redacted[key] = _redact_actor_input(nested)
    return redacted


def _call_actor(
    actor_id: str,
    candidate_inputs: list[dict],
    apify_token: str,
    accept_items: Callable[[list[dict]], bool] | None = None,
) -> list[dict]:
    """依次尝试多个 input schema 直到一个成功返回可接受的数据。"""
    return _call_actor_with_meta(actor_id, candidate_inputs, apify_token, accept_items)[0]


def _call_actor_with_meta(
    actor_id: str,
    candidate_inputs: list[dict],
    apify_token: str,
    accept_items: Callable[[list[dict]], bool] | None = None,
) -> tuple[list[dict], dict]:
    """Call an actor and preserve run/dataset metadata for diagnostics."""
    last_err = None
    last_items: list[dict] = []
    last_meta: dict = {}
    attempts: list[dict] = []
    for run_input in candidate_inputs:
        run_id = ""
        dataset_id = ""
        safe_input = _redact_actor_input(run_input)
        try:
            run = _start_actor(actor_id, run_input, apify_token)
            run_id = run.get("id")
            if not run_id:
                raise RuntimeError("actor 返回缺少 run id")
            final = _wait_actor(run_id, apify_token)
            if final.get("status") != "SUCCEEDED":
                raise RuntimeError(f"run 结束状态={final.get('status')}")
            dataset_id = final.get("defaultDatasetId")
            if not dataset_id:
                raise RuntimeError("run 缺少 defaultDatasetId")
            items = _fetch_dataset(dataset_id, apify_token)
            last_items = items
            last_meta = {
                "actor_id": actor_id,
                "run_id": run_id,
                "dataset_id": dataset_id,
                "status": final.get("status"),
                "started_at": final.get("startedAt") or run.get("startedAt"),
                "finished_at": final.get("finishedAt"),
                "input": safe_input,
            }
            if accept_items and not accept_items(items):
                last_err = f"run={run_id} dataset={dataset_id} 返回 {len(items)} 条但未通过校验"
                if not items:
                    last_err += "（dataset 为空）"
                last_meta = {**last_meta, "error": last_err}
                attempts.append({**last_meta, "item_count": len(items), "accepted": False})
                logger.warning(
                    f"⚠️ actor {actor_id} run={run_id} dataset={dataset_id} "
                    f"返回 {len(items)} 条但未通过校验，继续尝试下一个 input"
                )
                continue
            attempts.append({**last_meta, "item_count": len(items), "accepted": True})
            return items, {**last_meta, "attempts": attempts}
        except Exception as e:
            safe_error = _redact_sensitive_text(str(e), run_input)
            last_err = safe_error
            last_meta = {
                "actor_id": actor_id,
                "run_id": run_id,
                "dataset_id": dataset_id,
                "input": safe_input,
                "error": safe_error,
            }
            attempts.append({**last_meta, "item_count": 0, "accepted": False})
            logger.warning(
                f"⚠️ actor {actor_id} 调用失败（run={run_id or '-'}, input={list(run_input.keys())}）: {safe_error}"
            )
            continue
    if last_items or last_meta.get("dataset_id"):
        return last_items, {**last_meta, "attempts": attempts}
    raise RuntimeError(f"actor {actor_id} 所有候选 input 均失败: {last_err}")


def _dedupe_actor_inputs(candidate_inputs: list[dict]) -> list[dict]:
    """Keep candidate ordering stable while removing exact duplicate payloads."""
    seen: set[str] = set()
    result: list[dict] = []
    for item in candidate_inputs:
        key = json.dumps(item, sort_keys=True, ensure_ascii=False)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _dedupe_urls(urls: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for url in urls:
        cleaned = str(url or "").strip()
        if not cleaned or cleaned in seen:
            continue
        seen.add(cleaned)
        result.append(cleaned)
    return result


def _load_fb_comment_fallback_cookies() -> list[dict]:
    raw = os.environ.get("FB_COMMENT_FALLBACK_COOKIES_JSON") or ""
    raw = raw.strip()
    if not raw:
        return []
    try:
        cookies = json.loads(raw)
    except Exception as e:
        logger.warning(f"⚠️ FB cookie fallback 配置不是合法 JSON: {e}")
        return []
    if not isinstance(cookies, list):
        logger.warning("⚠️ FB cookie fallback 配置必须是 cookie 数组")
        return []
    cleaned: list[dict] = []
    for cookie in cookies:
        if not isinstance(cookie, dict):
            continue
        name = str(cookie.get("name") or "").strip()
        value = str(cookie.get("value") or "").strip()
        if not name or not value:
            continue
        normalized = dict(cookie)
        normalized["name"] = name
        normalized["value"] = value
        cleaned.append(normalized)
    return cleaned


def _merge_actor_meta(primary: dict, fallback: dict | None = None) -> dict:
    result = dict(primary or {})
    attempts: list[dict] = []
    for meta in (primary or {}, fallback or {}):
        if not isinstance(meta, dict):
            continue
        meta_attempts = meta.get("attempts")
        if isinstance(meta_attempts, list):
            attempts.extend([attempt for attempt in meta_attempts if isinstance(attempt, dict)])
        elif meta:
            attempts.append(meta)
    if fallback:
        result["fallback"] = fallback
        for key in ("actor_id", "run_id", "dataset_id", "status", "started_at", "finished_at", "input"):
            if fallback.get(key):
                result[key] = fallback.get(key)
        result["used_fallback"] = True
    if attempts:
        result["attempts"] = attempts
    return result


def _facebook_url_needs_resolution(url: str) -> bool:
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    host = (parsed.netloc or "").lower()
    path = (parsed.path or "").lower()
    return "fb.watch" in host or ("facebook.com" in host and path.startswith("/share/"))


def _clean_resolved_facebook_url(url: str) -> str:
    try:
        parsed = urlparse(url)
    except Exception:
        return url
    path = (parsed.path or "").rstrip("/") or parsed.path
    lower_path = path.lower()
    if any(part in lower_path for part in ("/posts/", "/videos/", "/reel/", "/watch/")) or "pfbid" in lower_path:
        return urlunparse((parsed.scheme, parsed.netloc, path, "", "", ""))
    if lower_path.endswith(("/story.php", "/permalink.php")):
        keep_keys = {"story_fbid", "id", "fbid", "v"}
        query = urlencode(
            [(k, v) for k, v in parse_qsl(parsed.query, keep_blank_values=True) if k in keep_keys]
        )
        return urlunparse((parsed.scheme, parsed.netloc, path, "", query, ""))
    return url


def _facebook_redirect_location(current_url: str, location: str) -> str:
    if not location:
        return ""
    next_url = urljoin(current_url, location.strip())
    try:
        parsed = urlparse(next_url)
    except Exception:
        return next_url
    for key, value in parse_qsl(parsed.query, keep_blank_values=True):
        if key.lower() in {"u", "url", "next", "target", "href"} and value.startswith(("http://", "https://")):
            return value
    return next_url


def _resolve_facebook_url(url: str) -> str:
    """Expand FB share/fb.watch URLs before sending them to Apify actors."""
    if not _facebook_url_needs_resolution(url):
        return url
    headers = {"Accept": "*/*"}

    def _follow_once(method: str) -> str:
        current = url
        seen = {url}
        for _ in range(4):
            resp = None
            try:
                if method == "HEAD":
                    resp = requests.head(current, headers=headers, allow_redirects=False, timeout=15)
                else:
                    resp = requests.get(current, headers=headers, allow_redirects=False, timeout=15, stream=True)
                redirect_url = _facebook_redirect_location(current, resp.headers.get("location") or "")
            except requests.RequestException as e:
                logger.warning(f"⚠️ FB share URL {method} 展开失败: {e}")
                return ""
            finally:
                if resp is not None and method != "HEAD":
                    resp.close()
            if not redirect_url:
                return ""
            resolved = _clean_resolved_facebook_url(redirect_url)
            if resolved in seen:
                return resolved
            seen.add(resolved)
            if not _facebook_url_needs_resolution(resolved):
                return resolved
            current = resolved
        return _clean_resolved_facebook_url(current)

    for method in ("HEAD", "GET"):
        resolved = _follow_once(method)
        if resolved and resolved != url:
            logger.info(f"🔁 FB share URL 已展开: {url} -> {resolved}")
            return resolved
    return url


# ============================================
# 字段提取（通用 + 平台特化）
# ============================================


def _first_str(item: dict, keys: Iterable[str]) -> str:
    for k in keys:
        v = _deep_get(item, k)
        if isinstance(v, str) and v.strip():
            return v.strip()
        if isinstance(v, dict):
            for sub in ("name", "username", "userName", "screen_name", "screenName", "text", "title"):
                vv = v.get(sub)
                if isinstance(vv, str) and vv.strip():
                    return vv.strip()
    return ""


def _deep_get(item: dict, path: str):
    current = item
    for part in str(path).split("."):
        if not isinstance(current, dict):
            return None
        current = current.get(part)
    return current


def _first_value_as_str(item: dict, keys: Iterable[str]) -> str:
    for k in keys:
        v = _deep_get(item, k)
        if v is None or isinstance(v, bool):
            continue
        if isinstance(v, dict):
            for sub in ("id", "commentId", "pk", "cid", "url", "link", "permalink"):
                vv = v.get(sub)
                if vv is not None and not isinstance(vv, bool) and str(vv).strip():
                    return str(vv).strip()
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


_COUNT_RE = re.compile(r"(-?\d+(?:\.\d+)?)\s*([kKmMbBwW万]?)")


def _safe_count(value, default: int = 0) -> int:
    """Parse social count fields, including compact values like 1.2K / 3万."""
    if value is None or isinstance(value, bool):
        return default
    try:
        if isinstance(value, (int, float)):
            return max(0, int(value))
        if isinstance(value, dict):
            for sub in ("count", "total", "value", "likes", "likeCount", "likesCount"):
                if sub in value:
                    parsed = _safe_count(value.get(sub), None)
                    if parsed is not None:
                        return parsed
            return default
        s = str(value).strip().replace(",", "")
        if not s:
            return default
        m = _COUNT_RE.search(s)
        if not m:
            cleaned = re.sub(r"[^\d-]", "", s)
            return max(0, int(cleaned)) if cleaned else default
        number = float(m.group(1))
        suffix = (m.group(2) or "").lower()
        multiplier = {
            "k": 1_000,
            "m": 1_000_000,
            "b": 1_000_000_000,
            "w": 10_000,
            "万": 10_000,
        }.get(suffix, 1)
        return max(0, int(number * multiplier))
    except Exception:
        return default


def _first_count(item: dict, keys: Iterable[str]) -> int:
    for k in keys:
        v = _deep_get(item, k)
        parsed = _safe_count(v, None)
        if parsed is not None:
            return parsed
    return 0


def _is_http_url(value: str) -> bool:
    return isinstance(value, str) and value.strip().startswith(("http://", "https://"))


def _first_url(item: dict, keys: Iterable[str]) -> str:
    for k in keys:
        v = _deep_get(item, k)
        if isinstance(v, dict):
            for sub in ("url", "link", "permalink", "webVideoUrl", "postUrl"):
                vv = v.get(sub)
                if _is_http_url(vv):
                    return vv.strip()
            continue
        if _is_http_url(v):
            return v.strip()
    return ""


_COMMENT_MEDIA_HINTS = (
    "attachment",
    "attachments",
    "image",
    "photo",
    "picture",
    "media",
    "thumbnail",
)
_COMMENT_PROFILE_MEDIA_HINTS = ("avatar", "profile", "author", "owner", "user")


def _extract_comment_media_url(item: dict) -> str:
    """Best-effort image/attachment URL extraction for media-only comments."""
    explicit = _first_url(
        item,
        [
            "imageUrl",
            "imageURL",
            "image.url",
            "image.uri",
            "photoUrl",
            "photoURL",
            "photo.url",
            "picture",
            "picture.url",
            "mediaUrl",
            "mediaURL",
            "media.url",
            "attachmentUrl",
            "attachmentURL",
            "attachment.url",
            "attachment.media.image.uri",
            "attachment.media.image.url",
            "attachments.media.image.uri",
            "attachments.media.image.url",
            "thumbnailUrl",
            "thumbnailURL",
        ],
    )
    if explicit:
        return explicit

    def _walk(value, in_media_context: bool = False) -> str:
        if isinstance(value, list):
            for entry in value:
                found = _walk(entry, in_media_context)
                if found:
                    return found
            return ""
        if not isinstance(value, dict):
            return ""

        for key, nested in value.items():
            key_lower = str(key).lower()
            has_media_hint = any(hint in key_lower for hint in _COMMENT_MEDIA_HINTS)
            is_profile_media = any(hint in key_lower for hint in _COMMENT_PROFILE_MEDIA_HINTS)
            next_media_context = in_media_context or (has_media_hint and not is_profile_media)

            if _is_http_url(nested) and next_media_context:
                return nested.strip()
            found = _walk(nested, next_media_context)
            if found:
                return found
        return ""

    return _walk(item)


def _is_probable_post_url(url: str, platform: str) -> bool:
    if not _is_http_url(url):
        return False
    u = url.lower()
    if platform == "TT":
        return "tiktok.com" in u and "/video/" in u
    if platform == "YTB":
        return "youtu.be/" in u or ("youtube.com" in u and ("/watch" in u or "/shorts/" in u))
    if platform == "IG":
        return "instagram.com" in u and any(p in u for p in ("/p/", "/reel/", "/tv/"))
    if platform == "FB":
        return any(h in u for h in ("facebook.com", "fb.watch")) and any(
            p in u for p in ("/posts/", "/videos/", "/reel/", "/watch", "story.php", "permalink.php")
        )
    if platform == "X":
        return any(h in u for h in ("twitter.com", "x.com")) and any(p in u for p in ("/status/", "/statuses/"))
    return True


def _extract_item_post_url(item: dict, platform: str, fallback_url: str) -> str:
    direct = _first_url(
        item,
        [
            "source.url",
            "postUrl",
            "postURL",
            "post_url",
            "postLink",
            "post_link",
            "postPermalink",
            "post.permalink",
            "post.url",
            "post.link",
            "facebookUrl",
            "webVideoUrl",
            "videoWebUrl",
            "video.webVideoUrl",
            "video.url",
            "tweetUrl",
            "tweet.url",
            "inputUrl",
            "sourceUrl",
        ],
    )
    if direct:
        return direct
    for key in ("url", "link", "permalink", "canonicalUrl"):
        candidate = _first_url(item, [key])
        if candidate and _is_probable_post_url(candidate, platform):
            return candidate
    return fallback_url


def _extract_comment_url(item: dict) -> str:
    return _first_url(
        item,
        [
            "comment.url",
            "commentUrl",
            "commentURL",
            "comment_url",
            "commentLink",
            "comment_link",
            "commentPermalink",
            "comment.permalink",
        ],
    )


def _extract_comment_like_count(item: dict) -> int:
    return _first_count(
        item,
        [
            "comment.like_count",
            "comment.total_reactions",
            "commentLikeCount",
            "commentLikesCount",
            "comment_likes_count",
            "commentLikes",
            "likesCount",
            "likeCount",
            "likes_count",
            "likes",
            "diggCount",
            "upvoteCount",
            "voteCount",
            "votes",
            "reactionCount",
            "reactionsCount",
            "feedback.reaction_count.count",
            "feedback.reactionCount",
        ],
    )


def _has_comment_identity(item: dict) -> bool:
    if not isinstance(item, dict):
        return False
    if _first_value_as_str(
        item,
        [
            "commentId",
            "comment_id",
            "commentPk",
            "replyId",
            "comment.id",
            "comment.parent_id",
            "comment.commentId",
            "commentUrl",
            "commentURL",
            "comment_url",
            "commentLink",
            "comment_link",
        ],
    ):
        return True
    return bool(
        _first_str(
            item,
            [
                "profileName",
                "authorName",
                "author.name",
                "from.name",
                "user.name",
            ],
        )
    )


def _stable_comment_id(item: dict, platform: str, post_url: str, author: str, text: str, created_str: str) -> str:
    raw_id = _first_value_as_str(
        item,
        [
            "commentId",
            "comment_id",
            "comment.id",
            "id",
            "cid",
            "pk",
            "commentPk",
            "replyId",
            "uid",
        ],
    )
    if raw_id:
        return f"{platform}:{raw_id}"
    fingerprint = f"{platform}|{post_url}|{author}|{created_str}|{text}"
    digest = hashlib.sha1(fingerprint.encode("utf-8", errors="ignore")).hexdigest()[:16]
    return f"{platform}:hash:{digest}"


def _is_x_original_tweet(item: dict) -> bool:
    """判断 X 平台一条记录是否是原帖（主推文），用于在评论列表里剔除。"""
    if not isinstance(item, dict):
        return False
    tid = item.get("id") or item.get("tweetId") or item.get("rest_id")
    conv = item.get("conversationId") or item.get("conversation_id") or item.get("conversationIdStr")
    if tid and conv and str(tid) == str(conv):
        return True
    if item.get("isReply") is False and item.get("inReplyToId") is None:
        # 严格条件：明确不是回复
        return True
    return False


def _to_beijing_dt(value) -> datetime.datetime | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, (int, float)):
            ts = float(value)
            # 秒/毫秒兼容
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).astimezone(BEIJING_TZ)
        s = str(value).strip()
        if not s:
            return None
        if s.isdigit():
            ts = float(s)
            if ts > 1e12:
                ts = ts / 1000.0
            return datetime.datetime.fromtimestamp(ts, tz=datetime.timezone.utc).astimezone(BEIJING_TZ)
        # ISO 8601
        dt = datetime.datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=datetime.timezone.utc)
        return dt.astimezone(BEIJING_TZ)
    except Exception:
        return None


def _fmt_dt(dt: datetime.datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y-%m-%d %H:%M")


def _bucket_for(dt: datetime.datetime | None) -> str:
    """评论时间归属时段。"""
    if not dt:
        return ""
    h = dt.hour
    if 0 <= h < 6:
        return "凌晨"
    if 6 <= h < 12:
        return "上午"
    if 12 <= h < 14:
        return "中午"
    if 14 <= h < 18:
        return "下午"
    return "晚上"


def _extract_post_meta(items: list[dict], platform: str, fallback_url: str) -> dict:
    """从 dataset items 里尽量抽出主贴标题/发布时间。"""
    title = ""
    post_dt = None

    title_keys = [
        "postTitle",
        "postText",
        "postCaption",
        "caption",
        "videoCaption",
        "videoDescription",
        "videoTitle",
        "title",
        "description",
        "postContent",
    ]
    date_keys = [
        "postPublishedAt",
        "postDate",
        "postPublishedTime",
        "videoCreateTime",
        "videoCreateTimeISO",
        "publishedTime",
        "publishedAt",
        "publishedTimeText",
        "createTimeOfPost",
    ]

    # X 平台：标题取「原推文 text」、发布时间取「原推文 createdAt」
    iter_items = items
    if platform == "X":
        originals = [it for it in items if isinstance(it, dict) and _is_x_original_tweet(it)]
        if originals:
            iter_items = originals + [it for it in items if it not in originals]
        title_keys = ["text", "fullText", "full_text"] + title_keys
        date_keys = ["createdAt", "created_at", "createdAtTimestamp"] + date_keys

    for it in iter_items:
        if not isinstance(it, dict):
            continue
        if not title:
            title = _first_str(it, title_keys)
        if not post_dt:
            for k in date_keys:
                cand = _to_beijing_dt(it.get(k))
                if cand:
                    post_dt = cand
                    break
        if title and post_dt:
            break
    # 折行去多余空白，限长
    title = re.sub(r"\s+", " ", title or "").strip()
    if len(title) > 180:
        title = title[:180] + "…"
    return {
        "post_title": title,
        "post_date": _fmt_dt(post_dt),
        "post_url": fallback_url,
    }


def _extract_comment(item: dict, platform: str, post_url: str = "") -> dict | None:
    """统一抽出评论字段。返回 None 表示这条不是评论（或要跳过）。"""
    if not isinstance(item, dict):
        return None

    # X 平台：原推文当作"主贴"，不进评论列表
    if platform == "X" and _is_x_original_tweet(item):
        return None

    text = _first_str(
        item,
        [
            "comment.text",
            "text",
            "fullText",
            "full_text",
            "content",
            "comment",
            "message",
            "commentText",
        ],
    )
    media_url = _extract_comment_media_url(item)
    if not text:
        if not media_url or not _has_comment_identity(item):
            return None
        text = f"（图片评论）{media_url}"
    elif media_url and media_url not in text:
        text = f"{text}\n（图片）{media_url}"

    author = _first_str(
        item,
        [
            "comment.author.name",
            "authorName",
            "ownerUsername",
            "username",
            "userName",
            "uniqueId",
            "name",
            "author",
            "profileName",
        ],
    )

    # 评论时间
    created_dt = None
    if platform == "FB":
        # FB 评论几乎没时间戳；尝试多个键，没有就 None
        pass
    for k in (
        "commentDate",
        "commentTime",
        "comment.created_at",
        "comment.created_time_unix",
        "createdAtTimestamp",
        "createdAtTimestampSeconds",
        "createTime",
        "createTimeISO",
        "createdAt",
        "created_at",
        "createdTime",
        "timestamp",
        "publishedAt",
        "publishedTimeText",
        "time",
        "date",
    ):
        cand = _to_beijing_dt(_deep_get(item, k))
        if cand:
            created_dt = cand
            break

    return {
        "comment_id": _stable_comment_id(item, platform, post_url, author or "", text, _fmt_dt(created_dt)),
        "comment_url": _extract_comment_url(item),
        "like_count": _extract_comment_like_count(item),
        "text": text,
        "author": author or "",
        "created_dt": created_dt,
        "created_str": _fmt_dt(created_dt),
        "bucket": _bucket_for(created_dt),
    }


_COMMENT_CHILD_KEYS = (
    "comments",
    "commentList",
    "comment_list",
    "replies",
    "replyComments",
    "commentReplies",
    "children",
    "edges",
    "nodes",
)

_COMMENT_CONTEXT_KEYS = (
    "facebookUrl",
    "postUrl",
    "postURL",
    "post_url",
    "postLink",
    "postPermalink",
    "inputUrl",
    "sourceUrl",
    "postTitle",
    "postText",
    "postCaption",
    "caption",
    "facebookId",
    "pageAdLibrary",
)

_RAW_ERROR_KEYS = (
    "error",
    "errorMessage",
    "error_message",
    "failedReason",
    "failReason",
    "reason",
    "message",
    "statusMessage",
    "debugMessage",
    "requestErrorMessages",
    "errors",
)


def _brief_raw_value(value, limit: int = 180) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, str):
        text = value
    else:
        try:
            text = json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            text = str(value)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit] + ("..." if len(text) > limit else "")


def _extract_raw_item_error(item: dict) -> str:
    if not isinstance(item, dict):
        return ""
    for key in _RAW_ERROR_KEYS:
        value = _deep_get(item, key)
        brief = _brief_raw_value(value)
        if brief:
            return brief

    for key, value in item.items():
        key_lower = str(key).lower()
        if any(marker in key_lower for marker in ("error", "fail", "reason")):
            brief = _brief_raw_value(value)
            if brief:
                return brief
    return ""


def _summarize_non_comment_items(items: list[dict], platform: str, post_url: str = "") -> str:
    if not isinstance(items, list) or not items:
        return ""
    if _flatten_comment_items(items, platform, post_url):
        return ""

    details: list[str] = []
    for idx, item in enumerate(items[:3], 1):
        if not isinstance(item, dict):
            continue
        parts: list[str] = []
        item_url = _first_url(item, ["inputUrl", "sourceUrl", "url", "facebookUrl", "postUrl", "post.url"])
        if item_url:
            short_url = item_url[:90] + ("..." if len(item_url) > 90 else "")
            parts.append(f"第{idx}条URL={short_url}")
        raw_error = _extract_raw_item_error(item)
        if raw_error:
            parts.append(f"错误={raw_error}")
        else:
            keys = ", ".join(list(item.keys())[:12])
            if keys:
                parts.append(f"字段={keys}")
        if parts:
            details.append("，".join(parts))

    if not details:
        return "actor 原始返回未包含可识别评论字段"
    return "actor 原始返回未包含可识别评论字段；" + "；".join(details)


def _with_comment_parent_context(parent: dict, child: dict) -> dict:
    merged = dict(child)
    for key in _COMMENT_CONTEXT_KEYS:
        if key not in merged and parent.get(key) not in (None, ""):
            merged[key] = parent.get(key)
    return merged


def _nested_comment_candidates(parent: dict) -> list[dict]:
    nested: list[dict] = []

    def _collect(value, context: dict):
        if isinstance(value, list):
            for entry in value:
                _collect(entry, context)
            return
        if not isinstance(value, dict):
            return

        node = value.get("node")
        if isinstance(node, dict):
            value = _with_comment_parent_context(value, node)

        nested.append(_with_comment_parent_context(context, value))
        for key in _COMMENT_CHILD_KEYS:
            child_value = value.get(key)
            if child_value:
                _collect(child_value, value)

    for key in _COMMENT_CHILD_KEYS:
        value = parent.get(key)
        if value:
            _collect(value, parent)
    return nested


def _flatten_comment_items(items: list[dict], platform: str, post_url: str = "") -> list[tuple[int, dict]]:
    """Return top-level comment rows plus nested replies/comments from post-shell rows."""
    flattened: list[tuple[int, dict]] = []
    seen: set[str] = set()

    def _add(raw_index: int, item: dict):
        row_post_url = _extract_item_post_url(item, platform, post_url)
        comment = _extract_comment(item, platform, row_post_url)
        if not comment:
            return
        key = comment.get("comment_id") or json.dumps(item, sort_keys=True, ensure_ascii=False, default=str)
        if key in seen:
            return
        seen.add(key)
        flattened.append((raw_index, item))

    for raw_index, item in enumerate(items or []):
        if not isinstance(item, dict):
            continue
        _add(raw_index, item)
        for child in _nested_comment_candidates(item):
            _add(raw_index, child)
    return flattened


# ============================================
# 平台抓取入口
# ============================================


def _items_have_comments(items: list[dict], platform: str, post_url: str = "") -> bool:
    return bool(_flatten_comment_items(items or [], platform, post_url))


def _facebook_cookie_fallback_input(url: str, limit: int, cookies: list[dict]) -> dict:
    fallback_limit = 0 if limit >= UNLIMITED_COMMENTS_PER_POST_LIMIT else limit
    return {
        "urls": [{"url": url}],
        "maxCommentsPerUrl": fallback_limit,
        "fetchReplies": True,
        "commentsIntentToken": FB_COOKIE_FALLBACK_SORT,
        "customCookies": cookies,
    }


def _facebook_cookie_fallback_inputs(urls: Iterable[str], limit: int, cookies: list[dict]) -> list[dict]:
    fallback_limit = 0 if limit >= UNLIMITED_COMMENTS_PER_POST_LIMIT else limit
    inputs: list[dict] = []
    for url in _dedupe_urls(urls):
        base = {
            "maxCommentsPerUrl": fallback_limit,
            "fetchReplies": True,
            "commentsIntentToken": FB_COOKIE_FALLBACK_SORT,
            "customCookies": cookies,
        }
        inputs.append({"urls": [{"url": url}], **base})
        inputs.append({"urls": [url], **base})
    return _dedupe_actor_inputs(inputs)


def _scrape_facebook_with_cookie_fallback(
    original_url: str,
    url: str,
    apify_token: str,
    limit: int,
    primary_items: list[dict],
    primary_meta: dict,
    primary_error: str,
) -> tuple[list[dict], dict, str]:
    cookies = _load_fb_comment_fallback_cookies()
    if not cookies:
        return primary_items, primary_meta, primary_error

    fallback_actor = FB_COOKIE_FALLBACK_ACTOR_ID
    fallback_inputs = _facebook_cookie_fallback_inputs([original_url, url], limit, cookies)
    fallback_input = fallback_inputs[0] if fallback_inputs else {}
    try:
        fallback_items, fallback_meta = _call_actor_with_meta(
            fallback_actor,
            fallback_inputs,
            apify_token,
            accept_items=lambda rows: _items_have_comments(rows, "FB", url),
        )
    except Exception as e:
        safe_error = _redact_sensitive_text(str(e), fallback_input)
        fallback_meta = {"actor_id": fallback_actor, "input": _redact_actor_input(fallback_input), "error": safe_error}
        merged_meta = _merge_actor_meta(primary_meta, fallback_meta)
        detail = f"{primary_error}; cookie fallback 失败: {safe_error}" if primary_error else f"cookie fallback 失败: {safe_error}"
        return primary_items, merged_meta, detail[:300]

    fallback_error = _summarize_non_comment_items(fallback_items, "FB", url)
    merged_meta = _merge_actor_meta(primary_meta, fallback_meta)
    if _items_have_comments(fallback_items, "FB", url):
        logger.info(f"✅ FB cookie fallback 成功: actor={fallback_actor}, rows={len(fallback_items)}")
        return fallback_items, merged_meta, ""

    fallback_detail = fallback_error or fallback_meta.get("error") or f"返回 {len(fallback_items or [])} 条/0 条评论"
    detail_parts = [part for part in (primary_error, f"cookie fallback: {fallback_detail}" if fallback_detail else "") if part]
    return primary_items, merged_meta, "; ".join(detail_parts)[:300]


def _scrape_facebook(url: str, apify_token: str, comments_per_post_limit: int | None = None) -> dict:
    actor = DEFAULT_ACTORS["FB"]
    limit = normalize_comments_per_post_limit(comments_per_post_limit)
    scrape_url = _resolve_facebook_url(url)
    candidates = [
        {
            "startUrls": [{"url": candidate_url}],
            "resultsLimit": limit,
            "includeNestedComments": True,
            "viewOption": "RANKED_UNFILTERED",
        }
        for candidate_url in _dedupe_urls([url, scrape_url])
    ]
    try:
        items, meta = _call_actor_with_meta(
            actor,
            candidates,
            apify_token,
            accept_items=lambda rows: _items_have_comments(rows, "FB", scrape_url),
        )
        error = _summarize_non_comment_items(items, "FB", scrape_url)
    except Exception as e:
        items = []
        meta = {"actor_id": actor, "error": _redact_sensitive_text(str(e))}
        error = f"primary actor 失败: {meta['error']}"
    if not _items_have_comments(items, "FB", scrape_url):
        items, meta, error = _scrape_facebook_with_cookie_fallback(
            url,
            scrape_url,
            apify_token,
            limit,
            items,
            meta,
            error,
        )
    return {
        "items": items,
        "platform": "FB",
        "url": url,
        "resolved_url": scrape_url,
        "actor_meta": meta,
        "error": error,
    }


def _scrape_instagram(url: str, apify_token: str, comments_per_post_limit: int | None = None) -> dict:
    actor = DEFAULT_ACTORS["IG"]
    limit = normalize_comments_per_post_limit(comments_per_post_limit)
    candidates = [
        {"directUrls": [url], "resultsLimit": limit},
        {"postUrls": [url], "resultsLimit": limit},
        {"startUrls": [{"url": url}], "resultsLimit": limit},
    ]
    items = _call_actor(actor, candidates, apify_token)
    return {"items": items, "platform": "IG", "url": url}


def _scrape_tiktok(url: str, apify_token: str, comments_per_post_limit: int | None = None) -> dict:
    actor = DEFAULT_ACTORS["TT"]
    limit = normalize_comments_per_post_limit(comments_per_post_limit)
    candidates = [
        {"postURLs": [url], "commentsPerPost": limit, "maxRepliesPerComment": 0},
        {"postUrls": [url], "commentsPerPost": limit},
        {"startUrls": [{"url": url}], "maxItems": limit},
    ]
    items = _call_actor(actor, candidates, apify_token)
    return {"items": items, "platform": "TT", "url": url}


def _scrape_youtube(url: str, apify_token: str, comments_per_post_limit: int | None = None) -> dict:
    actor = DEFAULT_ACTORS["YTB"]
    limit = normalize_comments_per_post_limit(comments_per_post_limit)
    candidates = [
        {"startUrls": [{"url": url}], "maxComments": limit, "includeReplies": False},
        {"videoUrls": [url], "maxComments": limit},
        {"startUrls": [{"url": url}], "maxResults": limit},
    ]
    items = _call_actor(actor, candidates, apify_token)
    return {"items": items, "platform": "YTB", "url": url}


def _scrape_x(url: str, apify_token: str, comments_per_post_limit: int | None = None) -> dict:
    """X（Twitter）回复抓取。

    默认 actor 是 apidojo/tweet-scraper，它通过 startUrls 直接抓主贴 + 回复。
    """
    actor = DEFAULT_ACTORS["X"]
    limit = normalize_comments_per_post_limit(comments_per_post_limit)
    candidates = [
        {
            "startUrls": [url],
            "maxItems": limit,
            "onlyImage": False,
            "onlyVerifiedUsers": False,
        },
        {"startUrls": [{"url": url}], "maxItems": limit},
        {"tweetUrls": [url], "maxItems": limit},
        {"conversationIds": [url], "maxItems": limit},
        {"searchTerms": [url], "maxItems": limit},
    ]
    items = _call_actor(actor, candidates, apify_token)
    return {"items": items, "platform": "X", "url": url}


PLATFORM_SCRAPERS: dict[str, Callable[..., dict]] = {
    "FB": _scrape_facebook,
    "IG": _scrape_instagram,
    "TT": _scrape_tiktok,
    "YTB": _scrape_youtube,
    "X": _scrape_x,
}


# ============================================
# AI 分析（翻译 + 情感 + 分类）
# ============================================


def _safe_json_array(text: str) -> list:
    if not text:
        return []
    # 去掉可能的 ```json fenced block
    cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip(), flags=re.M | re.I)
    # 取第一个 [ 到最后一个 ] 的子串，保险
    lb = cleaned.find("[")
    rb = cleaned.rfind("]")
    if lb != -1 and rb != -1 and rb > lb:
        cleaned = cleaned[lb : rb + 1]
    try:
        data = json.loads(cleaned)
        return data if isinstance(data, list) else []
    except Exception as e:
        logger.warning(f"⚠️ AI 输出 JSON 解析失败: {e} | head={text[:200]}")
        return []


_CJK_RE = re.compile(r"[\u4e00-\u9fff]")


def _post_check_translation(translation: str, original: str) -> str:
    """检测 AI 摆烂：当原文非中文，但译文与原文几乎一致（AI 没翻就照抄），
    返回空字符串以暴露问题，避免误导。
    """
    tr = (translation or "").strip()
    orig = (original or "").strip()
    if not tr:
        return ""
    # 原文已有中文 → 直接返回
    if _CJK_RE.search(orig):
        return tr
    # 译文没有任何中文字符 → 视为没翻
    if not _CJK_RE.search(tr):
        # 但如果原文+译文完全一致（比如 "Pubgggg" 这种灌水），允许
        if tr == orig:
            return tr
        # 否则记一次警告并清空，让导出表里那一格留空，比错位/原样照抄要好
        logger.warning(
            f"⚠️ 译文疑似未翻译（无中文字符）：原文={orig[:60]!r} 译文={tr[:60]!r}"
        )
        return ""
    return tr


def _normalize_ai_label(value: str, allowed: set[str], fallback: str) -> str:
    if not isinstance(value, str):
        return fallback
    v = value.strip()
    if v in allowed:
        return v
    # 兼容旧标签
    mapping = {
        "正面": "正向",
        "积极": "正向",
        "中性": "中立",
        "中间": "中立",
        "消极": "负面",
        "负向": "负面",
    }
    if v in mapping and mapping[v] in allowed:
        return mapping[v]
    return fallback


def _build_ai_prompt(batch_lines: str, expected_count: int) -> str:
    return (
        "你是中国头部手游社区的资深玩家与舆情分析师，同时负责把海外社媒评论翻成"
        "「中国玩家圈口语化中文」。请逐条处理下面的评论，并输出 JSON 数组（不要 markdown）。\n\n"
        f"⚠ 极重要：输入正好 {expected_count} 条评论（编号 0 ~ {expected_count - 1}），\n"
        f"   你必须严格输出 {expected_count} 个对象，按输入顺序排列，一条不能漏、一条不能加，\n"
        "   idx 必须等于输入编号，id 必须原样返回，且 translation_zh 必须是中文"
        "（非中文原文绝对不要原样照抄到译文）。\n\n"
        "每个对象字段：\n"
        "  - idx: 整数，等于输入编号\n"
        "  - id: 字符串，必须原样返回输入评论的 id\n"
        "  - translation_zh: 中文译文，按下面《翻译规范》执行；若原文已是中文，直接照搬，不要再润色\n"
        "  - sentiment: 严格三选一：正向 / 中立 / 负面\n"
        "  - category: 严格七选一：产品体验 / 功能建议 / 账号&充值 / 外挂作弊 / 活动运营 / 客服投诉 / 其他\n\n"
        "《输入格式》\n"
        "每行都是 JSON，字段包含 idx/id/likes/text。你只处理 text 字段；idx、id、likes 是定位元数据，"
        "不得混入 translation_zh。\n\n"
        "《翻译规范（重要，全部遵守）》\n"
        "1. 目标语感：像中国玩家在贴吧/B站/微博/TapTap 上聊天，自然口语、可短可碎，不要书面腔。\n"
        "2. 严禁逐字直译。要把英语/阿拉伯语/西语等的语序、虚词重组成符合中文习惯的表达。\n"
        "3. 专有名词保留英文：PUBG / PUBG Mobile / UC / Royale Pass / Erangle / Miramar / Rondo 等不要翻成中文俗称。\n"
        "4. 玩家黑话本土化：bruh→老哥/兄弟、pls/please→求/拜托、reply me→求回复/官方回个话、"
        "thx→谢谢/感谢、compensation→补偿/赔偿、skin→皮肤、event→活动、bug→BUG/卡顿、laggy→卡爆/掉帧。\n"
        "5. 抱怨/讽刺/反话要传达情绪，不要翻成中性。例：'nice, now give us 1000UC as compensation' "
        "→ '行行行，那赔1000UC吧'；'where is the update bruh😭' → '更新呢老哥😭'。\n"
        "6. 请求类要保留请求感：'please add X' → '求加X'；'we want X back' → '把X弄回来吧'。\n"
        "7. emoji / 表情贴纸照抄原位，不要描述。\n"
        "8. 评论里出现的话题标签（如 #cancel_card）需译成中文话题：'#取消该卡牌'。\n"
        "9. 极短或灌水内容（'Hi'、'Reply'、'Pubgggg'、纯 emoji）：分类一律「其他」，情感按字面取中立"
        "（除非含明显情绪词）。\n"
        "10. 文本里不要出现「我作为AI…」「以下是翻译…」之类说明语，只给纯净的中文译文。\n\n"
        "《翻译示例（务必按此风格）》\n"
        "  原: Give me the glacier and we Are all good Bro\n"
        "  译: 把冰川效果还给我们就行，老哥\n"
        "  原: please pubg mobile add tha flag\n"
        "  译: 求 PUBG Mobile 加一下那个国旗\n"
        "  原: nice,now give us 1000uc as compensation\n"
        "  译: 行行行，那赔我们 1000UC 吧\n"
        "  原: where is the update bruh 😭😭\n"
        "  译: 更新呢老哥 😭😭\n"
        "  原: low device\n"
        "  译: 我这设备配置低\n"
        "  原: i want 50 material and i will apologise\n"
        "  译: 给我 50 个材料我就当没事\n"
        "  原: PUBG Mobile please make a good one emulator for pc player 🙏🏻\n"
        "  译: 求 PUBG Mobile 给 PC 玩家做一个好用的模拟器 🙏🏻\n"
        "  原: Thanks for listening to us !\n"
        "  译: 谢谢你们能听玩家的意见！\n"
        "  原: Remove the promotion thing\n"
        "  译: 把那个推广的东西撤了吧\n\n"
        "《分类口径》\n"
        "  - 产品体验：画质、卡顿、闪退、操作手感、平衡性、设备兼容\n"
        "  - 功能建议：希望新增/调整模式、地图、英雄、玩法\n"
        "  - 账号&充值：登录、充值、退款、UC、皮肤、礼包、商城\n"
        "  - 外挂作弊：外挂、脚本、代练、开挂、举报无果\n"
        "  - 活动运营：联动、节日、福利、赛事、宣传、推广物料\n"
        "  - 客服投诉：客服响应、官方处理态度、申诉、邮件无回复\n"
        "  - 其他：无法归类、纯灌水、纯表情、宗教/文化敏感诉求等\n\n"
        f"《待处理评论》\n{batch_lines}\n\n"
        '只输出 JSON 数组，例如：[{"idx":0,"id":"C0","translation_zh":"...","sentiment":"正向","category":"产品体验"}]\n'
    )


def _run_ai_for_comments(
    comments: list[dict],
    ai_call: Callable[[str, int], tuple[str, int]],
    progress: Callable[[str], None] | None = None,
) -> tuple[list[dict], int]:
    """对评论调用 AI；返回 (按原顺序对齐的 AI 结果列表, 累计 token 数)。

    会先做预处理：剥离 [Sticker] 标记 + 识别灌水/纯表情类评论直接给占位翻译，
    剩余的才发给 AI，能显著降低机翻噪声 + 节省 token。
    """
    n = len(comments)
    results: list[dict] = [None] * n  # type: ignore
    total_tokens = 0
    if not comments:
        return [], 0

    # 预处理：为每条评论计算 (净化文本, 占位翻译, 是否需要 AI)
    preprocessed: list[tuple[str, str, bool]] = []
    for c in comments:
        cleaned, placeholder = _preprocess_comment(c.get("text", ""))
        # 占位非空 → 不送 AI，直接用占位作为译文
        needs_ai = not bool(placeholder)
        preprocessed.append((cleaned, placeholder, needs_ai))

    # 把需要 AI 的部分按 batch 处理
    ai_indexes = [i for i, p in enumerate(preprocessed) if p[2]]
    total_batches = (len(ai_indexes) + AI_BATCH_SIZE - 1) // AI_BATCH_SIZE if ai_indexes else 0
    for start in range(0, len(ai_indexes), AI_BATCH_SIZE):
        idx_chunk = ai_indexes[start : start + AI_BATCH_SIZE]
        expected = len(idx_chunk)
        batch_num = start // AI_BATCH_SIZE + 1
        if progress:
            progress(f"AI 翻译/分类中：第 {batch_num}/{total_batches} 批（共 {n} 条评论）")
        lines = "\n".join(
            json.dumps(
                {
                    "idx": k,
                    "id": str(comments[i].get("_analysis_id") or comments[i].get("comment_id") or f"C{i}"),
                    "likes": _safe_count(
                        comments[i].get("comment_like_count") or comments[i].get("like_count")
                    ),
                    "text": preprocessed[i][0].replace(chr(10), " ")[:1000],
                },
                ensure_ascii=False,
            )
            for k, i in enumerate(idx_chunk)
        )
        prompt = _build_ai_prompt(lines, expected)
        try:
            text, tokens = ai_call(prompt, 90)
            total_tokens += int(tokens or 0)
        except Exception as e:
            logger.error(f"❌ AI 调用失败: {e}")
            text = ""
        parsed = _safe_json_array(text)
        parsed_objs = [o for o in parsed if isinstance(o, dict)]
        if len(parsed_objs) != expected:
            logger.warning(
                f"⚠️ AI 返回数量不匹配：期望 {expected}，实际 {len(parsed_objs)}；"
                f"将按 id/idx 对齐，缺失项留白"
            )

        by_id: dict[str, dict] = {}
        by_idx: dict[int, dict] = {}
        for obj in parsed_objs:
            obj_id = str(obj.get("id") or "").strip()
            if obj_id and obj_id not in by_id:
                by_id[obj_id] = obj
            try:
                obj_idx = int(obj.get("idx"))
            except Exception:
                obj_idx = None
            if obj_idx is not None and 0 <= obj_idx < expected and obj_idx not in by_idx:
                by_idx[obj_idx] = obj

        for k, i in enumerate(idx_chunk):
            # 主对齐方式：稳定 id；其次严格 idx。只有对象完全没有定位字段且数量匹配时才按位置兜底。
            obj = {}
            target_id = str(comments[i].get("_analysis_id") or comments[i].get("comment_id") or f"C{i}")
            if target_id in by_id:
                obj = by_id[target_id]
            elif k in by_idx:
                obj = by_idx[k]
            elif len(parsed_objs) == expected and k < len(parsed_objs):
                cand = parsed_objs[k]
                if not str(cand.get("id") or "").strip() and cand.get("idx") is None:
                    obj = cand
            if not obj:
                logger.warning(f"⚠️ AI 结果缺失：comment_id={target_id}，该条译文留空以避免错位")

            tr = obj.get("translation_zh") or obj.get("translation") or ""
            original = preprocessed[i][0]
            tr_clean = _post_check_translation(tr, original)
            sent = _normalize_ai_label(obj.get("sentiment"), SENTIMENT_LABELS, "中立")
            cat = _normalize_ai_label(obj.get("category"), CATEGORY_LABELS, "其他")
            results[i] = {
                "translation_zh": tr_clean,
                "sentiment": sent,
                "category": cat,
            }

    # 填充不需要 AI 的占位结果（灌水、纯表情等）
    for i, (cleaned, placeholder, needs_ai) in enumerate(preprocessed):
        if needs_ai and results[i] is not None:
            continue
        if results[i] is not None:
            continue
        results[i] = {
            "translation_zh": placeholder or "",
            "sentiment": "中立",
            "category": "其他",
        }
    return results, total_tokens


# ============================================
# 主管线
# ============================================


def _scrape_summary_item(idx: int, payload: dict) -> dict:
    items = payload.get("items") or []
    platform = payload.get("platform") or "UNKNOWN"
    url = payload.get("url") or ""
    comment_count = 0
    if isinstance(items, list):
        comment_count = len(_flatten_comment_items(items, platform, url))
    actor_meta = payload.get("actor_meta") if isinstance(payload.get("actor_meta"), dict) else {}
    error = str(payload.get("error") or "")[:300]
    if not error and isinstance(items, list) and items and comment_count == 0:
        error = _summarize_non_comment_items(items, platform, url)[:300]
    return {
        "source_index": idx,
        "platform": platform,
        "url": url,
        "resolved_url": payload.get("resolved_url") or "",
        "item_count": len(items) if isinstance(items, list) else 0,
        "comment_count": comment_count,
        "error": error,
        "actor_run_id": actor_meta.get("run_id") or "",
        "actor_dataset_id": actor_meta.get("dataset_id") or "",
    }


def run_insight_pipeline(
    urls: list[str],
    apify_token: str,
    ai_call: Callable[[str, int], tuple[str, int]],
    progress: Callable[[str], None] | None = None,
    comments_per_post_limit: int | None = None,
) -> dict:
    """跑完整的：多平台抓取 → AI 翻译/分类 → 结构化结果 + HTML 表格。

    Args:
        urls: 帖子链接列表
        apify_token: Apify Token
        ai_call: 用于调用大模型的函数 (prompt, timeout) -> (text, tokens)
        progress: 可选，进度回调
        comments_per_post_limit: 单条链接评论抓取上限；None 时默认 500
    Returns:
        dict 含 structured / html / total_comments / total_tokens
    """
    if not apify_token:
        raise RuntimeError("APIFY_TOKEN 未配置")

    structured: list[dict] = []
    total_tokens = 0
    comment_limit = normalize_comments_per_post_limit(comments_per_post_limit)

    def _p(msg: str):
        logger.info(f"[insight] {msg}")
        if progress:
            try:
                progress(msg)
            except Exception:
                pass

    # 1. 抓取各平台
    per_url_payloads: list[dict] = []
    scrape_summary: list[dict] = []
    for idx, url in enumerate(urls, 1):
        platform = detect_platform(url)
        _p(f"抓取 {idx}/{len(urls)} [{platform}] {url[:60]}（单条上限 {comment_limit}）")
        if platform == "UNKNOWN" or platform not in PLATFORM_SCRAPERS:
            logger.warning(f"⚠️ 不支持的平台，跳过: {url}")
            payload = {"platform": platform or "UNKNOWN", "url": url, "items": [], "error": "不支持的平台"}
            per_url_payloads.append(payload)
            scrape_summary.append(_scrape_summary_item(idx, payload))
            continue
        try:
            scraper = PLATFORM_SCRAPERS[platform]
            if len(inspect.signature(scraper).parameters) >= 3:
                payload = scraper(url, apify_token, comment_limit)
            else:
                payload = scraper(url, apify_token)
        except Exception as e:
            logger.error(f"❌ 抓取失败 {url}: {e}")
            payload = {"platform": platform, "url": url, "items": [], "error": str(e)}
        per_url_payloads.append(payload)
        scrape_summary.append(_scrape_summary_item(idx, payload))

    # 2. 收集评论列表（带主贴元信息），准备给 AI
    all_comments_for_ai: list[dict] = []
    for payload_index, payload in enumerate(per_url_payloads):
        platform = payload.get("platform", "UNKNOWN")
        url = payload.get("url", "")
        items = payload.get("items") or []
        if not items:
            continue
        meta = _extract_post_meta(items, platform, url)
        for raw_index, it in _flatten_comment_items(items, platform, url):
            row_post_url = _extract_item_post_url(it, platform, url)
            c = _extract_comment(it, platform, row_post_url)
            if not c:
                continue
            analysis_id = f"C{len(all_comments_for_ai)}"
            all_comments_for_ai.append(
                {
                    "_source_index": raw_index,
                    "_analysis_id": analysis_id,
                    "platform": platform,
                    "comment_id": c["comment_id"],
                    "post_url": row_post_url,
                    "post_title": meta["post_title"],
                    "post_date": meta["post_date"],
                    "author": c["author"],
                    "text": c["text"],
                    "created_str": c["created_str"],
                    "bucket": c["bucket"],
                    "comment_like_count": c["like_count"],
                    "comment_url": c["comment_url"],
                }
            )

    if len(all_comments_for_ai) > MAX_AI_COMMENTS:
        all_comments_for_ai = sorted(
            all_comments_for_ai,
            key=lambda c: (int(c.get("comment_like_count") or 0), -int(c.get("_source_index") or 0)),
            reverse=True,
        )
        logger.warning(
            f"⚠️ 评论数 {len(all_comments_for_ai)} 超过上限 {MAX_AI_COMMENTS}，"
            f"优先保留高点赞评论后截取前 {MAX_AI_COMMENTS} 条"
        )
        all_comments_for_ai = all_comments_for_ai[:MAX_AI_COMMENTS]

    _p(f"AI 翻译/分类中（共 {len(all_comments_for_ai)} 条评论）")
    ai_results, ai_tokens = _run_ai_for_comments(all_comments_for_ai, ai_call, progress=_p)
    total_tokens += ai_tokens

    # 3. 合并结构化输出
    for c, r in zip(all_comments_for_ai, ai_results):
        structured.append(
            {
                "_schema": SCHEMA_VERSION,
                "source_index": int(c.get("_source_index") or 0) + 1,
                "platform": c["platform"],
                "post_date": c["post_date"],
                "post_url": c["post_url"],
                "post_title": c["post_title"],
                "comment_time": c["created_str"],
                "time_bucket": c["bucket"],
                "comment_like_count": int(c.get("comment_like_count") or 0),
                "comment_url": c.get("comment_url", ""),
                "comment_id": c.get("comment_id", ""),
                "author": c["author"],
                "content": c["text"],
                "translation_zh": r.get("translation_zh", ""),
                "sentiment_ai": r.get("sentiment", "中立"),
                "sentiment_manual": "",
                "category": r.get("category", "其他"),
                "scrape_status": "成功",
            }
        )

    html_table = build_html_table(structured)

    return {
        "structured": structured,
        "html": html_table,
        "total_comments": len(structured),
        "total_tokens": total_tokens,
        "scrape_summary": scrape_summary,
        "per_url": per_url_payloads,  # 调试用
    }


# ============================================
# HTML / Excel 输出
# ============================================


def _html_escape(s: str) -> str:
    if s is None:
        return ""
    return (
        str(s)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def _sentiment_color(label: str) -> str:
    return {"正向": "#2e7d32", "中立": "#616161", "负面": "#c62828"}.get(label, "#616161")


def build_html_table(rows: list[dict]) -> str:
    if not rows:
        return (
            "<div style='padding:24px;text-align:center;color:#666;'>未抓取到可分析的评论。</div>"
        )

    head_cells = "".join(
        f"<th>{_html_escape(h)}</th>" for h in INSIGHT_HEADERS
    )

    body_parts: list[str] = []
    for r in rows:
        link = _html_escape(r.get("post_url", ""))
        link_cell = (
            f"<a href=\"{link}\" target=\"_blank\" rel=\"noopener\">{link}</a>" if link else ""
        )
        sentiment = _html_escape(r.get("sentiment_ai", ""))
        color = _sentiment_color(r.get("sentiment_ai", ""))
        cells = [
            _html_escape(r.get("source_index", "")),
            _html_escape(r.get("platform", "")),
            _html_escape(r.get("post_date", "")),
            link_cell,
            _html_escape(r.get("post_title", "")),
            _html_escape(r.get("comment_time", "")),
            _html_escape(r.get("time_bucket", "")),
            str(_safe_count(r.get("comment_like_count"))),
            _html_escape(r.get("author", "")),
            _html_escape(r.get("content", "")),
            _html_escape(r.get("translation_zh", "")),
            f"<span style='color:{color};font-weight:600;'>{sentiment}</span>",
            _html_escape(r.get("sentiment_manual", "")),
            _html_escape(r.get("category", "")),
            _html_escape(r.get("comment_id", "")),
            _html_escape(r.get("comment_url", "")),
            _html_escape(r.get("scrape_status", "成功")),
        ]
        body_parts.append("<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>")

    return (
        "<table class='insight-table'>"
        f"<thead><tr>{head_cells}</tr></thead>"
        f"<tbody>{''.join(body_parts)}</tbody>"
        "</table>"
    )


_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sanitize_sheet_name(name: str, used: set[str]) -> str:
    """Excel sheet 名约束：≤31 字符，且不含 [ ] : * ? / \\，且全工作簿唯一。"""
    cleaned = _INVALID_SHEET_CHARS.sub("·", name or "").strip()
    if not cleaned:
        cleaned = "Sheet"
    cleaned = cleaned[:31]
    base = cleaned
    suffix = 2
    while cleaned in used:
        tail = f" ({suffix})"
        cleaned = (base[: 31 - len(tail)] + tail)
        suffix += 1
    used.add(cleaned)
    return cleaned


def _write_insight_sheet(ws, rows: list[dict]) -> None:
    """把若干行写入一个 worksheet：表头 + 内容 + 样式。"""
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="top", wrap_text=True)

    ws.append(INSIGHT_HEADERS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for r in rows:
        ws.append([
            r.get("source_index", ""),
            r.get("platform", ""),
            r.get("post_date", ""),
            r.get("post_url", ""),
            r.get("post_title", ""),
            r.get("comment_time", ""),
            r.get("time_bucket", ""),
            _safe_count(r.get("comment_like_count")),
            r.get("author", ""),
            r.get("content", ""),
            r.get("translation_zh", ""),
            r.get("sentiment_ai", ""),
            r.get("sentiment_manual", ""),
            r.get("category", ""),
            r.get("comment_id", ""),
            r.get("comment_url", ""),
            r.get("scrape_status", "成功"),
        ])

    widths = [10, 8, 18, 50, 30, 18, 12, 12, 18, 50, 50, 18, 18, 16, 34, 50, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    ws.freeze_panes = "A2"


def build_excel(rows: list[dict]) -> Workbook:
    """生成 Excel：每个帖文一个 sheet，按出现顺序排列。

    Sheet 命名：`{平台}-{序号} {标题前几个字}`，自动截到 31 字符且唯一。
    """
    wb = Workbook()
    wb.remove(wb.active)

    if not rows:
        ws = wb.create_sheet("舆情洞察")
        _write_insight_sheet(ws, [])
        return wb

    # 按帖子分组，保留首次出现的顺序
    groups: dict[str, list[dict]] = {}
    order: list[str] = []
    for r in rows:
        key = r.get("post_url") or f"_no_url_{len(order)}"
        if key not in groups:
            groups[key] = []
            order.append(key)
        groups[key].append(r)

    used_names: set[str] = set()
    platform_counter: dict[str, int] = {}
    for key in order:
        post_rows = groups[key]
        first = post_rows[0]
        platform = (first.get("platform") or "POST").strip() or "POST"
        platform_counter[platform] = platform_counter.get(platform, 0) + 1
        seq = platform_counter[platform]
        title = (first.get("post_title") or "").strip()
        # 去掉换行/制表，截一段当 sheet 名后缀
        short_title = re.sub(r"\s+", " ", title)[:18]
        base = f"{platform}-{seq}"
        raw_name = f"{base} {short_title}".strip() if short_title else base
        sheet_name = _sanitize_sheet_name(raw_name, used_names)
        ws = wb.create_sheet(sheet_name)
        _write_insight_sheet(ws, post_rows)

    return wb


def parse_urls_text(text: str) -> list[str]:
    """从用户输入（多行/逗号分隔）解析出 URL 列表，去重去空。"""
    if not text:
        return []
    seen: set[str] = set()
    result: list[str] = []
    for token in re.split(r"[\s,;\n]+", text.strip()):
        token = token.strip()
        if not token:
            continue
        if not (token.startswith("http://") or token.startswith("https://")):
            continue
        if token in seen:
            continue
        seen.add(token)
        result.append(token)
    return result
