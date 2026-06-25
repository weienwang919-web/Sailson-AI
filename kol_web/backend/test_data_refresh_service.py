from __future__ import annotations

import sys
import tempfile
import unittest
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker


sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.database import Base
from app.models import DataRefreshJob, KOLRecord
from app.services import data_refresh_service


class DataRefreshServiceTests(unittest.TestCase):
    def setUp(self):
        engine = create_engine("sqlite:///:memory:", future=True)
        Base.metadata.create_all(bind=engine)
        self.Session = sessionmaker(bind=engine, future=True)
        self.tmp = tempfile.TemporaryDirectory()
        self.old_output_dir = data_refresh_service.OUTPUT_DIR
        data_refresh_service.OUTPUT_DIR = Path(self.tmp.name)

    def tearDown(self):
        data_refresh_service.OUTPUT_DIR = self.old_output_dir
        self.tmp.cleanup()

    def test_excel_refresh_preserves_original_columns_and_does_not_sync_by_default(self):
        db = self.Session()
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Input"
            ws.append(["达人", "Link", "备注"])
            ws.append(["Alice", "https://www.tiktok.com/@alice", "keep me"])
            rows = data_refresh_service.rows_from_workbook(wb)
            job = DataRefreshJob(input_type="excel", status="running", total=len(rows), include_acv=1)
            db.add(job)
            db.commit()
            db.refresh(job)

            old_fetch = data_refresh_service._fetch_raw
            data_refresh_service._fetch_raw = lambda _records, _n: (
                {
                    "tiktok": [
                        {
                            "authorMeta": {"name": "alice", "fans": 1000},
                            "webVideoUrl": "https://www.tiktok.com/@alice/video/1",
                            "playCount": 200,
                            "avgLiveViewers": 55,
                        }
                    ],
                    "ins": [],
                    "youtube": [],
                },
                {},
            )
            try:
                data_refresh_service._run_refresh(db, job, rows, wb, False, True, 10)
            finally:
                data_refresh_service._fetch_raw = old_fetch

            self.assertEqual(db.query(KOLRecord).count(), 0)
            out = load_workbook(job.output_path)
            out_ws = out["Input"]
            self.assertEqual(out_ws.cell(2, 1).value, "Alice")
            self.assertEqual(out_ws.cell(2, 3).value, "keep me")
            headers = [out_ws.cell(1, col).value for col in range(1, out_ws.max_column + 1)]
            for header in data_refresh_service.OUTPUT_COLUMNS:
                self.assertIn(header, headers)
            self.assertEqual(out_ws.cell(2, headers.index("粉丝数") + 1).value, 1000)
            self.assertEqual(out_ws.cell(2, headers.index("AVV / 均观看量") + 1).value, 200)
            self.assertEqual(out_ws.cell(2, headers.index("ACV / 平均直播观看人数") + 1).value, 55)
            self.assertEqual(out_ws.cell(2, headers.index("抓取状态") + 1).value, "成功")
        finally:
            db.close()

    def test_rows_from_workbook_detects_header_below_first_row(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Input"
        ws.append(["达人数据更新模板"])
        ws.append(["请勿删除说明行"])
        ws.append(["达人", "主页链接", "备注"])
        ws.append(["Alice", "https://www.tiktok.com/@alice", "keep me"])

        rows = data_refresh_service.rows_from_workbook(wb)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["row_idx"], 4)
        self.assertEqual(rows[0]["platform"], "tiktok")
        self.assertEqual(rows[0]["link"], "https://www.tiktok.com/@alice")

    def test_rows_from_workbook_reads_hyperlink_target(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Input"
        ws.append(["达人", "TikTok", "备注"])
        ws.append(["Alice", "用户附件", "keep me"])
        ws.cell(2, 2).hyperlink = "https://www.tiktok.com/@alice"

        rows = data_refresh_service.rows_from_workbook(wb)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["raw_link"], "https://www.tiktok.com/@alice")
        self.assertEqual(rows[0]["platform"], "tiktok")

    def test_rows_from_workbook_skips_non_profile_values_in_platform_columns(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Input"
        ws.append(["达人", "Instagram", "YouTube", "备注"])
        ws.append(["Noise", "Instagram", "https://www.youtube.com/watch?v=abc123", "skip these"])
        ws.append(["Alice", "https://www.instagram.com/alice/", "https://www.youtube.com/channel/UCabc123", "keep"])

        rows = data_refresh_service.rows_from_workbook(wb)

        self.assertEqual(len(rows), 2)
        self.assertEqual({row["platform"] for row in rows}, {"ins", "youtube"})
        self.assertEqual(rows[0]["link"], "https://www.instagram.com/alice/")
        self.assertEqual(rows[1]["link"], "https://www.youtube.com/channel/UCabc123")
        self.assertFalse(any(row.get("error") == "链接格式无效" for row in rows))

    def test_sync_to_pool_adds_record_and_acv_only_when_explicitly_returned(self):
        db = self.Session()
        try:
            rows = data_refresh_service.rows_from_links("https://www.youtube.com/@creator")
            job = DataRefreshJob(input_type="links", status="running", total=len(rows), include_acv=1)
            db.add(job)
            db.commit()
            db.refresh(job)

            old_fetch = data_refresh_service._fetch_raw
            data_refresh_service._fetch_raw = lambda _records, _n: (
                {
                    "youtube": [
                        {
                            "channelUsername": "@creator",
                            "numberOfSubscribers": 5000,
                            "viewCount": 300,
                        }
                    ],
                    "tiktok": [],
                    "ins": [],
                },
                {},
            )
            try:
                data_refresh_service._run_refresh(db, job, rows, None, True, True, 10)
            finally:
                data_refresh_service._fetch_raw = old_fetch

            record = db.query(KOLRecord).one()
            self.assertEqual(record.yt_follower, 5000)
            self.assertEqual(record.yt_avv, 300)
            self.assertIsNone(record.yt_acv)
        finally:
            db.close()

    def test_youtube_channel_id_links_match_raw_input_channel_url(self):
        db = self.Session()
        try:
            rows = data_refresh_service.rows_from_links("https://www.youtube.com/channel/UCHePEXTqWF-lexZaIkFHTMw")
            job = DataRefreshJob(input_type="links", status="running", total=len(rows), include_acv=0)
            db.add(job)
            db.commit()
            db.refresh(job)

            old_fetch = data_refresh_service._fetch_raw
            data_refresh_service._fetch_raw = lambda _records, _n: (
                {
                    "youtube": [
                        {
                            "inputChannelUrl": "https://www.youtube.com/channel/UCHePEXTqWF-lexZaIkFHTMw",
                            "numberOfSubscribers": 12000,
                            "viewCount": 900,
                        }
                    ],
                    "tiktok": [],
                    "ins": [],
                },
                {},
            )
            try:
                data_refresh_service._run_refresh(db, job, rows, None, False, False, 10)
            finally:
                data_refresh_service._fetch_raw = old_fetch

            self.assertEqual(job.success_count, 1)
            self.assertEqual(job.failed_count, 0)
            self.assertEqual(rows[0]["followers"], 12000)
            self.assertEqual(rows[0]["avv"], 900)
        finally:
            db.close()

    def test_empty_excel_rows_fail_with_clear_error(self):
        db = self.Session()
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Input"
            ws.append(["达人", "备注"])
            ws.append(["Alice", "no link"])
            rows = data_refresh_service.rows_from_workbook(wb)
            job = DataRefreshJob(input_type="excel", status="running", total=0, include_acv=1)
            db.add(job)
            db.commit()
            db.refresh(job)

            data_refresh_service._run_refresh(db, job, rows, wb, False, True, 10)

            self.assertEqual(job.status, "failed")
            self.assertIn("未识别到达人主页链接", job.error)
        finally:
            db.close()

    def test_refresh_job_summarizes_failed_rows_with_examples(self):
        db = self.Session()
        try:
            rows = data_refresh_service.rows_from_links("https://www.tiktok.com/@alice")
            job = DataRefreshJob(input_type="links", status="running", total=len(rows), include_acv=1)
            db.add(job)
            db.commit()
            db.refresh(job)

            old_fetch = data_refresh_service._fetch_raw
            data_refresh_service._fetch_raw = lambda _records, _n: (
                {
                    "tiktok": [
                        {
                            "authorMeta": {"name": "bob", "fans": 1000},
                            "webVideoUrl": "https://www.tiktok.com/@bob/video/1",
                            "playCount": 200,
                        }
                    ],
                    "ins": [],
                    "youtube": [],
                },
                {},
            )
            try:
                data_refresh_service._run_refresh(db, job, rows, None, False, True, 10)
            finally:
                data_refresh_service._fetch_raw = old_fetch

            self.assertEqual(job.status, "completed")
            self.assertEqual(job.success_count, 0)
            self.assertEqual(job.failed_count, 1)
            self.assertIn("未匹配到该达人数据", rows[0]["error"])
            self.assertIn("alice", rows[0]["error"])
            self.assertIn("bob", rows[0]["error"])
            summary = json.loads(job.summary_json)
            self.assertEqual(summary["failure_summary"]["平台返回数据未匹配到该链接"], 1)
            self.assertEqual(summary["error_examples"][0]["link"], "https://www.tiktok.com/@alice")
            self.assertIn("未匹配到该达人数据", summary["error_examples"][0]["error"])
        finally:
            db.close()


if __name__ == "__main__":
    unittest.main()
