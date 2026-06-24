from __future__ import annotations

import json
import re
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.core.category import normalize_category
from app.core.pipeline import (
    build_apify_inputs,
    call_apify,
    clean_text,
    instagram_username,
    metrics_from_raw,
    normalize_link,
    tiktok_username,
    to_int,
    youtube_channel_seed,
)
from app.models import DataRefreshJob, KOLRecord

OUTPUT_DIR = Path(tempfile.gettempdir()) / "kol-data-refresh"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

LINK_ALIASES = (
    "链接", "link", "url", "channel", "主页链接", "达人主页链接",
    "profile", "profile url", "主页", "homepage", "profile link",
    "tiktok", "instagram", "youtube", "yt", "ins", "tt",
)
NAME_ALIASES = ("达人", "达人名称", "kol", "kol name", "name", "creator", "渠道名", "资源名称")
PLATFORM_ALIASES = ("平台", "platform", "渠道")
FOLLOWER_ALIASES = ("粉丝数", "followers", "follower", "➡️ followers")
AVV_ALIASES = ("avv", "avg view", "均观看量", "average views", "avg view count")
ACV_ALIASES = ("acv", "avg live viewers", "average live viewers", "平均直播观看人数", "直播均观")
OUTPUT_COLUMNS = [
    "平台",
    "达人主页链接",
    "粉丝数",
    "AVV / 均观看量",
    "ACV / 平均直播观看人数",
    "数据更新时间",
    "抓取状态",
    "失败原因",
]
JOB_STEPS = ["解析输入", "抓取平台数据", "写回 Excel", "可选同步资源池", "完成下载"]


def create_job(db: Session, input_type: str, total: int, sync_to_pool: bool, include_acv: bool, videos_per_profile: int) -> DataRefreshJob:
    job = DataRefreshJob(
        input_type=input_type,
        status="pending",
        total=total,
        sync_to_pool=1 if sync_to_pool else 0,
        include_acv=1 if include_acv else 0,
        videos_per_profile=max(1, int(videos_per_profile or 10)),
        summary_json=_job_summary_json(step=JOB_STEPS[0], total=total),
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def run_link_refresh_job(job_id: int, text: str, sync_to_pool: bool, include_acv: bool, videos_per_profile: int) -> None:
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        job = db.get(DataRefreshJob, job_id)
        if not job:
            return
        job.status = "running"
        job.summary_json = _job_summary_json(step=JOB_STEPS[0])
        db.commit()
        rows = rows_from_links(text)
        _run_refresh(db, job, rows, None, sync_to_pool, include_acv, videos_per_profile)
    except Exception as exc:  # noqa: BLE001
        _fail_job(db, job_id, str(exc))
    finally:
        db.close()


def run_excel_refresh_job(
    job_id: int,
    input_path: str,
    sync_to_pool: bool,
    include_acv: bool,
    videos_per_profile: int,
) -> None:
    from app.database import SessionLocal

    db = SessionLocal()
    try:
        job = db.get(DataRefreshJob, job_id)
        if not job:
            return
        job.status = "running"
        job.summary_json = _job_summary_json(step=JOB_STEPS[0])
        db.commit()
        wb = load_workbook(input_path)
        rows = rows_from_workbook(wb)
        _run_refresh(db, job, rows, wb, sync_to_pool, include_acv, videos_per_profile)
    except Exception as exc:  # noqa: BLE001
        _fail_job(db, job_id, str(exc))
    finally:
        db.close()


def rows_from_links(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for raw in re.findall(r"https?://[^\s,;，；]+", text or ""):
        raw = raw.rstrip(").]}>\"'")
        platform = detect_link_platform(raw)
        if not platform:
            rows.append({"raw_link": raw, "status": "失败", "error": "无法识别平台"})
            continue
        normalized = normalize_link(platform, raw)
        if not normalized:
            rows.append({"raw_link": raw, "status": "失败", "error": "链接格式无效"})
            continue
        key = (platform, normalized)
        if key in seen:
            continue
        seen.add(key)
        rows.append({
            "name": display_name_for_link(platform, normalized),
            "platform": platform,
            "link": normalized,
            "raw_link": raw,
        })
    return rows


def rows_from_workbook(wb) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        header_row, headers, indexes = _find_header_row(ws)
        if not header_row:
            continue
        link_cols = indexes["links"]
        if not link_cols:
            continue
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row_values = {headers[col - 1]: ws.cell(row_idx, col).value for col in range(1, ws.max_column + 1) if headers[col - 1]}
            for col in link_cols:
                raw = _cell_link_text(ws.cell(row_idx, col))
                if not raw:
                    continue
                platform = detect_link_platform(raw) or detect_platform_text(_cell_text(row_values.get(indexes.get("platform") or "")))
                if not platform:
                    platform = _guess_platform_from_header(headers[col - 1])
                normalized = normalize_link(platform, raw) if platform else ""
                if not normalized:
                    rows.append({"worksheet": ws.title, "row_idx": row_idx, "raw_link": raw, "status": "失败", "error": "链接格式无效"})
                    continue
                rows.append({
                    "worksheet": ws.title,
                    "row_idx": row_idx,
                    "name": _cell_text(row_values.get(indexes.get("name") or "")) or display_name_for_link(platform, normalized),
                    "platform": platform,
                    "link": normalized,
                    "raw_link": raw,
                    "existing": row_values,
                })
    return rows


def _run_refresh(
    db: Session,
    job: DataRefreshJob,
    rows: list[dict[str, Any]],
    wb,
    sync_to_pool: bool,
    include_acv: bool,
    videos_per_profile: int,
) -> None:
    job.total = len(rows)
    job.summary_json = _job_summary_json(step=JOB_STEPS[1], total=len(rows))
    db.commit()
    if not rows:
        job.status = "failed"
        job.error = "未识别到达人主页链接，请确认 Excel 表头包含链接/Link/URL/主页链接，或单元格超链接指向 TikTok/Instagram/YouTube 主页"
        job.failed_count = 0
        job.summary_json = _job_summary_json(
            step=JOB_STEPS[0],
            total=0,
            success=0,
            failed=0,
            errors=[job.error],
        )
        job.updated_at = datetime.utcnow()
        db.commit()
        return
    valid_rows = [row for row in rows if row.get("platform") and row.get("link")]
    records = [_row_to_record_stub(row) for row in valid_rows]
    raw, platform_errors = _fetch_raw(records, videos_per_profile)
    metrics = metrics_from_raw(raw, videos_per_profile)
    acv_metrics = extract_acv_metrics(raw) if include_acv else {}

    success = 0
    failed = 0
    added = 0
    updated = 0
    for row in rows:
        platform = row.get("platform")
        link = row.get("link")
        if not platform or not link:
            row["status"] = row.get("status") or "失败"
            row["error"] = row.get("error") or "无法识别平台或链接"
            failed += 1
            continue
        if platform_errors.get(platform):
            row["status"] = "失败"
            row["error"] = f"抓取失败：{platform_errors[platform][:220]}"
            failed += 1
            continue
        metric = _lookup_metrics(platform, link, metrics)
        acv = _lookup_acv(platform, link, acv_metrics)
        if not metric and acv is None:
            row["status"] = "失败"
            row["error"] = _missing_metrics_reason(row, raw, metrics, acv_metrics)
            failed += 1
        else:
            row["followers"] = metric.get("followers") if metric else None
            row["avv"] = metric.get("avv") if metric else None
            row["acv"] = acv
            row["status"] = "成功"
            row["error"] = ""
            success += 1
            if sync_to_pool:
                was_update = upsert_pool_record(db, row)
                if was_update:
                    updated += 1
                else:
                    added += 1
    db.commit()

    failure_summary = _failure_summary(rows)
    job.summary_json = _job_summary_json(
        step=JOB_STEPS[2],
        total=len(rows),
        success=success,
        failed=failed,
        added=added,
        updated=updated,
        failure_summary=failure_summary,
    )
    db.commit()
    out_path = write_output(rows, wb, job.id)
    job.status = "completed"
    job.success_count = success
    job.failed_count = failed
    job.added_count = added
    job.updated_count = updated
    job.output_path = str(out_path)
    job.output_filename = out_path.name
    failure_summary = _failure_summary(rows)
    job.summary_json = _job_summary_json(
        step=JOB_STEPS[4],
        total=len(rows),
        success=success,
        failed=failed,
        added=added,
        updated=updated,
        failure_summary=failure_summary,
        errors=[row.get("error") for row in rows if row.get("error")][:20],
        error_examples=_failure_examples(rows),
        download_file_id=job.id,
        output_filename=out_path.name,
    )
    job.updated_at = datetime.utcnow()
    db.commit()


def _fetch_raw(records: list[KOLRecord], videos_per_profile: int) -> tuple[dict[str, list[dict[str, Any]]], dict[str, str]]:
    inputs = build_apify_inputs(records, videos_per_profile)
    raw: dict[str, list[dict[str, Any]]] = {}
    errors: dict[str, str] = {}
    for platform, config in inputs.items():
        run_input = config["input"]
        if platform == "tiktok" and not run_input.get("profiles"):
            raw[platform] = []
            continue
        if platform == "ins" and not run_input.get("directUrls"):
            raw[platform] = []
            continue
        if platform == "youtube" and not run_input.get("startUrls"):
            raw[platform] = []
            continue
        try:
            raw[platform] = call_apify(config["actor"], run_input)
        except Exception as exc:  # noqa: BLE001 - one platform should not block the whole workbook
            raw[platform] = []
            errors[platform] = str(exc)
    return raw, errors


def _missing_metrics_reason(
    row: dict[str, Any],
    raw: dict[str, list[dict[str, Any]]],
    metrics: dict[str, dict[str, dict[str, Any]]],
    acv_metrics: dict[str, dict[str, int]],
) -> str:
    platform = row.get("platform") or ""
    link = row.get("link") or row.get("raw_link") or ""
    label = platform_label(platform)
    platform_raw = raw.get(platform) or []
    if not platform_raw:
        return f"{label} 抓取无返回：请检查主页是否公开/存在，或稍后重试"
    key = _metric_key(platform, link)
    metric_keys = sorted((metrics.get(platform) or {}).keys())
    acv_keys = sorted((acv_metrics.get(platform) or {}).keys())
    if key:
        sample_keys = metric_keys[:5] or acv_keys[:5]
        if sample_keys:
            return f"{label} 未匹配到该达人数据：链接标识 {key}，平台返回标识示例 {', '.join(sample_keys)}"
        return f"{label} 返回了 {len(platform_raw)} 条数据，但没有可匹配的粉丝/AVV/ACV 字段"
    return f"{label} 链接无法提取达人标识：{link}"


def _failure_summary(rows: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for row in rows:
        if not row.get("error"):
            continue
        reason = _compact_failure_reason(row.get("error") or "")
        summary[reason] = summary.get(reason, 0) + 1
    return summary


def _failure_examples(rows: list[dict[str, Any]], limit: int = 20) -> list[dict[str, Any]]:
    examples = []
    for row in rows:
        if not row.get("error"):
            continue
        examples.append({
            "name": row.get("name") or "",
            "platform": platform_label(row.get("platform") or ""),
            "link": row.get("link") or row.get("raw_link") or "",
            "error": row.get("error") or "",
        })
        if len(examples) >= limit:
            break
    return examples


def _compact_failure_reason(error: str) -> str:
    text = clean_text(error)
    if not text:
        return "未知原因"
    if "抓取无返回" in text:
        return "平台抓取无返回"
    if "未匹配到该达人数据" in text:
        return "平台返回数据未匹配到该链接"
    if "没有可匹配的粉丝/AVV/ACV字段" in text or "没有可匹配的粉丝/AVV/ACV 字段" in text:
        return "平台返回缺少粉丝/AVV/ACV字段"
    if "链接格式无效" in text:
        return "链接格式无效"
    if "无法识别平台" in text:
        return "无法识别平台"
    if "抓取失败" in text:
        return "平台抓取失败"
    return text[:80]


def write_output(rows: list[dict[str, Any]], wb, job_id: int) -> Path:
    if wb is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "达人数据更新"
        ws.append(OUTPUT_COLUMNS)
        for row in rows:
            ws.append(_output_values(row))
    else:
        by_sheet_row: dict[tuple[str, int], list[dict[str, Any]]] = {}
        for row in rows:
            if row.get("worksheet") and row.get("row_idx"):
                by_sheet_row.setdefault((row["worksheet"], int(row["row_idx"])), []).append(row)
        for ws in wb.worksheets:
            if ws.max_row < 2:
                continue
            header_row, _headers, _indexes = _find_header_row(ws)
            if not header_row:
                continue
            col_map = ensure_output_columns(ws, header_row=header_row)
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row_group = by_sheet_row.get((ws.title, row_idx))
                if not row_group:
                    continue
                row = _merge_sheet_row_results(row_group)
                _write_row(ws, row_idx, col_map, row)
    out_path = OUTPUT_DIR / f"kol_data_refresh_{job_id}.xlsx"
    wb.save(out_path)
    return out_path


def ensure_output_columns(ws: Worksheet, header_row: int = 1) -> dict[str, int]:
    headers = {_cell_text(ws.cell(header_row, col).value): col for col in range(1, ws.max_column + 1)}
    for name in OUTPUT_COLUMNS:
        if name not in headers:
            col = ws.max_column + 1
            ws.cell(header_row, col).value = name
            headers[name] = col
    return headers


def upsert_pool_record(db: Session, row: dict[str, Any]) -> bool:
    platform = row.get("platform")
    link = row.get("link")
    record = find_record(db, platform, link)
    is_update = record is not None
    if not record:
        record = KOLRecord(
            name=row.get("name") or display_name_for_link(platform, link),
            category="Data Refresh",
            normalized_category=normalize_category("Data Refresh"),
            source_file="达人数据更新",
            platform_text=platform_label(platform),
        )
        db.add(record)
        db.flush()
    assign_link(record, platform, link)
    assign_metrics(record, platform, row)
    record.last_scraped_at = datetime.utcnow()
    return is_update


def find_record(db: Session, platform: str, link: str) -> KOLRecord | None:
    if platform == "tiktok":
        return db.query(KOLRecord).filter(KOLRecord.tt_link == link).order_by(KOLRecord.id.asc()).first()
    if platform == "ins":
        return db.query(KOLRecord).filter(KOLRecord.ins_link == link).order_by(KOLRecord.id.asc()).first()
    return db.query(KOLRecord).filter(KOLRecord.yt_link == link).order_by(KOLRecord.id.asc()).first()


def assign_link(record: KOLRecord, platform: str, link: str) -> None:
    if platform == "tiktok":
        record.tt_link = link
    elif platform == "ins":
        record.ins_link = link
    else:
        record.yt_link = link


def assign_metrics(record: KOLRecord, platform: str, row: dict[str, Any]) -> None:
    follower = row.get("followers")
    avv = row.get("avv")
    acv = row.get("acv")
    if platform == "tiktok":
        if follower is not None:
            record.tt_follower = follower
        if avv is not None:
            record.tt_avv = avv
        if acv is not None:
            record.tt_acv = acv
    elif platform == "ins":
        if follower is not None:
            record.ins_follower = follower
        if acv is not None:
            record.ins_acv = acv
    else:
        if follower is not None:
            record.yt_follower = follower
        if avv is not None:
            record.yt_avv = avv
        if acv is not None:
            record.yt_acv = acv


def extract_acv_metrics(raw: dict[str, list[dict[str, Any]]]) -> dict[str, dict[str, int]]:
    out: dict[str, dict[str, int]] = {"tiktok": {}, "ins": {}, "youtube": {}}
    for item in raw.get("tiktok", []):
        user = ""
        author = item.get("authorMeta") or {}
        if isinstance(author, dict):
            user = clean_text(author.get("name") or author.get("uniqueId")).lower()
        if not user:
            user = tiktok_username(clean_text(item.get("webVideoUrl"))).lower()
        acv = _extract_acv(item)
        if user and acv is not None:
            out["tiktok"][user] = acv
    for item in raw.get("ins", []):
        user = clean_text(item.get("ownerUsername") or (item.get("owner") or {}).get("username")).lower()
        acv = _extract_acv(item)
        if user and acv is not None:
            out["ins"][user] = acv
    for item in raw.get("youtube", []):
        handle = clean_text(item.get("channelUsername")).lstrip("@").lower() or clean_text(item.get("channelName")).lower()
        acv = _extract_acv(item)
        if handle and acv is not None:
            out["youtube"][handle] = acv
    return out


def _extract_acv(item: dict[str, Any]) -> int | None:
    keys = (
        "avgLiveViewers", "averageLiveViewers", "averageConcurrentViewers", "concurrentViewers",
        "liveAvgViewers", "liveAverageViewers", "avg_viewers", "average_viewers",
        "直播均观", "平均直播观看人数", "ACV",
    )
    for key in keys:
        value = _dig_value(item, key)
        parsed = to_int(value)
        if parsed is not None:
            return parsed
    return None


def _dig_value(obj: Any, key: str) -> Any:
    if isinstance(obj, dict):
        if key in obj:
            return obj[key]
        lower_key = key.lower()
        for k, v in obj.items():
            if str(k).lower() == lower_key:
                return v
            found = _dig_value(v, key)
            if found is not None:
                return found
    elif isinstance(obj, list):
        for v in obj:
            found = _dig_value(v, key)
            if found is not None:
                return found
    return None


def _lookup_metrics(platform: str, link: str, metrics: dict[str, dict[str, dict[str, Any]]]) -> dict[str, Any] | None:
    key = _metric_key(platform, link)
    if not key:
        return None
    return metrics.get(platform, {}).get(key)


def _lookup_acv(platform: str, link: str, metrics: dict[str, dict[str, int]]) -> int | None:
    key = _metric_key(platform, link)
    if not key:
        return None
    return metrics.get(platform, {}).get(key)


def _metric_key(platform: str, link: str) -> str:
    if platform == "tiktok":
        return tiktok_username(link).lower()
    if platform == "ins":
        return instagram_username(link).lower()
    value = youtube_channel_seed(link).lower()
    if "/@" in value:
        return value.split("/@")[-1].split("/")[0].lower()
    return value.rstrip("/").split("/")[-1]


def _row_to_record_stub(row: dict[str, Any]) -> KOLRecord:
    record = KOLRecord(name=row.get("name") or row.get("link"), category="Data Refresh")
    assign_link(record, row.get("platform"), row.get("link"))
    return record


def detect_link_platform(link: str) -> str | None:
    lower = (link or "").lower()
    if "tiktok.com" in lower:
        return "tiktok"
    if "instagram.com" in lower:
        return "ins"
    if "youtube.com" in lower or "youtu.be" in lower:
        return "youtube"
    return None


def detect_platform_text(text: str) -> str | None:
    lower = (text or "").lower()
    if "tiktok" in lower or lower == "tt":
        return "tiktok"
    if "instagram" in lower or lower in {"ig", "ins"}:
        return "ins"
    if "youtube" in lower or lower in {"yt", "ytb"}:
        return "youtube"
    return None


def display_name_for_link(platform: str, link: str) -> str:
    if platform == "tiktok":
        user = tiktok_username(link)
        return f"TikTok @{user}" if user else link
    if platform == "ins":
        user = instagram_username(link)
        return f"Instagram @{user}" if user else link
    handle = youtube_channel_seed(link).rstrip("/").split("/")[-1]
    return f"YouTube {handle}" if handle else link


def platform_label(platform: str) -> str:
    return {"tiktok": "TikTok", "ins": "Instagram", "youtube": "YouTube", "multi": "多平台"}.get(platform, platform)


def _detect_indexes(headers: list[str]) -> dict[str, Any]:
    out: dict[str, Any] = {"links": []}
    for idx, header in enumerate(headers, start=1):
        norm = _norm(header)
        if _matches_alias(norm, LINK_ALIASES):
            out["links"].append(idx)
        if "name" not in out and _matches_alias(norm, NAME_ALIASES):
            out["name"] = header
        if "platform" not in out and _matches_alias(norm, PLATFORM_ALIASES):
            out["platform"] = header
    return out


def _find_header_row(ws: Worksheet) -> tuple[int | None, list[str], dict[str, Any]]:
    best: tuple[int | None, list[str], dict[str, Any], int] = (None, [], {"links": []}, -1)
    max_scan = min(ws.max_row, 20)
    for row_idx in range(1, max_scan + 1):
        headers = [_cell_text(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        if not any(headers):
            continue
        indexes = _detect_indexes(headers)
        link_cols = indexes.get("links") or []
        if not link_cols:
            continue
        data_hits = 0
        for data_row in range(row_idx + 1, min(ws.max_row, row_idx + 25) + 1):
            for col in link_cols:
                if _looks_like_supported_link(_cell_link_text(ws.cell(data_row, col))):
                    data_hits += 1
        strong_alias = any(
            _matches_alias(_norm(headers[col - 1]), ("链接", "link", "url", "channel", "主页链接", "达人主页链接", "profile", "profile url", "homepage", "profile link"))
            for col in link_cols
        )
        score = data_hits * 10 + len(link_cols) * 2 + (2 if "name" in indexes else 0) + (1 if "platform" in indexes else 0)
        if not data_hits and not strong_alias:
            continue
        if score > best[3]:
            best = (row_idx, headers, indexes, score)
    return best[0], best[1], best[2]


def _cell_link_text(cell) -> str:
    hyperlink = getattr(cell, "hyperlink", None)
    target = getattr(hyperlink, "target", None) if hyperlink is not None else None
    if _looks_like_supported_link(target):
        return clean_text(target)
    value = _cell_text(cell.value)
    if value.upper().startswith("=HYPERLINK("):
        match = re.search(r"https?://[^\"')\s]+", value, re.I)
        if match:
            return match.group(0)
    return value


def _looks_like_supported_link(value: Any) -> bool:
    text = clean_text(value)
    if not text:
        return False
    if detect_link_platform(text):
        return True
    return bool(re.search(r"https?://[^\s,;，；)）]+", text, re.I))


def _guess_platform_from_header(header: str) -> str | None:
    text = _norm(header)
    if "tiktok" in text or text.startswith("tt"):
        return "tiktok"
    if "instagram" in text or text in {"ins", "ig"}:
        return "ins"
    if "youtube" in text or text.startswith("yt"):
        return "youtube"
    return None


def _write_row(ws: Worksheet, row_idx: int, col_map: dict[str, int], row: dict[str, Any]) -> None:
    for col_name, value in zip(OUTPUT_COLUMNS, _output_values(row)):
        ws.cell(row_idx, col_map[col_name]).value = value


def _output_values(row: dict[str, Any]) -> list[Any]:
    return [
        platform_label(row.get("platform") or ""),
        row.get("link") or row.get("raw_link") or "",
        row.get("followers"),
        row.get("avv"),
        row.get("acv"),
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        row.get("status") or "",
        row.get("error") or "",
    ]


def _cell_text(value: Any) -> str:
    return clean_text(value)


def _norm(value: str) -> str:
    return clean_text(value).strip().lower().replace("\n", " ")


def _matches_alias(norm_header: str, aliases: tuple[str, ...]) -> bool:
    return any(_norm(alias) in norm_header for alias in aliases)


def _merge_sheet_row_results(rows: list[dict[str, Any]]) -> dict[str, Any]:
    if len(rows) == 1:
        return rows[0]
    successful = [row for row in rows if row.get("status") == "成功"]
    base = dict((successful or rows)[0])
    base["platform"] = "multi"
    base["link"] = "\n".join(dict.fromkeys(row.get("link") or row.get("raw_link") or "" for row in rows if row.get("link") or row.get("raw_link")))
    base["followers"] = _join_metric_values(rows, "followers")
    base["avv"] = _join_metric_values(rows, "avv")
    base["acv"] = _join_metric_values(rows, "acv")
    base["status"] = "成功" if len(successful) == len(rows) else ("部分成功" if successful else "失败")
    base["error"] = "；".join(
        f"{platform_label(row.get('platform') or '')}: {row.get('error')}"
        for row in rows
        if row.get("error")
    )
    return base


def _join_metric_values(rows: list[dict[str, Any]], key: str) -> str | int | None:
    parts = []
    for row in rows:
        value = row.get(key)
        if value is None or value == "":
            continue
        parts.append(f"{platform_label(row.get('platform') or '')}: {value}")
    if not parts:
        return None
    if len(parts) == 1:
        raw = parts[0].split(": ", 1)[-1]
        try:
            return int(raw)
        except (TypeError, ValueError):
            return raw
    return "\n".join(parts)


def _job_summary_json(**kwargs: Any) -> str:
    payload = {"steps": JOB_STEPS}
    payload.update(kwargs)
    return json.dumps(payload, ensure_ascii=False, default=str)


def _fail_job(db: Session, job_id: int, error: str) -> None:
    job = db.get(DataRefreshJob, job_id)
    if job:
        job.status = "failed"
        job.error = error[:1000]
        job.updated_at = datetime.utcnow()
        db.commit()
