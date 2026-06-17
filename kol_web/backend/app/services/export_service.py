from __future__ import annotations

import json
import tempfile
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.core.display_format import audience_value, format_all_prices_cell, unified_platform_row
from app.core.pipeline import clean_text, normalize_link, parse_platform_blocks
from app.models import KOLRecord, SheetTemplate


def export_records_restored(db: Session, records: list[KOLRecord]) -> Path:
    templates = {t.category: t for t in db.query(SheetTemplate).all()}
    template_path = _first_template_path(templates)
    style_cache: dict[str, list[Any]] = {}
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        style_cache = _capture_data_row_styles(wb)
        _clear_template_data_rows(wb)
    else:
        wb = Workbook()
        wb.active.title = _safe_sheet_title("KOL 列表 List")

    by_category: dict[str, list[KOLRecord]] = {}
    for record in records:
        by_category.setdefault(record.category or "Uncategorized", []).append(record)

    used_categories: set[str] = set()
    for category, rows in by_category.items():
        if category in templates and category in wb.sheetnames:
            metadata = json.loads(templates[category].metadata_json)
            _append_hok_rows(wb[category], metadata, rows, style_cache.get(category))
            used_categories.add(category)

    flat_rows = [r for r in records if (r.category or "Uncategorized") not in used_categories]
    if flat_rows:
        _append_flat_sheet(wb, flat_rows)

    out_dir = Path(tempfile.gettempdir()) / "kol-web-exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "kol_export_restored.xlsx"
    wb.save(out_path)
    return out_path


def export_source_workbook_updated(source_file: str, records: list[KOLRecord]) -> Path:
    source_path = Path(__file__).resolve().parents[2] / "data" / "uploads" / source_file
    if not source_path.exists():
        raise FileNotFoundError(f"Source workbook not found: {source_file}")
    wb = load_workbook(source_path)
    record_index = _build_record_index(records)
    for ws in wb.worksheets:
        if ws.max_row < 1:
            continue
        row0 = [_cell_value(ws, 1, col) for col in range(1, ws.max_column + 1)]
        row1 = [_cell_value(ws, 2, col) for col in range(1, ws.max_column + 1)] if ws.max_row >= 2 else []
        blocks = parse_platform_blocks(row0, row1) if row1 else {}
        if blocks:
            _update_hok_sheet_metrics(ws, blocks, record_index)
        else:
            _update_flat_sheet_metrics(ws, record_index)
        _trim_far_empty_columns(ws)
    out_dir = Path(tempfile.gettempdir()) / "kol-web-exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"{Path(source_file).stem}_metrics_updated.xlsx"
    wb.save(out_path)
    return out_path


def _build_record_index(records: list[KOLRecord]) -> dict[tuple[str, str], KOLRecord]:
    index: dict[tuple[str, str], KOLRecord] = {}
    for record in records:
        for platform, link in (
            ("tiktok", record.tt_link),
            ("ins", record.ins_link),
            ("youtube", record.yt_link),
        ):
            normalized = normalize_link(platform, link)
            if normalized:
                index[(platform, normalized)] = record
    return index


def _update_hok_sheet_metrics(
    ws: Worksheet, blocks: dict[str, dict[str, Any]], record_index: dict[tuple[str, str], KOLRecord]
) -> None:
    for row in range(3, ws.max_row + 1):
        for platform, info in blocks.items():
            cols = info.get("cols", {})
            link_col = cols.get("link")
            if link_col is None:
                continue
            link = _cell_value(ws, row, link_col + 1)
            record = record_index.get((platform, normalize_link(platform, link)))
            if not record:
                continue
            _write_metric_cells(ws, row, cols, platform, record)


def _update_flat_sheet_metrics(ws: Worksheet, record_index: dict[tuple[str, str], KOLRecord]) -> None:
    max_scan_col = min(ws.max_column, 80)
    header_map: dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        values = [_cell_value(ws, row, col) for col in range(1, max_scan_col + 1)]
        if _looks_like_flat_header(values):
            headers = [clean_text(value) for value in values]
            header_map = {header.lower(): idx + 1 for idx, header in enumerate(headers) if header}
            continue
        link_col = _first_col(header_map, "链接", "➡️ link", "link", "url")
        if not link_col and "creator" in header_map and "channel" in header_map:
            link_col = header_map["channel"]
        if not link_col:
            continue
        link = _cell_value(ws, row, link_col)
        platform_col = _first_col(header_map, "平台", "➡️ 平台", "platform")
        platform = _detect_platform(clean_text(_cell_value(ws, row, platform_col or 0)), link)
        record = record_index.get((platform, normalize_link(platform, link)))
        if not record:
            continue
        follower_col = _first_col(header_map, "粉丝数\n(自然数)", "➡️ followers", "followers")
        avv_col = _first_col(
            header_map,
            "均观看量\n(自然数)",
            "近一个月均观看量\n(自然数)",
            "均观看量\n(自然数/ccv)",
            "均观看量\n(视频)",
            "均观看量\n(自然数)/ccv",
            "➡️ avg view",
            "avg view",
            "avg view count (3m)",
            "expected\n30dviews/ccv",
            "expected\n30d views/ccv",
            "expected \n30d views/ccv",
        )
        _set_flat_metrics(ws, row, platform, record, follower_col, avv_col)


def _write_metric_cells(ws: Worksheet, row: int, cols: dict, platform: str, record: KOLRecord) -> None:
    follower_col = cols.get("follower")
    avv_col = cols.get("avv")
    if platform == "tiktok":
        _set_cell(ws, row, follower_col, record.tt_follower, zero_based=True)
        _set_cell(ws, row, avv_col, record.tt_avv, zero_based=True)
    elif platform == "ins":
        _set_cell(ws, row, follower_col, record.ins_follower, zero_based=True)
    elif platform == "youtube":
        _set_cell(ws, row, follower_col, record.yt_follower, zero_based=True)
        _set_cell(ws, row, avv_col, record.yt_avv, zero_based=True)


def _set_flat_metrics(
    ws: Worksheet, row: int, platform: str, record: KOLRecord, follower_col: int | None, avv_col: int | None
) -> None:
    if platform == "tiktok":
        _set_cell(ws, row, follower_col, record.tt_follower)
        _set_cell(ws, row, avv_col, record.tt_avv)
    elif platform == "ins":
        _set_cell(ws, row, follower_col, record.ins_follower)
    elif platform == "youtube":
        _set_cell(ws, row, follower_col, record.yt_follower)
        _set_cell(ws, row, avv_col, record.yt_avv)


def _set_cell(ws: Worksheet, row: int, col: int | None, value: Any, zero_based: bool = False) -> None:
    if col is None or value is None:
        return
    ws.cell(row=row, column=col + 1 if zero_based else col, value=value)


def _first_col(header_map: dict[str, int], *names: str) -> int | None:
    for name in names:
        col = header_map.get(name.lower())
        if col:
            return col
    return None


def _detect_platform(platform_text: str, link: Any) -> str:
    platform = platform_text.lower()
    text = f"{platform} {clean_text(link)}".lower()
    tokens = {token for token in platform.replace("/", " ").replace(",", " ").replace("&", " ").split() if token}
    if "tiktok" in text or "tik tok" in text or "tt" in tokens:
        return "tiktok"
    if "instagram" in text or "ins" in tokens or "ig" in tokens:
        return "ins"
    if "youtube" in text or "youtu.be" in text or "ytb" in tokens or "yt" in tokens:
        return "youtube"
    return ""


def _looks_like_flat_header(values: list[Any]) -> bool:
    labels = {clean_text(value).lower() for value in values if clean_text(value)}
    has_link = bool(labels & {"链接", "➡️ link", "link", "url", "channel link", "channel"})
    has_name = bool(labels & {"渠道名", "资源名称", "creator", "kol name", "name"})
    has_metric = bool(labels & {"粉丝数\n(自然数)", "followers", "➡️ followers"})
    return has_link and (has_name or has_metric)


def _cell_value(ws: Worksheet, row: int, col: int) -> Any:
    if col < 1 or col > ws.max_column:
        return None
    return ws.cell(row=row, column=col).value


def _value_at(values: list[Any], idx: int) -> Any:
    return values[idx] if idx < len(values) else None


def _trim_far_empty_columns(ws: Worksheet, keep_at_least: int = 80) -> None:
    if ws.max_column <= keep_at_least:
        return
    for key, cell in list(ws._cells.items()):
        if key[1] > keep_at_least and cell.value is None:
            del ws._cells[key]


def _first_template_path(templates: dict[str, SheetTemplate]) -> Path | None:
    for template in templates.values():
        if template.source_file:
            return Path(template.source_file)
    return None


def _clear_template_data_rows(wb: Workbook) -> None:
    for ws in wb.worksheets:
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)


def _capture_data_row_styles(wb: Workbook) -> dict[str, list[Any]]:
    cache: dict[str, list[Any]] = {}
    for ws in wb.worksheets:
        if ws.max_row >= 3:
            cache[ws.title] = [copy(ws.cell(row=3, column=col)._style) for col in range(1, ws.max_column + 1)]
    return cache


def _append_hok_rows(
    ws: Worksheet, metadata: dict, records: list[KOLRecord], row_style: list[Any] | None
) -> None:
    max_cols = int(metadata.get("max_cols") or ws.max_column)
    platform_blocks = metadata.get("platform_blocks", {})
    start_row = 3
    for idx, record in enumerate(records):
        row_num = start_row + idx
        ws.cell(row=row_num, column=1, value=record.name)
        ws.cell(row=row_num, column=2, value=record.category)
        for platform, info in platform_blocks.items():
            cols = info.get("cols", {})
            _write_platform(ws, row_num, cols, platform, record)
        for col in range(1, max_cols + 1):
            if row_style and col <= len(row_style):
                ws.cell(row=row_num, column=col)._style = copy(row_style[col - 1])


def _write_platform(ws: Worksheet, row: int, cols: dict, platform: str, record: KOLRecord) -> None:
    def set_col(key: str, value):
        col = cols.get(key)
        if col is not None:
            ws.cell(row=row, column=col + 1, value=value if value is not None else "/")

    if platform == "tiktok":
        set_col("link", record.tt_link)
        set_col("follower", record.tt_follower)
        set_col("avv", record.tt_avv)
        set_col("short video", record.tt_short_video_price)
        set_col("anchor link", record.tt_anchor_link_price)
    elif platform == "ins":
        set_col("link", record.ins_link)
        set_col("follower", record.ins_follower)
        set_col("post", record.ins_post_price)
        set_col("ig reels", record.ins_reels_price)
    elif platform == "youtube":
        set_col("link", record.yt_link)
        set_col("follower", record.yt_follower)
        set_col("avv", record.yt_avv)
        set_col("full-video", record.yt_full_video_price)
        set_col("live 2hr", record.yt_live_2hr_price)
        set_col("pre-roll", record.yt_pre_roll_price)
        set_col("short video", record.yt_short_video_price)


_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
_NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
_EVEN_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_INVALID_SHEET_CHARS = set("\\/[]?*")


def _safe_sheet_title(title: str, max_len: int = 31) -> str:
    cleaned = "".join(" " if ch in _INVALID_SHEET_CHARS else ch for ch in title).strip()
    cleaned = " ".join(cleaned.split())
    return (cleaned[:max_len] if cleaned else "Sheet").strip() or "Sheet"


def _append_flat_sheet(wb: Workbook, records: list[KOLRecord]) -> None:
    title = _safe_sheet_title("KOL 数据 Data")
    if title in wb.sheetnames:
        ws = wb[title]
    elif len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active["A1"].value is None:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    from app.core.category import major_category

    headers = [
        ("KOL 名称/Name", 20),
        ("大类/Category", 12),
        ("国家/Country", 10),
        ("语言/Language", 10),
        ("平台/Platform", 12),
        ("链接/Link", 30),
        ("粉丝/Followers", 14),
        ("AVV/均观看量", 14),
        ("CPM", 12),
        ("合作模式/Collaboration", 18),
        ("报价汇总/All Prices", 28),
        ("受众地区/Audience Region", 24),
        ("受众性别/Audience Gender", 20),
        ("受众年龄/Audience Age", 24),
        ("备注/Notes", 24),
    ]
    header_labels = [h[0] for h in headers]
    col_widths = [h[1] for h in headers]

    ws.append(header_labels)
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
    ws.row_dimensions[1].height = 28

    for record_idx, record in enumerate(records):
        platform_row = unified_platform_row(record)
        row_num = record_idx + 2
        row_data = [
            record.name,
            major_category(record.normalized_category or ""),
            record.country,
            record.language,
            platform_row["platform"],
            platform_row["link"],
            platform_row["follower"],
            platform_row["avv"],
            platform_row["cpm"],
            platform_row["collaboration"],
            format_all_prices_cell(record),
            audience_value(record, "audience_region", "受众地区"),
            audience_value(record, "audience_gender", "受众性别"),
            audience_value(record, "audience_age", "受众年龄"),
            record.notes,
        ]
        ws.append(row_data)

        is_even = record_idx % 2 == 1
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = _THIN_BORDER
            if is_even:
                cell.fill = _EVEN_ROW_FILL
            value = cell.value
            if isinstance(value, (int, float)):
                cell.alignment = _NUMBER_ALIGNMENT
                cell.number_format = "#,##0" if isinstance(value, int) else "#,##0.00"
            elif isinstance(value, str) and value.startswith("http"):
                cell.alignment = _DATA_ALIGNMENT
                try:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                except Exception:
                    pass
            else:
                cell.alignment = _DATA_ALIGNMENT
        ws.row_dimensions[row_num].height = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
