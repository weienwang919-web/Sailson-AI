from __future__ import annotations

import json
import os
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.models import OfficialAccount, OfficialMonitorJob, OfficialProfileDailyMetric, OfficialVideoSnapshot
from app.services.tiktok_business_service import TikTokBusinessClient, TikTokBusinessError


def load_configured_accounts(db: Session) -> list[OfficialAccount]:
    raw = os.getenv("TIKTOK_OFFICIAL_ACCOUNTS_JSON", "").strip()
    if raw:
        try:
            configs = json.loads(raw)
        except json.JSONDecodeError as exc:
            raise RuntimeError("TIKTOK_OFFICIAL_ACCOUNTS_JSON is not valid JSON") from exc
        for config in configs if isinstance(configs, list) else []:
            business_id = str(config.get("business_id") or config.get("open_id") or "").strip()
            if not business_id:
                continue
            account = db.query(OfficialAccount).filter(OfficialAccount.business_id == business_id).one_or_none()
            if not account:
                account = OfficialAccount(business_id=business_id)
                db.add(account)
            account.username = config.get("username") or account.username
            account.display_name = config.get("display_name") or account.display_name
            account.profile_deep_link = config.get("profile_deep_link") or account.profile_deep_link
            account.notes = config.get("notes") or account.notes
            account.enabled = bool(config.get("enabled", True))
        db.commit()
    accounts = db.query(OfficialAccount).filter(OfficialAccount.enabled.is_(True)).order_by(OfficialAccount.id).all()
    if not accounts:
        account = OfficialAccount(
            business_id="mock_open_id",
            username="sailson_official",
            display_name="Sailson Official Mock",
            profile_deep_link="https://www.tiktok.com/@sailson_official",
            notes="Mock account. Set TIKTOK_OFFICIAL_ACCOUNTS_JSON and TIKTOK_BUSINESS_ACCESS_TOKEN for real data.",
        )
        db.add(account)
        db.commit()
        db.refresh(account)
        accounts = [account]
    return accounts


def create_official_job(db: Session, total: int) -> OfficialMonitorJob:
    job = OfficialMonitorJob(status="pending", total=total, done=0)
    db.add(job)
    db.commit()
    db.refresh(job)
    return job


def refresh_official_accounts(db: Session, account_ids: list[int] | None = None, days: int = 30) -> OfficialMonitorJob:
    accounts = load_configured_accounts(db)
    if account_ids:
        id_set = set(account_ids)
        accounts = [account for account in accounts if account.id in id_set]
    job = create_official_job(db, len(accounts))
    client = TikTokBusinessClient()
    job.status = "running"
    db.commit()
    try:
        total_videos = 0
        for account in accounts:
            before = db.query(OfficialVideoSnapshot).filter(OfficialVideoSnapshot.account_id == account.id).count()
            _refresh_account(db, client, account, days)
            after = db.query(OfficialVideoSnapshot).filter(OfficialVideoSnapshot.account_id == account.id).count()
            total_videos += max(0, after - before)
            job.done += 1
            job.updated_at = datetime.utcnow()
            db.commit()
        job.status = "completed"
        job.item_count = total_videos
        job.crawler_items = 0
        job.api_calls = len(accounts) * 2
        job.crawler_cost_usd = 0
        job.crawler_cost_cny = 0
        job.total_cost_cny = 0
        job.usage_detail_json = json.dumps({"pricing": "Official TikTok Business API, crawler cost not applied", "days": days}, ensure_ascii=False)
    except TikTokBusinessError as exc:
        job.status = "failed"
        job.error = str(exc)
        job.request_id = exc.request_id
        job.log_id = exc.log_id
    except Exception as exc:  # noqa: BLE001 - returned to dashboard for operator action.
        job.status = "failed"
        job.error = str(exc)
    finally:
        job.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(job)
    return job


def _refresh_account(db: Session, client: TikTokBusinessClient, account: OfficialAccount, days: int) -> None:
    end = date.today() - timedelta(days=1)
    start = end - timedelta(days=max(1, min(days, 60)) - 1)
    profile_response = client.get_profile(account.business_id, start, end)
    profile = profile_response.data.get("data") or {}
    _upsert_profile(db, account, profile)
    _upsert_profile_daily_metrics(db, account, profile, profile_response.request_id, profile_response.log_id)

    video_response = client.list_videos(account.business_id)
    videos = video_response.data.get("videos") or []
    for video in videos:
        _upsert_video(db, account, video, video_response.request_id, video_response.log_id)
    account.last_refreshed_at = datetime.utcnow()
    db.commit()


def _upsert_profile(db: Session, account: OfficialAccount, profile: dict[str, Any]) -> None:
    for field in (
        "username",
        "display_name",
        "profile_image",
        "profile_deep_link",
        "bio_description",
        "is_business_account",
        "is_verified",
        "following_count",
        "followers_count",
        "total_likes",
        "videos_count",
    ):
        if field in profile:
            setattr(account, field, profile.get(field))
    account.raw_json = json.dumps(profile, ensure_ascii=False, default=str)


def _upsert_video(db: Session, account: OfficialAccount, video: dict[str, Any], request_id: str | None, log_id: str | None) -> None:
    item_id = str(video.get("item_id") or "").strip()
    if not item_id:
        return
    snapshot = (
        db.query(OfficialVideoSnapshot)
        .filter(OfficialVideoSnapshot.account_id == account.id, OfficialVideoSnapshot.item_id == item_id)
        .one_or_none()
    )
    if not snapshot:
        snapshot = OfficialVideoSnapshot(account_id=account.id, business_id=account.business_id, item_id=item_id)
        db.add(snapshot)
    scalar_fields = [
        "media_type",
        "is_ad",
        "thumbnail_url",
        "share_url",
        "embed_url",
        "caption",
        "video_duration",
        "reach",
        "video_views",
        "likes",
        "comments",
        "shares",
        "favorites",
        "total_time_watched",
        "average_time_watched",
        "full_video_watched_rate",
        "new_followers",
        "profile_views",
        "website_clicks",
        "phone_number_clicks",
        "lead_submissions",
        "app_download_clicks",
        "email_clicks",
        "address_clicks",
    ]
    for field in scalar_fields:
        if field in video:
            setattr(snapshot, field, video.get(field))
    snapshot.create_time = _epoch_to_datetime(video.get("create_time"))
    for field in (
        "engagement_likes",
        "video_view_retention",
        "impression_sources",
        "audience_genders",
        "audience_countries",
        "audience_cities",
        "audience_types",
    ):
        if field in video:
            setattr(snapshot, f"{field}_json", json.dumps(video.get(field), ensure_ascii=False, default=str))
    snapshot.request_id = request_id
    snapshot.log_id = log_id
    snapshot.raw_json = json.dumps(video, ensure_ascii=False, default=str)
    snapshot.fetched_at = datetime.utcnow()


def _upsert_profile_daily_metrics(
    db: Session, account: OfficialAccount, profile: dict[str, Any], request_id: str | None, log_id: str | None
) -> None:
    daily_metrics = _daily_rows(profile)
    for row in daily_metrics:
        metric_date = _parse_date(row.get("date") or row.get("stat_time") or row.get("metric_date"))
        if not metric_date:
            continue
        metric = (
            db.query(OfficialProfileDailyMetric)
            .filter(OfficialProfileDailyMetric.account_id == account.id, OfficialProfileDailyMetric.metric_date == metric_date)
            .one_or_none()
        )
        if not metric:
            metric = OfficialProfileDailyMetric(
                account_id=account.id,
                business_id=account.business_id,
                metric_date=metric_date,
            )
            db.add(metric)
        for field in (
            "followers_count",
            "video_views",
            "unique_video_views",
            "profile_views",
            "likes",
            "comments",
            "shares",
            "daily_total_followers",
            "daily_new_followers",
            "daily_lost_followers",
            "engaged_audience",
        ):
            if field in row:
                setattr(metric, field, row.get(field))
        for field in ("audience_activity", "audience_ages", "audience_genders", "audience_countries", "audience_cities"):
            if field in profile:
                setattr(metric, f"{field}_json", json.dumps(profile.get(field), ensure_ascii=False, default=str))
        metric.raw_json = json.dumps({"daily": row, "profile_request_id": request_id, "profile_log_id": log_id}, ensure_ascii=False, default=str)


def _daily_rows(profile: dict[str, Any]) -> list[dict[str, Any]]:
    if isinstance(profile.get("daily_metrics"), list):
        return profile["daily_metrics"]
    dates: set[str] = set()
    series_fields = [
        "video_views",
        "unique_video_views",
        "profile_views",
        "likes",
        "comments",
        "shares",
        "daily_total_followers",
        "daily_new_followers",
        "daily_lost_followers",
        "engaged_audience",
    ]
    buckets: dict[str, dict[str, Any]] = {}
    for field in series_fields:
        value = profile.get(field)
        if not isinstance(value, list):
            continue
        for entry in value:
            if not isinstance(entry, dict):
                continue
            day = entry.get("date") or entry.get("stat_time") or entry.get("metric_date")
            if not day:
                continue
            dates.add(day)
            row = buckets.setdefault(day, {"date": day})
            row[field] = entry.get("value") if "value" in entry else entry.get(field)
    return [buckets[day] for day in sorted(dates)]


def list_videos(db: Session, page: int, page_size: int, account_id: int | None = None) -> tuple[int, list[OfficialVideoSnapshot]]:
    query = db.query(OfficialVideoSnapshot)
    if account_id:
        query = query.filter(OfficialVideoSnapshot.account_id == account_id)
    total = query.count()
    rows = (
        query.order_by(desc(OfficialVideoSnapshot.create_time), desc(OfficialVideoSnapshot.fetched_at))
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return total, rows


def list_profile_metrics(db: Session, account_id: int | None = None) -> list[OfficialProfileDailyMetric]:
    query = db.query(OfficialProfileDailyMetric)
    if account_id:
        query = query.filter(OfficialProfileDailyMetric.account_id == account_id)
    return query.order_by(OfficialProfileDailyMetric.metric_date).all()


def export_official_videos(db: Session) -> Path:
    accounts = {account.id: account for account in db.query(OfficialAccount).all()}
    videos = db.query(OfficialVideoSnapshot).order_by(OfficialVideoSnapshot.create_time.desc()).all()
    metrics = db.query(OfficialProfileDailyMetric).order_by(OfficialProfileDailyMetric.metric_date).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Videos"
    headers = [
        "账号",
        "caption",
        "发布时间",
        "视频链接",
        "视频观看次数",
        "平均观看时长(s)",
        "点赞数",
        "评论数",
        "分享次数",
        "收藏次数",
        "单条带来新增粉",
        "视频时长",
        "3秒留存率",
        "完播率",
        "是否广告视频",
        "媒体类型",
    ]
    ws.append(headers)
    for video in videos:
        account = accounts.get(video.account_id)
        retention = _json_list(video.video_view_retention_json)
        ws.append(
            [
                account.display_name or account.username or video.business_id if account else video.business_id,
                video.caption,
                video.create_time,
                video.share_url,
                video.video_views,
                video.average_time_watched,
                video.likes,
                video.comments,
                video.shares,
                video.favorites,
                video.new_followers,
                video.video_duration,
                _retention_at_second(retention, "3"),
                video.full_video_watched_rate,
                video.is_ad,
                video.media_type,
            ]
        )
    _append_distribution_sheet(wb, "Engagement Likes", videos, "engagement_likes_json", ["item_id", "second", "percentage"])
    _append_distribution_sheet(wb, "Retention", videos, "video_view_retention_json", ["item_id", "second", "percentage"])
    _append_distribution_sheet(
        wb, "Impression Sources", videos, "impression_sources_json", ["item_id", "impression_source", "percentage"]
    )
    _append_distribution_sheet(
        wb, "Video Audience Countries", videos, "audience_countries_json", ["item_id", "country", "percentage"]
    )
    profile_ws = wb.create_sheet("Profile Daily Metrics")
    profile_ws.append(
        [
            "business_id",
            "date",
            "followers_count",
            "video_views",
            "unique_video_views",
            "profile_views",
            "likes",
            "comments",
            "shares",
            "daily_new_followers",
            "daily_lost_followers",
            "engaged_audience",
        ]
    )
    for metric in metrics:
        profile_ws.append(
            [
                metric.business_id,
                metric.metric_date,
                metric.followers_count,
                metric.video_views,
                metric.unique_video_views,
                metric.profile_views,
                metric.likes,
                metric.comments,
                metric.shares,
                metric.daily_new_followers,
                metric.daily_lost_followers,
                metric.engaged_audience,
            ]
        )
    out_dir = Path(tempfile.gettempdir()) / "kol-web-exports"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "official_tiktok_monitor_export.xlsx"
    wb.save(out_path)
    return out_path


def _append_distribution_sheet(
    wb: Workbook, title: str, videos: list[OfficialVideoSnapshot], json_attr: str, headers: list[str]
) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for video in videos:
        for entry in _json_list(getattr(video, json_attr)):
            if not isinstance(entry, dict):
                continue
            ws.append([video.item_id] + [entry.get(header) for header in headers[1:]])


def _json_list(raw: str | None) -> list[Any]:
    if not raw:
        return []
    try:
        value = json.loads(raw)
    except json.JSONDecodeError:
        return []
    return value if isinstance(value, list) else []


def _retention_at_second(rows: list[Any], second: str) -> Any:
    for row in rows:
        if isinstance(row, dict) and str(row.get("second")) == second:
            return row.get("percentage")
    return None


def _epoch_to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    try:
        return datetime.utcfromtimestamp(int(float(value)))
    except (TypeError, ValueError, OSError):
        return None


def _parse_date(value: Any) -> date | None:
    if not value:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None
